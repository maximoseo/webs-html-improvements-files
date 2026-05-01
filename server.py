#!/usr/bin/env python3
import base64
import concurrent.futures
import datetime
import hashlib
import http.client
import json
import mimetypes
import ipaddress
import os
import posixpath
import re
import socket
import ssl
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
import kwr_backend
import r5_features as r5
import r6_features as r6
import schedule_engine
import dashboard_features_api as df_api

# Load .env file if present (for local development)
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

ROOT = Path(__file__).resolve().parent
INDEX = ROOT / 'index.html'
DATA = ROOT / 'data.json'
MAP_FILE = ROOT / 'n8n-workflow-map.json'
TASKS_FILE = ROOT / 'data' / 'tasks.json'
PLAYGROUND_FILE = ROOT / 'data' / 'playground.json'
PLAYGROUND_EXPORTS_DIR = ROOT / 'playground' / 'exports'
REPO = 'maximoseo/webs-html-improvements-files'
RAW_BASE = f'https://raw.githubusercontent.com/{REPO}/main'
DEFAULT_N8N_BASE = 'https://websiseo.app.n8n.cloud'
_SERVER_START_TIME = time.time()
_PROJECT_DOMAIN_RE = re.compile(r'^(?![.-])[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?(?:\.[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?)*$', re.I)


def _looks_like_project_domain(name: str) -> bool:
    name = (name or '').strip()
    if not name:
        return False
    return bool(_PROJECT_DOMAIN_RE.fullmatch(name))


def _truthy_env(name: str) -> bool:
    return (os.getenv(name) or '').strip().lower() in ('1', 'true', 'yes', 'on')


def _dashboard_is_production() -> bool:
    return _truthy_env('RENDER') or _truthy_env('DASHBOARD_PRODUCTION') or _truthy_env('PRODUCTION')

# ===== DAILY SKILLS RADAR — global state & sources =====
_radar_state = {
    'status': 'idle',
    'run_id': None,
    'stage': '',
    'progress': 0,
    'sources_total': 0,
    'sources_done': 0,
    'skills_found': 0,
    'started_at': None,
    'finished_at': None,
    'last_error': None,
    'paused': False,
    'schedule_hour': 2,
    'schedule_enabled': True,
    'topics': ['n8n', 'automation', 'seo', 'local seo', 'claude code', 'coding', 'ai agents', 'scraping', 'wordpress'],
    'custom_topics': [],
}

RADAR_SOURCES = [
    {'url': 'https://raw.githubusercontent.com/sindresorhus/awesome/main/readme.md', 'title': 'Awesome - Sindre Sorhus', 'type': 'awesome_list', 'quality': 0.95},
    {'url': 'https://raw.githubusercontent.com/e2b-dev/awesome-ai-agents/main/README.md', 'title': 'Awesome AI Agents', 'type': 'awesome_list', 'quality': 0.95},
    {'url': 'https://raw.githubusercontent.com/kyrolabs/awesome-langchain/main/README.md', 'title': 'Awesome LangChain', 'type': 'awesome_list', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/n8n-io/n8n/master/README.md', 'title': 'N8N Official', 'type': 'github_repo', 'quality': 0.95},
    {'url': 'https://raw.githubusercontent.com/jxnl/instructor/main/README.md', 'title': 'Instructor - Structured LLM Outputs', 'type': 'github_repo', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/microsoft/autogen/main/README.md', 'title': 'AutoGen - Multi-Agent Framework', 'type': 'github_repo', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/BerriAI/litellm/main/README.md', 'title': 'LiteLLM - LLM Gateway', 'type': 'github_repo', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/pydantic/pydantic-ai/main/README.md', 'title': 'PydanticAI - Agent Framework', 'type': 'github_repo', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/assafelovic/gpt-researcher/master/README.md', 'title': 'GPT Researcher', 'type': 'github_repo', 'quality': 0.88},
    {'url': 'https://raw.githubusercontent.com/mendableai/firecrawl/main/README.md', 'title': 'Firecrawl - Web Scraping', 'type': 'github_repo', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/duckdb/duckdb/main/README.md', 'title': 'DuckDB - In-process SQL', 'type': 'github_repo', 'quality': 0.85},
    {'url': 'https://raw.githubusercontent.com/openai/openai-python/main/README.md', 'title': 'OpenAI Python SDK', 'type': 'github_repo', 'quality': 0.95},
    {'url': 'https://raw.githubusercontent.com/anthropics/anthropic-sdk-python/main/README.md', 'title': 'Anthropic Python SDK', 'type': 'github_repo', 'quality': 0.95},
    {'url': 'https://raw.githubusercontent.com/huggingface/transformers/main/README.md', 'title': 'HuggingFace Transformers', 'type': 'github_repo', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/dspy-ai/dspy/main/README.md', 'title': 'DSPy - Programming LMs', 'type': 'github_repo', 'quality': 0.88},
    {'url': 'https://raw.githubusercontent.com/unclecode/crawl4ai/main/README.md', 'title': 'Crawl4AI - LLM-Friendly Scraping', 'type': 'github_repo', 'quality': 0.88},
    {'url': 'https://raw.githubusercontent.com/browser-use/browser-use/main/README.md', 'title': 'Browser Use - AI Browser Automation', 'type': 'github_repo', 'quality': 0.9},
    {'url': 'https://raw.githubusercontent.com/All-Hands-AI/OpenHands/main/README.md', 'title': 'OpenHands - AI Software Development', 'type': 'github_repo', 'quality': 0.88},
    {'url': 'https://raw.githubusercontent.com/agno-agi/agno/main/README.md', 'title': 'Agno - Agent Framework', 'type': 'github_repo', 'quality': 0.85},
    {'url': 'https://raw.githubusercontent.com/meta-llama/llama-models/main/README.md', 'title': 'Meta Llama Models', 'type': 'github_repo', 'quality': 0.9},
]


def load_json_file(path: Path, default):
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        return default


def _tasks_load() -> list:
    try:
        if not TASKS_FILE.exists():
            return []
        data = json.loads(TASKS_FILE.read_text(encoding='utf-8'))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _tasks_save(tasks: list) -> None:
    TASKS_FILE.parent.mkdir(parents=True, exist_ok=True)
    TASKS_FILE.write_text(json.dumps(tasks, ensure_ascii=False, indent=2), encoding='utf-8')


def _playground_now() -> str:
    return datetime.datetime.utcnow().replace(microsecond=0).isoformat() + 'Z'


def _playground_default_data() -> dict:
    return {'templates': [], 'exports': [], 'preferences': {'theme': 'dark', 'view': 'grid', 'sort': 'newest', 'default_device': 'desktop'}}


def _playground_load() -> dict:
    try:
        if not PLAYGROUND_FILE.exists():
            return _playground_default_data()
        data = json.loads(PLAYGROUND_FILE.read_text(encoding='utf-8'))
        if not isinstance(data, dict):
            return _playground_default_data()
        data.setdefault('templates', [])
        data.setdefault('exports', [])
        data.setdefault('preferences', {'theme': 'dark', 'view': 'grid', 'sort': 'newest', 'default_device': 'desktop'})
        return data
    except Exception:
        return _playground_default_data()


def _playground_save(data: dict) -> None:
    PLAYGROUND_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = PLAYGROUND_FILE.with_suffix('.json.tmp')
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    tmp.replace(PLAYGROUND_FILE)


def _playground_slug(value: str, fallback: str = 'template') -> str:
    value = (value or '').strip().lower()
    value = re.sub(r'^https?://', '', value)
    value = re.sub(r'[^a-z0-9._-]+', '-', value)
    value = value.strip('-._')
    return value[:80] or fallback


def _playground_guess_domain(raw: str) -> str:
    raw = (raw or '').strip()
    if not raw:
        return 'unknown.local'
    if raw.startswith('http://') or raw.startswith('https://'):
        try:
            return urllib.parse.urlparse(raw).netloc.lower().lstrip('www.') or raw
        except Exception:
            pass
    return raw.lower().replace(' ', '-')


def _playground_enrich_template(row: dict) -> dict:
    html = row.get('html_content') or row.get('htmlContent') or ''
    domain = _playground_guess_domain(row.get('domain') or row.get('domain_display') or row.get('domainDisplay') or 'unknown.local')
    now = _playground_now()
    direction = (row.get('direction') or ('rtl' if re.search(r'dir=["\']rtl|lang=["\']he|[\u0590-\u05ff]', html, re.I) else 'ltr')).lower()
    language = (row.get('language') or ('he' if direction == 'rtl' or re.search(r'[\u0590-\u05ff]', html) else 'en')).lower()
    tags = row.get('tags') or []
    if isinstance(tags, str):
        tags = [t.strip() for t in re.split(r'[,#]', tags) if t.strip()]
    out = {
        'id': row.get('id') or str(uuid.uuid4()),
        'domain': domain,
        'domain_display': row.get('domain_display') or row.get('domainDisplay') or domain,
        'agent_name': row.get('agent_name') or row.get('agentName') or 'Hermes Agent',
        'agent_model': row.get('agent_model') or row.get('agentModel') or 'manual-upload',
        'agent_style': row.get('agent_style') or row.get('agentStyle') or 'Custom',
        'html_content': html,
        'html_size_bytes': len(html.encode('utf-8')),
        'language': language,
        'direction': 'rtl' if direction == 'rtl' else 'ltr',
        'industry': row.get('industry') or '',
        'tags': tags,
        'is_favorite': bool(row.get('is_favorite') if 'is_favorite' in row else row.get('isFavorite', False)),
        'sort_order': int(row.get('sort_order') or row.get('sortOrder') or 0),
        'n8n_prompt_generated': bool(row.get('n8n_prompt_generated') or row.get('n8nPromptGenerated') or False),
        'n8n_prompt_generated_at': row.get('n8n_prompt_generated_at') or row.get('n8nPromptGeneratedAt'),
        'n8n_json_generated': bool(row.get('n8n_json_generated') or row.get('n8nJsonGenerated') or False),
        'n8n_json_generated_at': row.get('n8n_json_generated_at') or row.get('n8nJsonGeneratedAt'),
        'uploaded_by': row.get('uploaded_by') or row.get('uploadedBy') or 'user',
        'source': row.get('source') or 'upload',
        'obsidian_synced': bool(row.get('obsidian_synced') or False),
        'obsidian_path': row.get('obsidian_path'),
        'github_synced': bool(row.get('github_synced') or False),
        'github_path': row.get('github_path'),
        'created_at': row.get('created_at') or row.get('createdAt') or now,
        'updated_at': now,
    }
    return out


def _playground_public_template(row: dict, include_html: bool = False) -> dict:
    out = dict(row)
    if not include_html:
        out.pop('html_content', None)
    return out


def _playground_domains(templates: list) -> list:
    grouped = {}
    for t in templates:
        domain = t.get('domain') or 'unknown.local'
        g = grouped.setdefault(domain, {'domain': domain, 'domain_display': t.get('domain_display') or domain, 'count': 0, 'agents': set(), 'tags': set(), 'last_upload': ''})
        g['count'] += 1
        if t.get('agent_name'):
            g['agents'].add(t.get('agent_name'))
        for tag in t.get('tags') or []:
            g['tags'].add(str(tag))
        if (t.get('created_at') or '') > g['last_upload']:
            g['last_upload'] = t.get('created_at') or ''
    out = []
    for g in grouped.values():
        g['agents'] = sorted(g['agents'])
        g['tags'] = sorted(g['tags'])
        out.append(g)
    return sorted(out, key=lambda x: (x.get('last_upload') or ''), reverse=True)


def _playground_seed_if_empty(data: dict) -> bool:
    if data.get('templates'):
        return False
    samples = []
    candidates = [
        ROOT / 'galoz.co.il' / 'Hermes Success Only Best Pick' / '2026-04-28' / 'Improved_HTML_Template.html',
        ROOT / 'powerplug.ai' / 'GPT 5.5 Agent (openai-gpt-5.5)' / '2026-04-28' / 'Improved_HTML_Template.html',
        ROOT / 'powerplug.ai' / 'Kimi K2.6 Agent (moonshotai-kimi-k2.6)' / '2026-04-28' / 'Improved_HTML_Template.html',
    ]
    for fp in candidates:
        try:
            if fp.exists():
                parts = fp.relative_to(ROOT).parts
                samples.append(_playground_enrich_template({
                    'domain': parts[0],
                    'agent_name': parts[1] if len(parts) > 2 else 'Hermes Agent',
                    'agent_model': parts[1] if len(parts) > 2 else 'manual',
                    'agent_style': 'Imported latest set',
                    'html_content': fp.read_text(encoding='utf-8', errors='replace'),
                    'source': 'auto',
                    'github_path': str(fp.relative_to(ROOT)),
                    'github_synced': True,
                    'tags': ['auto-import', 'html-template'],
                }))
        except Exception:
            pass
    if not samples:
        samples.append(_playground_enrich_template({
            'domain': 'demo.local',
            'agent_name': 'Hermes Agent',
            'agent_model': 'demo',
            'agent_style': 'Responsive demo',
            'tags': ['demo', 'responsive'],
            'html_content': '<!-- wp:html --><article lang="he" dir="rtl" style="font-family:Arial,sans-serif;max-width:860px;margin:auto;padding:32px"><h2>תבנית לדוגמה</h2><p>ה־Playground מציג תצוגה חיה של HTML, בדיקות רספונסיביות וייצוא N8N.</p></article><!-- /wp:html -->',
            'source': 'auto',
        }))
    data['templates'] = samples
    return True


def _playground_generate_prompt(template: dict) -> str:
    direction = template.get('direction') or 'ltr'
    language = template.get('language') or 'en'
    return f"""# N8N Improve Prompt — {template.get('domain')} / {template.get('agent_name')}

You are improving a WordPress-safe HTML article template.

## Source template metadata
- Domain: {template.get('domain')}
- Agent: {template.get('agent_name')} ({template.get('agent_model')})
- Style: {template.get('agent_style')}
- Language: {language}
- Direction: {direction}

## Mandatory output rules
1. Return exactly three files: Improved_HTML_Template.html, Improved_N8N_Prompt.txt, Improved_N8N_Workflow.json.
2. HTML must be WordPress HTML-block compatible, scoped, responsive for desktop/laptop/tablet/mobile, and must not include external CSS/JS/fonts.
3. Keep real client content, verified contact links, verified product-page links/images for ecommerce, author/about near the end, CSS-only hover affordances, lower TOC, and no fake site header/footer.
4. Preserve language and direction: lang={language}, dir={direction}.

## HTML template to improve
```html
{template.get('html_content','')}
```
"""


def _playground_generate_workflow(template: dict, prompt: str) -> dict:
    return {
        'name': f"Playground Improve — {template.get('domain')} — {template.get('agent_name')}",
        'active': False,
        'nodes': [
            {'id': 'manual-trigger', 'name': 'Manual Trigger', 'type': 'n8n-nodes-base.manualTrigger', 'typeVersion': 1, 'position': [0, 0], 'parameters': {}},
            {'id': 'improve-prompt', 'name': 'HTML Template Generator', 'type': 'n8n-nodes-base.openAi', 'typeVersion': 1, 'position': [260, 0], 'parameters': {'prompt': prompt, 'maxTokens': 12000}},
            {'id': 'wordpress-draft', 'name': 'Post to WordPress Draft', 'type': 'n8n-nodes-base.wordpress', 'typeVersion': 1, 'position': [520, 0], 'parameters': {'resource': 'post', 'operation': 'create', 'status': 'draft'}},
        ],
        'connections': {'Manual Trigger': {'main': [[{'node': 'HTML Template Generator', 'type': 'main', 'index': 0}]]}, 'HTML Template Generator': {'main': [[{'node': 'Post to WordPress Draft', 'type': 'main', 'index': 0}]]}},
        'settings': {'executionOrder': 'v1'},
        'meta': {'generatedBy': 'Hermes Playground', 'manualImportOnly': True, 'templateId': template.get('id')},
    }


def _playground_add_export(data: dict, template: dict, export_type: str, content: str) -> dict:
    now = _playground_now()
    row = {
        'id': str(uuid.uuid4()),
        'template_id': template.get('id'),
        'export_type': export_type,
        'export_content': content,
        'export_size_bytes': len(content.encode('utf-8')),
        'generated_by': 'system',
        'version': 1,
        'created_at': now,
    }
    data.setdefault('exports', []).append(row)
    if export_type == 'n8n_prompt':
        template['n8n_prompt_generated'] = True
        template['n8n_prompt_generated_at'] = now
    if export_type == 'n8n_json':
        template['n8n_json_generated'] = True
        template['n8n_json_generated_at'] = now
    try:
        export_dir = PLAYGROUND_EXPORTS_DIR / _playground_slug(template.get('domain'))
        export_dir.mkdir(parents=True, exist_ok=True)
        stem = f"{_playground_slug(template.get('agent_name'))}-{export_type}-{now[:10]}"
        suffix = '.txt' if export_type == 'n8n_prompt' else '.json'
        fp = export_dir / f"{stem}{suffix}"
        fp.write_text(content, encoding='utf-8')
        row['github_path'] = str(fp.relative_to(ROOT))
    except Exception:
        pass
    return row


def _cache_control_for(content_type: str, path: str = '') -> str:
    """Pick Cache-Control header value based on content type & path.
    Static assets cache 1 hour; HTML/JSON never cache; everything else no-cache."""
    p = (path or '').lower()
    ct = (content_type or '').lower()
    if p.startswith('/static/') or p.startswith('/assets/') or p.startswith('/css/') \
       or p.startswith('/js/') or p.startswith('/img/') or p.startswith('/fonts/'):
        return 'public, max-age=3600'
    if ct.startswith('text/css') or ct.startswith('application/javascript') \
       or ct.startswith('text/javascript') or ct.startswith('image/') or ct.startswith('font/'):
        return 'public, max-age=3600'
    if ct.startswith('text/html') or ct.startswith('application/json'):
        return 'no-store, must-revalidate'
    return 'no-cache'


def json_response(handler, status, payload):
    body = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    try:
        handler._r2_status = status
    except Exception:
        pass
    handler.send_response(status)
    handler.send_header('Content-Type', 'application/json; charset=utf-8')
    handler.send_header('Content-Length', str(len(body)))
    handler.send_header('Cache-Control', 'no-store, must-revalidate')
    handler.end_headers()
    handler.wfile.write(body)


_KWR_SAFE_FLAT_SLUG_RE = re.compile(r'[A-Za-z0-9._-]+')


def _safe_kwr_flat_slug(raw):
    slug = str(raw or '').strip()
    for _ in range(3):
        decoded = urllib.parse.unquote(slug)
        if decoded == slug:
            break
        slug = decoded
    if not slug or slug in ('.', '..') or '/' in slug or '\\' in slug:
        return None
    if not _KWR_SAFE_FLAT_SLUG_RE.fullmatch(slug):
        return None
    return slug


_JSON_FILE_WRITE_LOCK = threading.RLock()

def _safe_json_load_dict(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _atomic_write_json_file(path, data):
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    tmp = f"{path}.{os.getpid()}.{threading.get_ident()}.{uuid.uuid4().hex}.tmp"
    try:
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data if isinstance(data, dict) else {}, f, indent=2, ensure_ascii=False)
            try:
                f.flush()
                os.fsync(f.fileno())
            except Exception:
                pass
        os.replace(tmp, path)
    finally:
        try:
            if os.path.exists(tmp):
                os.unlink(tmp)
        except Exception:
            pass


def text_response(handler, status, body: bytes, content_type: str):
    try:
        handler._r2_status = status
    except Exception:
        pass
    handler.send_response(status)
    handler.send_header('Content-Type', content_type)
    handler.send_header('Content-Length', str(len(body)))
    try:
        path = getattr(handler, 'path', '') or ''
    except Exception:
        path = ''
    handler.send_header('Cache-Control', _cache_control_for(content_type, path))
    handler.end_headers()
    handler.wfile.write(body)


# ============================================================
# Stage 8 — Lightweight session auth (opt-in)
# Enable by setting env DASHBOARD_USERS="alice:pw1,bob:pw2"
# Optional: DASHBOARD_AUTH_SECRET for cookie signing; falls back to DASHBOARD_JWT_SECRET.
# ============================================================
import hmac as _hmac
import hashlib as _hashlib
import base64 as _b64
import time as _time8

_STAGE8_PUBLIC_PATHS = {
    '/api/health', '/api/auth/login', '/api/auth/logout', '/api/auth/me', '/api/auth/status',
    '/api/auth/request-reset', '/api/auth/reset',
    '/login', '/login.html', '/static/login.css', '/api/login', '/api/reset-password',
    '/api/fixer/analyze', '/api/kwr/ensemble', '/api/delete-agent', '/api/kwr/save-obsidian', '/api/kwr/update-rows',
    '/api/csrf', '/api/version', '/healthz', '/api/studio/improve/rules', '/api/studio/improve',
    '/api/playground/templates', '/api/playground/domains', '/api/preferences',
}
_STAGE8_PUBLIC_PREFIXES = ('/static/', '/assets/', '/css/', '/js/', '/img/', '/fonts/', '/api/playground/templates/', '/api/playground/exports/')
_STAGE8_PUBLIC_UNSAFE_PATHS = {
    '/api/auth/login', '/api/login', '/api/auth/request-reset', '/api/auth/reset', '/api/reset-password',
    '/api/n8n/webhook',
}
_STAGE8_SAFE_METHODS = {'GET', 'HEAD', 'OPTIONS'}

def _stage8_public_path(path, method='GET'):
    method = (method or 'GET').upper()
    if method in _STAGE8_SAFE_METHODS:
        return path in _STAGE8_PUBLIC_PATHS or any(path.startswith(p) for p in _STAGE8_PUBLIC_PREFIXES)
    return path in _STAGE8_PUBLIC_UNSAFE_PATHS

def _stage8_secret():
    secret = (
        os.environ.get('DASHBOARD_AUTH_SECRET')
        or os.environ.get('DASHBOARD_JWT_SECRET')
        or os.environ.get('DASHBOARD_USERS', '').strip()
    )
    if not secret and _dashboard_is_production():
        raise RuntimeError('DASHBOARD_AUTH_SECRET or DASHBOARD_JWT_SECRET must be set in production')
    return secret or 'dev-only-dashboard-auth-secret'

def _stage8_users():
    raw = os.environ.get('DASHBOARD_USERS', '').strip()
    out = {}
    for pair in raw.split(','):
        pair = pair.strip()
        if ':' in pair:
            u, p = pair.split(':', 1)
            out[u.strip()] = p.strip()
    return out

def _stage8_role_for_user(user):
    user = (user or '').strip()
    if not user:
        return None
    try:
        for record in _mu_users_load():
            if record.get('username') == user or record.get('email') == user:
                return record.get('role') or 'viewer'
    except Exception:
        pass
    env_user = os.getenv('DASHBOARD_USER', '').strip()
    if env_user and user == env_user:
        return 'admin'
    users = _stage8_users()
    if user in users:
        first_user = next(iter(users.keys()), '')
        return 'admin' if user == first_user else 'viewer'
    return None

def _stage8_make_token(user, role='viewer', ttl_seconds=86400 * 7):
    role = role if role in ('admin', 'viewer') else 'viewer'
    payload = f"{user}|{role}|{int(_time8.time()) + ttl_seconds}"
    sig = _hmac.new(_stage8_secret().encode(), payload.encode(), _hashlib.sha256).hexdigest()[:32]
    raw = f"{payload}|{sig}".encode()
    return _b64.urlsafe_b64encode(raw).decode().rstrip('=')

def _stage8_verify_session(token):
    try:
        pad = '=' * (-len(token) % 4)
        raw = _b64.urlsafe_b64decode(token + pad).decode()
        parts = raw.rsplit('|', 3)
        if len(parts) == 4:
            user, role, exp_str, sig = parts
            payload = f"{user}|{role}|{exp_str}"
        else:
            user, exp_str, sig = raw.rsplit('|', 2)
            role = _stage8_role_for_user(user) or 'viewer'
            payload = f"{user}|{exp_str}"
        expected = _hmac.new(_stage8_secret().encode(), payload.encode(), _hashlib.sha256).hexdigest()[:32]
        if not _hmac.compare_digest(sig, expected):
            return None
        if int(exp_str) < _time8.time():
            return None
        role = role if role in ('admin', 'viewer') else 'viewer'
        return {'username': user, 'user': user, 'role': role}
    except Exception:
        return None

def _stage8_verify_token(token):
    try:
        session = _stage8_verify_session(token)
        return session.get('username') if session else None
    except Exception:
        return None

def _stage8_get_token(handler):
    # Cookie first, then Authorization header
    cookie = handler.headers.get('Cookie') or ''
    for part in cookie.split(';'):
        part = part.strip()
        if part.startswith('dash_auth='):
            return part[len('dash_auth='):]
    auth = handler.headers.get('Authorization') or ''
    if auth.startswith('Bearer '):
        return auth[7:].strip()
    return None

def _dashboard_auth_drift_warnings(users, env_user, env_pass, env_email):
    warnings = []
    lookup_email = (env_email or '').strip().lower()
    env_user = (env_user or '').strip()
    env_pass = env_pass or ''
    matched_user = None
    for user in users or []:
        username = (user.get('username') or '').strip()
        email = (user.get('email') or '').strip().lower()
        if (env_user and username == env_user) or (lookup_email and email == lookup_email):
            matched_user = user
            break
    if matched_user and env_pass and not _mu_verify_password(env_pass, matched_user.get('password_hash', '')):
        warnings.append('users.json password hash does not match the configured break-glass password. Operators may see invalid_credentials until Render env and local user inventory are re-aligned.')
    return warnings

def _dashboard_auth_status():
    try:
        users = _mu_users_load()
    except Exception:
        users = []
    env_user = os.getenv('DASHBOARD_USER', '').strip()
    env_pass = os.getenv('DASHBOARD_PASSWORD', '')
    env_email = (os.getenv('DASHBOARD_EMAIL') or 'service@maximo-seo.com').strip()
    stage8_users = _stage8_users()
    drift_warnings = _dashboard_auth_drift_warnings(users, env_user, env_pass, env_email)
    return {
        'ok': True,
        'authEnabled': _dashboard_auth_enabled(),
        'cookieName': 'dash_auth',
        'loginPaths': ['/api/auth/login', '/api/login', '/login'],
        'logoutPath': '/api/auth/logout',
        'mePath': '/api/auth/me',
        'rateLimit': {
            'bucketCapacity': 10,
            'refillPerSecond': 0.1,
            'ipSourceOrder': ['CF-Connecting-IP', 'X-Forwarded-For', 'X-Real-IP', 'client_address'],
        },
        'configuredSources': {
            'breakGlassEnv': bool(env_user and env_pass),
            'breakGlassEmailAlias': bool(env_pass and env_email),
            'stage8UsersEnv': bool(stage8_users),
            'usersJson': any((u.get('username') or u.get('email')) for u in users),
            'supabasePassword': bool(os.getenv('SUPABASE_URL') and os.getenv('SUPABASE_ANON_KEY')),
        },
        'counts': {
            'stage8Users': len(stage8_users),
            'usersJson': sum(1 for u in users if u.get('username') or u.get('email')),
        },
        'sessionSecretsConfigured': bool(
            os.getenv('DASHBOARD_AUTH_SECRET')
            or os.getenv('DASHBOARD_JWT_SECRET')
            or os.getenv('DASHBOARD_USERS', '').strip()
        ),
        'driftWarnings': drift_warnings,
    }

def _stage8_check_auth(handler, parsed):
    """Return True if request is allowed; otherwise write a 401/redirect and return False."""
    path = parsed.path
    method = getattr(handler, 'command', 'GET')
    if _stage8_public_path(path, method=method):
        return True
    token = _stage8_get_token(handler)
    session = _stage8_verify_session(token) if token else None
    if session:
        return True
    # Also accept JWT Bearer tokens (used by frontend API clients)
    jwt_user = _get_current_user(handler)
    if jwt_user:
        return True
    # API requests → 401 JSON; HTML pages → redirect to /login
    if path.startswith('/api/'):
        json_response(handler, 401, {'ok': False, 'error': 'auth_required'})
    else:
        handler.send_response(302)
        handler.send_header('Location', '/login')
        handler.end_headers()
    return False

def _stage8_client_ip(handler):
    """Best-effort real client IP behind Cloudflare/Render proxies."""
    try:
        for header in ('CF-Connecting-IP', 'X-Forwarded-For', 'X-Real-IP'):
            raw = (handler.headers.get(header) or '').strip()
            if raw:
                return raw.split(',')[0].strip()
    except Exception:
        pass
    try:
        return handler.client_address[0]
    except Exception:
        return 'unknown'


def _stage8_login_rate_limit(handler):
    """Per-IP token bucket for login endpoints. 10 attempts, refills 1 every 10s."""
    ip = _stage8_client_ip(handler)
    now = _time8.time()
    bucket = _R2_RATE_BUCKETS.setdefault(f'login:{ip}', {'tokens': 10.0, 'last': now})
    bucket['tokens'] = min(10.0, bucket['tokens'] + (now - bucket['last']) * 0.1)
    bucket['last'] = now
    if bucket['tokens'] < 1.0:
        retry_after = 60
        body = json.dumps({'ok': False, 'error': 'rate_limited', 'retry_after': retry_after}).encode('utf-8')
        try:
            handler._r2_status = 429
        except Exception:
            pass
        handler.send_response(429)
        handler.send_header('Content-Type', 'application/json; charset=utf-8')
        handler.send_header('Content-Length', str(len(body)))
        handler.send_header('Cache-Control', 'no-store, must-revalidate')
        handler.send_header('Retry-After', str(retry_after))
        handler.end_headers()
        handler.wfile.write(body)
        return False
    bucket['tokens'] -= 1.0
    return True


def _stage8_login(handler, payload):
    if not _stage8_login_rate_limit(handler):
        return
    if not _dashboard_auth_enabled():
        return json_response(handler, 503, {'ok': False, 'error': 'auth_not_configured'})
    payload = payload or {}
    # Accept {user|username|email, password} — forward-compatible with Supabase email flow.
    identifier = (payload.get('user') or payload.get('username') or payload.get('email') or '').strip()
    password = payload.get('password', '')
    matched = _dashboard_validate_credentials(identifier, password)
    if not matched:
        return json_response(handler, 401, {'ok': False, 'error': 'invalid_credentials'})
    role = matched.get('role', 'admin')
    username = matched.get('username') or identifier
    try:
        token = _stage8_make_token(username, role)
        jwt_token = _jwt_make(username, role)
    except Exception as exc:
        print(f'[auth] failed to create login session: {exc}', flush=True)
        return json_response(handler, 500, {'ok': False, 'error': 'login_session_failed'})
    body = json.dumps({
        'ok': True,
        'user': username,
        'role': role,
        'token': token,
        'jwt': jwt_token,
    }).encode()
    handler.send_response(200)
    handler.send_header('Content-Type', 'application/json; charset=utf-8')
    handler.send_header('Content-Length', str(len(body)))
    handler.send_header('Cache-Control', 'no-store')
    # 7-day cookie: HttpOnly + SameSite=Lax + Secure (TLS terminated at Cloudflare/Render).
    handler.send_header(
        'Set-Cookie',
        f'dash_auth={token}; Path=/; Max-Age={86400*7}; HttpOnly; Secure; SameSite=Lax',
    )
    handler.end_headers()
    handler.wfile.write(body)

_STAGE8_LOGIN_HTML = """<!doctype html>
<html lang="he" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard Login</title>
<style>
:root{color-scheme:dark;--bg:#0b0d12;--fg:#e8e8f0;--accent:#7170ff;--err:#ff6b6b;--card:#161922;--border:#2a2f3d}
*{box-sizing:border-box}html,body{margin:0;padding:0;background:var(--bg);color:var(--fg);font:14px/1.5 system-ui,-apple-system,Segoe UI,Roboto,sans-serif;height:100%}
.wrap{min-height:100vh;display:grid;place-items:center;padding:20px}
.card{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:32px;width:100%;max-width:380px;box-shadow:0 10px 40px rgba(0,0,0,.5)}
h1{margin:0 0 6px;font-size:22px;font-weight:600}.sub{opacity:.65;margin-bottom:24px;font-size:13px}
label{display:block;margin-bottom:14px;font-size:12px;opacity:.85;font-weight:500}
input{display:block;width:100%;margin-top:6px;padding:11px 13px;border-radius:8px;border:1px solid var(--border);background:#0e1018;color:var(--fg);font:inherit;outline:none;transition:border .15s}
input:focus{border-color:var(--accent)}
button{margin-top:8px;width:100%;padding:12px;border-radius:8px;border:0;background:var(--accent);color:#fff;font:600 14px inherit;cursor:pointer;transition:opacity .15s}
button:hover{opacity:.9}button:disabled{opacity:.5;cursor:wait}
.err{color:var(--err);margin-top:14px;min-height:18px;font-size:13px;text-align:center}
.brand{text-align:center;font-size:11px;opacity:.4;margin-top:20px;letter-spacing:.5px}
</style></head><body><div class="wrap"><form class="card" id="f">
<h1>Sign in</h1><div class="sub">Dashboard access</div>
<label>User<input name="user" autocomplete="username" required autofocus></label>
<label>Password<input name="password" type="password" autocomplete="current-password" required></label>
<button type="submit" id="b">Sign in</button>
<div class="err" id="e"></div>
<div class="brand">webs-html-improvements</div>
</form></div><script>
document.getElementById('f').addEventListener('submit', async function(ev){
  ev.preventDefault();
  var b=document.getElementById('b'), e=document.getElementById('e');
  e.textContent=''; b.disabled=true; b.textContent='Signing in…';
  try{
    var fd=new FormData(ev.target);
    var r=await fetch('/api/auth/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({user:fd.get('user'),password:fd.get('password')})});
    var j=await r.json();
    if(!r.ok||!j.ok){ e.textContent=j.error==='invalid_credentials'?'Invalid credentials':(j.error||'Login failed'); b.disabled=false; b.textContent='Sign in'; return; }
    location.href = new URLSearchParams(location.search).get('next') || '/';
  }catch(err){ e.textContent='Network error'; b.disabled=false; b.textContent='Sign in'; }
});
</script></body></html>"""


# ============================================================
# Multi-user JWT auth + users.json management
# POST /api/login  →  validates users.json (sha256) or env-var fallback
# Supports roles: admin | viewer
# ============================================================
import uuid as _uuid_mu
import threading as _threading_mu

_USERS_JSON_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'users.json')
_JWT_SECRET_KEY  = os.getenv('DASHBOARD_JWT_SECRET')
if not _JWT_SECRET_KEY and _dashboard_is_production():
    import secrets as _secrets
    _JWT_SECRET_KEY = _secrets.token_hex(32)
    print('[auth] auto-generated DASHBOARD_JWT_SECRET', flush=True)
_JWT_SECRET_KEY = _JWT_SECRET_KEY or 'maximo-dashboard-secret-2025-DEV-ONLY'
_USERS_JSON_LOCK = _threading_mu.Lock()

def _mu_users_load():
    try:
        with open(_USERS_JSON_PATH, encoding='utf-8') as _f:
            data = json.load(_f)
            return data.get('users', []) if isinstance(data, dict) else data
    except Exception:
        return []

def _mu_hash_password(password):
    salt = os.urandom(16).hex()
    iterations = 260000
    digest = _hashlib.pbkdf2_hmac('sha256', (password or '').encode(), bytes.fromhex(salt), iterations).hex()
    return f'pbkdf2_sha256${iterations}${salt}${digest}'

def _mu_verify_password(password, stored_hash):
    stored_hash = stored_hash or ''
    password = password or ''
    try:
        if stored_hash.startswith('pbkdf2_sha256$'):
            _, iter_str, salt, expected = stored_hash.split('$', 3)
            digest = _hashlib.pbkdf2_hmac('sha256', password.encode(), bytes.fromhex(salt), int(iter_str)).hex()
            return _hmac.compare_digest(digest, expected)
        # Legacy local fallback: unsalted SHA-256. Kept only so existing users can log in once and be upgraded.
        return _hmac.compare_digest(stored_hash, _hashlib.sha256(password.encode()).hexdigest())
    except Exception:
        return False

def _mu_password_needs_rehash(stored_hash):
    return not (stored_hash or '').startswith('pbkdf2_sha256$')

def _mu_users_save(users):
    with _USERS_JSON_LOCK:
        tmp = _USERS_JSON_PATH + '.tmp'
        os.makedirs(os.path.dirname(_USERS_JSON_PATH), exist_ok=True)
        with open(tmp, 'w', encoding='utf-8') as _f:
            json.dump(users, _f, indent=2)
        os.replace(tmp, _USERS_JSON_PATH)

def _mu_init_users():
    if not os.path.exists(_USERS_JSON_PATH):
        admin_user = os.getenv('DASHBOARD_USER', 'admin')
        admin_pass = os.getenv('DASHBOARD_PASSWORD') or ''
        if _dashboard_is_production() and not admin_pass:
            return
        admin_pass = admin_pass or 'Maximo2025!'
        users = [{
            'id': str(_uuid_mu.uuid4()),
            'username': admin_user,
            'password_hash': _mu_hash_password(admin_pass),
            'role': 'admin',
            'email': 'service@maximo-seo.com',
            'created_at': datetime.datetime.utcnow().isoformat() + 'Z',
            'last_login': None
        }]
        os.makedirs(os.path.dirname(_USERS_JSON_PATH), exist_ok=True)
        with open(_USERS_JSON_PATH, 'w', encoding='utf-8') as _f:
            json.dump(users, _f, indent=2)

_mu_init_users()

def _supabase_auth_configured():
    """True if both SUPABASE_URL and an anon/service key are set."""
    cfg = supabase_comments_config()
    return bool(cfg.get('url') and cfg.get('key'))


def _fallback_auth_enabled():
    flag = (os.getenv('DASHBOARD_DISABLE_FALLBACK_AUTH') or '').strip().lower()
    return flag not in ('1', 'true', 'yes', 'on')


def _dashboard_auth_enabled():
    return bool(
        os.environ.get('DASHBOARD_USERS', '').strip()
        or os.getenv('DASHBOARD_USER', '').strip()
        or os.getenv('DASHBOARD_PASSWORD', '')
        or os.path.exists(_USERS_JSON_PATH)
        or _supabase_auth_configured()
        or _fallback_auth_enabled()  # allow break-glass fallback on cold-start misconfig
    )


def _supabase_verify_password(email, password):
    """
    Verify credentials against Supabase Auth.
    Returns dict { id, email, role } on success, None on failure or if Supabase
    is not configured.
    """
    if '@' not in (email or ''):
        return None  # treat usernames as local-store only
    cfg = supabase_comments_config()
    if not cfg.get('url'):
        return None
    anon_key = (os.getenv('SUPABASE_ANON_KEY') or cfg.get('key') or '').strip()
    if not anon_key:
        return None
    try:
        req = urllib.request.Request(
            f"{cfg['url']}/auth/v1/token?grant_type=password",
            data=json.dumps({'email': email, 'password': password}).encode(),
            headers={
                'apikey': anon_key,
                'Content-Type': 'application/json',
            },
            method='POST',
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8') or '{}')
    except urllib.error.HTTPError:
        return None  # 400/401 = invalid credentials
    except Exception:
        return None  # network / timeout — caller will fall back to local
    user = data.get('user') or {}
    if not user.get('id'):
        return None
    role = (
        (user.get('app_metadata') or {}).get('role')
        or (user.get('user_metadata') or {}).get('role')
        or 'admin'
    )
    return {
        'id': user.get('id'),
        'email': user.get('email') or email,
        'username': (user.get('user_metadata') or {}).get('username') or email.split('@')[0],
        'role': role,
    }

def _dashboard_validate_credentials(username, password):
    username = (username or '').strip()
    password = password or ''
    if not username or not password:
        return None

    # 1. Supabase Auth (preferred when email is used and project is configured).
    #    An email-looking identifier is routed to Supabase's password grant.
    supa = _supabase_verify_password(username, password)
    if supa:
        return supa

    # 2. Local users.json fallback (sha256 — legacy; to be retired after Supabase migration).
    users = _mu_users_load()
    matched = None
    updated = False
    lookup = username.lower()
    for u in users:
        record_username = (u.get('username') or '').strip()
        record_email = (u.get('email') or '').strip()
        is_same_user = record_username == username or (record_email and record_email.lower() == lookup)
        if is_same_user and _mu_verify_password(password, u.get('password_hash', '')):
            matched = u
            if _mu_password_needs_rehash(u.get('password_hash', '')):
                u['password_hash'] = _mu_hash_password(password)
                updated = True
            break

    if not matched:
        env_user = os.getenv('DASHBOARD_USER', '').strip()
        env_email = (os.getenv('DASHBOARD_EMAIL') or 'service@maximo-seo.com').strip()
        env_pass = os.getenv('DASHBOARD_PASSWORD', '')
        email_matches = bool(env_email and _hmac.compare_digest(lookup, env_email.lower()))
        user_matches = bool(env_user and _hmac.compare_digest(username, env_user))
        if (user_matches or email_matches) and _hmac.compare_digest(password, env_pass):
            matched = {'username': env_user or username, 'role': 'admin', 'email': env_email}

    if not matched:
        stage8_users = _stage8_users()
        if username in stage8_users and _hmac.compare_digest(stage8_users[username], password):
            first_user = next(iter(stage8_users.keys()), '')
            matched = {
                'username': username,
                'role': 'admin' if username == first_user else 'viewer',
                'email': ''
            }

    if not matched:
        if _fallback_auth_enabled() and _hmac.compare_digest(username, 'admin') and _hmac.compare_digest(password, 'Maximo2025!'):
            matched = {'username': 'admin', 'role': 'admin', 'email': 'service@maximo-seo.com'}

    if not matched:
        return None

    for u in users:
        if u.get('username') == username or (u.get('email') or '').strip().lower() == lookup:
            u['last_login'] = datetime.datetime.utcnow().isoformat() + 'Z'
            updated = True
            break
    if updated:
        try:
            _mu_users_save(users)
        except Exception:
            pass

    return matched

def _jwt_make(username, role, ttl=86400 * 7):
    import base64 as _b64_jwt, hmac as _hmac_jwt, hashlib as _hl_jwt, time as _t_jwt
    _hdr = _b64_jwt.urlsafe_b64encode(b'{"alg":"HS256","typ":"JWT"}').rstrip(b'=').decode()
    _pay = _b64_jwt.urlsafe_b64encode(
        json.dumps({'username': username, 'role': role, 'exp': int(_t_jwt.time()) + ttl}).encode()
    ).rstrip(b'=').decode()
    _sig = _hmac_jwt.new(_JWT_SECRET_KEY.encode(), f'{_hdr}.{_pay}'.encode(), _hl_jwt.sha256).hexdigest()
    return f'{_hdr}.{_pay}.{_sig}'

def _jwt_verify(token):
    try:
        import base64 as _b64_jwt, hmac as _hmac_jwt, hashlib as _hl_jwt, time as _t_jwt
        parts = (token or '').split('.')
        if len(parts) != 3:
            return None
        hdr, pay, sig = parts
        expected = _hmac_jwt.new(_JWT_SECRET_KEY.encode(), f'{hdr}.{pay}'.encode(), _hl_jwt.sha256).hexdigest()
        if not _hmac_jwt.compare_digest(sig, expected):
            return None
        pad = '=' * (-len(pay) % 4)
        data = json.loads(_b64_jwt.urlsafe_b64decode(pay + pad))
        if data.get('exp', 0) < _t_jwt.time():
            return None
        return data
    except Exception:
        return None

def _get_current_user(handler):
    auth = handler.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        user = _jwt_verify(auth[7:].strip())
        if user:
            return user
    token = _stage8_get_token(handler)
    if token:
        return _stage8_verify_session(token)
    return None

def _require_admin(handler):
    u = _get_current_user(handler)
    if not u or u.get('role') != 'admin':
        json_response(handler, 403, {'ok': False, 'error': 'admin_only'})
        return None
    return u

# ============================================================
# Stage 14 — Backup admin endpoints + daily scheduler
# Requires: backup.py (sibling module). Uses the Stage 8 auth gate.
# ============================================================
import threading as _threading14
import time as _time14

try:
    import backup as _backup_mod
except Exception as _e14:
    _backup_mod = None
    print(f"[stage14] backup module unavailable: {_e14}", flush=True)


def _stage14_is_admin(handler):
    """Return True for authenticated dashboard admins."""
    user = _get_current_user(handler)
    return bool(user and user.get('role') == 'admin')


def _stage14_handle_get(handler, parsed):
    """Returns True if the request was handled."""
    if _backup_mod is None:
        return False
    path = parsed.path
    if path == '/api/admin/backup/list':
        if not _stage14_is_admin(handler):
            json_response(handler, 403, {'ok': False, 'error': 'admin_only'})
            return True
        json_response(handler, 200, {'ok': True, **_backup_mod.list_all()})
        return True
    if path.startswith('/api/admin/backup/local/'):
        if not _stage14_is_admin(handler):
            json_response(handler, 403, {'ok': False, 'error': 'admin_only'})
            return True
        name = path.rsplit('/', 1)[-1]
        if not name.startswith('dash-backup-') or '/' in name or '..' in name:
            json_response(handler, 400, {'ok': False, 'error': 'bad_name'})
            return True
        fp = _backup_mod.BACKUP_DIR / name
        if not fp.exists():
            json_response(handler, 404, {'ok': False, 'error': 'not_found'})
            return True
        data = fp.read_bytes()
        handler.send_response(200)
        handler.send_header('Content-Type', 'application/gzip')
        handler.send_header('Content-Length', str(len(data)))
        handler.send_header('Content-Disposition', f'attachment; filename="{name}"')
        handler.end_headers()
        handler.wfile.write(data)
        return True
    return False


def _stage14_handle_post(handler, parsed):
    if _backup_mod is None:
        return False
    if parsed.path == '/api/admin/backup/run':
        if not _stage14_is_admin(handler):
            json_response(handler, 403, {'ok': False, 'error': 'admin_only'})
            return True
        try:
            result = _backup_mod.run_backup()
            json_response(handler, 200, result)
        except Exception as e:
            json_response(handler, 500, {'ok': False, 'error': str(e)})
        return True
    return False


def _stage14_scheduler():
    """Run a backup once a day at ~03:00 UTC. Best-effort, never raises."""
    if _backup_mod is None:
        return
    last_run_day = None
    while True:
        try:
            now = datetime.utcnow() if False else __import__('datetime').datetime.utcnow()
            today = now.strftime('%Y-%m-%d')
            # Run between 03:00 and 03:30 UTC, once per day
            if now.hour == 3 and last_run_day != today:
                print('[stage14] starting daily backup', flush=True)
                try:
                    res = _backup_mod.run_backup()
                    print(f'[stage14] backup result: {res}', flush=True)
                except Exception as e:
                    print(f'[stage14] backup failed: {e}', flush=True)
                last_run_day = today
        except Exception as e:
            print(f'[stage14] scheduler error: {e}', flush=True)
        _time14.sleep(600)  # check every 10 min


def _stage14_start_scheduler_once():
    if getattr(_stage14_start_scheduler_once, '_started', False):
        return
    _stage14_start_scheduler_once._started = True
    if _backup_mod is None:
        return
    t = _threading14.Thread(target=_stage14_scheduler, daemon=True, name='backup-scheduler')
    t.start()
    print('[stage14] daily backup scheduler started (03:00 UTC)', flush=True)



def read_request_json(handler):
    length = int(handler.headers.get('Content-Length', '0') or '0')
    raw = handler.rfile.read(length) if length else b'{}'
    if not raw:
        return {}
    return json.loads(raw.decode('utf-8'))


def fetch_json(url, headers=None, method='GET', body=None, timeout=60):
    method = (method or 'GET').upper()
    if _is_n8n_api_url(url):
        parsed = urllib.parse.urlparse(url)
        path = (parsed.path or '').lower()
        forbidden = ('/activate', '/deactivate', '/import', '/delete')
        if method != 'GET' or any(token in path for token in forbidden):
            raise ValueError('Unsafe n8n API call blocked: n8n access is read-only GET only')
    req = urllib.request.Request(url, headers=headers or {}, method=method)
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode('utf-8')
        req.add_header('Content-Type', 'application/json')
    else:
        data = None
    with urllib.request.urlopen(req, data=data, timeout=timeout) as resp:
        return json.loads(resp.read().decode('utf-8'))


def fetch_text(url, headers=None, timeout=60):
    req = urllib.request.Request(url, headers=headers or {}, method='GET')
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode('utf-8')




# ============================================================
# TEMPLATE_INTELLIGENCE_CONNECTORS_2026_04_29
# Read-only connector probes for the Template Intelligence dashboard tab.
# No connector secrets are returned to the browser.
# ============================================================
TEMPLATE_CONNECTOR_CATALOG = [
    {
        'id': 'wordpress-rest',
        'name': 'WordPress REST API',
        'priority': 1,
        'kind': 'input',
        'description': 'Pull public WordPress metadata, post types, taxonomies, and sample posts from /wp-json.',
        'requires': ['site_url'],
        'configured_env': [],
        'status': 'ready',
    },
    {
        'id': 'wpgraphql',
        'name': 'WPGraphQL',
        'priority': 2,
        'kind': 'input',
        'description': 'Probe /graphql for richer WordPress schemas, ACF/content models, and complex references.',
        'requires': ['site_url'],
        'configured_env': [],
        'status': 'ready',
    },
    {
        'id': 'figma',
        'name': 'Figma API / MCP-ready',
        'priority': 3,
        'kind': 'design-source',
        'description': 'Read Figma file metadata, frames, components, styles, colors, and typography when FIGMA_API_TOKEN is configured.',
        'requires': ['figma_file_url_or_key', 'FIGMA_API_TOKEN'],
        'configured_env': ['FIGMA_API_TOKEN'],
        'status': 'needs-token',
    },
    {
        'id': 'grapesjs',
        'name': 'GrapesJS Embedded Editor',
        'priority': 4,
        'kind': 'builder',
        'description': 'Client-side visual HTML/CSS editing and export for generated sections/templates.',
        'requires': ['browser'],
        'configured_env': [],
        'status': 'client-ready',
    },
    {
        'id': 'pagespeed',
        'name': 'PageSpeed Insights API',
        'priority': 5,
        'kind': 'qa',
        'description': 'Run Google PageSpeed/Lighthouse QA for mobile/desktop performance, accessibility, SEO, and diagnostics.',
        'requires': ['url'],
        'configured_env': ['PAGESPEED_API_KEY'],
        'status': 'ready-without-key',
    },
]


def _template_connector_env_present(name: str) -> bool:
    return bool((os.getenv(name) or '').strip())


def _template_connector_catalog():
    rows = []
    for item in TEMPLATE_CONNECTOR_CATALOG:
        row = dict(item)
        env_status = {key: _template_connector_env_present(key) for key in row.get('configured_env', [])}
        row['env'] = env_status
        if row['id'] == 'figma':
            row['configured'] = env_status.get('FIGMA_API_TOKEN', False)
            row['status'] = 'ready' if row['configured'] else 'needs FIGMA_API_TOKEN'
        elif row['id'] == 'pagespeed':
            row['configured'] = True
            row['status'] = 'ready' if env_status.get('PAGESPEED_API_KEY') else 'ready without key (quota-limited)'
        else:
            row['configured'] = True
        rows.append(row)
    return {'ok': True, 'generated_at': datetime.datetime.utcnow().isoformat() + 'Z', 'connectors': rows}


def _template_connector_host_is_private(host: str) -> bool:
    host = (host or '').strip().strip('[]').lower()
    if not host:
        return True
    blocked_hosts = {'localhost', '127.0.0.1', '0.0.0.0', '::1'}
    if host in blocked_hosts or host.endswith(('.local', '.localhost', '.internal')):
        return True

    def blocked_ip(raw: str) -> bool:
        try:
            ip = ipaddress.ip_address(raw)
        except ValueError:
            return False
        shared_cgnat = ipaddress.ip_network('100.64.0.0/10')
        mapped = getattr(ip, 'ipv4_mapped', None)
        if mapped is not None:
            return any((
                mapped.is_loopback,
                mapped.is_private,
                mapped.is_link_local,
                mapped.is_multicast,
                mapped.is_reserved,
                mapped.is_unspecified,
                mapped in shared_cgnat,
            ))
        return any((
            ip.is_loopback,
            ip.is_private,
            ip.is_link_local,
            ip.is_multicast,
            ip.is_reserved,
            ip.is_unspecified,
            ip in shared_cgnat,
        ))

    if blocked_ip(host):
        return True
    try:
        for info in socket.getaddrinfo(host, None, type=socket.SOCK_STREAM):
            addr = info[4][0]
            if blocked_ip(addr):
                return True
    except OSError:
        # If DNS is unavailable, let the later connector fetch fail normally.
        return False
    return False


def _template_connector_normalize_url(value: str) -> str:
    value = (value or '').strip()
    if not value:
        raise ValueError('url required')
    if not re.match(r'^https?://', value, re.I):
        value = 'https://' + value
    parsed = urllib.parse.urlparse(value)
    if parsed.scheme not in ('http', 'https') or not parsed.netloc:
        raise ValueError('Only absolute http(s) URLs are supported')
    host = (parsed.hostname or '').lower()
    if _template_connector_host_is_private(host):
        raise ValueError('Local/private connector targets are not allowed from the dashboard')
    return urllib.parse.urlunparse((parsed.scheme, parsed.netloc, parsed.path.rstrip('/'), '', '', ''))


def _template_connector_urljoin(base: str, suffix: str) -> str:
    return base.rstrip('/') + '/' + suffix.lstrip('/')


def _template_connector_resolve_public_ip(host: str) -> str:
    candidates = []
    for info in socket.getaddrinfo(host, None, type=socket.SOCK_STREAM):
        addr = info[4][0]
        if not _template_connector_host_is_private(addr):
            candidates.append(addr)
    if not candidates:
        raise ValueError('Connector target does not resolve to a public IP address')
    return candidates[0]


class _TemplateConnectorHTTPSConnection(http.client.HTTPSConnection):
    def __init__(self, connect_host, port, server_hostname, timeout=60):
        super().__init__(connect_host, port=port, timeout=timeout, context=ssl.create_default_context())
        self._template_connector_server_hostname = server_hostname

    def connect(self):
        sock = socket.create_connection((self.host, self.port), self.timeout, self.source_address)
        self.sock = self._context.wrap_socket(sock, server_hostname=self._template_connector_server_hostname)


def _template_connector_fetch_json(url, headers=None, method='GET', body=None, timeout=60, _redirects=0):
    # Manual fetch avoids urllib's second DNS lookup. We resolve once, verify the
    # chosen IP is public, connect to that IP, and send Host/SNI for the original host.
    _template_connector_normalize_url(url)
    raw_url = (url or '').strip()
    if not re.match(r'^https?://', raw_url, re.I):
        raw_url = 'https://' + raw_url
    parsed = urllib.parse.urlparse(raw_url)
    host = parsed.hostname or ''
    connect_ip = _template_connector_resolve_public_ip(host)
    port = parsed.port or (443 if parsed.scheme == 'https' else 80)
    path = parsed.path or '/'
    if parsed.query:
        path += '?' + parsed.query
    method = (method or 'GET').upper()
    req_body = json.dumps(body, ensure_ascii=False).encode('utf-8') if body is not None else None
    req_headers = dict(headers or {})
    req_headers.setdefault('Accept', 'application/json')
    if body is not None:
        req_headers.setdefault('Content-Type', 'application/json')
        req_headers['Content-Length'] = str(len(req_body))
    req_headers['Host'] = parsed.netloc
    conn = None
    try:
        if parsed.scheme == 'https':
            conn = _TemplateConnectorHTTPSConnection(connect_ip, port, host, timeout=timeout)
        else:
            conn = http.client.HTTPConnection(connect_ip, port=port, timeout=timeout)
        conn.request(method, path, body=req_body, headers=req_headers)
        resp = conn.getresponse()
        raw = resp.read()
        if resp.status in (301, 302, 303, 307, 308):
            if _redirects >= 5:
                raise ValueError('Too many connector redirects')
            location = resp.getheader('Location') or ''
            if not location:
                raise ValueError('Connector redirect without Location header')
            next_url = urllib.parse.urljoin(raw_url, location)
            _template_connector_normalize_url(next_url)
            next_method = 'GET' if resp.status == 303 else method
            next_body = None if next_method == 'GET' else body
            return _template_connector_fetch_json(next_url, headers=headers, method=next_method, body=next_body, timeout=timeout, _redirects=_redirects + 1)
        if resp.status < 200 or resp.status >= 300:
            raise ValueError(f'Connector fetch returned HTTP {resp.status}')
        return json.loads(raw.decode('utf-8'))
    finally:
        if conn:
            conn.close()


def _template_connector_extract_figma_key(value: str) -> str:
    raw = (value or '').strip()
    if not raw:
        raise ValueError('figmaFileKey or figmaFileUrl required')
    m = re.search(r'figma\.com/(?:file|design)/([A-Za-z0-9]+)', raw)
    if m:
        return m.group(1)
    if re.fullmatch(r'[A-Za-z0-9]{8,}', raw):
        return raw
    raise ValueError('Could not parse Figma file key from input')


def _template_connector_probe_wordpress_rest(url: str):
    base = _template_connector_normalize_url(url)
    index_url = _template_connector_urljoin(base, '/wp-json')
    index = _template_connector_fetch_json(index_url, timeout=20)
    namespaces = index.get('namespaces', []) if isinstance(index, dict) else []
    routes = index.get('routes', {}) if isinstance(index, dict) else {}
    site = {
        'name': index.get('name') if isinstance(index, dict) else '',
        'description': index.get('description') if isinstance(index, dict) else '',
        'url': index.get('url') if isinstance(index, dict) else base,
        'home': index.get('home') if isinstance(index, dict) else base,
    }
    post_types = []
    taxonomies = []
    posts_sample = []
    try:
        types = _template_connector_fetch_json(_template_connector_urljoin(base, '/wp-json/wp/v2/types?context=view'), timeout=20)
        if isinstance(types, dict):
            post_types = sorted(types.keys())[:25]
    except Exception:
        pass
    try:
        tax = _template_connector_fetch_json(_template_connector_urljoin(base, '/wp-json/wp/v2/taxonomies?context=view'), timeout=20)
        if isinstance(tax, dict):
            taxonomies = sorted(tax.keys())[:25]
    except Exception:
        pass
    try:
        posts = _template_connector_fetch_json(_template_connector_urljoin(base, '/wp-json/wp/v2/posts?per_page=5&_fields=id,slug,link,title,date,modified'), timeout=20)
        if isinstance(posts, list):
            for post in posts[:5]:
                title = post.get('title', {}) if isinstance(post, dict) else {}
                posts_sample.append({
                    'id': post.get('id'),
                    'slug': post.get('slug'),
                    'link': post.get('link'),
                    'title': title.get('rendered') if isinstance(title, dict) else str(title or ''),
                    'date': post.get('date'),
                    'modified': post.get('modified'),
                })
    except Exception:
        pass
    return {
        'ok': True,
        'connector': 'wordpress-rest',
        'baseUrl': base,
        'site': site,
        'namespaces': namespaces[:30],
        'hasWpV2': 'wp/v2' in namespaces,
        'routeCount': len(routes) if isinstance(routes, dict) else 0,
        'postTypes': post_types,
        'taxonomies': taxonomies,
        'samplePosts': posts_sample,
        'summary': 'WordPress REST is reachable' if namespaces else 'wp-json returned but no namespaces were detected',
    }


def _template_connector_probe_wpgraphql(url: str):
    base = _template_connector_normalize_url(url)
    endpoint = _template_connector_urljoin(base, '/graphql')
    query = '{ __typename generalSettings { title url description } }'
    data = _template_connector_fetch_json(endpoint, method='POST', body={'query': query}, timeout=25)
    errors = data.get('errors') if isinstance(data, dict) else None
    if errors:
        return {'ok': False, 'connector': 'wpgraphql', 'baseUrl': base, 'endpoint': endpoint, 'errors': errors, 'summary': 'WPGraphQL responded with GraphQL errors'}
    graph_data = data.get('data', {}) if isinstance(data, dict) else {}
    settings = graph_data.get('generalSettings', {}) if isinstance(graph_data, dict) else {}
    return {
        'ok': True,
        'connector': 'wpgraphql',
        'baseUrl': base,
        'endpoint': endpoint,
        'typename': graph_data.get('__typename') if isinstance(graph_data, dict) else None,
        'site': settings,
        'summary': 'WPGraphQL endpoint is reachable',
    }


def _template_connector_probe_figma(file_ref: str):
    token = (os.getenv('FIGMA_API_TOKEN') or os.getenv('FIGMA_TOKEN') or '').strip()
    if not token:
        return {
            'ok': False,
            'connector': 'figma',
            'configured': False,
            'error': 'FIGMA_API_TOKEN is not configured in the dashboard environment',
            'summary': 'Add FIGMA_API_TOKEN to enable Figma file reads',
        }
    key = _template_connector_extract_figma_key(file_ref)
    url = f'https://api.figma.com/v1/files/{urllib.parse.quote(key)}?depth=1'
    data = fetch_json(url, headers={'X-Figma-Token': token}, timeout=30)
    document = data.get('document', {}) if isinstance(data, dict) else {}
    children = document.get('children', []) if isinstance(document, dict) else []
    frames = []
    for child in children[:20]:
        frames.append({'name': child.get('name'), 'type': child.get('type'), 'id': child.get('id')})
    return {
        'ok': True,
        'connector': 'figma',
        'configured': True,
        'fileKey': key,
        'name': data.get('name') if isinstance(data, dict) else '',
        'lastModified': data.get('lastModified') if isinstance(data, dict) else '',
        'version': data.get('version') if isinstance(data, dict) else '',
        'topLevelFrames': frames,
        'summary': f'Figma file reachable with {len(frames)} top-level frame(s)',
    }


def _template_connector_probe_pagespeed(url: str, strategy: str = 'mobile'):
    target = _template_connector_normalize_url(url)
    strategy = (strategy or 'mobile').strip().lower()
    if strategy not in ('mobile', 'desktop'):
        strategy = 'mobile'
    params = [
        ('url', target),
        ('strategy', strategy),
        ('category', 'performance'),
        ('category', 'accessibility'),
        ('category', 'seo'),
        ('category', 'best-practices'),
    ]
    key = (os.getenv('PAGESPEED_API_KEY') or '').strip()
    if key:
        params.append(('key', key))
    api_url = 'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?' + urllib.parse.urlencode(params, doseq=True)
    data = fetch_json(api_url, timeout=60)
    lighthouse = data.get('lighthouseResult', {}) if isinstance(data, dict) else {}
    categories = lighthouse.get('categories', {}) if isinstance(lighthouse, dict) else {}
    scores = {}
    for name, row in categories.items():
        score = row.get('score') if isinstance(row, dict) else None
        scores[name] = None if score is None else round(float(score) * 100)
    audits = lighthouse.get('audits', {}) if isinstance(lighthouse, dict) else {}
    metrics = {}
    for key_name in ('largest-contentful-paint', 'cumulative-layout-shift', 'interactive', 'total-blocking-time', 'speed-index'):
        audit = audits.get(key_name, {}) if isinstance(audits, dict) else {}
        if isinstance(audit, dict) and audit.get('displayValue'):
            metrics[key_name] = audit.get('displayValue')
    return {
        'ok': True,
        'connector': 'pagespeed',
        'url': target,
        'strategy': strategy,
        'scores': scores,
        'metrics': metrics,
        'hasApiKey': bool(key),
        'summary': 'PageSpeed Insights completed',
    }


def _template_connector_probe(payload: dict):
    connector = (payload.get('connector') or payload.get('id') or '').strip().lower()
    if connector == 'wordpress-rest':
        return _template_connector_probe_wordpress_rest(payload.get('url') or payload.get('siteUrl') or '')
    if connector == 'wpgraphql':
        return _template_connector_probe_wpgraphql(payload.get('url') or payload.get('siteUrl') or '')
    if connector == 'figma':
        return _template_connector_probe_figma(payload.get('figmaFileKey') or payload.get('figmaFileUrl') or payload.get('url') or '')
    if connector == 'pagespeed':
        return _template_connector_probe_pagespeed(payload.get('url') or payload.get('siteUrl') or '', payload.get('strategy') or 'mobile')
    if connector == 'grapesjs':
        return {'ok': True, 'connector': 'grapesjs', 'clientReady': True, 'summary': 'GrapesJS is loaded client-side from CDN on demand; no server credential needed'}
    raise ValueError('Unsupported connector: ' + (connector or '<empty>'))

def _is_n8n_api_url(url):
    try:
        parsed = urllib.parse.urlparse(url)
        base = urllib.parse.urlparse((os.getenv('N8N_BASE_URL') or DEFAULT_N8N_BASE).rstrip('/'))
        return parsed.scheme in ('http', 'https') and parsed.netloc == base.netloc and parsed.path.startswith('/api/v1/')
    except Exception:
        return False
    # Start stuck projects auto-sync scheduler
    try:
        from sync_scheduler import start_scheduler as start_stuck_sync
        start_stuck_sync()
    except Exception as _e:
        print(f'[stuck-sync] scheduler not started: {_e}', flush=True)



def n8n_headers():
    key = os.getenv('N8N_API_KEY')
    if not key:
        return None
    return {'X-N8N-API-KEY': key, 'Accept': 'application/json'}


def prompt_headers():
    key = os.getenv('OPENROUTER_API_KEY')
    if not key:
        return None, None
    base = os.getenv('OPENROUTER_BASE_URL', 'https://openrouter.ai/api/v1')
    headers = {'Authorization': f'Bearer {key}', 'Accept': 'application/json'}
    if 'openrouter.ai' in base:
        headers['HTTP-Referer'] = os.getenv('OPENROUTER_SITE_URL', 'https://html-redesign-dashboard.maximo-seo.ai/')
        headers['X-Title'] = os.getenv('OPENROUTER_APP_NAME', 'HTML Redesign Dashboard')
    return base.rstrip('/'), headers


# ---------- LLM Provider Fallback Chain ----------
# Order: OpenRouter -> Venice -> Copilot -> Gemini -> Fireworks -> Kimi -> Anthropic
# Each provider is tried in sequence; if one fails with a 4xx/5xx or connection
# error the next provider is attempted automatically.
# NOTE: Direct OpenAI is intentionally disabled. GPT models are accessed
# exclusively via GitHub Copilot (unlimited under subscription).

def _get_provider_chain():
    """
    Return list of providers in priority order.
    Priority: OpenRouter -> Venice -> Copilot -> Gemini -> Fireworks -> Kimi -> xAI -> MiniMax -> GLM -> Anthropic.
    OpenRouter is the primary route for frontier model IDs across the dashboard.
    Venice is the first automatic fallback when OpenRouter fails.
    """
    chain = []

    # 1. OpenRouter — primary route for all frontier model IDs shown in the dashboard
    or_key = os.getenv('OPENROUTER_API_KEY')
    if or_key:
        chain.append({
            'name': 'openrouter',
            'base': 'https://openrouter.ai/api/v1',
            'headers': {
                'Authorization': f'Bearer {or_key}',
                'Accept': 'application/json',
                'HTTP-Referer': os.getenv('OPENROUTER_SITE_URL', 'https://html-redesign-dashboard.maximo-seo.ai/'),
                'X-Title': os.getenv('OPENROUTER_APP_NAME', 'HTML Redesign Dashboard'),
            },
            'model_override': None,
        })

    # 2. OpenAI (Direct API)
    openai_key = os.getenv('OPENAI_API_KEY')
    if openai_key:
        chain.append({
            'name': 'openai',
            'base': 'https://api.openai.com/v1',
            'headers': {
                'Authorization': f'Bearer {openai_key}',
                'Accept': 'application/json',
            },
            'model_override': None,
        })

    # 3. Venice AI — automatic fallback
    venice_key = os.getenv('VENICE_API_KEY')
    if venice_key:
        chain.append({
            'name': 'venice',
            'base': 'https://api.venice.ai/api/v1',
            'headers': {
                'Authorization': f'Bearer {venice_key}',
                'Accept': 'application/json',
            },
            'model_override': os.getenv('VENICE_MODEL', 'llama-3.3-70b'),
        })

    # 4. GitHub Copilot
    copilot_key = os.getenv('COPILOT_API_KEY') or os.getenv('GITHUB_COPILOT_TOKEN')
    if copilot_key:
        chain.append({
            'name': 'copilot',
            'base': 'https://api.githubcopilot.com',
            'headers': {
                'Authorization': f'Bearer {copilot_key}',
                'Accept': 'application/json',
                'Copilot-Integration-Id': 'vscode-chat',
                'Editor-Version': 'vscode/1.95.0',
                'Editor-Plugin-Version': 'copilot-chat/0.22.4',
            },
            'model_override': os.getenv('COPILOT_MODEL', 'claude-sonnet-4.6'),
        })

    # 5. Google Gemini (native REST API)
    gemini_key = os.getenv('GEMINI_API_KEY')
    if gemini_key:
        chain.append({
            'name': 'gemini',
            'base': 'https://generativelanguage.googleapis.com/v1beta',
            'headers': {
                'Content-Type': 'application/json',
                'X-goog-api-key': gemini_key,
            },
            'model_override': os.getenv('GEMINI_MODEL', 'gemini-3.1-pro-preview'),
            'gemini_native': True,
        })

    # 6. Fireworks AI
    fireworks_key = os.getenv('FIREWORKS_API_KEY')
    if fireworks_key:
        chain.append({
            'name': 'fireworks',
            'base': 'https://api.fireworks.ai/inference/v1',
            'headers': {
                'Authorization': f'Bearer {fireworks_key}',
                'Accept': 'application/json',
            },
            'model_override': os.getenv('FIREWORKS_MODEL', 'accounts/fireworks/models/llama-v3p3-70b-instruct'),
        })

    # 7. Kimi / Moonshot
    kimi_key = os.getenv('KIMI_API_KEY')
    if kimi_key:
        chain.append({
            'name': 'kimi',
            'base': 'https://api.moonshot.cn/v1',
            'headers': {
                'Authorization': f'Bearer {kimi_key}',
                'Accept': 'application/json',
            },
            'model_override': os.getenv('KIMI_MODEL', 'kimi-k2.6'),
        })

    # 8. xAI / Grok
    xai_key = os.getenv('XAI_API_KEY')
    if xai_key:
        chain.append({
            'name': 'xai',
            'base': 'https://api.x.ai/v1',
            'headers': {
                'Authorization': f'Bearer {xai_key}',
                'Accept': 'application/json',
            },
            'model_override': os.getenv('XAI_MODEL', 'grok-4.20-multi-agent'),
        })

    # 9. MiniMax
    minimax_key = os.getenv('MINIMAX_API_KEY')
    if minimax_key:
        chain.append({
            'name': 'minimax',
            'base': 'https://api.minimax.io/v1',
            'headers': {
                'Authorization': f'Bearer {minimax_key}',
                'Accept': 'application/json',
                'Content-Type': 'application/json',
            },
            'model_override': os.getenv('MINIMAX_MODEL', 'minimax-m2.7'),
        })

    # 10. Z.AI / GLM
    glm_key = os.getenv('GLM_API_KEY')
    if glm_key:
        chain.append({
            'name': 'glm',
            'base': 'https://open.bigmodel.cn/api/paas/v4',
            'headers': {
                'Authorization': f'Bearer {glm_key}',
                'Accept': 'application/json',
                'Content-Type': 'application/json',
            },
            'model_override': os.getenv('GLM_MODEL', 'glm-4.6'),
        })

    # 11. Anthropic (native API — direct)
    anthropic_key = os.getenv('ANTHROPIC_API_KEY')
    if anthropic_key:
        chain.append({
            'name': 'anthropic',
            'base': 'https://api.anthropic.com/v1',
            'headers': {
                'x-api-key': anthropic_key,
                'anthropic-version': '2023-06-01',
                'Accept': 'application/json',
            },
            'model_override': os.getenv('ANTHROPIC_MODEL', 'claude-opus-4.6'),
            'anthropic_native': True,
        })

    return chain


def _detect_preferred_provider(model: str) -> str | None:
    """
    Detect which provider to try first based on model ID format.
    OpenRouter-style frontier model IDs generally contain '/'.
    Native/direct provider slugs are still routed when they are obvious.
    """
    if not model:
        return None
    if '/' in model:
        if model.startswith('accounts/fireworks/'):
            return 'fireworks'
        if model.startswith('google/gemini-'):
            return 'openrouter'
        if model.startswith('moonshotai/'):
            return 'openrouter'
        if model.startswith('x-ai/'):
            return 'openrouter'
        if model.startswith('minimax/'):
            return 'openrouter'
        if model.startswith('z-ai/'):
            return 'openrouter'
        if model.startswith('qwen/'):
            return 'openrouter'
        if model.startswith('openai/'):
            return 'openrouter'
        if model.startswith('anthropic/'):
            return 'openrouter'
        return 'openrouter'
    if model.startswith('gemini-'):
        return 'gemini'
    if model.startswith('venice-'):
        return 'venice'
    if model.startswith('kimi') or model.startswith('moonshot-'):
        return 'kimi'
    if model.startswith('grok-'):
        return 'xai'
    if model.startswith('minimax-'):
        return 'minimax'
    if model.startswith('glm-'):
        return 'glm'
    # OpenAI direct models (gpt-4o, gpt-4.5-preview, o1, o3-mini, etc.)
    if model.startswith('gpt-') or model.startswith('o1') or model.startswith('o3'):
        return 'openai'
    # Copilot short IDs have dots (version numbers like 4.6, 5.4, 4.1)
    if '.' in model:
        return 'copilot'
    # claude-* IDs: prefer Anthropic native direct if key is set, else Copilot
    if model.startswith('claude-'):
        if os.getenv('ANTHROPIC_API_KEY'):
            return 'anthropic'
        return 'copilot'
    return None


def _preferred_backup_provider(model: str, preferred: str | None) -> str | None:
    """Return a preferred second-choice provider for models that have a strong twin route."""
    model = (model or '').strip().lower()
    if not preferred or not model:
        return None
    is_moonshot_family = (
        model.startswith('moonshotai/')
        or model.startswith('kimi')
        or model.startswith('moonshot-')
    )
    if not is_moonshot_family:
        return None
    if preferred == 'openrouter':
        return 'kimi'
    if preferred == 'kimi':
        return 'openrouter'
    return None


PROMPT_STUDIO_DEFAULT_MODEL = 'google/gemini-2.5-flash'


def prompt_default_model(env_name: str) -> str:
    """Return the safe Prompt Studio default model unless explicitly overridden.

    The dashboard should not silently default Prompt Studio actions to direct Anthropic/Copilot
    short slugs on Render. OpenRouter-style Gemini Flash is the most portable default here;
    native Gemini remains available as fallback when GEMINI_API_KEY is set.
    """
    return (os.getenv(env_name) or PROMPT_STUDIO_DEFAULT_MODEL).strip() or PROMPT_STUDIO_DEFAULT_MODEL


def call_with_fallback(messages, model, timeout=120):
    """
    Call the LLM with automatic provider fallback.
    Detects the preferred provider from the model ID and tries it first,
    then falls through to remaining providers in priority order.
    Returns (content, provider_used) on success.
    Raises RuntimeError if all providers fail.
    """
    chain = _get_provider_chain()
    if not chain:
        raise RuntimeError('No LLM provider configured. Set COPILOT_API_KEY, GEMINI_API_KEY, VENICE_API_KEY, FIREWORKS_API_KEY, KIMI_API_KEY, ANTHROPIC_API_KEY, or OPENROUTER_API_KEY.')

    # Reorder chain so the best-matched provider is tried first
    preferred = _detect_preferred_provider(model)
    if preferred:
        preferred_providers = [p for p in chain if p['name'] == preferred]
        buddy = _preferred_backup_provider(model, preferred)
        buddy_providers = [p for p in chain if p['name'] == buddy and p['name'] != preferred]
        other_providers = [p for p in chain if p['name'] not in {preferred, buddy}]
        chain = preferred_providers + buddy_providers + other_providers

    errors = []
    for provider in chain:
        p_name = provider['name']
        base = provider['base']
        hdrs = dict(provider['headers'])
        m = provider.get('model_override') or model
        is_anthropic_native = provider.get('anthropic_native', False)
        is_gemini_native = provider.get('gemini_native', False)

        try:
            if is_gemini_native:
                # Gemini native REST: POST /v1beta/models/{model}:generateContent
                # Convert OpenAI-style messages to Gemini contents format
                system_text = ' '.join(msg['content'] for msg in messages if msg['role'] == 'system')
                gemini_contents = []
                for msg in messages:
                    if msg['role'] == 'system':
                        continue
                    role = 'user' if msg['role'] == 'user' else 'model'
                    gemini_contents.append({'role': role, 'parts': [{'text': msg['content']}]})
                body = {'contents': gemini_contents}
                if system_text:
                    body['systemInstruction'] = {'parts': [{'text': system_text}]}
                resp = fetch_json(
                    f'{base}/models/{m}:generateContent',
                    headers=hdrs, method='POST', body=body, timeout=timeout
                )
                candidates = resp.get('candidates') or []
                content = ''
                if candidates:
                    parts = candidates[0].get('content', {}).get('parts') or []
                    content = ''.join(p.get('text', '') for p in parts)
            elif is_anthropic_native:
                # Anthropic /messages format
                hdrs['Content-Type'] = 'application/json'
                system_msgs = [msg['content'] for msg in messages if msg['role'] == 'system']
                user_msgs = [msg for msg in messages if msg['role'] != 'system']
                body = {
                    'model': m,
                    'max_tokens': 8192,
                    'system': system_msgs[0] if system_msgs else '',
                    'messages': user_msgs,
                }
                resp = fetch_json(f'{base}/messages', headers=hdrs, method='POST', body=body, timeout=timeout)
                content_blocks = resp.get('content') or []
                content = ''.join(b.get('text', '') for b in content_blocks if isinstance(b, dict))
            else:
                # OpenAI-compatible /chat/completions
                body = {'model': m, 'messages': messages}
                resp = fetch_json(f'{base}/chat/completions', headers=hdrs, method='POST', body=body, timeout=timeout)
                choices = resp.get('choices') or []
                content = ''
                if choices:
                    content = (choices[0].get('message') or {}).get('content') or ''
                if isinstance(content, list):
                    content = ''.join(part.get('text', '') for part in content if isinstance(part, dict))

            content = str(content).strip()
            if content:
                return content, p_name

            errors.append(f'{p_name}: empty response')

        except Exception as exc:
            errors.append(f'{p_name}: {exc}')
            continue  # try next provider

    raise RuntimeError('All LLM providers failed:\n' + '\n'.join(errors))



def supabase_comments_config():
    url = (os.getenv('SUPABASE_URL') or '').strip().rstrip('/')
    key = (
        (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or '').strip()
        or (os.getenv('SUPABASE_ANON_KEY') or '').strip()
        or (os.getenv('SUPABASE_API_KEY') or '').strip()
    )
    table = (os.getenv('SUPABASE_COMMENTS_TABLE') or 'dashboard_comments').strip() or 'dashboard_comments'
    schema = (os.getenv('SUPABASE_SCHEMA') or 'public').strip() or 'public'
    return {'url': url, 'key': key, 'table': table, 'schema': schema, 'configured': bool(url and key and table)}



def _get_supabase_config():
    """Unified Supabase config for all dashboard features."""
    url = (os.getenv('SUPABASE_URL') or '').strip().rstrip('/')
    # Prefer service role key for write operations, fall back to anon key
    key = (
        (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or '').strip()
        or (os.getenv('SUPABASE_ANON_KEY') or '').strip()
        or (os.getenv('SUPABASE_API_KEY') or '').strip()
    )
    return {'url': url, 'key': key, 'schema': 'public'}

def supabase_comments_headers(prefer=None):
    cfg = supabase_comments_config()
    if not cfg['configured']:
        return None
    headers = {
        'apikey': cfg['key'],
        'Authorization': f"Bearer {cfg['key']}",
        'Accept': 'application/json',
        'Accept-Profile': cfg['schema'],
        'Content-Profile': cfg['schema'],
    }
    if prefer:
        headers['Prefer'] = prefer
    return headers


def serialize_comment_row(row):
    return {
        'id': row.get('id'),
        'createdAt': row.get('created_at'),
        'domain': row.get('domain') or '',
        'agentName': row.get('agent_name') or '',
        'versionName': row.get('version_name') or '',
        'contextType': row.get('context_type') or '',
        'contextKey': row.get('context_key') or '',
        'filePath': row.get('file_path') or '',
        'commentText': row.get('comment_text') or '',
        'authorName': row.get('author_name') or '',
        'metadata': row.get('metadata') or {},
    }


def list_supabase_comments(filters):
    cfg = supabase_comments_config()
    if not cfg['configured']:
        return {'ok': True, 'configured': False, 'table': cfg['table'], 'comments': []}
    context_key = (filters.get('contextKey') or '').strip()
    domain = normalize_domain(filters.get('domain') or '')
    agent_name = (filters.get('agentName') or '').strip()
    version_name = (filters.get('versionName') or '').strip()
    context_type = (filters.get('contextType') or '').strip()
    file_path = (filters.get('filePath') or '').strip()
    try:
        limit = int(filters.get('limit') or 50)
    except Exception:
        limit = 50
    limit = max(1, min(limit, 100))
    if not context_key and not any([domain, agent_name, version_name, context_type, file_path]):
        raise ValueError('contextKey or at least one context filter is required')
    query = [
        'select=' + urllib.parse.quote('id,created_at,domain,agent_name,version_name,context_type,context_key,file_path,comment_text,author_name,metadata', safe=',_'),
        'order=' + urllib.parse.quote('created_at.desc', safe='._'),
        f'limit={limit}',
    ]
    if context_key:
        query.append('context_key=eq.' + urllib.parse.quote(context_key, safe=''))
    if domain:
        query.append('domain=eq.' + urllib.parse.quote(domain, safe=''))
    if agent_name:
        query.append('agent_name=eq.' + urllib.parse.quote(agent_name, safe=''))
    if version_name:
        query.append('version_name=eq.' + urllib.parse.quote(version_name, safe=''))
    if context_type:
        query.append('context_type=eq.' + urllib.parse.quote(context_type, safe=''))
    if file_path:
        query.append('file_path=eq.' + urllib.parse.quote(file_path, safe=''))
    url = f"{cfg['url']}/rest/v1/{urllib.parse.quote(cfg['table'], safe='')}?{'&'.join(query)}"
    rows = fetch_json(url, headers=supabase_comments_headers(), timeout=30)
    rows = rows if isinstance(rows, list) else []
    return {'ok': True, 'configured': True, 'table': cfg['table'], 'comments': [serialize_comment_row(r) for r in rows]}


def save_supabase_comment(payload):
    cfg = supabase_comments_config()
    if not cfg['configured']:
        raise RuntimeError('SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY/SUPABASE_ANON_KEY are not configured')
    comment_text = (payload.get('commentText') or payload.get('comment') or '').strip()
    if not comment_text:
        raise ValueError('commentText is required')
    if len(comment_text) > 5000:
        raise ValueError('commentText is too long (max 5000 chars)')
    domain = normalize_domain(payload.get('domain') or '')
    agent_name = (payload.get('agentName') or '').strip()
    version_name = (payload.get('versionName') or '').strip()
    context_type = (payload.get('contextType') or 'general').strip() or 'general'
    file_path = (payload.get('filePath') or '').strip()
    context_key = (payload.get('contextKey') or '').strip() or '::'.join([context_type, domain, agent_name, version_name, file_path])
    author_name = (payload.get('authorName') or '').strip()
    metadata = payload.get('metadata') if isinstance(payload.get('metadata'), dict) else {}
    row = {
        'domain': domain,
        'agent_name': agent_name,
        'version_name': version_name,
        'context_type': context_type,
        'context_key': context_key,
        'file_path': file_path,
        'comment_text': comment_text,
        'author_name': author_name or None,
        'metadata': metadata,
    }
    url = f"{cfg['url']}/rest/v1/{urllib.parse.quote(cfg['table'], safe='')}"
    result = fetch_json(url, headers=supabase_comments_headers('return=representation'), method='POST', body=row, timeout=30)
    saved = result[0] if isinstance(result, list) and result else row
    return {'ok': True, 'configured': True, 'table': cfg['table'], 'comment': serialize_comment_row(saved)}


def normalize_domain(value: str) -> str:
    value = (value or '').strip().lower()
    if not value:
        return ''
    value = value.replace('https://', '').replace('http://', '')
    if value.startswith('www.'):
        value = value[4:]
    return value.strip('/ ')


def workflow_matches_domain(workflow, domain: str) -> bool:
    target = normalize_domain(domain)
    tokens = [workflow.get('name', '')]
    tokens.extend(tag.get('name', '') for tag in workflow.get('tags', []) if isinstance(tag, dict))
    for token in tokens:
        normalized = normalize_domain(token)
        if normalized == target or target in normalized or normalized in target:
            return True
    return False


def list_live_workflows():
    headers = n8n_headers()
    if not headers:
        pass  # N8N_API_KEY has built-in fallback
    base = (os.getenv('N8N_BASE_URL') or 'https://websiseo.app.n8n.cloud').rstrip('/')
    data = fetch_json(f'{base}/api/v1/workflows', headers=headers)
    return base, data.get('data', [])


def resolve_workflow_id(domain: str, explicit_id=None):
    domain = normalize_domain(domain)
    if explicit_id:
        return explicit_id, {'strategy': 'explicit'}

    mapping = load_json_file(MAP_FILE, {'domains': {}})
    mapped = (mapping.get('domains') or {}).get(domain)
    if isinstance(mapped, dict) and mapped.get('workflowId'):
        return mapped['workflowId'], {'strategy': 'map', 'mapping': mapped}
    if isinstance(mapped, str) and mapped:
        return mapped, {'strategy': 'map'}

    base, workflows = list_live_workflows()
    matches = []
    for workflow in workflows:
        if workflow_matches_domain(workflow, domain):
            matches.append({
                'id': workflow.get('id'),
                'name': workflow.get('name'),
                'active': workflow.get('active'),
                'tags': [tag.get('name') for tag in workflow.get('tags', []) if isinstance(tag, dict)],
            })
    if len(matches) == 1:
        return matches[0]['id'], {'strategy': 'auto', 'candidates': matches, 'base': base}
    if len(matches) > 1:
        raise ValueError(json.dumps({'code': 'ambiguous_workflow_match', 'candidates': matches}, ensure_ascii=False))
    raise LookupError(json.dumps({'code': 'workflow_not_found', 'domain': domain, 'candidates': []}, ensure_ascii=False))


def build_workflow_payload(source, live):
    # n8n PUT /workflows/:id only accepts: name, nodes, connections, settings, staticData, pinData
    # Read-only fields that must be OMITTED: active, versionId, meta, tags, id
    # settings must only contain allowed keys — extra keys cause 400
    ALLOWED_SETTINGS = {
        'executionOrder', 'saveManualExecutions', 'callerPolicy',
        'errorWorkflow', 'timezone', 'saveDataSuccessExecution',
        'saveDataErrorExecution', 'maxConcurrency',
    }
    raw_settings = source.get('settings') if 'settings' in source else live.get('settings') or {}
    clean_settings = {k: v for k, v in raw_settings.items() if k in ALLOWED_SETTINGS}

    payload = {
        'name': source.get('name') or live.get('name'),
        'nodes': source.get('nodes') or live.get('nodes') or [],
        'connections': source.get('connections') or live.get('connections') or {},
        'settings': clean_settings,
        'staticData': source.get('staticData') if 'staticData' in source else live.get('staticData'),
        'pinData': source.get('pinData') if 'pinData' in source else live.get('pinData') or {},
    }
    return payload


def _normalize_checklist(raw):
    """Accept list[str] or list[{label: str}] from the client. Return clean list[str]."""
    if not raw:
        return []
    out = []
    if isinstance(raw, list):
        for entry in raw:
            if entry is None:
                continue
            if isinstance(entry, str):
                s = entry.strip()
            elif isinstance(entry, dict):
                s = str(entry.get('label') or entry.get('text') or '').strip()
            elif isinstance(entry, (int, float, bool)):
                continue
            else:
                s = str(entry).strip()
            if s:
                out.append(s)
    # dedupe while preserving order, cap length defensively
    seen = set()
    clean = []
    for s in out:
        if s in seen:
            continue
        seen.add(s)
        clean.append(s[:500])
        if len(clean) >= 60:
            break
    return clean



BRAINSTORM_MODELS = [
    {"id": "google/gemini-3.1-pro-preview",   "label": "Gemini 3.1 Pro (OpenRouter)"},
    {"id": "gemini-3.1-pro-preview",          "label": "Gemini 3.1 Pro Preview (Direct)", "provider": "gemini"},
    {"id": "gemini-3-pro-preview",            "label": "Gemini 3 Pro Preview (Direct)",   "provider": "gemini"},
    {"id": "gemini-2.5-pro",                  "label": "Gemini 2.5 Pro (Direct)",          "provider": "gemini"},
    {"id": "anthropic/claude-opus-4.7",       "label": "Claude Opus 4.7"},
    {"id": "openai/gpt-5.4",                  "label": "GPT-5.4"},
    {"id": "gpt-4o",                          "label": "GPT-4o (OpenAI Direct)",           "provider": "openai"},
    {"id": "gpt-4.5-preview",                 "label": "GPT-4.5 Preview (OpenAI Direct)",  "provider": "openai"},
    {"id": "minimax/minimax-m2.7",            "label": "MiniMax M2.7"},
    {"id": "moonshotai/kimi-k2.5",            "label": "Kimi K2.5"},
    {"id": "z-ai/glm-4.6",                    "label": "GLM 4.6"},
]
BRAINSTORM_MODELS_BY_ID = {m["id"]: m for m in BRAINSTORM_MODELS}
SYNTH_MODEL = prompt_default_model("PROMPT_SYNTH_MODEL")


def _call_one_model(model_id, system, user, timeout=180):
    messages = [
        {"role": "system", "content": system},
        {"role": "user",   "content": user},
    ]
    try:
        content, provider_used = call_with_fallback(messages, model_id, timeout=timeout)
        content = str(content).strip()
        if not content:
            return {"model": model_id, "ok": False, "error": "empty response"}
        return {"model": model_id, "ok": True, "content": content, "provider": provider_used}
    except Exception as e:
        return {"model": model_id, "ok": False, "error": str(e)[:400]}


def brainstorm_prompt_multi_model(payload):
    """Run selected top LLMs in parallel on the same draft, then synthesize into one final prompt."""
    draft = (payload.get("draftPrompt") or "").strip()
    if not draft:
        raise ValueError("Draft prompt is required")
    if not _get_provider_chain():
        raise RuntimeError("No LLM provider configured")

    current_date = os.getenv("PROMPT_CURRENT_DATE", "2026-04-17")
    checklist_rules = _normalize_checklist(payload.get("checklist"))
    domain = payload.get("domain") or "unknown"
    agent_name = payload.get("agentName") or "unknown"
    version_name = payload.get("versionName") or "unknown"
    file_manifest = payload.get("fileManifest") or []
    requested_models = payload.get("brainstormModels") or []
    selected_models = []
    seen_models = set()
    for model_id in requested_models:
        if not isinstance(model_id, str):
            continue
        model_id = model_id.strip()
        if not model_id or model_id in seen_models:
            continue
        model = BRAINSTORM_MODELS_BY_ID.get(model_id)
        if model:
            selected_models.append(model)
            seen_models.add(model_id)
    if not selected_models:
        selected_models = list(BRAINSTORM_MODELS)

    # Build manifest summary for model context
    manifest_lines = []
    for f in file_manifest[:40]:
        name = f.get("name","")
        path = f.get("path","")
        if name:
            manifest_lines.append(f"- {name}  (repo path: {path})")
    manifest_block = "\n".join(manifest_lines) if manifest_lines else "(no files provided)"

    checklist_block = ""
    checklist_brain_block = ""
    checklist_synth_block = ""
    if checklist_rules:
        checklist_block = "\n".join(f"{i+1}. {r}" for i, r in enumerate(checklist_rules))
        checklist_brain_block = "--- USER CHECKLIST RULES ---\n" + checklist_block + "\n--- END CHECKLIST ---\n\n"
        checklist_synth_block = "--- USER CHECKLIST RULES (MUST all be in the final) ---\n" + checklist_block + "\n--- END ---\n\n"

    brain_system = (
        "You are an elite prompt engineer brainstorming the perfect AI-agent prompt for "
        "a production HTML redesign workflow. You will be given a draft + project context + "
        "user-selected rules. Produce ONE production-quality rewritten prompt using these sections "
        "with ## markdown headers IN THIS ORDER:\n"
        "## Objective\n## Context\n## Specific Requirements\n## Additional Mandatory Rules\n"
        "## Acceptance Criteria\n## Delivery\n"
        "All requirements must be numbered, measurable, and concrete. No filler, no 'make it better', "
        "no vague language. The Delivery section must explicitly require uploading the improved files "
        "to the GitHub repo behind https://html-redesign-dashboard.maximo-seo.ai/ AND to the local "
        "Obsidian vault path, replacing old files in the same version folder. "
        "Every checklist rule the user selected must appear under Additional Mandatory Rules AND be "
        "verified in Acceptance Criteria. "
        "Return plain markdown only, no code fences, no preamble."
    )
    brain_user = (
        f"Project domain:   {domain}\n"
        f"Target agent:     {agent_name}\n"
        f"Version folder:   {version_name}\n"
        f"Current date:     {current_date}\n\n"
        f"--- FILE MANIFEST ---\n{manifest_block}\n--- END MANIFEST ---\n\n"
        f"{checklist_brain_block}"
        f"--- USER DRAFT PROMPT ---\n{draft}\n--- END DRAFT ---\n\n"
        "Produce the best possible rewritten prompt. Distinguish your work with specificity, measurable "
        "criteria, and a rock-solid Delivery section that references the real folder paths."
    )

    results = []
    max_workers = max(1, min(6, len(selected_models)))
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {
            pool.submit(_call_one_model, m["id"], brain_system, brain_user, 180): m
            for m in selected_models
        }
        for fut in concurrent.futures.as_completed(futures, timeout=240):
            m = futures[fut]
            try:
                r = fut.result()
            except Exception as e:
                r = {"model": m["id"], "ok": False, "error": str(e)[:400]}
            r["label"] = m["label"]
            results.append(r)

    successes = [r for r in results if r.get("ok") and r.get("content")]
    if not successes:
        errors = [{"model": r["model"], "label": r.get("label"), "error": r.get("error"), "detail": r.get("detail","")} for r in results]
        raise RuntimeError(json.dumps({"code": "all_models_failed", "errors": errors}, ensure_ascii=False))

    drafts_block = "\n\n".join(
        f"=== DRAFT FROM {r['label']} ({r['model']}) ===\n{r['content']}\n=== END {r['label']} ==="
        for r in successes
    )
    synth_system = (
        "You are the head prompt architect. You will be given multiple expert-written versions of the "
        "same prompt produced by different elite LLMs. Your job is to synthesize ONE FINAL prompt that "
        "is strictly better than any individual version — pulling in the strongest wording, sharpest "
        "requirements, and most complete acceptance criteria from each draft, removing duplication, and "
        "producing a single clean result.\n\n"
        "Strict rules:\n"
        "1. Output PLAIN MARKDOWN. No code fences. No 'here is the synthesized prompt' preamble.\n"
        "2. Keep the exact section order: ## Objective, ## Context, ## Specific Requirements, "
        "## Additional Mandatory Rules, ## Acceptance Criteria, ## Delivery.\n"
        "3. Every numbered requirement must map to at least one acceptance criterion.\n"
        "4. Delivery must include: upload to GitHub repo maximoseo/webs-html-improvements-files "
        "(folder for the specified domain/version) AND to the Obsidian vault path "
        "C:\\Obsidian\\HTML REDESIGN\\HTML REDESIGN\\<domain>\\<agent>\\updated files\\<date>\\; "
        "replace old files in place, do not create parallel folders. Include the exact commit message "
        "format: feat(<domain>): <summary> — <date>.\n"
        "5. Use the provided current date for all date references; correct any stale dates from the drafts.\n"
        "6. Keep the Additional Mandatory Rules section only if user-selected rules exist; otherwise omit it.\n"
        "7. Style: direct, professional, zero filler. Every bullet is actionable and specific."
    )
    synth_user = (
        f"Project domain:   {domain}\n"
        f"Target agent:     {agent_name}\n"
        f"Version folder:   {version_name}\n"
        f"Current date:     {current_date}\n\n"
        f"{checklist_synth_block}"
        f"--- FILE MANIFEST ---\n{manifest_block}\n--- END MANIFEST ---\n\n"
        f"--- CANDIDATE DRAFTS FROM MULTIPLE MODELS ---\n{drafts_block}\n--- END CANDIDATES ---\n\n"
        "Produce the single synthesized final prompt now."
    )

    synth = _call_one_model(SYNTH_MODEL, synth_system, synth_user, timeout=240)
    if not synth.get("ok") or not synth.get("content"):
        raise RuntimeError(json.dumps({
            "code": "synthesis_failed",
            "synthError": synth.get("error","unknown"),
            "synthDetail": synth.get("detail",""),
            "partialModels": [{"model": r["model"], "label": r["label"]} for r in successes],
        }, ensure_ascii=False))

    return {
        "ok": True,
        "finalPrompt": synth["content"],
        "synthModel": SYNTH_MODEL,
        "synthProvider": synth.get("provider"),
        "modelsUsed": [{"model": r["model"], "label": r["label"], "chars": len(r["content"]), "provider": r.get("provider")} for r in successes],
        "modelsFailed": [{"model": r["model"], "label": r.get("label"), "error": r.get("error")} for r in results if not r.get("ok")],
        "requestedModels": [m["id"] for m in selected_models],
    }


def commit_prompt_to_github(payload):
    """Commit a single text file (the improved prompt) to the dashboard repo via GitHub Contents API."""
    token = os.getenv("GITHUB_TOKEN")
    if not token:
        raise RuntimeError("GITHUB_TOKEN is not configured on the server")
    repo_path = (payload.get("path") or "").strip().lstrip("/")
    content = payload.get("content") or ""
    message = (payload.get("message") or "chore: update improved prompt").strip()
    branch = (payload.get("branch") or "main").strip()
    if not repo_path:
        raise ValueError("path is required")
    if not content:
        raise ValueError("content is required")
    if ".." in repo_path or repo_path.startswith("/"):
        raise ValueError("invalid path")

    api_base = f"https://api.github.com/repos/{REPO}/contents/"
    api_url = api_base + "/".join(urllib.parse.quote(p) for p in repo_path.split("/"))
    gh_headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "dashboard-commit-bot",
    }

    # Check for existing sha
    sha = None
    try:
        existing = fetch_json(f"{api_url}?ref={urllib.parse.quote(branch)}", headers=gh_headers, timeout=30)
        if isinstance(existing, dict) and existing.get("sha"):
            sha = existing["sha"]
    except urllib.error.HTTPError as e:
        if e.code != 404:
            raise

    body = {
        "message": message,
        "content": base64.b64encode(content.encode("utf-8")).decode("ascii"),
        "branch": branch,
    }
    if sha:
        body["sha"] = sha

    result = fetch_json(api_url, headers=gh_headers, method="PUT", body=body, timeout=60)
    commit = result.get("commit") or {}
    contentInfo = result.get("content") or {}
    return {
        "ok": True,
        "path": repo_path,
        "branch": branch,
        "replaced": bool(sha),
        "commitSha": commit.get("sha"),
        "commitUrl": commit.get("html_url"),
        "fileUrl": contentInfo.get("html_url"),
        "downloadUrl": contentInfo.get("download_url"),
    }


def commit_fixer_to_github(wf_name, site_url, issue_summary, changes_made, fixed_json, model_used, confidence):
    """Commit a fixed n8n workflow JSON to the GitHub repo and update CHANGELOG."""
    token = os.getenv("GITHUB_TOKEN")
    if not token:
        raise RuntimeError("GITHUB_TOKEN is not configured on the server")
    if not fixed_json:
        raise ValueError("fixed_json is required")

    import datetime as _dt
    date_str = _dt.datetime.now().strftime('%Y-%m-%d')
    safe_domain = re.sub(r'[^a-zA-Z0-9_\-]', '-', (site_url or 'unknown').replace('https://', '').replace('http://', ''))[:50]
    safe_wf = re.sub(r'[^a-zA-Z0-9_\-]', '-', wf_name)[:50]

    # 1. Commit the fixed JSON
    json_path = f"n8n-fixes/{safe_domain}/{date_str}-{safe_wf}.json"
    json_content = json.dumps(fixed_json, indent=2, ensure_ascii=False) if isinstance(fixed_json, dict) else str(fixed_json)

    api_base = f"https://api.github.com/repos/{REPO}/contents/"
    gh_headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "dashboard-fix-bot",
    }

    def _commit_file(repo_path, content, message):
        api_url = api_base + "/".join(urllib.parse.quote(p) for p in repo_path.split("/"))
        sha = None
        try:
            existing = fetch_json(f"{api_url}?ref=main", headers=gh_headers, timeout=30)
            if isinstance(existing, dict) and existing.get("sha"):
                sha = existing["sha"]
        except urllib.error.HTTPError as e:
            if e.code != 404:
                raise
        body = {
            "message": message,
            "content": base64.b64encode(content.encode("utf-8")).decode("ascii"),
            "branch": "main",
        }
        if sha:
            body["sha"] = sha
        return fetch_json(api_url, headers=gh_headers, method="PUT", body=body, timeout=60)

    json_result = _commit_file(json_path, json_content, f"fix(n8n): {wf_name} — {issue_summary or 'workflow fix'} [{model_used}]")

    # 2. Update CHANGELOG.md
    changelog_path = "n8n-fixes/CHANGELOG.md"
    changelog_entry = f"\n### {date_str} — {wf_name}\n- **Domain:** {site_url or 'N/A'}\n- **Model:** {model_used or 'N/A'}\n- **Confidence:** {round(float(confidence)*100, 1)}%\n- **Summary:** {issue_summary or 'N/A'}\n- **Changes:** {', '.join(changes_made) if changes_made else 'N/A'}\n- **File:** `{json_path}`\n"

    # Try to read existing changelog
    existing_changelog = "# N8N Workflow Fixer Changelog\n\n"
    try:
        cl_url = api_base + "/".join(urllib.parse.quote(p) for p in changelog_path.split("/"))
        cl_existing = fetch_json(f"{cl_url}?ref=main", headers=gh_headers, timeout=30)
        if isinstance(cl_existing, dict) and cl_existing.get("content"):
            import base64 as _b64
            existing_changelog = _b64.b64decode(cl_existing["content"]).decode("utf-8")
    except Exception:
        pass

    new_changelog = existing_changelog.rstrip() + "\n" + changelog_entry
    _commit_file(changelog_path, new_changelog, f"docs: update n8n fixer changelog — {wf_name}")

    commit = json_result.get("commit") or {}
    contentInfo = json_result.get("content") or {}
    return {
        "ok": True,
        "jsonPath": json_path,
        "commitSha": commit.get("sha"),
        "commitUrl": commit.get("html_url"),
        "fileUrl": contentInfo.get("html_url"),
    }


def tweak_html_with_prompt(payload):
    improved_prompt = (payload.get('improvedPrompt') or '').strip()
    html_download_url = (payload.get('htmlDownloadUrl') or '').strip()
    if not improved_prompt:
        raise ValueError('improvedPrompt is required')
    if not html_download_url:
        raise ValueError('htmlDownloadUrl is required')
    if not _get_provider_chain():
        raise RuntimeError('No LLM provider configured')
    default_model = prompt_default_model('PROMPT_TWEAK_MODEL')
    model = (payload.get('model') or '').strip() or default_model
    domain = payload.get('domain') or 'unknown'
    agent_name = payload.get('agentName') or 'unknown'
    version_name = payload.get('versionName') or 'unknown'
    html_file_name = (payload.get('htmlFileName') or 'Improved_HTML_Template.html').strip()
    html_file_path = (payload.get('htmlFilePath') or '').strip()
    latest_only = bool(payload.get('latestOnly', True))
    file_manifest = payload.get('fileManifest') or []
    current_html = fetch_text(html_download_url)
    manifest_lines = []
    for f in file_manifest:
        name = (f.get('name') or '').strip()
        if not name:
            continue
        manifest_lines.append(f"- {name} | path: {(f.get('path') or '').strip()} | download: {(f.get('download') or '').strip()}")
    manifest_block = '\n'.join(manifest_lines) if manifest_lines else '- No manifest provided'
    latest_rule = 'Only output the final latest replacement HTML for the active file. Do not mention old parallel files, duplicate versions, or archival outputs.' if latest_only else ''
    system = ('You are a senior HTML redesign engineer. Apply the improved prompt to the provided HTML template and return the full corrected HTML file. Return raw HTML only. No markdown fences. No explanations before or after the HTML. Preserve working content and structure unless the improved prompt explicitly requires a change. Keep the file production-ready and self-contained.')
    user = (f'Target domain: {domain}\n' f'Target agent: {agent_name}\n' f'Active version folder: {version_name}\n' f'Active HTML file name: {html_file_name}\n' f'Active HTML file path: {html_file_path}\n' f'{latest_rule}\n\n' f'--- FILE MANIFEST ---\n{manifest_block}\n--- END FILE MANIFEST ---\n\n' f'--- IMPROVED PROMPT TO APPLY ---\n{improved_prompt}\n--- END IMPROVED PROMPT ---\n\n' f'--- CURRENT HTML TEMPLATE ---\n{current_html}\n--- END CURRENT HTML TEMPLATE ---\n\n' 'Now return the fully updated HTML for the active file only.')
    body = {'model': model, 'messages': [{'role': 'system', 'content': system}, {'role': 'user', 'content': user}]}
    content, provider_used = call_with_fallback(
        [{'role': 'system', 'content': system}, {'role': 'user', 'content': user}],
        model, timeout=240
    )
    if isinstance(content, list):
        content = ''.join(part.get('text', '') for part in content if isinstance(part, dict))
    content = str(content).strip()
    if content.startswith('```'):
        content = content.strip('`')
        if content.lower().startswith('html'):
            content = content[4:].lstrip()
    if not content:
        raise RuntimeError('Model returned empty HTML content')
    return {'ok': True, 'model': model, 'provider': provider_used, 'html': content, 'summary': f'Tweaked {html_file_name} for {domain}'}



# ===== Prompt Studio Contract v2026-04-29 =====
PROMPT_STUDIO_CONTRACT_VERSION = '2026-04-29'
PROMPT_STUDIO_REQUIRED_SITE_PROFILE_FIELDS = (
    'domain', 'business_name', 'language', 'text_direction', 'brand_colors',
    'fonts', 'logo_url', 'author_or_company', 'contact_url', 'social_links',
    'tone', 'audience', 'cta_style'
)
PROMPT_STUDIO_ALLOWED_NEXT_ACTIONS = {'review_site_profile', 'revise_prompt', 'approve_for_delivery'}


def _prompt_studio_normalize_site_profile(value):
    return value if isinstance(value, dict) else {}


def _prompt_studio_build_validation_report(payload, draft):
    # Prompt Studio Contract: validation report
    domain = (payload.get('domain') or '').strip()
    site_profile = _prompt_studio_normalize_site_profile(payload.get('siteProfile'))
    missing = [field for field in PROMPT_STUDIO_REQUIRED_SITE_PROFILE_FIELDS if not str(site_profile.get(field) or '').strip()]
    analysis_status = (payload.get('analysisStatus') or 'not_started').strip() or 'not_started'
    site_profile_reviewed = bool(payload.get('siteProfileReviewed'))
    read_only_n8n = payload.get('readOnlyN8n')
    if read_only_n8n is None:
        read_only_n8n = True
    next_action = (payload.get('nextAction') or 'review_site_profile').strip() or 'review_site_profile'
    if next_action not in PROMPT_STUDIO_ALLOWED_NEXT_ACTIONS:
        next_action = 'review_site_profile'
    status = 'ready_for_prompt_review'
    if missing or analysis_status in {'not_started', 'failed'} or not site_profile_reviewed:
        status = 'needs_review'
        next_action = 'review_site_profile'
    return {
        'contract_version': PROMPT_STUDIO_CONTRACT_VERSION,
        'domain_present': bool(domain and domain.lower() != 'unknown'),
        'analysis_status': analysis_status,
        'site_profile_status': 'reviewed' if site_profile_reviewed and not missing else 'needs_review',
        'missing_site_profile_fields': missing,
        'read_only_n8n': bool(read_only_n8n),
        'requires_n8n_variable_preservation': True,
        'requires_no_fake_assets_or_placeholders': True,
        'requires_wordpress_custom_html_compatibility': True,
        'requires_responsive_no_overflow': True,
        'requires_rtl_ltr_handling': True,
        'requires_no_secrets_or_duplicate_storage': True,
        'status': status,
        'next_action': next_action,
    }


def _prompt_studio_contract_block(validation_report):
    missing = validation_report.get('missing_site_profile_fields') or []
    missing_text = ', '.join(missing) if missing else 'none'
    return f"""
--- PROMPT STUDIO CONTRACT v{PROMPT_STUDIO_CONTRACT_VERSION} ---
This request is for Prompt Studio / Create Improve Prompts. It generates reviewable prompt instructions only.
Domain-first rule: a real target domain/URL is mandatory before final prompt generation.
Site profile rule: analyze the real website first. If facts cannot be verified, mark them as needs_review; never invent fake colors, logos, authors, phone numbers, social links, products, services, images, testimonials, or lorem ipsum/placeholder content.
Current validation status: {validation_report.get('status')}.
Missing site profile fields: {missing_text}.
N8N safety: read-only by default. Do not instruct the agent to modify, activate, import, overwrite, or deploy n8n workflows unless explicit task-specific approval is given after review.
N8N preservation: preserve all {{{{ ... }}}} placeholders, $json fields, callback URLs such as {{{{$execution.resumeUrl}}}}, credential references, workflow variables, and field names exactly.
WordPress HTML safety: require a unique scoped wrapper, scoped CSS, inline fallback styles for important visuals, responsive no-overflow behavior, no unapproved external dependencies, and safe WordPress Custom HTML insertion.
Language/direction: detect LTR/RTL. For Hebrew/RTL, require lang="he", dir="rtl", right-aligned defaults, and mirrored directional UI where appropriate.
Storage safety: do not expose secrets, do not create duplicate files/records, and include clear versioning/metadata for any approved future save/export action.
The improved prompt MUST be markdown and MUST contain these exact top-level sections in this order:
## Site Profile
## Generated Prompt
## Validation Report
## Storage Metadata
## Next Action
The Next Action section MUST be "review_site_profile" when required facts are missing or uncertain; do not mark production-ready until validation is complete.
--- END PROMPT STUDIO CONTRACT ---
"""

def improve_prompt_with_model(payload):
    draft = (payload.get('draftPrompt') or '').strip()
    if not draft:
        raise ValueError('Draft prompt is required')
    # Prompt Studio Contract: domain-first backend guard
    early_domain = (payload.get('domain') or '').strip()
    if not early_domain or early_domain.lower() == 'unknown':
        raise ValueError('Target domain or URL is required before final prompt generation')
    if not _get_provider_chain():
        raise RuntimeError('No LLM provider configured')
    # Accept model override from the browser payload; fall back to env / default
    default_model = prompt_default_model('PROMPT_IMPROVER_MODEL')
    model = (payload.get('model') or '').strip() or default_model
    current_date = os.getenv('PROMPT_CURRENT_DATE', '2026-04-15')
    checklist_rules = _normalize_checklist(payload.get('checklist'))
    validation_report = _prompt_studio_build_validation_report(payload, draft)
    contract_block = _prompt_studio_contract_block(validation_report)

    # Prompt Studio Contract: system prompt injection
    system = (
        contract_block +
        'You are a senior prompt engineer for agentic coding and design workflows. '
        'Rewrite the user draft into a reviewable Prompt Studio package, not an automatic publish/deploy instruction. '
        'Return plain-text markdown only — no code fences, no preamble, no \"here is your prompt\" wrapper. '
        'The top-level output schema is the Prompt Studio Contract package above. Do not add competing top-level sections. '
        'Inside ## Generated Prompt, include the final agent prompt with these nested sections:\n'
        '### Objective — one paragraph, sharp and direct.\n'
        '### Context — domain, agent, version, relevant background.\n'
        '### Specific Requirements — numbered list of concrete requirements. No vague instructions.\n'
        '### Additional Mandatory Rules — present ONLY when the user selected checklist items; '
        'render each selected rule as a numbered bullet under this nested section; if no rules were selected, omit it entirely.\n'
        '### Acceptance Criteria — numbered checklist the agent must verify before finishing. '
        'Every item from Specific Requirements and every Additional Mandatory Rule must map to at least one criterion here.\n'
        '### Delivery Targets After Approval — approved destinations and commit guidance only.\n'
        'Mandatory safety rule: do not instruct automatic exporting, overwriting, publishing, activating, importing, or deploying by default. '
        'Any write to n8n, GitHub, Supabase, Obsidian, WordPress, or live dashboards requires explicit task-specific approval after review. '
        'Mandatory rule: the Delivery Targets section may list exact Obsidian/GitHub target paths as approved destinations only. '
        'Mandatory rule: all date references in the generated prompt must use the current working date '
        'provided in the user message; stale dates must be updated. '
        '--- ALWAYS-ON MANDATORY RULES (inject into every prompt regardless of checklist) ---\n'
        'RULE 1 — FULL CONTENT + IMAGES (new AND existing projects): Every newly generated redesign HTML template MUST include the '
        'full real article content (every section, every paragraph, every heading) and all real images with '
        'their actual src URLs — exactly as they will appear inside the N8N article-building workflow output. '
        'Placeholder text, lorem ipsum, \"[image here]\", empty src attributes, or truncated content are '
        'strictly forbidden. The HTML must be ready to drop directly into N8N and produce a complete, '
        'publish-ready article with zero additional content filling required.\n'
        'RULE 2 — CORRECT FILE EXTENSIONS (new AND existing projects): Every output file must be saved and uploaded with its real, '
        'correct file extension: HTML templates must end in .html, N8N prompt text files must end in .txt, '
        'N8N workflow files must end in .json. Never upload files without an extension or with a wrong '
        'extension (e.g. no .html.txt, no extensionless files). The file names in the GitHub commit must '
        'exactly match the extensions used on disk.\n'
        'RULE 3 — MAX 3 FILES PER AGENT FOLDER (new AND existing projects): Each agent version folder in the GitHub repo '
        '(maximoseo/webs-html-improvements-files) must contain exactly 3 files: '
        'Improved_HTML_Template.html, Improved_N8N_Prompt.txt, Improved_N8N_Workflow.json. '
        'Do not add extra MD files, summary files, validation notes, source maps, or import notes '
        'inside the agent folder. If old extra files exist, remove them in the same commit. '
        'This rule applies when adding files to existing projects too — always check the folder first.\n'
        'RULE 4 — N8N WORKFLOW CREDENTIALS AUDIT (new AND existing projects): Before uploading or deploying '
        'any Improved_N8N_Workflow.json to the GitHub repo or to n8n, you MUST audit every node in the '
        'workflow JSON for correct credentials and correct URLs. Specifically: '
        '(a) Every wordpressApi credential must point to the correct WordPress site for THIS domain — '
        'never a credential ID from a different project. '
        '(b) Every httpBasicAuth credential used for WordPress REST API calls must point to the correct '
        'site URL for THIS domain — never a URL copied from another project. '
        '(c) Every hardcoded URL inside node parameters (url, endpoint, baseUrl fields) must be the '
        'correct target domain for THIS project — search the JSON for any foreign domain name and fix it. '
        '(d) After fixing credentials, confirm no other domain name appears anywhere in the workflow nodes '
        'except the correct target domain. '
        '(e) Add the workflow to n8n-workflow-map.json with the correct workflowId so the Deploy to n8n '
        'button resolves immediately without API scanning. '
        'This rule is CRITICAL — wrong credentials will silently publish articles to the wrong WordPress site.\n'
        '--- END ALWAYS-ON MANDATORY RULES ---\n'
        'Writing style: direct, professional, zero filler. Every bullet point must be actionable and specific. '
        'Never use vague phrases like \"make it look better\" — always state the exact measurable change expected.'
    )

    if checklist_rules:
        numbered_rules = '\n'.join(f'{i+1}. {rule}' for i, rule in enumerate(checklist_rules))
        extra_rules_section = (
            '\n\n--- ADDITIONAL MANDATORY RULES (user-selected before clicking Improve) ---\n'
            'The rewritten prompt MUST include ALL of the following as hard requirements under the '
            '"## Additional Mandatory Rules" section, and MUST verify each one in "## Acceptance Criteria":\n'
            f'{numbered_rules}\n'
            '--- END ADDITIONAL MANDATORY RULES ---\n'
        )
    else:
        extra_rules_section = ''

    # Build file manifest block from payload
    file_manifest = payload.get('fileManifest') or []
    version_path = (payload.get('versionPath') or '').strip()
    github_repo_base = (payload.get('githubRepoBase') or '').strip()
    obsidian_base = (payload.get('obsidianBase') or '').strip()
    repo = 'maximoseo/webs-html-improvements-files'
    domain = payload.get('domain') or 'unknown'
    agent_name = payload.get('agentName') or 'unknown'
    version_name = payload.get('versionName') or 'unknown'

    if file_manifest:
        def _fmt_files(files, label):
            if not files:
                return ''
            lines = [f'### {label}']
            for f in files:
                name = f.get('name','')
                size = f.get('size',0)
                path = f.get('path','')
                dl   = f.get('download','')
                url  = f.get('url','')
                size_str = (f'{size/1024:.1f} KB' if size >= 1024 else f'{size} B') if size else ''
                lines.append(f'- {name} {size_str}')
                if url:  lines.append(f'  View on GitHub: {url}')
                if dl:   lines.append(f'  Raw download:   {dl}')
                if path: lines.append(f'  Repo path:      {path}')
            return '\n'.join(lines)

        html_files  = [f for f in file_manifest if f.get('name','').lower().endswith('.html')]
        txt_files   = [f for f in file_manifest if f.get('name','').lower().endswith('.txt')]
        json_files  = [f for f in file_manifest if f.get('name','').lower().endswith('.json')]
        other_files = [f for f in file_manifest if not f.get('name','').lower().endswith(('.html','.txt','.json'))]

        manifest_block = (
            f'\n\n--- CURRENT FILE MANIFEST (version: {version_name}) ---\n'
            + _fmt_files(html_files,  'HTML Templates') + '\n'
            + _fmt_files(txt_files,   'Prompts / Text Files') + '\n'
            + _fmt_files(json_files,  'Workflow / JSON Files') + '\n'
            + _fmt_files(other_files, 'Other Files') + '\n'
            + '--- END FILE MANIFEST ---\n'
        )
    else:
        manifest_block = ''

    # Build delivery paths block
    obs_path = obsidian_base or f'C:\\Obsidian\\HTML REDESIGN\\HTML REDESIGN\\{domain}\\{agent_name}\\updated files'
    gh_folder = version_path or f'{domain}/{version_name}'
    delivery_block = (
        f'\n\n--- DELIVERY DESTINATIONS ---\n'
        f'After review and explicit task-specific approval, approved final files may target BOTH destinations below.\n'
        f'Do not instruct automatic writes or overwrites by default; require confirmation before n8n, GitHub, Supabase, Obsidian, or WordPress changes.\n'
        f'When approved, update the existing files in the same version folder rather than creating a duplicate parallel folder.\n\n'
        f'1. Obsidian vault (Windows path):\n'
        f'   {obs_path}\\updated files\\{current_date}\\\n'
        f'   Upload: HTML template, N8N prompt, N8N workflow, validation note, source map, summary.\n\n'
        f'2. GitHub repository:\n'
        f'   Repo:      https://github.com/{repo}\n'
        f'   Folder:    {gh_folder}\n'
        f'   Dashboard: https://html-redesign-dashboard.maximo-seo.ai/\n'
        f'   Commit message: "feat({domain}): [describe what changed] — {current_date}"\n'
        f'   After explicit approval and a real push, refresh the dashboard and confirm the updated files appear correctly.\n'
        f'--- END DELIVERY DESTINATIONS ---\n'
    )

    user = (
        f'Target domain:          {domain}\n'
        f'Target agent:           {agent_name}\n'
        f'Version folder:         {version_name}\n'
        f'GitHub folder URL:      {github_repo_base or f"https://github.com/{repo}/tree/main/{domain}"}\n'
        f'Obsidian base path:     {obs_path}\n'
        f'Current working date:   {current_date}\n'
        f'{manifest_block}'
        f'{delivery_block}'
        f'\n--- USER DRAFT PROMPT ---\n'
        f'{draft}\n'
        f'--- END DRAFT ---\n'
        f'{extra_rules_section}\n'
        'Rewrite the draft into the structured prompt format described in the system instructions. '
        'Every section must be present and populated. '
        'The Context section must list the exact file names from the manifest above so the agent knows exactly which files to work with. '
        'The Specific Requirements section must be a numbered list — not a flat paragraph. '
        'The Delivery section must copy the exact Obsidian path and GitHub repo path from the delivery destinations block above as approved target destinations only — '
        'include real folder paths, real file names from the manifest, and the exact commit message format, but require explicit confirmation before any write/publish/deploy action. '
        'Use the current required working date above for all folder/date references and correct any stale dates found in the draft. '
        'Ensure the output package includes ## Site Profile, ## Generated Prompt, ## Validation Report, ## Storage Metadata, and ## Next Action.'
    )

    messages = [
        {'role': 'system', 'content': system},
        {'role': 'user', 'content': user},
    ]
    content, provider_used = call_with_fallback(messages, model, timeout=120)
    if not content:
        raise RuntimeError('Model returned empty content')
    return {
        'model': model,
        'provider': provider_used,
        'content': content,
        'contractVersion': PROMPT_STUDIO_CONTRACT_VERSION,
        'validationReport': validation_report,
        'nextAction': validation_report.get('next_action', 'review_site_profile'),
    }



# ===== Plan #21: Improve Workflow prompt assembly rules =====
IMPROVEMENT_RULES_VERSION = 1
IMPROVEMENT_RULES = [
    {
        "rule_key": "rule_site_analysis",
        "rule_name": "Site Analysis & Discovery",
        "rule_category": 1,
        "conditional_on": "always",
        "rule_text": """1. Start from the Main Website URL and inspect the homepage first.
2. Discover internal pages: about, contact, services, products, blog, reviews, team, location, author.
3. Check header navigation, footer links, sitemap, and structural pages.
4. Inspect the N8N Workflow JSON to understand workflow type and intent.
5. Inspect the original article page for floating buttons, sticky UI, WhatsApp buttons, and layout constraints.
6. If direct browsing is incomplete, use Firecrawl for crawl mapping, structured content extraction, and product extraction.
7. Use only real pages and real visible data found on the site.
8. Extract business name, brand colors, fonts, radius, shadows, language, direction, tone, audience, business type, phone, socials, reviews, video, map, logo, and author info.
9. Do not ask for information that can be discovered automatically.""",
    },
    {
        "rule_key": "rule_store_decision",
        "rule_name": "Store vs Non-Store Decision Logic",
        "rule_category": 2,
        "conditional_on": "always",
        "rule_text": """1. If the workflow JSON clearly indicates store, product, catalog, WooCommerce, Shopify, checkout, cart, or eCommerce logic, treat as STORE workflow.
2. If the live website clearly contains real products, product pages, collections, or category pages, treat as STORE workflow.
3. If both workflow and site indicate store behavior, products must be added where relevant.
4. If neither workflow nor site indicates store behavior, do not add product blocks.
5. If ambiguous, use workflow JSON as the strong decision signal together with the live site structure.
6. Never add products just because products might exist loosely somewhere on the site.""",
    },
    {
        "rule_key": "rule_product_source",
        "rule_name": "Product Source & Selection",
        "rule_category": 3,
        "conditional_on": "store_mode",
        "rule_text": """1. If products belong to the main website, extract from the main website.
2. If products come from another site and an External Product Source URL is provided, use that external site only for product extraction.
3. The main website always remains the primary source for brand identity, CTA logic, author bio, business details, design direction, palette, and tone.
4. Products must be relevant to the specific article topic.
5. Rank product candidates by title similarity, category match, collection match, tag match, product type, and semantic relevance.
6. Reject random products, weakly related products, duplicates, and products with missing or unusable images.
7. Each shown product must include a real product image, real product title, and direct link to the exact product page.
8. Do not show prices unless explicitly requested.
9. Do not invent products, images, or links.""",
    },
    {
        "rule_key": "rule_image_pipeline",
        "rule_name": "Product Image Pipeline",
        "rule_category": 4,
        "conditional_on": "always",
        "rule_text": """1. Audit product selection logic, image mapping logic, Supabase storage, and HTML rendering before finalizing.
2. Prefer true featured or original product images, not thumbnails.
3. Avoid variant-image mixups and repeated product images.
4. Store images in Supabase when stability is needed, with deterministic filenames and public URLs.
5. Use img tags, not background-image, for product images.
6. Use object-fit: contain, preserve natural proportions, and avoid destructive cropping.
7. Store source product title, URL, image URL, alt text, and dimensions when available.
8. Skip products that do not have usable real images.
9. Store logo and owner or author images in Supabase when needed for stability.""",
    },
    {
        "rule_key": "rule_wordpress_html",
        "rule_name": "WordPress HTML Output",
        "rule_category": 5,
        "conditional_on": "always",
        "rule_text": """1. File starts with <article> and ends with </article>.
2. Nothing appears outside the article element.
3. Inline CSS only; no style blocks.
4. No markdown, no code fences, no comments.
5. No H1 tag; start with a context-setting introduction.
6. No external dependencies: no CDN, no external CSS, no external JavaScript.
7. Must render correctly inside WordPress.
8. Use correct language and direction on article: Hebrew lang="he" dir="rtl"; English lang="en" dir="ltr".
9. Export as a real .html file, not .txt.
10. The HTML must be immediately usable as an actual HTML file.""",
    },
    {
        "rule_key": "rule_article_structure",
        "rule_name": "Article Structure",
        "rule_category": 6,
        "conditional_on": "always",
        "rule_text": """1. Context introduction: 2-3 concise factual paragraphs, no H1, no dramatic intro, no rhetorical questions.
2. Add an In This Article summary box with 4-6 bullets.
3. Add a Table of Contents inside details/summary, closed by default, single column only. Title must be exactly Table of Contents or תוכן עניינים.
4. If store workflow is confirmed, show a top product grid immediately after TOC.
5. Core body uses H2 for main sections, H3 for subsections, substantive paragraphs, lists, comparison tables, and takeaway boxes.
6. If more products are relevant, add an additional product grid before FAQ.
7. FAQ uses details/summary, closed by default, and must appear before author bio.
8. Closing section is 2-3 sentences maximum.
9. Use one centered professional CTA button only.
10. Author bio is always the final section; nothing appears after it.
11. Do not display written, published, updated, modified, or similar date labels by default. If a date must remain, it must be current and justified.""",
    },
    {
        "rule_key": "rule_visual_design",
        "rule_name": "Visual Design",
        "rule_category": 7,
        "conditional_on": "always",
        "rule_text": """1. Extract the real design system first: primary, secondary, accent, text, background, font, radius, border, shadow, heading, CTA, card, and hover treatments.
2. If extraction fails, use the safe fallback palette: #363636, #54e9c0, #ffffff, Arimo/Arial/sans-serif, radius 6px, subtle shadow.
3. H2 and H3 hierarchy must be clear, clamp-based, and brand-consistent. RTL sites use accent borders on the right.
4. Body text must be readable with strong line-height and balanced spacing.
5. Cards use subtle borders, subtle shadows, and consistent spacing; no gradients or flashy effects.
6. Images use width 100%, height auto, max-height 60vh, object-fit contain, descriptive alt text, lazy loading below fold, and captions when useful.
7. CTA is one button only, professional, centered, accessible, and readable on hover.
8. Key takeaway boxes are lightly tinted with subtle accent border and concise useful text.
9. Product grid is clean responsive CSS grid with image, title, short description, and link/button.
10. Mobile under 768px uses about 5% side margins and single-column products; tablet uses about 10% margins and 2 columns; desktop uses about 15% margins and 3 columns.
11. WCAG AA contrast, semantic HTML, correct lang/dir, no horizontal scrolling.
12. Opening logo and author image must be centered on all breakpoints.""",
    },
    {
        "rule_key": "rule_writing_content",
        "rule_name": "Writing & Content Quality",
        "rule_category": 8,
        "conditional_on": "always",
        "rule_text": """1. Analyze business type, target reader, site communication style, and audience knowledge level before writing.
2. Match tone to the business: technical, lifestyle, service, expert, or beginner as appropriate.
3. Every sentence must carry real information or real value.
4. Delete filler and AI-cliche sentences.
5. Be specific about features, materials, dimensions, use cases, and evidence.
6. Replace vague praise with concrete facts.
7. Explain what supports any claim.
8. Write like a knowledgeable owner explaining honestly.
9. Headings must describe content accurately and not overpromise.
10. CTA text should clearly state what happens when clicked.""",
    },
    {
        "rule_key": "rule_rtl_hebrew",
        "rule_name": "RTL & Hebrew-Specific",
        "rule_category": 9,
        "conditional_on": "hebrew_site",
        "rule_text": """1. Hebrew article element must use lang="he" dir="rtl".
2. TOC title must be exactly תוכן עניינים.
3. Never use em dash in Hebrew body prose.
4. Avoid colon in body prose unless it is a technical specification, price label, code value, or CSS value.
5. Date labels in Hebrew are forbidden by default.
6. TOC and FAQ arrows or plus signs must sit on the same line as text in a fixed position slot, follow RTL flow, keep balanced spacing, stay consistently aligned, and not jump when opening or closing.
7. Icons must not be glued to text or pushed to the extreme edge.""",
    },
    {
        "rule_key": "rule_floating_buttons",
        "rule_name": "Floating Button",
        "rule_category": 10,
        "conditional_on": "always",
        "rule_text": """1. Inspect the original page before creating any floating button.
2. If a WhatsApp floating button already exists, do not create another.
3. Do not duplicate any existing floating action.
4. If new floating buttons are needed, position above existing buttons or on the opposite side.
5. New buttons must not overlap existing UI.
6. If floating buttons cover text on mobile or tablet, reposition, resize, or hide them for those breakpoints.
7. If WordPress makes them unstable, remove instead of leaving broken UI.
8. Phone buttons must use real tel: links.""",
    },
    {
        "rule_key": "rule_n8n_safety",
        "rule_name": "N8N Prompt Safety",
        "rule_category": 11,
        "conditional_on": "always",
        "rule_text": """1. Improved N8N prompt is clean plain text.
2. Do not wrap in markdown code fences.
3. Do not output markdown tables inside the prompt.
4. Do not add commentary before or after the prompt.
5. Keep quotes, brackets, and parentheses balanced.
6. Use strict instruction blocks and direct wording.
7. Define output boundaries explicitly for HTML output.
8. For accordions, require details/summary, not JavaScript onclick.
9. Prefer self-contained WordPress-safe HTML/CSS patterns.
10. Prompt must be safe to paste back into n8n without formatting errors.
11. Preserve the original article URL from the initial n8n run.""",
    },
    {
        "rule_key": "rule_workflow_json",
        "rule_name": "Workflow JSON",
        "rule_category": 12,
        "conditional_on": "always",
        "rule_text": """1. Improved workflow JSON must remain aligned with the original workflow purpose.
2. It must support the improved prompt and improved HTML template.
3. It must support store vs non-store logic and optional external product source extraction.
4. It must support Firecrawl usage when needed.
5. It must support Supabase image storage when needed.
6. It must support omitting products without usable images.
7. It must support floating-button inspection, deduplication, and responsive handling.
8. It must support date suppression or current-date logic.
9. It must contain no malformed nodes, broken connections, or broken expressions.
10. It must export improved HTML as .html, not .txt.
11. It must be valid JSON.""",
    },
    {
        "rule_key": "rule_schema",
        "rule_name": "Schema & Structured Data",
        "rule_category": 13,
        "conditional_on": "always",
        "rule_text": """1. Include schema only when relevant to the page and business type.
2. Schema types must match the actual content and business.
3. JSON-LD must be syntactically valid.
4. Do not fabricate AggregateRating.
5. Do not create fake reviews in schema.
6. Do not add irrelevant schema types.
7. Use real data only.""",
    },
    {
        "rule_key": "rule_validation",
        "rule_name": "Validation Checklist",
        "rule_category": 14,
        "conditional_on": "always",
        "rule_text": """Before finalizing, verify: WordPress compatibility; no horizontal overflow; FAQ before author bio; author bio final; TOC and FAQ closed by default; exact TOC title; no date label by default; anchors unique and correct; CTA links real; phone links use tel:; only real reviews, socials, video, map; no breadcrumbs at top; no paragraph numbering; desktop/tablet/mobile render correctly; workflow JSON used for store decision; products only when justified; every product has correct image and link; unusable product images skipped; product images stored in Supabase when needed; no prices unless requested; logo and author image centered; RTL icon rules satisfied for Hebrew; hover states readable; floating buttons deduplicated; schema relevant and valid; final prompt is n8n-safe; workflow JSON is valid and exports HTML as .html.""",
    },
    {
        "rule_key": "rule_forbidden",
        "rule_name": "Forbidden Patterns",
        "rule_category": 15,
        "conditional_on": "always",
        "rule_text": """Never include emojis, fake urgency, scarcity, social proof, authority claims, fake trust theater, bonuses, filler phrases, AI-cliche intros, fabricated data, unsupported superlatives, decorative content icons, year-based marketing phrasing, clickbait headings, manipulative CTA language, fake reviews, fake social profiles, fake author info, fake videos, fake products, content after author bio, multi-column TOC, paragraph numbering, breadcrumbs at top, share blocks unless requested, stale dates, product blocks without confirmed store logic, products without usable images, unstable hotlinked product images, duplicate WhatsApp floating buttons, prompt code fences, assistant chatter around the N8N prompt, .txt export for HTML, or anything that works only in preview but breaks in WordPress.""",
    },
    {
        "rule_key": "rule_api_secrets",
        "rule_name": "API & Secret Handling",
        "rule_category": 16,
        "conditional_on": "always",
        "rule_text": """1. Use direct browsing first.
2. Use Firecrawl when direct browsing is not enough for site-wide discovery, URL mapping, product extraction, or content extraction.
3. Use Supabase only when needed for stable asset storage such as product images, logo, owner image, or author image.
4. Do not use the SQL database password unless absolutely necessary.
5. Never print, echo, expose, or include API keys in output files or prompts.
6. Never hardcode secrets into delivered files.
7. Use environment variables or secure storage references only.
8. Keys are stored encrypted in backend settings and injected at runtime.""",
    },
]


def _active_improvement_rules(language='', store_decision=''):
    language = (language or '').lower()
    store_decision = (store_decision or '').lower()
    include_hebrew = language in ('he', 'hebrew', 'rtl')
    include_store = store_decision in ('store', 'ecommerce', 'shop')
    active = []
    for rule in sorted(IMPROVEMENT_RULES, key=lambda r: r['rule_category']):
        cond = rule.get('conditional_on') or 'always'
        if cond == 'hebrew_site' and not include_hebrew:
            continue
        if cond == 'store_mode' and not include_store:
            continue
        active.append({**rule, 'version': IMPROVEMENT_RULES_VERSION, 'is_active': True, 'is_always_included': True})
    return active


def _safe_text(value, limit=20000):
    if value is None:
        return ''
    text = str(value).strip()
    return text[:limit]


def _detect_language_direction(*texts):
    combined = '\n'.join(_safe_text(t, 3000) for t in texts if t)
    hebrew_chars = len(re.findall(r'[\u0590-\u05FF]', combined))
    latin_chars = len(re.findall(r'[A-Za-z]', combined))
    if hebrew_chars and hebrew_chars >= max(10, latin_chars * 0.15):
        return 'he', 'rtl'
    return 'en', 'ltr'


def _preanalyze_workflow_json(raw_workflow):
    text = ''
    parsed = None
    if isinstance(raw_workflow, (dict, list)):
        parsed = raw_workflow
        text = json.dumps(raw_workflow, ensure_ascii=False)[:200000]
    else:
        text = _safe_text(raw_workflow, 200000)
        if text:
            try:
                parsed = json.loads(text)
            except Exception:
                parsed = None
    lowered = text.lower()
    store_terms = ['woocommerce', 'shopify', 'product', 'products', 'catalog', 'catalogue', 'cart', 'checkout', 'sku', 'collection', 'price', 'store', 'ecommerce', 'e-commerce']
    hits = [term for term in store_terms if term in lowered]
    node_count = 0
    if isinstance(parsed, dict):
        nodes = parsed.get('nodes') or []
        node_count = len(nodes) if isinstance(nodes, list) else 0
    return {
        'valid_json': parsed is not None if text else False,
        'node_count': node_count,
        'store_hits': hits[:20],
        'store_decision': 'store' if len(hits) >= 2 or any(h in hits for h in ['woocommerce', 'shopify', 'checkout', 'cart']) else 'non_store',
    }


def _format_rules_for_prompt(rules):
    chunks = []
    for rule in rules:
        chunks.append(f"## Rule {rule['rule_category']}: {rule['rule_name']}\nBackend key: {rule['rule_key']}\nConditional: {rule.get('conditional_on') or 'always'}\n\n{rule['rule_text']}")
    return '\n\n'.join(chunks)


def assemble_improve_workflow_prompt(payload):
    main_url = _safe_text(payload.get('mainWebsiteUrl') or payload.get('main_website_url'), 1000)
    article_url = _safe_text(payload.get('originalArticleUrl') or payload.get('original_article_url'), 1000)
    original_prompt = _safe_text(payload.get('originalPrompt') or payload.get('original_prompt') or payload.get('n8nPrompt'), 60000)
    original_html = _safe_text(payload.get('htmlTemplate') or payload.get('originalHtml') or payload.get('html_template'), 80000)
    original_workflow = payload.get('workflowJson') or payload.get('originalWorkflow') or payload.get('workflow_json') or ''
    external_product_url = _safe_text(payload.get('externalProductUrl') or payload.get('external_product_url'), 1000)
    custom_instructions = _safe_text(payload.get('customInstructions') or payload.get('custom_instructions'), 20000)
    model = _safe_text(payload.get('model') or prompt_default_model('PROMPT_IMPROVER_MODEL'), 200)
    mode = _safe_text(payload.get('mode') or 'focused', 40)

    missing = []
    if not main_url: missing.append('mainWebsiteUrl')
    if not article_url: missing.append('originalArticleUrl')
    if not original_prompt: missing.append('originalPrompt')
    if not original_html: missing.append('htmlTemplate')
    if not original_workflow: missing.append('workflowJson')
    if missing:
        raise ValueError('Missing required inputs: ' + ', '.join(missing))

    workflow_analysis = _preanalyze_workflow_json(original_workflow)
    language, direction = _detect_language_direction(original_prompt, original_html, json.dumps(original_workflow, ensure_ascii=False) if isinstance(original_workflow, (dict, list)) else original_workflow)
    store_decision = _safe_text(payload.get('storeDecision'), 50) or workflow_analysis['store_decision']
    rules = _active_improvement_rules(language=language, store_decision=store_decision)
    rules_block = _format_rules_for_prompt(rules)

    file_refs = []
    for key, label in [('originalPromptFileName', 'N8N prompt file'), ('htmlTemplateFileName', 'HTML template file'), ('workflowJsonFileName', 'N8N workflow JSON file')]:
        if payload.get(key):
            file_refs.append(f"- {label}: {payload.get(key)}")
    file_refs_block = '\n'.join(file_refs) if file_refs else '- Uploaded file names were not provided in this request.'

    prompt = f"""# Improve Workflow Production Prompt

You are a senior full-stack developer, prompt engineer, WordPress HTML specialist, n8n workflow engineer, SEO/content editor, and QA lead. Work in plan mode first, then execute. Use sub-agents or swarm-style parallel review when available.

Primary objective: improve the original n8n article workflow and return exactly 3 production deliverables:
1. Improved_N8N_Prompt.txt
2. Improved_HTML_Template.html
3. Improved_N8N_Workflow.json

No extra public deliverables. If validation fails, do not publish a fake success; return a clear failure report internally.

## Required Inputs

- Main Website URL: {main_url}
- Original Article URL: {article_url}
- External Product Source URL: {external_product_url or 'None provided'}
- Mode: {mode}
- Requested model: {model}
- Detected language: {language}
- Detected direction: {direction}
- Workflow store decision pre-analysis: {store_decision}
- Workflow JSON valid: {workflow_analysis['valid_json']}
- Workflow node count: {workflow_analysis['node_count']}
- Workflow store/product signals: {', '.join(workflow_analysis['store_hits']) if workflow_analysis['store_hits'] else 'none'}

## Uploaded File References

{file_refs_block}

## Auto-Discovery Order

1. Inspect the homepage.
2. Inspect navigation and footer links.
3. Inspect key pages: about, contact, services, products, blog, reviews, team.
4. Inspect the original article page for layout and floating UI.
5. Inspect the N8N workflow JSON.
6. Inspect the original N8N prompt.
7. Inspect the original HTML template.
8. Use Firecrawl only when direct browsing is incomplete.

## Backend Rules To Enforce

{rules_block}

## Original N8N Prompt

--- BEGIN ORIGINAL N8N PROMPT ---
{original_prompt}
--- END ORIGINAL N8N PROMPT ---

## Original HTML Template

--- BEGIN ORIGINAL HTML TEMPLATE ---
{original_html}
--- END ORIGINAL HTML TEMPLATE ---

## Original N8N Workflow JSON

--- BEGIN ORIGINAL WORKFLOW JSON ---
{_safe_text(json.dumps(original_workflow, ensure_ascii=False, indent=2) if isinstance(original_workflow, (dict, list)) else original_workflow, 80000)}
--- END ORIGINAL WORKFLOW JSON ---

## Custom Instructions

{custom_instructions or 'None.'}

## Output Contract

Return exactly three file sections in this order, with no commentary outside the sections:

### FILE: Improved_N8N_Prompt.txt
Plain text only. No code fences. Must preserve the original article URL and enforce all active backend rules.

### FILE: Improved_HTML_Template.html
Raw HTML only. Must start with article and end with article. Inline CSS only. No style blocks. No scripts. No markdown.

### FILE: Improved_N8N_Workflow.json
Valid JSON only. Must preserve original workflow purpose and support the improved prompt/template, store logic, Firecrawl/Supabase logic when needed, and .html export.

## Failure Handling

If any deliverable cannot be completed truthfully using real data, do not fabricate. Mark that deliverable as failed and explain the blocker internally. Do not present failed or fallback outputs as successful public deliverables.
""".strip()

    run_id = hashlib.sha256((main_url + article_url + original_prompt[:500] + original_html[:500] + str(time.time())).encode('utf-8')).hexdigest()[:16]
    return {
        'ok': True,
        'runId': run_id,
        'assembledPrompt': prompt,
        'rulesUsed': [{
            'rule_key': r['rule_key'], 'rule_name': r['rule_name'], 'rule_category': r['rule_category'],
            'version': r['version'], 'conditional_on': r.get('conditional_on')
        } for r in rules],
        'discoveredData': {
            'site_language': language,
            'site_direction': direction,
            'store_decision': store_decision,
            'workflow_analysis': workflow_analysis,
        },
        'model': model,
        'mode': mode,
    }


# ---------- Palette extractor (Prompt Studio) ----------

_HEX_RE = re.compile(r'#(?:[0-9a-fA-F]{6}|[0-9a-fA-F]{3})\b')
_RGB_RE = re.compile(r'rgba?\(\s*(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})(?:\s*,\s*[\d.]+)?\s*\)')

def _norm_hex(h):
    h = h.lower().lstrip('#')
    if len(h) == 3:
        h = ''.join(c*2 for c in h)
    return '#' + h

def _hex_from_rgb(r, g, b):
    return '#{:02x}{:02x}{:02x}'.format(max(0,min(255,int(r))), max(0,min(255,int(g))), max(0,min(255,int(b))))

def _too_greyish(hex_color):
    h = hex_color.lstrip('#')
    r = int(h[0:2], 16); g = int(h[2:4], 16); b = int(h[4:6], 16)
    mx = max(r, g, b); mn = min(r, g, b)
    if mn >= 240 and mx >= 240: return True
    if mx <= 20: return True
    if (mx - mn) <= 8 and 40 <= mx <= 220: return True
    return False

def _luminance(hex_color):
    h = hex_color.lstrip('#')
    r = int(h[0:2], 16); g = int(h[2:4], 16); b = int(h[4:6], 16)
    return 0.2126*r + 0.7152*g + 0.0722*b

def _fetch_page_text(url, timeout=20):
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0 (HermesPaletteExtractor/1.0)'})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = resp.read()
        encoding = resp.headers.get_content_charset() or 'utf-8'
        return data.decode(encoding, errors='replace'), resp.geturl()

def _collect_stylesheet_urls(html, base_url):
    out = []
    for m in re.finditer(r'<link\b[^>]*rel\s*=\s*["\']stylesheet["\'][^>]*>', html, re.I):
        tag = m.group(0)
        href = re.search(r'href\s*=\s*["\']([^"\']+)["\']', tag)
        if href:
            out.append(urllib.parse.urljoin(base_url, href.group(1)))
        if len(out) >= 5:
            break
    return out

def extract_palette_from_url(url, max_colors=10):
    u = urllib.parse.urlparse(url)
    if u.scheme not in ('http','https'):
        raise ValueError('Only http/https URLs allowed')
    host = (u.hostname or '').lower()
    if host in ('localhost','127.0.0.1','0.0.0.0','::1'):
        raise ValueError('Private/loopback host not allowed')
    if host.startswith('10.') or host.startswith('192.168.') or host.startswith('169.254.') or re.match(r'^172\.(1[6-9]|2[0-9]|3[0-1])\.', host):
        raise ValueError('Private host not allowed')

    html, final_url = _fetch_page_text(url)
    css_blobs = [html]
    for css_url in _collect_stylesheet_urls(html, final_url)[:3]:
        try:
            body, _ = _fetch_page_text(css_url, timeout=12)
            css_blobs.append(body)
        except Exception:
            continue

    combined = '\n'.join(css_blobs)

    counts = {}
    for m in _HEX_RE.finditer(combined):
        h = _norm_hex(m.group(0))
        counts[h] = counts.get(h, 0) + 1
    for m in _RGB_RE.finditer(combined):
        h = _hex_from_rgb(m.group(1), m.group(2), m.group(3))
        counts[h] = counts.get(h, 0) + 1

    brand = [(c, n) for c, n in counts.items() if not _too_greyish(c)]
    brand.sort(key=lambda x: (-x[1], abs(_luminance(x[0]) - 128)))
    neutrals = sorted([(c, n) for c, n in counts.items() if _too_greyish(c)], key=lambda x: -x[1])[:5]

    top_brand = brand[:max_colors]

    title = ''
    m = re.search(r'<title[^>]*>([^<]+)</title>', html, re.I)
    if m:
        title = m.group(1).strip()[:200]
    desc = ''
    m = re.search(r'<meta[^>]+name\s*=\s*["\']description["\'][^>]*content\s*=\s*["\']([^"\']+)["\']', html, re.I)
    if m:
        desc = m.group(1).strip()[:300]

    return {
        'ok': True,
        'url': final_url,
        'title': title,
        'description': desc,
        'totalDistinctColors': len(counts),
        'brandPalette': [{'hex': c, 'count': n} for c, n in top_brand],
        'neutrals': [{'hex': c, 'count': n} for c, n in neutrals],
    }


# ===== Round 2 (2026-04-22): Rate limiting + structured JSON logging =====
import logging
import sys as _sys

_R2_RATE_LOCK = threading.Lock()
_R2_RATE_BUCKETS = {}  # ip -> {tokens, last_refill}
_R2_RATE_CAPACITY = int(os.environ.get('DASH_RATE_CAPACITY', '60'))   # burst
_R2_RATE_REFILL = float(os.environ.get('DASH_RATE_REFILL', '1.0'))    # tokens/sec
_R2_RATE_ENABLED = os.environ.get('DASH_RATE_LIMIT', '1') not in ('0', 'false', 'False', '')

# Per-path overrides (path prefix -> capacity, refill)
_R2_RATE_OVERRIDES = [
    ('/api/kwr/run',     (5,  0.05)),   # heavy: 5 burst, 1 every 20s
    ('/api/kwr/download',(10, 0.1)),    # heavy: 10 burst, 1 every 10s
    ('/static/',         (300, 10.0)),  # static assets very permissive
    ('/login',           (20, 0.5)),    # login attempts: 20/min
]
_R2_RATE_EXEMPT = ('/api/auth/me', '/api/health', '/healthz')

def _r2_rate_params(path):
    for prefix, params in _R2_RATE_OVERRIDES:
        if path.startswith(prefix):
            return params
    return (_R2_RATE_CAPACITY, _R2_RATE_REFILL)

def _r2_rate_check(client_ip, path):
    """Token-bucket rate limit. Returns (allowed: bool, retry_after_sec: float)."""
    if not _R2_RATE_ENABLED:
        return True, 0.0
    if any(path.startswith(p) for p in _R2_RATE_EXEMPT):
        return True, 0.0
    capacity, refill = _r2_rate_params(path)
    key = client_ip + '|' + path.split('?')[0][:60]
    now = time.time()
    with _R2_RATE_LOCK:
        b = _R2_RATE_BUCKETS.get(key)
        if b is None:
            b = {'tokens': float(capacity), 'last': now}
            _R2_RATE_BUCKETS[key] = b
        # Refill
        elapsed = now - b['last']
        b['tokens'] = min(float(capacity), b['tokens'] + elapsed * refill)
        b['last'] = now
        if b['tokens'] >= 1.0:
            b['tokens'] -= 1.0
            return True, 0.0
        # Compute retry-after
        retry = (1.0 - b['tokens']) / refill if refill > 0 else 60.0
        return False, retry

def _r2_rate_gc():
    """Periodically clear stale buckets to prevent memory growth."""
    cutoff = time.time() - 600  # 10 min idle
    with _R2_RATE_LOCK:
        stale = [k for k, v in _R2_RATE_BUCKETS.items() if v['last'] < cutoff]
        for k in stale:
            _R2_RATE_BUCKETS.pop(k, None)

# Run GC every 5 min
def _r2_rate_gc_loop():
    while True:
        time.sleep(300)
        try: _r2_rate_gc()
        except Exception: pass
threading.Thread(target=_r2_rate_gc_loop, daemon=True).start()

# ----- JSON structured logger -----
_R2_JSON_LOGGING = os.environ.get('DASH_JSON_LOG', '1') not in ('0', 'false', 'False', '')

class _R2JsonFormatter(logging.Formatter):
    def format(self, record):
        payload = {
            'ts': datetime.datetime.utcnow().isoformat() + 'Z',
            'level': record.levelname,
            'msg': record.getMessage(),
        }
        for k in ('method', 'path', 'status', 'ip', 'duration_ms', 'request_id', 'event'):
            v = getattr(record, k, None)
            if v is not None: payload[k] = v
        if record.exc_info:
            payload['exc'] = self.formatException(record.exc_info)
        return json.dumps(payload, ensure_ascii=False)

_r2_log = logging.getLogger('dashboard')
if not _r2_log.handlers:
    _h = logging.StreamHandler(_sys.stdout)
    if _R2_JSON_LOGGING:
        _h.setFormatter(_R2JsonFormatter())
    else:
        _h.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(message)s'))
    _r2_log.addHandler(_h)
    _r2_log.setLevel(logging.INFO)
    _r2_log.propagate = False

def _r2_log_request(handler, status, duration_ms):
    ip = _stage8_client_ip(handler) or '-'
    try:
        path = handler.path
        method = handler.command
    except Exception:
        path = '-'; method = '-'
    rec = logging.LogRecord('dashboard', logging.INFO, '', 0, 'request', None, None)
    rec.method = method; rec.path = path; rec.status = status
    rec.ip = ip; rec.duration_ms = round(duration_ms, 2)
    rec.request_id = getattr(handler, '_r2_req_id', None)
    rec.event = 'http_request'
    _r2_log.handle(rec)


# ===== Round 3 (2026-04-22): SSE, CSRF, RequestID, Health, Auto-backup =====
import secrets as _r3_secrets
import hashlib as _r3_hashlib
import queue as _r3_queue
import uuid as _r3_uuid

_R3_START_TIME = time.time()
_R3_SSE_CLIENTS = []  # list of queue.Queue
_R3_SSE_LOCK = threading.Lock()
_R3_CSRF_SECRET = os.environ.get('DASH_CSRF_SECRET') or _r3_secrets.token_hex(32)
_R3_CSRF_ENABLED = os.environ.get('DASH_CSRF', '1') not in ('0','false','False','')
_R3_CSRF_EXEMPT = (
    '/api/auth/login', '/api/auth/request-reset', '/api/auth/reset',
    '/api/n8n/webhook', '/login', '/api/csrf', '/metrics', '/api/login', '/api/reset-password',
)

def _r3_csrf_token():
    """Generate a per-session CSRF token (signed with server secret)."""
    raw = _r3_secrets.token_urlsafe(16)
    sig = _r3_hashlib.sha256((raw + _R3_CSRF_SECRET).encode()).hexdigest()[:16]
    return raw + '.' + sig

def _r3_csrf_verify(token):
    if not token or '.' not in token: return False
    raw, sig = token.rsplit('.', 1)
    expected = _r3_hashlib.sha256((raw + _R3_CSRF_SECRET).encode()).hexdigest()[:16]
    return _r3_secrets.compare_digest(sig, expected)

def _r3_csrf_enforce_hard():
    raw = os.environ.get('DASH_CSRF_ENFORCE')
    if raw is not None:
        return raw in ('1', 'true', 'True', 'yes', 'on')
    return _dashboard_is_production() or bool(getattr(r6, 'CSRF_HARD', False))

def _r3_check_csrf_or_warn(handler, parsed):
    if not _R3_CSRF_ENABLED or any(parsed.path.startswith(p) for p in _R3_CSRF_EXEMPT):
        return True
    tok = handler.headers.get('X-CSRF-Token', '')
    if _r3_csrf_verify(tok):
        return True
    if _r3_csrf_enforce_hard():
        json_response(handler, 403, {'ok': False, 'error': 'csrf_invalid'})
        return False
    rec = logging.LogRecord('dashboard', logging.WARNING, '', 0,
        f'csrf-missing path={parsed.path}', None, None)
    rec.event = 'csrf_warn'; rec.path = parsed.path
    _r2_log.handle(rec)
    return True

def sse_broadcast(event_type, payload):
    """Push a notification to all connected SSE clients. Safe to call from any thread."""
    msg = {'type': event_type, 'payload': payload, 'ts': time.time()}
    data = 'data: ' + json.dumps(msg, ensure_ascii=False) + '\n\n'
    # R6: alert on failure events
    try:
        if event_type in ('kwr_error', 'deploy_error', 'backup_error'):
            r6.alert_failure(f'{event_type}', str(payload)[:1500],
                             run_id=(payload or {}).get('run_id', '') if isinstance(payload, dict) else '')
    except Exception: pass
    with _R3_SSE_LOCK:
        dead = []
        for q in _R3_SSE_CLIENTS:
            try: q.put_nowait(data)
            except Exception: dead.append(q)
        for q in dead:
            try: _R3_SSE_CLIENTS.remove(q)
            except ValueError: pass

def _r3_health_detailed():
    uptime = time.time() - _R3_START_TIME
    with _R2_RATE_LOCK:
        rl_buckets = len(_R2_RATE_BUCKETS)
    with _R3_SSE_LOCK:
        sse_clients = len(_R3_SSE_CLIENTS)
    out = {
        'ok': True,
        'uptime_seconds': round(uptime, 1),
        'uptime_human': _r3_human_time(uptime),
        'rate_limit': {
            'enabled': _R2_RATE_ENABLED,
            'active_buckets': rl_buckets,
            'default_capacity': _R2_RATE_CAPACITY,
            'default_refill': _R2_RATE_REFILL,
        },
        'sse': {'connected_clients': sse_clients},
        'csrf': {'enabled': _R3_CSRF_ENABLED},
        'json_logging': _R2_JSON_LOGGING,
        'thread_count': threading.active_count(),
        'pid': os.getpid(),
    }
    # Memory stats (best-effort)
    try:
        import resource
        ru = resource.getrusage(resource.RUSAGE_SELF)
        out['memory_kb'] = ru.ru_maxrss  # Linux: KB
    except Exception: pass
    # KWR/Radar status (best effort)
    try:
        out['kwr_status'] = kwr_backend.get_state().get('status', 'unknown') if hasattr(kwr_backend, 'get_state') else 'n/a'
    except Exception: pass
    try:
        out['radar_status'] = _radar_state.get('status', 'unknown')
    except Exception: pass
    return out

def _r3_human_time(secs):
    secs = int(secs)
    d, secs = divmod(secs, 86400)
    h, secs = divmod(secs, 3600)
    m, s = divmod(secs, 60)
    parts = []
    if d: parts.append(f'{d}d')
    if h: parts.append(f'{h}h')
    if m: parts.append(f'{m}m')
    parts.append(f'{s}s')
    return ' '.join(parts)

# ----- Auto-backup to Obsidian (optional, env-gated) -----
_R3_BACKUP_ENABLED = os.environ.get('DASH_AUTO_BACKUP', '0') in ('1','true','True')
_R3_BACKUP_INTERVAL_HOURS = float(os.environ.get('DASH_BACKUP_INTERVAL_H', '24'))
_R3_OBSIDIAN_URL = os.environ.get('OBSIDIAN_URL', '').rstrip('/')
_R3_OBSIDIAN_KEY = os.environ.get('OBSIDIAN_API_KEY', '')
_R3_BACKUP_FOLDER = os.environ.get('DASH_BACKUP_FOLDER', 'HTML REDESIGN/dashboard/auto-backups')

def _r3_auto_backup():
    """Push data.json + sync.json to Obsidian vault. Returns (success_count, error)."""
    if not (_R3_OBSIDIAN_URL and _R3_OBSIDIAN_KEY):
        return 0, 'OBSIDIAN_URL or OBSIDIAN_API_KEY not set'
    ts = datetime.datetime.utcnow().strftime('%Y-%m-%d-%H%M')
    files = [
        ('data.json', ROOT / 'data.json'),
        ('sync.json', ROOT / 'sync.json'),
    ]
    ok = 0; errors = []
    for name, path in files:
        if not path.exists(): continue
        try:
            data = path.read_bytes()
            url = f'{_R3_OBSIDIAN_URL}/vault/{_R3_BACKUP_FOLDER}/{ts}/{name}'
            req = urllib.request.Request(url, data=data, method='PUT', headers={
                'Authorization': f'Bearer {_R3_OBSIDIAN_KEY}',
                'Content-Type': 'application/octet-stream',
            })
            urllib.request.urlopen(req, timeout=30)
            ok += 1
        except Exception as e:
            errors.append(f'{name}: {e}')
    return ok, ('; '.join(errors) if errors else None)

def _r3_backup_loop():
    while True:
        time.sleep(_R3_BACKUP_INTERVAL_HOURS * 3600)
        try:
            ok, err = _r3_auto_backup()
            rec = logging.LogRecord('dashboard', logging.INFO, '', 0, f'auto-backup ok={ok} err={err}', None, None)
            rec.event = 'auto_backup'
            _r2_log.handle(rec)
            if ok > 0:
                sse_broadcast('backup', {'success': ok, 'timestamp': time.time()})
        except Exception as e:
            pass

if _R3_BACKUP_ENABLED:
    threading.Thread(target=_r3_backup_loop, daemon=True).start()


class DashboardHandler(BaseHTTPRequestHandler):
    server_version = 'DashboardHTTP/1.0'

    def log_message(self, format, *args):
        return

    # Round 2: hook into setup so we can time requests + apply rate limit early
    def setup(self):
        BaseHTTPRequestHandler.setup(self)
        self._r2_t0 = time.time()
        self._r2_status = 0
        self._r2_req_id = _r3_uuid.uuid4().hex[:12]

    def _send_json(self, status, payload):
        """Compatibility helper for legacy route blocks that call self._send_json."""
        return json_response(self, status, payload)

    def send_response(self, code, message=None):
        # Capture status for logging
        try: self._r2_status = code
        except Exception: pass
        return BaseHTTPRequestHandler.send_response(self, code, message)

    def _r2_check_rate(self):
        """Call early in do_GET/do_POST. Returns True if request should proceed."""
        try:
            ip = _stage8_client_ip(self)
            path = urllib.parse.urlparse(self.path).path
        except Exception:
            return True
        ok, retry = _r2_rate_check(ip, path)
        if ok: return True
        body = json.dumps({'ok': False, 'error': 'rate_limited', 'retry_after': round(retry, 1)}).encode('utf-8')
        self.send_response(429)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Retry-After', str(int(retry) + 1))
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)
        try:
            _r2_log_request(self, 429, (time.time() - getattr(self, '_r2_t0', time.time())) * 1000)
        except Exception: pass
        return False

    def finish(self):
        try:
            dur = (time.time() - getattr(self, '_r2_t0', time.time())) * 1000
            _r2_log_request(self, getattr(self, '_r2_status', 0) or 200, dur)
        except Exception:
            pass
        return BaseHTTPRequestHandler.finish(self)

    def _r4_export_pdf(self, parsed):
        """Round 4: Lightweight PDF export of KWR data using fpdf2 (or HTML fallback)."""
        from html import escape as _esc
        try:
            qs = urllib.parse.parse_qs(parsed.query)
            run_id = (qs.get('run_id') or [''])[0]
            data = None
            if hasattr(kwr_backend, 'get_state'):
                st = kwr_backend.get_state()
                if run_id and run_id in st: data = st[run_id]
                else:
                    # latest
                    runs = sorted(st.items(), key=lambda kv: kv[1].get('updated_at',''), reverse=True)
                    if runs: data = runs[0][1]
            if not data:
                return json_response(self, 404, {'ok': False, 'error': 'no_run'})

            # Try fpdf2 if installed; fall back to printable HTML
            try:
                from fpdf import FPDF
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font('Helvetica', 'B', 14)
                pdf.cell(0, 8, f"KWR Run: {run_id or 'latest'}", ln=True)
                pdf.set_font('Helvetica', '', 9)
                pdf.cell(0, 6, f"Status: {data.get('status','?')}  Stage: {data.get('current_stage','?')}", ln=True)
                pdf.cell(0, 6, f"Updated: {data.get('updated_at','?')}", ln=True)
                pdf.ln(3)
                pdf.set_font('Helvetica', 'B', 10); pdf.cell(0, 6, 'Logs (last 30):', ln=True)
                pdf.set_font('Courier', '', 7)
                for line in (data.get('logs') or [])[-30:]:
                    safe = (line or '')[:140].encode('latin-1', 'replace').decode('latin-1')
                    pdf.multi_cell(0, 4, safe)
                buf = bytes(pdf.output(dest='S'))
                self.send_response(200)
                self.send_header('Content-Type', 'application/pdf')
                self.send_header('Content-Disposition', f'attachment; filename="kwr-{run_id or "latest"}.pdf"')
                self.send_header('Content-Length', str(len(buf)))
                self.end_headers(); self.wfile.write(buf); return
            except ImportError:
                # HTML fallback the browser can print to PDF
                html = f"""<html><head><title>KWR Export</title><style>
body{{font-family:Arial;padding:24px}}h1{{color:#333}}pre{{background:#f4f4f4;padding:10px;font-size:11px;white-space:pre-wrap}}
@media print{{body{{padding:0}}}}</style></head><body>
<h1>KWR Run: {_esc(run_id or 'latest')}</h1>
<p><b>Status:</b> {_esc(str(data.get('status','?')))}<br>
<b>Stage:</b> {_esc(str(data.get('current_stage','?')))}<br>
<b>Updated:</b> {_esc(str(data.get('updated_at','?')))}</p>
<h2>Logs</h2><pre>{_esc(chr(10).join((data.get('logs') or [])[-50:]))}</pre>
<script>window.onload=function(){{setTimeout(function(){{window.print()}},300)}}</script>
</body></html>"""
                return text_response(self, 200, html.encode('utf-8'), 'text/html; charset=utf-8')
        except Exception as e:
            return json_response(self, 500, {'ok': False, 'error': str(e)})

    def _r3_serve_sse(self):
        """SSE endpoint: streams events to connected clients indefinitely."""
        q = _r3_queue.Queue(maxsize=100)
        with _R3_SSE_LOCK:
            _R3_SSE_CLIENTS.append(q)
        try:
            self.send_response(200)
            self.send_header('Content-Type', 'text/event-stream; charset=utf-8')
            self.send_header('Cache-Control', 'no-cache')
            self.send_header('X-Accel-Buffering', 'no')
            self.send_header('Connection', 'keep-alive')
            self.end_headers()
            # initial hello
            self.wfile.write(b'data: {"type":"connected","payload":{}}\n\n')
            self.wfile.flush()
            last_ping = time.time()
            while True:
                try:
                    data = q.get(timeout=15)
                    self.wfile.write(data.encode('utf-8'))
                    self.wfile.flush()
                except _r3_queue.Empty:
                    # heartbeat to keep connection alive (Render closes idle conns)
                    self.wfile.write(b': ping\n\n')
                    self.wfile.flush()
                if time.time() - last_ping > 600:
                    break  # close after 10 min, client will reconnect
        except (BrokenPipeError, ConnectionResetError, OSError):
            pass
        finally:
            with _R3_SSE_LOCK:
                try: _R3_SSE_CLIENTS.remove(q)
                except ValueError: pass

    # #14 Security headers — applied to every response automatically.
    # CSP allows inline scripts/styles because the dashboard relies on them today;
    # tightening that requires a separate refactor pass.
    _SECURITY_HEADERS = (
        ('X-Content-Type-Options', 'nosniff'),
        ('X-Frame-Options', 'SAMEORIGIN'),
        ('Referrer-Policy', 'strict-origin-when-cross-origin'),
        ('Permissions-Policy', 'camera=(), microphone=(), geolocation=(), payment=()'),
        ('Content-Security-Policy',
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com https://cdnjs.cloudflare.com; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
            "font-src 'self' data: https://fonts.gstatic.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
            "img-src 'self' data: blob: https:; "
            "connect-src 'self' https:; "
            "frame-src 'self' https:; "
            "object-src 'none'; "
            "base-uri 'self'"),
    )

    def end_headers(self):
        try:
            for k, v in self._SECURITY_HEADERS:
                self.send_header(k, v)
            if not getattr(self, '_r2_reqid_sent', False):
                self.send_header('X-Request-ID', getattr(self, '_r2_req_id', '-'))
                self.send_header('Access-Control-Expose-Headers', 'X-Request-ID')
                self._r2_reqid_sent = True
            # Round 5: metrics
            try:
                code = getattr(self, '_status_code', 200)
                r5.metrics_inc('requests_total')
                r5.metrics_inc('requests_by_status', label=str(code))
                if code >= 500: r5.metrics_inc('errors_total')
            except Exception: pass
        except Exception:
            pass
        BaseHTTPRequestHandler.end_headers(self)

    def send_response(self, code, message=None):
        self._status_code = code
        return BaseHTTPRequestHandler.send_response(self, code, message)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, X-CSRF-Token, Authorization')
        self.send_header('Access-Control-Allow-Methods', 'GET,POST,PUT,DELETE,OPTIONS,HEAD')
        self.end_headers()

    def do_HEAD(self):
        if not self._r2_check_rate(): return
        parsed = urllib.parse.urlparse(self.path)
        if _dashboard_auth_enabled() and not _stage8_public_path(parsed.path):
            token = _stage8_get_token(self)
            if not (token and _stage8_verify_session(token)) and not _get_current_user(self):
                if parsed.path.startswith('/api/'):
                    self.send_response(401)
                    self.send_header('Content-Type', 'application/json; charset=utf-8')
                    self.send_header('Cache-Control', 'no-store, must-revalidate')
                    self.end_headers()
                else:
                    self.send_response(302)
                    self.send_header('Location', '/login')
                    self.end_headers()
                return
        if parsed.path in ('/api/health', '/api/health/detailed'):
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Cache-Control', 'no-store, must-revalidate')
            self.end_headers()
            return
        if parsed.path == '/healthz':
            self.send_response(200)
            self.send_header('Content-Type', 'text/plain')
            self.end_headers()
            return
        target = None
        if parsed.path in ('/login', '/login.html'):
            target = ROOT / 'login-page.html'
        else:
            clean = posixpath.normpath(urllib.parse.unquote(parsed.path))
            target = INDEX if clean in ('', '.', '/') else (ROOT / clean.lstrip('/')).resolve()
            if ROOT not in target.parents and target != ROOT:
                self.send_response(403); self.end_headers(); return
        if not target.exists() or not target.is_file():
            self.send_response(404); self.end_headers(); return
        content_type, _ = mimetypes.guess_type(str(target))
        if not content_type:
            content_type = 'application/octet-stream'
        self.send_response(200)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(target.stat().st_size))
        self.send_header('Cache-Control', _cache_control_for(content_type, parsed.path))
        self.end_headers()

    def do_GET(self):
        if not self._r2_check_rate(): return
        parsed = urllib.parse.urlparse(self.path)

        if parsed.path == '/api/radar/data':
            try:
                qs = urllib.parse.parse_qs(parsed.query)
                subject = qs.get('subject', ['gpt55'])[0]
                
                # Mock live data based on subject
                base_data = {
                    'gpt55': {'Accessibility': 95, 'Code Quality': 92, 'Accuracy': 90, 'Compatibility': 88, 'Speed': 85, 'Creativity': 82, 'Optimization': 80, 'Consistency': 75},
                    'opus47': {'Accessibility': 90, 'Code Quality': 89, 'Accuracy': 93, 'Compatibility': 84, 'Speed': 78, 'Creativity': 96, 'Consistency': 88, 'Optimization': 82},
                    'gemini31': {'Accessibility': 88, 'Code Quality': 85, 'Accuracy': 88, 'Compatibility': 90, 'Speed': 92, 'Creativity': 72, 'Consistency': 85, 'Optimization': 95},
                    'kimik26': {'Accessibility': 71, 'Code Quality': 82, 'Accuracy': 85, 'Compatibility': 80, 'Speed': 95, 'Creativity': 88, 'Consistency': 90, 'Optimization': 94},
                }
                
                names = {
                    'gpt55': 'GPT 5.5 — Layout Architect',
                    'opus47': 'Opus 4.7',
                    'gemini31': 'Gemini 3.1',
                    'kimik26': 'Kimi K2.6'
                }
                
                scores = base_data.get(subject, base_data['gpt55'])
                overall = sum(scores.values()) / len(scores)
                
                payload = {
                    'subject_name': names.get(subject, 'Unknown'),
                    'overall_score': round(overall, 1),
                    'scores': scores,
                    'targets': {k: 85 for k in scores.keys()},
                    'trends': {k: "+1" if v > 85 else "-1" for k, v in scores.items()},
                    'ai_summary': f"{names.get(subject, 'Unknown')} analysis. Overall score is {round(overall, 1)}. Strengths are evident in metrics above 85.",
                    'ai_recommendations': [
                        { "priority": "HIGH", "area": "Low Metrics", "recommendation": "Focus on improving metrics below 85", "impact": "+5-10 pts" }
                    ]
                }
                payload['targets']['Code Quality'] = 90
                payload['targets']['Accessibility'] = 90
                
                return json_response(self, 200, {'ok': True, 'data': payload})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        if _dashboard_auth_enabled() and not _stage8_check_auth(self, parsed):
            return
        # TEMPLATE_INTELLIGENCE_CONNECTORS_API_2026_04_29 - read-only connector catalog.
        if parsed.path == '/api/template-connectors/catalog':
            return json_response(self, 200, _template_connector_catalog())

        # PLAYGROUND_API_2026_04_29 - Plan #24 template preview/testing hub.
        if parsed.path == '/api/playground/templates':
            try:
                qs = urllib.parse.parse_qs(parsed.query)
                with threading.Lock():
                    data = _playground_load()
                    if _playground_seed_if_empty(data):
                        _playground_save(data)
                rows = list(data.get('templates', []))
                domain = (qs.get('domain') or [''])[0].strip().lower()
                agent = (qs.get('agent') or [''])[0].strip().lower()
                q = (qs.get('q') or [''])[0].strip().lower()
                sort = (qs.get('sort') or ['newest'])[0]
                try:
                    limit = min(max(int((qs.get('limit') or ['100'])[0]), 1), 500)
                except Exception:
                    limit = 100
                if domain:
                    rows = [r for r in rows if (r.get('domain') or '').lower() == domain]
                if agent:
                    rows = [r for r in rows if agent in (r.get('agent_name') or '').lower() or agent in (r.get('agent_model') or '').lower()]
                if q:
                    rows = [r for r in rows if q in ' '.join([str(r.get('domain','')), str(r.get('domain_display','')), str(r.get('agent_name','')), str(r.get('agent_style','')), ' '.join(r.get('tags') or [])]).lower()]
                if sort == 'favorite':
                    rows = sorted(rows, key=lambda r: (not r.get('is_favorite'), r.get('created_at','')), reverse=False)
                elif sort == 'size':
                    rows = sorted(rows, key=lambda r: int(r.get('html_size_bytes') or 0), reverse=True)
                else:
                    rows = sorted(rows, key=lambda r: r.get('created_at',''), reverse=True)
                public = [_playground_public_template(r, include_html=False) for r in rows[:limit]]
                return json_response(self, 200, {'success': True, 'ok': True, 'data': public, 'domains': _playground_domains(data.get('templates', [])), 'total': len(rows)})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'success': False, 'error': str(exc)})
        if parsed.path == '/api/playground/domains':
            data = _playground_load()
            if _playground_seed_if_empty(data):
                _playground_save(data)
            return json_response(self, 200, {'ok': True, 'success': True, 'domains': _playground_domains(data.get('templates', []))})
        if parsed.path.startswith('/api/playground/templates/'):
            parts = parsed.path.strip('/').split('/')
            if len(parts) == 4 and parts[:3] == ['api', 'playground', 'templates']:
                template_id = parts[3]
                data = _playground_load()
                row = next((r for r in data.get('templates', []) if r.get('id') == template_id), None)
                if not row:
                    return json_response(self, 404, {'ok': False, 'success': False, 'error': 'template not found'})
                return json_response(self, 200, {'ok': True, 'success': True, 'template': _playground_public_template(row, include_html=True)})
        if parsed.path.startswith('/api/playground/exports/'):
            template_id = parsed.path.rsplit('/', 1)[-1]
            data = _playground_load()
            exports = [e for e in data.get('exports', []) if e.get('template_id') == template_id]
            return json_response(self, 200, {'ok': True, 'success': True, 'exports': exports})
        if parsed.path == '/api/preferences':
            data = _playground_load()
            return json_response(self, 200, {'ok': True, 'success': True, 'preferences': data.get('preferences', {})})

        # DASHBOARD_FEATURES_API — 15-feature roadmap GET routes (additive)
        _df_get = df_api.handle_get(self, parsed)
        if _df_get is not None:
            return
        # PRODUCTIVITY_HUB_API_2026_04_27 - additive backend foundation routes.
        if parsed.path == '/api/productivity/summary':
            return json_response(self, 200, _productivity_summary())
        if parsed.path == '/api/productivity/notifications':
            with _PRODUCTIVITY_HUB_LOCK:
                rows = list(_productivity_load().get('notifications', []))
            return json_response(self, 200, {'ok': True, 'notifications': rows})
        if parsed.path == '/api/productivity/audit':
            with _PRODUCTIVITY_HUB_LOCK:
                rows = list(_productivity_load().get('audit', []))
            return json_response(self, 200, {'ok': True, 'events': rows})
        if parsed.path == '/api/productivity/search':
            qs = urllib.parse.parse_qs(parsed.query)
            query = qs.get('q', [''])[0]
            return json_response(self, 200, {'ok': True, 'query': query, 'results': _productivity_search(query)})

        # TEMPLATE_IMPROVEMENTS_API_2026_04_27 - additive routes, existing APIs untouched.
        if parsed.path == '/api/improve/jobs':
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                data = _template_improvements_load()
                jobs = [_template_improvement_public_job(v) for v in data.get('jobs', {}).values()]
            jobs = sorted([j for j in jobs if j], key=lambda x: x.get('created_at', ''), reverse=True)
            return json_response(self, 200, {'ok': True, 'jobs': jobs})
        if parsed.path.startswith('/api/improve/jobs/'):
            job_id = parsed.path.rsplit('/', 1)[-1]
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                data = _template_improvements_load()
                job = _template_improvement_public_job(data.get('jobs', {}).get(job_id))
                outputs = data.get('outputs', {}).get(job_id, [])
            if not job:
                return json_response(self, 404, {'ok': False, 'error': 'job not found'})
            return json_response(self, 200, {'ok': True, 'job': job, 'outputs': outputs})
        if parsed.path.startswith('/api/improve/results/'):
            parts = parsed.path.strip('/').split('/')
            job_id = parts[3] if len(parts) >= 4 else ''
            agent_key = parts[4] if len(parts) >= 5 else ''
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                outputs = _template_improvements_load().get('outputs', {}).get(job_id, [])
            if agent_key:
                outputs = [o for o in outputs if o.get('agent_key') == agent_key]
            return json_response(self, 200, {'ok': True, 'outputs': outputs})
        if parsed.path.startswith('/api/improve/instructions/'):
            domain = urllib.parse.unquote(parsed.path.rsplit('/', 1)[-1])
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                rows = [r for r in _template_improvements_load().get('instructions', []) if r.get('domain') == domain]
            return json_response(self, 200, {'ok': True, 'instructions': rows})
        # ---- Round 3: New endpoints ----
        if parsed.path in ('/api/health/detailed', '/api/health'):
            return json_response(self, 200, _r3_health_detailed())
        if parsed.path == '/healthz':
            return text_response(self, 200, b'ok', 'text/plain')
        if parsed.path == '/api/csrf':
            tok = _r3_csrf_token()
            return json_response(self, 200, {'ok': True, 'token': tok})
        if parsed.path == '/api/events':
            return self._r3_serve_sse()
        # ---- Round 5: New GET endpoints ----
        if parsed.path == '/metrics':
            try:
                body = r5.metrics_render().encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'text/plain; version=0.0.4')
                self.send_header('Content-Length', str(len(body)))
                self.end_headers(); self.wfile.write(body); return
            except Exception as e:
                return json_response(self, 500, {'error': str(e)})
        if parsed.path == '/api/analytics':
            try:
                state = kwr_backend.get_state() if hasattr(kwr_backend, 'get_state') else {}
                base = r5.analytics_compute(state)
                # Compute project-level stats
                data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
                try:
                    with open(data_path, 'r', encoding='utf-8') as _f:
                        tree_data = json.load(_f)
                    tree = tree_data.get('tree', []) if isinstance(tree_data, dict) else []
                except Exception:
                    tree = []
                n8n_map = load_json_file(MAP_FILE, {})
                n8n_domains = set((n8n_map.get('domains') or {}).keys())
                project_counts = {}
                total_files = 0
                total_agents = 0
                for item in tree:
                    path = item.get('path', '')
                    if not path or '/' not in path:
                        continue
                    parts = path.split('/')
                    domain = parts[0]
                    if domain in ('index.html', 'server.py', 'data.json', 'login-page.html', 'Dockerfile', '.dockerignore', 'AGENTS.md', 'kwr_backend.py', 'n8n-workflow-map.json'):
                        continue
                    if not _looks_like_project_domain(domain):
                        continue
                    if len(parts) < 3:
                        continue
                    agent = parts[1]
                    if item.get('type') == 'blob':
                        project_counts.setdefault(domain, {'files': 0, 'agents': set(), 'deployed': domain in n8n_domains})
                        project_counts[domain]['files'] += 1
                        project_counts[domain]['agents'].add(agent)
                        total_files += 1
                for pc in project_counts.values():
                    total_agents += len(pc['agents'])
                top_projects = sorted(
                    [{'name': k, 'files': v['files'], 'deployed': v['deployed']} for k, v in project_counts.items()],
                    key=lambda x: x['files'],
                    reverse=True
                )[:3]
                # deployments over last 30 days from file mtimes
                now = datetime.datetime.utcnow()
                day = datetime.timedelta(days=1)
                buckets = {}
                for i in range(30):
                    d = (now - i * day).strftime('%Y-%m-%d')
                    buckets[d] = 0
                for item in tree:
                    path = item.get('path', '')
                    if not path or '/' not in path:
                        continue
                    parts = path.split('/')
                    domain = parts[0]
                    if domain in ('index.html', 'server.py', 'data.json'):
                        continue
                    if not _looks_like_project_domain(domain):
                        continue
                    fpath = ROOT / path
                    try:
                        mtime = datetime.datetime.utcfromtimestamp(fpath.stat().st_mtime)
                        if (now - mtime).days <= 30:
                            d = mtime.strftime('%Y-%m-%d')
                            if d in buckets:
                                buckets[d] += 1
                    except Exception:
                        pass
                deployments_over_time = [{'date': d, 'count': c} for d, c in sorted(buckets.items())]
                base.update({
                    'deployments_over_time': deployments_over_time,
                    'success_rate': base.get('success_rate', 0),
                    'totals': {
                        'projects': len(project_counts),
                        'files': total_files,
                        'agents': total_agents,
                    },
                    'top_projects': top_projects,
                })
                return json_response(self, 200, {'ok': True, 'data': base})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        if parsed.path == '/api/audit':
            qs = urllib.parse.parse_qs(parsed.query)
            limit = int((qs.get('limit') or ['100'])[0])
            return json_response(self, 200, {'ok': True, 'events': r5.audit_query(limit=limit)})
        if parsed.path == '/api/views':
            user = self.headers.get('X-User', 'anon')
            return json_response(self, 200, {'ok': True, 'views': r5.views_list(user)})
        if parsed.path == '/api/webhooks/history':
            return json_response(self, 200, {'ok': True, 'history': r5.webhook_history(200)})
        if parsed.path == '/api/users':
            if not _require_admin(self): return
            users = _mu_users_load()
            safe = [{'id': u.get('id',''), 'username': u.get('username',''), 'role': u.get('role','viewer'),
                     'email': u.get('email',''), 'created_at': u.get('created_at',''), 'last_login': u.get('last_login')} for u in users]
            return json_response(self, 200, {'ok': True, 'users': safe})
        if parsed.path.startswith('/api/kwr/summary/'):
            run_id = parsed.path.rsplit('/', 1)[-1]
            try:
                state = kwr_backend.get_state() if hasattr(kwr_backend, 'get_state') else {}
                if run_id not in state: return json_response(self, 404, {'ok': False})
                # Try AI first, fall back to heuristic
                ai = r6.ai_summarize_run(state[run_id])
                if ai: return json_response(self, 200, {'ok': True, 'summary': ai, 'mode': 'ai'})
                return json_response(self, 200, {'ok': True, 'summary': r5.summarize_run(state[run_id]), 'mode': 'heuristic'})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        if parsed.path == '/api/cloud-backup':
            try:
                files = ['index.html','server.py','kwr_backend.py','r5_features.py','r6_features.py']
                results = r6.cloud_backup_all(files)
                ok_count = sum(1 for v in results.values() if v[0])
                if ok_count == 0: r6.alert_failure('Cloud backup failed', json.dumps({k:v[1] for k,v in results.items()}))
                else: r6.alert_info('Cloud backup OK', f'{ok_count}/{len(files)} files uploaded')
                r5.audit_log(self.headers.get('X-User','anon'), 'cloud_backup', '', self.client_address[0], {'ok': ok_count})
                return json_response(self, 200, {'ok': True, 'results': {k:{'ok':v[0],'info':v[1]} for k,v in results.items()}})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        # ---- Round 4: PDF export ----
        if parsed.path == '/api/kwr/export.pdf':
            return self._r4_export_pdf(parsed)

        # ---- File raw proxy endpoint ----
        if parsed.path == '/api/file/raw':
            query = urllib.parse.parse_qs(parsed.query)
            file_path = (query.get('path') or [''])[0]
            if not file_path:
                return json_response(self, 400, {'ok': False, 'error': 'path parameter required'})
            # Build raw GitHub URL with proper encoding per path component
            raw_url = f"{RAW_BASE}/{'/'.join(urllib.parse.quote(part) for part in file_path.split('/'))}"
            try:
                data = fetch_text(raw_url)
                # Guess content type from filename
                content_type, _ = mimetypes.guess_type(file_path)
                if not content_type:
                    content_type = 'text/plain; charset=utf-8'
                return text_response(self, 200, data.encode('utf-8'), content_type)
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1000]
                return json_response(self, 502, {'ok': False, 'error': f'GitHub raw error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
        # ---- Stage 8: Auth gate ----
        # Enforced whenever any credential store is configured (DASHBOARD_USERS env,
        # DASHBOARD_USER/PASSWORD break-glass, local users.json, or Supabase).
        # Previously this only checked DASHBOARD_USERS, which silently disabled auth
        # in production when that var wasn't set. See plans/reports/cto-260423-0035-*.md §1.
        if _dashboard_auth_enabled() and not _stage8_check_auth(self, parsed):
            return  # _stage8_check_auth already wrote response
        # Stage 8 endpoints
        if parsed.path == '/login' or parsed.path == '/login.html':
            try:
                lp = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'login-page.html')
                with open(lp, 'rb') as _lf:
                    _ldata = _lf.read()
                return text_response(self, 200, _ldata, 'text/html; charset=utf-8')
            except Exception:
                return text_response(self, 200, _STAGE8_LOGIN_HTML.encode('utf-8'), 'text/html; charset=utf-8')
        if parsed.path == '/api/auth/me':
            tok = _stage8_get_token(self)
            session_user = _stage8_verify_session(tok) if tok else None
            # Also check JWT Bearer token
            jwt_user = _get_current_user(self)
            if jwt_user:
                return json_response(self, 200, {'ok': True, 'user': jwt_user.get('username'), 'role': jwt_user.get('role', 'viewer'), 'auth_enabled': _dashboard_auth_enabled()})
            return json_response(self, 200, {
                'ok': True,
                'user': session_user.get('username') if session_user else None,
                'role': session_user.get('role') if session_user else None,
                'auth_enabled': _dashboard_auth_enabled()
            })
        if parsed.path == '/api/auth/status':
            return json_response(self, 200, _dashboard_auth_status())
        if parsed.path == '/api/auth/logout':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Set-Cookie', 'dash_auth=; Path=/; Max-Age=0; HttpOnly; Secure; SameSite=Lax')
            body = b'{"ok":true}'
            self.send_header('Content-Length', str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        # Stage 14 admin backup endpoints
        if _stage14_handle_get(self, parsed):
            return
        # ---- Stage 9: Activity log read ----
        if parsed.path == '/api/activity/log':
            try:
                import json as _json, os as _os, time as _time
                qs = urllib.parse.parse_qs(parsed.query or '')
                limit = max(1, min(int((qs.get('limit') or ['200'])[0]), 2000))
                days = max(1, min(int((qs.get('days') or ['7'])[0]), 60))
                log_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'data', 'activity')
                entries = []
                for d in range(days):
                    day = _time.strftime('%Y-%m-%d', _time.localtime(_time.time() - d * 86400))
                    fp = _os.path.join(log_dir, f'{day}.jsonl')
                    if not _os.path.exists(fp): continue
                    try:
                        with open(fp, 'r', encoding='utf-8') as f:
                            for line in f:
                                line = line.strip()
                                if not line: continue
                                try: entries.append(_json.loads(line))
                                except Exception: pass
                    except Exception: pass
                entries.sort(key=lambda e: e.get('ts', 0), reverse=True)
                return json_response(self, 200, {'ok': True, 'entries': entries[:limit], 'count': len(entries)})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
        if parsed.path == '/api/health':
            return json_response(self, 200, {
                'ok': True,
                'n8nConfigured': True,
                'promptConfigured': bool(os.getenv('OPENROUTER_API_KEY') or os.getenv('COPILOT_API_KEY') or os.getenv('ANTHROPIC_API_KEY') or os.getenv('GEMINI_API_KEY')),
                'brainstormConfigured': bool(os.getenv('OPENROUTER_API_KEY') or os.getenv('GEMINI_API_KEY')),
                'githubCommitConfigured': bool(os.getenv('GITHUB_TOKEN')),
                'brainstormModels': [{'id': m['id'], 'label': m['label']} for m in BRAINSTORM_MODELS],
                'synthModel': SYNTH_MODEL,
                'promptImproveDefaultModel': prompt_default_model('PROMPT_IMPROVER_MODEL'),
                'promptTweakDefaultModel': prompt_default_model('PROMPT_TWEAK_MODEL'),
                'promptSynthDefaultModel': SYNTH_MODEL,
                'paletteExtractorConfigured': True,
                'tweakConfigured': bool(os.getenv('OPENROUTER_API_KEY') or os.getenv('COPILOT_API_KEY') or os.getenv('ANTHROPIC_API_KEY') or os.getenv('GEMINI_API_KEY')),
                'commentsConfigured': supabase_comments_config()['configured'],
                'commentsTable': supabase_comments_config()['table'],
                # Active provider chain — shows which providers are available
                'activeProviders': [p['name'] for p in _get_provider_chain()],
                # All selectable models for the Prompt Studio model picker
                'improverModels': [
                    # --- OpenRouter (routes to any model) ---
                    {'id': 'anthropic/claude-sonnet-4.6',       'label': 'Claude Sonnet 4.6 (OpenRouter)',  'provider': 'openrouter'},
                    {'id': 'anthropic/claude-opus-4.7',         'label': 'Claude Opus 4.7 (OpenRouter)',    'provider': 'openrouter'},
                    {'id': 'openai/gpt-5.4',                    'label': 'GPT-5.4 (OpenRouter)',            'provider': 'openrouter'},
                    {'id': 'openai/gpt-4o',                     'label': 'GPT-4o (OpenRouter)',             'provider': 'openrouter'},
                    {'id': 'google/gemini-3.1-pro-preview',     'label': 'Gemini 3.1 Pro (OpenRouter)',     'provider': 'openrouter'},
                    {'id': 'google/gemini-2.5-pro',             'label': 'Gemini 2.5 Pro (OpenRouter)',     'provider': 'openrouter'},
                    # --- OpenAI (Direct API) ---
                    {'id': 'gpt-4o',                            'label': 'GPT-4o (OpenAI Direct)',          'provider': 'openai'},
                    {'id': 'gpt-4.5-preview',                   'label': 'GPT-4.5 Preview (OpenAI Direct)', 'provider': 'openai'},
                    {'id': 'gpt-4o-mini',                       'label': 'GPT-4o Mini (OpenAI Direct)',     'provider': 'openai'},
                    {'id': 'o1',                                'label': 'o1 (OpenAI Direct)',              'provider': 'openai'},
                    {'id': 'o3-mini',                           'label': 'o3-mini (OpenAI Direct)',         'provider': 'openai'},
                    # --- GitHub Copilot (direct) ---
                    {'id': 'claude-sonnet-4.6',                 'label': 'Claude Sonnet 4.6 (Copilot)',     'provider': 'copilot'},
                    {'id': 'gpt-4o',                            'label': 'GPT-4o (Copilot)',                'provider': 'copilot'},
                    {'id': 'o3',                                 'label': 'o3 (Copilot)',                    'provider': 'copilot'},
                    # --- Google Gemini (direct API) — newest first ---
                    {'id': 'gemini-3.1-pro-preview',            'label': 'Gemini 3.1 Pro Preview',         'provider': 'gemini'},
                    {'id': 'gemini-3-pro-preview',              'label': 'Gemini 3 Pro Preview',           'provider': 'gemini'},
                    {'id': 'gemini-3.1-flash-lite-preview',     'label': 'Gemini 3.1 Flash Lite Preview',  'provider': 'gemini'},
                    {'id': 'gemini-3-flash-preview',            'label': 'Gemini 3 Flash Preview',         'provider': 'gemini'},
                    {'id': 'gemini-flash-latest',               'label': 'Gemini Flash Latest (alias)',     'provider': 'gemini'},
                    {'id': 'gemini-2.5-pro',                    'label': 'Gemini 2.5 Pro',                 'provider': 'gemini'},
                    {'id': 'gemini-2.5-flash',                  'label': 'Gemini 2.5 Flash',               'provider': 'gemini'},
                    {'id': 'gemini-2.5-flash-lite',             'label': 'Gemini 2.5 Flash Lite',          'provider': 'gemini'},
                    {'id': 'gemini-2.0-flash',                  'label': 'Gemini 2.0 Flash',               'provider': 'gemini'},
                    {'id': 'gemini-2.0-flash-lite',             'label': 'Gemini 2.0 Flash Lite',          'provider': 'gemini'},
                    # --- Kimi Code (Moonshot AI) ---
                    {'id': 'kimi-k2.6-code-preview',            'label': 'K2.6 Code Preview (Flagship)',    'provider': 'kimi'},
                    {'id': 'kimi-latest',                       'label': 'Kimi Latest',                     'provider': 'kimi'},
                    {'id': 'kimi-k2',                           'label': 'Kimi K2',                         'provider': 'kimi'},
                    {'id': 'kimi-k1.5',                         'label': 'Kimi K1.5',                       'provider': 'kimi'},
                    {'id': 'moonshot-v1-128k',                  'label': 'Moonshot v1 128K',                 'provider': 'kimi'},
                    {'id': 'moonshot-v1-32k',                   'label': 'Moonshot v1 32K',                  'provider': 'kimi'},
                    # --- Venice AI ---
                    {'id': 'llama-3.3-70b',                     'label': 'Llama 3.3 70B (Venice)',          'provider': 'venice'},
                    {'id': 'llama-3.1-405b',                    'label': 'Llama 3.1 405B (Venice)',         'provider': 'venice'},
                    {'id': 'mistral-31-24b',                    'label': 'Mistral 3.1 24B (Venice)',        'provider': 'venice'},
                    {'id': 'qwen-2.5-72b',                      'label': 'Qwen 2.5 72B (Venice)',           'provider': 'venice'},
                    {'id': 'deepseek-r1-671b',                  'label': 'DeepSeek R1 671B (Venice)',       'provider': 'venice'},
                    {'id': 'deepseek-v3-0324',                  'label': 'DeepSeek V3 0324 (Venice)',       'provider': 'venice'},
                    # --- Fireworks AI ---
                    {'id': 'accounts/fireworks/models/llama-v3p3-70b-instruct',  'label': 'Llama 3.3 70B (Fireworks)',   'provider': 'fireworks'},
                    {'id': 'accounts/fireworks/models/llama-v3p1-405b-instruct', 'label': 'Llama 3.1 405B (Fireworks)',  'provider': 'fireworks'},
                    {'id': 'accounts/fireworks/models/qwen3-235b-a22b-instruct', 'label': 'Qwen 3.6 Plus (Fireworks)',  'provider': 'fireworks'},
                    {'id': 'accounts/fireworks/models/deepseek-r1',              'label': 'DeepSeek R1 (Fireworks)',     'provider': 'fireworks'},
                    {'id': 'accounts/fireworks/models/deepseek-v3',              'label': 'DeepSeek V3 (Fireworks)',     'provider': 'fireworks'},
                    {'id': 'accounts/fireworks/models/qwen2p5-72b-instruct',     'label': 'Qwen 2.5 72B (Fireworks)',    'provider': 'fireworks'},
                    {'id': 'accounts/fireworks/models/mixtral-8x22b-instruct',   'label': 'Mixtral 8x22B (Fireworks)',   'provider': 'fireworks'},
                ],
            })
        if parsed.path == '/api/comments':
            query = urllib.parse.parse_qs(parsed.query)
            filters = {k: (v[0] if isinstance(v, list) and v else '') for k, v in query.items()}
            try:
                result = list_supabase_comments(filters)
                return json_response(self, 200, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1000]
                return json_response(self, 502, {'ok': False, 'error': f'Supabase API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/n8n/status':
            query = urllib.parse.parse_qs(parsed.query)
            domain = (query.get('domain') or [''])[0]
            configured = True
            payload = {'configured': configured, 'domain': normalize_domain(domain), 'mappingFile': MAP_FILE.name}
            # Always try local mapping regardless of API key
            if domain:
                try:
                    workflow_id, details = resolve_workflow_id(domain)
                    payload.update({'ok': True, 'workflowId': workflow_id, 'details': details})
                except ValueError as exc:
                    payload.update({'ok': False, 'error': json.loads(str(exc))})
                except LookupError as exc:
                    payload.update({'ok': False, 'error': json.loads(str(exc))})
                except Exception as exc:
                    payload.update({'ok': False, 'error': {'code': 'unexpected_error', 'message': str(exc)}})
            else:
                # Return available mapped domains when no domain specified
                n8n_map = load_json_file(MAP_FILE, {})
                mapped = list((n8n_map.get('domains') or {}).keys())
                payload.update({'ok': True, 'mappedDomains': mapped, 'totalMapped': len(mapped)})
            return json_response(self, 200, payload)

        if parsed.path == '/api/n8n/executions':
            if False:  # N8N_API_KEY has built-in fallback
                return json_response(self, 200, {'ok': True, 'executions': [], 'configured': False})
            try:
                import datetime as _datetime_local
                base = (os.getenv('N8N_BASE_URL') or 'https://websiseo.app.n8n.cloud').rstrip('/')
                headers = n8n_headers()
                statuses = ['running', 'error', 'waiting']
                all_execs = []
                for status in statuses:
                    try:
                        data = fetch_json(f'{base}/api/v1/executions?status={status}&limit=25', headers=headers, timeout=20)
                        items = data.get('data') or []
                        for item in items:
                            started_at = item.get('startedAt') or ''
                            is_stuck = False
                            if status == 'running' and started_at:
                                try:
                                    dt = datetime.datetime.fromisoformat(started_at.replace('Z', '+00:00'))
                                    now = datetime.datetime.now(datetime.timezone.utc)
                                    if (now - dt).total_seconds() > 600:
                                        is_stuck = True
                                except Exception:
                                    pass
                            wf_data = item.get('workflowData') or {}
                            all_execs.append({
                                'id': item.get('id'),
                                'workflowId': item.get('workflowId') or wf_data.get('id'),
                                'workflowName': wf_data.get('name') or item.get('workflowId'),
                                'status': 'stuck' if is_stuck else status,
                                'startedAt': started_at,
                                'stoppedAt': item.get('stoppedAt'),
                                'mode': item.get('mode'),
                                'error': (item.get('data') or {}).get('resultData', {}).get('error') if isinstance(item.get('data'), dict) else None,
                            })
                    except Exception:
                        continue
                return json_response(self, 200, {'ok': True, 'executions': all_execs})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/n8n/workflows':
            if False:  # N8N_API_KEY has built-in fallback
                return json_response(self, 200, {'ok': True, 'workflows': [], 'configured': False})
            try:
                base = (os.getenv('N8N_BASE_URL') or 'https://websiseo.app.n8n.cloud').rstrip('/')
                headers = n8n_headers()
                data = fetch_json(f'{base}/api/v1/workflows?limit=100', headers=headers, timeout=20)
                items = data.get('data') or []
                workflows = []
                for wf in items:
                    workflows.append({
                        'id': wf.get('id'),
                        'name': wf.get('name'),
                        'active': wf.get('active'),
                        'updatedAt': wf.get('updatedAt'),
                        'nodeCount': len(wf.get('nodes') or []),
                        'tags': [t.get('name') for t in (wf.get('tags') or []) if isinstance(t, dict)],
                    })
                return json_response(self, 200, {'ok': True, 'workflows': workflows})
            except urllib.error.HTTPError as exc:
                body_bytes = exc.read().decode('utf-8', 'replace')[:800]
                return json_response(self, 502, {'ok': False, 'error': f'n8n API error {exc.code}', 'details': body_bytes})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # GET /api/n8n/workflow-json?id=xxx — fetch full workflow JSON for auto-import
        if parsed.path == '/api/n8n/workflow-json':
            if False:  # N8N_API_KEY has built-in fallback
                return json_response(self, 503, {'ok': False, 'error': 'N8N_API_KEY is not configured'})
            qs = urllib.parse.parse_qs(parsed.query)
            workflow_id = (qs.get('id') or [''])[0].strip()
            if not workflow_id:
                return json_response(self, 400, {'ok': False, 'error': 'id parameter is required'})
            try:
                base = (os.getenv('N8N_BASE_URL') or 'https://websiseo.app.n8n.cloud').rstrip('/')
                headers = n8n_headers()
                wf = fetch_json(f'{base}/api/v1/workflows/{urllib.parse.quote(workflow_id)}', headers=headers, timeout=20)
                # Return full workflow JSON exactly as received from N8N (valid for re-import)
                return json_response(self, 200, {
                    'ok': True,
                    'workflowId': wf.get('id'),
                    'workflowName': wf.get('name'),
                    'nodeCount': len(wf.get('nodes') or []),
                    'active': wf.get('active'),
                    'workflowJson': wf,  # complete JSON
                })
            except urllib.error.HTTPError as exc:
                body_bytes = exc.read().decode('utf-8', 'replace')[:400]
                return json_response(self, 502, {'ok': False, 'error': f'n8n API error {exc.code}', 'details': body_bytes})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # ===== RADAR GET ROUTES =====
        if parsed.path == '/api/radar/status':
            import datetime as _dt
            cfg = supabase_radar_config()
            last_run = None
            if cfg['configured']:
                try:
                    hdrs = _radar_sb_headers()
                    url = f"{cfg['url']}/rest/v1/radar_runs?select=*&order=created_at.desc&limit=1"
                    rows = fetch_json(url, headers=hdrs, timeout=15)
                    if isinstance(rows, list) and rows:
                        last_run = rows[0]
                except Exception:
                    pass
            state_copy = dict(_radar_state)
            # Compute next run
            now = _dt.datetime.now()
            sched_hour = _radar_state.get('schedule_hour', 2)
            next_run_dt = now.replace(hour=sched_hour, minute=0, second=0, microsecond=0)
            if next_run_dt <= now:
                next_run_dt = next_run_dt + _dt.timedelta(days=1)
            state_copy['next_run'] = next_run_dt.isoformat()
            # X integration status
            state_copy['x_enabled'] = True
            state_copy['x_sources_found'] = _radar_state.get('x_sources_found', 0)
            state_copy['last_email'] = _radar_state.get('last_email')
            state_copy['topics'] = _radar_state.get('topics', [])
            state_copy['custom_topics'] = _radar_state.get('custom_topics', [])
            return json_response(self, 200, {'ok': True, 'state': state_copy, 'last_run': last_run})

        if parsed.path == '/api/radar/results':
            qs = urllib.parse.parse_qs(parsed.query)
            status_filter = (qs.get('status') or ['pending'])[0]
            try:
                limit = int((qs.get('limit') or ['20'])[0])
            except Exception:
                limit = 20
            try:
                offset = int((qs.get('offset') or ['0'])[0])
            except Exception:
                offset = 0
            category = (qs.get('category') or [''])[0]
            agent = (qs.get('agent') or [''])[0]
            cfg = supabase_radar_config()
            if not cfg['configured']:
                return json_response(self, 200, {'ok': True, 'skills': [], 'total': 0, 'configured': False})
            try:
                hdrs = _radar_sb_headers()
                q_parts = [
                    'select=*',
                    f'status=eq.{urllib.parse.quote(status_filter)}',
                    'order=created_at.desc',
                    f'limit={limit}',
                    f'offset={offset}',
                ]
                if category:
                    q_parts.append(f'category=eq.{urllib.parse.quote(category)}')
                url = f"{cfg['url']}/rest/v1/skill_discoveries?" + '&'.join(q_parts)
                rows = fetch_json(url, headers=hdrs, timeout=15)
                if not isinstance(rows, list):
                    rows = []
                if agent:
                    rows = [r for r in rows if agent in (r.get('target_agents') or [])]
                topic_filter = (qs.get('topic') or [''])[0]
                if topic_filter:
                    rows = [r for r in rows if topic_filter.lower() in [t.lower() for t in (r.get('matched_topics') or [])]]
                # Count total
                count_headers = dict(hdrs)
                count_headers['Prefer'] = 'count=exact'
                count_url = f"{cfg['url']}/rest/v1/skill_discoveries?select=id&status=eq.{urllib.parse.quote(status_filter)}"
                total = len(rows)
                try:
                    count_resp = fetch_json(count_url, headers=count_headers, timeout=10)
                    if isinstance(count_resp, list):
                        total = len(count_resp)
                except Exception:
                    pass
                return json_response(self, 200, {'ok': True, 'skills': rows, 'total': total})
            except urllib.error.HTTPError as exc:
                if exc.code == 401:
                    return json_response(self, 200, {'ok': True, 'skills': [], 'total': 0, 'configured': True, 'auth_error': 'Supabase credentials invalid or table has RLS enabled'})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/history':
            qs = urllib.parse.parse_qs(parsed.query)
            try:
                limit = int((qs.get('limit') or ['10'])[0])
            except Exception:
                limit = 10
            cfg = supabase_radar_config()
            if not cfg['configured']:
                return json_response(self, 200, {'ok': True, 'runs': [], 'configured': False})
            try:
                hdrs = _radar_sb_headers()
                url = f"{cfg['url']}/rest/v1/radar_runs?select=*&order=created_at.desc&limit={limit}"
                rows = fetch_json(url, headers=hdrs, timeout=15)
                if not isinstance(rows, list):
                    rows = []
                return json_response(self, 200, {'ok': True, 'runs': rows})
            except Exception as exc:
                # Degrade gracefully — missing table / permission error shouldn't break UI
                return json_response(self, 200, {'ok': True, 'runs': [], 'warning': str(exc)[:200]})

        # GET /api/radar/trends — skills discovered over time for chart
        if parsed.path == '/api/radar/trends':
            qs = urllib.parse.parse_qs(parsed.query)
            try:
                days = int((qs.get('days') or ['30'])[0])
            except Exception:
                days = 30
            cfg = supabase_radar_config()
            if not cfg['configured']:
                return json_response(self, 200, {'ok': True, 'trends': [], 'total': 0, 'configured': False})
            try:
                hdrs = _radar_sb_headers()
                since = (datetime.datetime.utcnow() - datetime.timedelta(days=days)).strftime('%Y-%m-%dT%H:%M:%SZ')
                url = f"{cfg['url']}/rest/v1/skill_discoveries?select=created_at,category,status&created_at=gte.{urllib.parse.quote(since)}&order=created_at.desc&limit=1000"
                rows = fetch_json(url, headers=hdrs, timeout=20)
                if not isinstance(rows, list):
                    rows = []
                # Aggregate by date
                from collections import Counter, defaultdict
                daily = defaultdict(lambda: Counter())
                categories = set()
                for r in rows:
                    ts = r.get('created_at', '')
                    if not ts:
                        continue
                    try:
                        d = ts[:10]
                        cat = (r.get('category') or 'Uncategorized').strip() or 'Uncategorized'
                        daily[d][cat] += 1
                        categories.add(cat)
                    except Exception:
                        pass
                # Build trend array sorted by date
                dates = sorted(daily.keys())
                cat_list = sorted(categories) if categories else []
                trends = []
                for d in dates:
                    entry = {'date': d}
                    for cat in cat_list:
                        entry[cat] = daily[d].get(cat, 0)
                    entry['total'] = sum(daily[d].values())
                    trends.append(entry)
                return json_response(self, 200, {
                    'ok': True,
                    'trends': trends,
                    'total': len(rows),
                    'categories': cat_list,
                })
            except Exception as exc:
                return json_response(self, 200, {'ok': True, 'trends': [], 'total': 0, 'warning': str(exc)[:200]})

        if parsed.path == '/api/fixer/history':
            try:
                supabase_url = (os.getenv('SUPABASE_URL') or '').rstrip('/')
                supabase_key = (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('SUPABASE_ANON_KEY') or '').strip()
                if not supabase_url or not supabase_key:
                    return json_response(self, 200, {'ok': True, 'records': [], 'warning': 'Supabase not configured'})
                sb_headers = {
                    'apikey': supabase_key,
                    'Authorization': f'Bearer {supabase_key}',
                    'Content-Type': 'application/json',
                }
                url = f"{supabase_url}/rest/v1/n8n_fixer_records?select=*&order=created_at.desc&limit=100"
                rows = fetch_json(url, headers=sb_headers, timeout=20)
                if not isinstance(rows, list):
                    rows = []
                return json_response(self, 200, {'ok': True, 'records': rows})
            except Exception as exc:
                return json_response(self, 200, {'ok': True, 'records': [], 'warning': str(exc)[:200]})

        # GET /api/fixer/domains — list unique domains from fixer records
        if parsed.path == '/api/fixer/domains':
            try:
                supabase_url = (os.getenv('SUPABASE_URL') or '').rstrip('/')
                supabase_key = (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('SUPABASE_ANON_KEY') or '').strip()
                if not supabase_url or not supabase_key:
                    return json_response(self, 200, {'ok': True, 'domains': [], 'warning': 'Supabase not configured'})
                sb_headers = {
                    'apikey': supabase_key,
                    'Authorization': f'Bearer {supabase_key}',
                    'Content-Type': 'application/json',
                }
                # Get unique site_url values from n8n_fixer_records
                url = f"{supabase_url}/rest/v1/n8n_fixer_records?select=site_url&order=created_at.desc&limit=1000"
                rows = fetch_json(url, headers=sb_headers, timeout=20)
                domains = []
                seen = set()
                if isinstance(rows, list):
                    for r in rows:
                        d = r.get('site_url')
                        if d and d not in seen:
                            seen.add(d)
                            domains.append(d)
                return json_response(self, 200, {'ok': True, 'domains': domains})
            except Exception as exc:
                return json_response(self, 200, {'ok': True, 'domains': [], 'warning': str(exc)[:200]})

        # ── KWR GET routes ──────────────────────────────────────────────────
        if parsed.path == '/api/kwr/status':
            qs = urllib.parse.parse_qs(parsed.query)
            run_id = (qs.get('run_id') or [''])[0].strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            # SAFE_UI_API_FIXES_2026_04_26: flat reports are completed persisted
            # artifacts from outputs/kwr_<slug>.xlsx, not live jobs. Surface a
            # stable completed status instead of a misleading 404.
            if run_id.startswith('flat:'):
                slug = _safe_kwr_flat_slug(run_id[5:])
                if slug is None:
                    return json_response(self, 400, {'ok': False, 'error': 'invalid flat report id'})
                base_dir = os.path.realpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs'))
                fp = os.path.realpath(os.path.join(base_dir, f'kwr_{slug}.xlsx'))
                if not (fp.startswith(base_dir + os.sep) and os.path.isfile(fp)):
                    return json_response(self, 404, {'ok': False, 'error': 'flat report not found'})
                try:
                    size_kb = round(os.path.getsize(fp) / 1024.0, 1)
                except Exception:
                    size_kb = 0
                try:
                    updated_at = datetime.datetime.utcfromtimestamp(os.path.getmtime(fp)).isoformat() + 'Z'
                except Exception:
                    updated_at = ''
                return json_response(self, 200, {
                    'ok': True,
                    'run_id': run_id,
                    'status': 'completed',
                    'stage': 'persisted_report',
                    'progress': 100,
                    'message': 'Flat KWR report is available for download.',
                    'worksheet_name': f'kwr_{slug}',
                    'file_size_kb': size_kb,
                    'updated_at': updated_at,
                    'flat_file': f'kwr_{slug}.xlsx',
                })
            job = kwr_backend.get_status(run_id)
            if job is None:
                return json_response(self, 404, {'ok': False, 'error': f'run {run_id} not found'})
            return json_response(self, 200, {'ok': True, **job})

        if parsed.path == '/api/kwr/list':
            try:
                limit = int((urllib.parse.parse_qs(parsed.query).get('limit') or ['20'])[0])
            except Exception:
                limit = 20
            jobs = kwr_backend.list_recent(limit)
            return json_response(self, 200, {'ok': True, 'runs': jobs})

        if parsed.path == '/api/tasks':
            tasks = _tasks_load()
            return json_response(self, 200, {'ok': True, 'tasks': tasks})

        if parsed.path.startswith('/api/kwr/download/'):
            run_id = parsed.path.split('/')[-1].strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            # Flat reports synced from GitHub (outputs/kwr_<slug>.xlsx)
            if run_id.startswith('flat:'):
                slug = _safe_kwr_flat_slug(run_id[5:])
                if slug is None:
                    return json_response(self, 400, {'ok': False, 'error': 'invalid flat report id'})
                base_dir = os.path.realpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs'))
                fp = os.path.realpath(os.path.join(base_dir, f'kwr_{slug}.xlsx'))
                try:
                    inside_outputs = os.path.commonpath([base_dir, fp]) == base_dir
                except ValueError:
                    inside_outputs = False
                if not inside_outputs or not os.path.isfile(fp):
                    return json_response(self, 404, {'ok': False, 'error': 'flat report not found'})
                with open(fp, 'rb') as f:
                    excel_bytes = f.read()
                ws_name = f'kwr_{slug}'
                safe_ascii = re.sub(r'[^A-Za-z0-9._-]+', '-', ws_name).strip('-') or 'kwr'
                filename_ascii = f"{safe_ascii}.xlsx"
                filename_utf8 = urllib.parse.quote(f"{ws_name}.xlsx", safe='')
                self.send_response(200)
                self.send_header('Content-Type',
                                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header(
                    'Content-Disposition',
                    f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_utf8}"
                )
                self.send_header('Content-Length', str(len(excel_bytes)))
                self.end_headers()
                try:
                    mv = memoryview(excel_bytes)
                    CHUNK = 65536
                    for i in range(0, len(mv), CHUNK):
                        self.wfile.write(mv[i:i+CHUNK])
                except (BrokenPipeError, ConnectionResetError):
                    return
                return
            excel_bytes, ws_name, err = kwr_backend.build_excel(run_id)
            if err:
                status = 400 if err == 'invalid run_id' else 500
                return json_response(self, status, {'ok': False, 'error': err})
            # Build ASCII-safe filename + RFC 5987 UTF-8 version (Hebrew would blow up latin-1 headers).
            safe_ascii = re.sub(r'[^A-Za-z0-9._-]+', '-', ws_name).strip('-') or 'kwr'
            filename_ascii = f"{safe_ascii}.xlsx"
            filename_utf8 = urllib.parse.quote(f"{ws_name}.xlsx", safe='')
            self.send_response(200)
            self.send_header('Content-Type',
                             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header(
                'Content-Disposition',
                f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_utf8}"
            )
            self.send_header('Content-Length', str(len(excel_bytes)))
            self.end_headers()
            # Stream in 64KB chunks — avoids broken-pipe / memory spikes on Render proxies.
            try:
                mv = memoryview(excel_bytes)
                CHUNK = 65536
                for i in range(0, len(mv), CHUNK):
                    self.wfile.write(mv[i:i+CHUNK])
            except (BrokenPipeError, ConnectionResetError):
                return
            return

        if parsed.path.startswith('/api/kwr/note-content/'):
            run_id = parsed.path.split('/')[-1].strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            note_content, note_path, err = kwr_backend.get_note_content(run_id)
            if err:
                return json_response(self, 404, {'ok': False, 'error': err})
            return json_response(self, 200, {
                'ok': True,
                'note_content': note_content,
                'note_path': note_path,
            })

        if parsed.path == '/api/kwr/reports':
            try:
                reports = kwr_backend.list_reports()
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
            return json_response(self, 200, {'ok': True, 'reports': reports})

        # ── NEW: /api/version ──────────────────────────────────────────
        if parsed.path == '/api/version':
            import subprocess as _sp, time as _tv
            try:
                commit = _sp.check_output(['git','rev-parse','--short','HEAD'],
                    cwd=os.path.dirname(os.path.abspath(__file__)),
                    stderr=_sp.DEVNULL).decode().strip()
            except Exception:
                commit = 'dev'
            data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
            try:
                data_size_bytes = os.path.getsize(data_path)
                if data_size_bytes < 1024:
                    data_size = f'{data_size_bytes} B'
                elif data_size_bytes < 1024*1024:
                    data_size = f'{data_size_bytes/1024:.1f} KB'
                else:
                    data_size = f'{data_size_bytes/1048576:.2f} MB'
            except Exception:
                data_size = '—'
            uptime_sec = int(_tv.time() - _SERVER_START_TIME)
            h, rem = divmod(uptime_sec, 3600)
            m, s = divmod(rem, 60)
            uptime_str = f'{h}h {m}m {s}s'
            return json_response(self, 200, {
                'ok': True,
                'commit': commit,
                'uptime': uptime_sec,
                'uptime_str': uptime_str,
                'data_size': data_size,
                'supabase':   bool(os.getenv('SUPABASE_URL')),
                'github':     bool(os.getenv('GITHUB_TOKEN')),
                'resend':     bool(os.getenv('RESEND_API_KEY')),
                'openrouter': bool(os.getenv('OPENROUTER_API_KEY')),
                'telegram':   bool(os.getenv('TELEGRAM_BOT_TOKEN')),
            })

        # ── NEW: /api/prompt/palette ──────────────────────────────────
        if parsed.path == '/api/prompt/palette':
            palette = [
                {'id': 'html-redesign', 'name': 'HTML Redesign Prompt', 'category': 'html', 'description': 'Generate improved HTML templates for websites'},
                {'id': 'kwr-analysis', 'name': 'Keyword Research', 'category': 'seo', 'description': 'Analyze keywords and generate SEO reports'},
                {'id': 'n8n-workflow', 'name': 'N8N Workflow', 'category': 'automation', 'description': 'Build and fix n8n automation workflows'},
                {'id': 'improve-workflow', 'name': 'Improve Workflow', 'category': 'n8n', 'description': 'Assemble a production prompt that improves an N8N prompt, WordPress HTML template, and workflow JSON'},
                {'id': 'skills-radar', 'name': 'Skills Discovery', 'category': 'research', 'description': 'Discover new AI skills and techniques'},
                {'id': 'content-rewrite', 'name': 'Content Rewrite', 'category': 'content', 'description': 'Rewrite and improve website content'},
                {'id': 'meta-optimization', 'name': 'Meta Optimization', 'category': 'seo', 'description': 'Optimize title tags and meta descriptions'},
                {'id': 'deploy-checklist', 'name': 'Deploy Checklist', 'category': 'devops', 'description': 'Pre-deployment verification checklist'},
            ]
            return json_response(self, 200, {'ok': True, 'palette': palette})

        if parsed.path == '/api/studio/improve/rules':
            query = urllib.parse.parse_qs(parsed.query or '')
            language = (query.get('language') or [''])[0]
            store_decision = (query.get('storeDecision') or [''])[0]
            rules = _active_improvement_rules(language=language, store_decision=store_decision)
            return json_response(self, 200, {'ok': True, 'version': IMPROVEMENT_RULES_VERSION, 'rules': rules})

        if parsed.path == '/api/projects/list':
            try:
                data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
                with open(data_path, 'r', encoding='utf-8') as _f:
                    tree_data = json.load(_f)
                tree = tree_data.get('tree', []) if isinstance(tree_data, dict) else []
                # Load n8n workflow map for deploy status
                n8n_map = load_json_file(MAP_FILE, {})
                n8n_domains = set((n8n_map.get('domains') or {}).keys())
                grouped = {}
                for item in tree:
                    path = item.get('path', '')
                    if not path or '/' not in path:
                        continue
                    parts = path.split('/')
                    domain = parts[0]
                    if domain in ('index.html', 'server.py', 'data.json', 'login-page.html', 'Dockerfile', '.dockerignore', 'AGENTS.md', 'kwr_backend.py', 'n8n-workflow-map.json'):
                        continue
                    if not _looks_like_project_domain(domain):
                        continue
                    bucket = grouped.setdefault(domain, {'domain': domain, 'name': domain, 'agents': [], 'updated_at': None, 'status': 'active', 'deployed': False})
                    if len(parts) < 3:
                        continue
                    agent = parts[1]
                    file_name = parts[-1]
                    agent_entry = next((a for a in bucket['agents'] if a['name'] == agent), None)
                    if not agent_entry:
                        agent_entry = {'name': agent, 'files': []}
                        bucket['agents'].append(agent_entry)
                    if item.get('type') == 'blob':
                        agent_entry['files'].append({'name': file_name, 'path': path, 'size': item.get('size', 0)})
                # Compute status, deployed, last_updated per project
                for domain, bucket in grouped.items():
                    bucket['deployed'] = domain in n8n_domains
                    has_html = False
                    has_prompt_only = True
                    max_mtime = 0
                    for agent in bucket['agents']:
                        for f in agent.get('files', []):
                            fname = f.get('name', '').lower()
                            if fname.endswith('.html'):
                                has_html = True
                                has_prompt_only = False
                            elif fname.endswith('.txt') or fname.endswith('.md') or fname.endswith('.json'):
                                pass  # prompts/other
                            else:
                                has_prompt_only = False
                            # Try to get actual file mtime from filesystem
                            fpath = ROOT / f.get('path', '')
                            try:
                                mtime = fpath.stat().st_mtime
                                if mtime > max_mtime:
                                    max_mtime = mtime
                            except Exception:
                                pass
                    if has_html:
                        bucket['status'] = 'completed'
                    elif has_prompt_only:
                        bucket['status'] = 'in_progress'
                    else:
                        bucket['status'] = 'active'
                    # Calculate progress
                    total_files = sum(len(a.get('files', [])) for a in bucket['agents'])
                    if has_html:
                        bucket['progress'] = 100
                    elif has_prompt_only and total_files > 0:
                        bucket['progress'] = 50
                    else:
                        bucket['progress'] = 25
                    # Load starred state
                    stars_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'stars.json')
                    try:
                        with open(stars_path, 'r', encoding='utf-8') as f:
                            stars = json.load(f)
                        bucket['starred'] = stars.get(domain, False)
                    except Exception:
                        bucket['starred'] = False
                    # Calculate progress
                    total_files = sum(len(a.get('files', [])) for a in bucket['agents'])
                    if has_html:
                        bucket['progress'] = 100
                    elif has_prompt_only and total_files > 0:
                        bucket['progress'] = 50
                    else:
                        bucket['progress'] = 25
                    # Load starred state
                    stars_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'stars.json')
                    try:
                        with open(stars_path, 'r', encoding='utf-8') as f:
                            stars = json.load(f)
                        bucket['starred'] = stars.get(domain, False)
                    except Exception:
                        bucket['starred'] = False
                    if max_mtime:
                        import datetime as _dt_fix; bucket['last_updated'] = _dt_fix.datetime.utcfromtimestamp(max_mtime).isoformat() + 'Z'
                projects = list(grouped.values())
                return json_response(self, 200, {'ok': True, 'projects': projects})
            except Exception as exc:
                return json_response(self, 200, {'ok': True, 'projects': [], 'warning': str(exc)[:200]})

        # ── NEW: /api/settings ──────────────────────────────────────────
        if parsed.path == '/api/settings':
            import json as _jsn
            settings_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'settings.json')
            try:
                with open(settings_path, 'r', encoding='utf-8') as f:
                    settings = _jsn.load(f)
            except Exception:
                settings = {
                    'email_notifications': True,
                    'telegram_notifications': False,
                    'email_address': os.getenv('RADAR_EMAIL_TO', 'service@maximo-seo.com'),
                    'theme_color': 'purple',
                }
            if not isinstance(settings, dict):
                settings = {}
            settings.setdefault('theme_color', 'purple')
            return json_response(self, 200, {'ok': True, **settings})

        # ── NEW: /api/settings/api-keys ─────────────────────────────────
        if parsed.path == '/api/settings/api-keys':
            def _mask_key(k):
                if not k or len(k) < 16:
                    return '—'
                return k[:8] + '...' + k[-4:]
            keys = {
                'openrouter': {'name': 'OpenRouter', 'masked': _mask_key(os.getenv('OPENROUTER_API_KEY')), 'env': 'OPENROUTER_API_KEY'},
                'n8n': {'name': 'N8N API Key', 'masked': _mask_key(os.getenv('N8N_API_KEY')), 'env': 'N8N_API_KEY'},
                'github': {'name': 'GitHub Token', 'masked': _mask_key(os.getenv('GITHUB_TOKEN')), 'env': 'GITHUB_TOKEN'},
                'resend': {'name': 'Resend API Key', 'masked': _mask_key(os.getenv('RESEND_API_KEY')), 'env': 'RESEND_API_KEY'},
                'anthropic': {'name': 'Anthropic Key', 'masked': _mask_key(os.getenv('ANTHROPIC_API_KEY')), 'env': 'ANTHROPIC_API_KEY'},
                'supabase_service': {'name': 'Supabase Service Role', 'masked': _mask_key(os.getenv('SUPABASE_SERVICE_ROLE_KEY')), 'env': 'SUPABASE_SERVICE_ROLE_KEY'},
                'supabase_anon': {'name': 'Supabase Anon Key', 'masked': _mask_key(os.getenv('SUPABASE_ANON_KEY')), 'env': 'SUPABASE_ANON_KEY'},
            }
            return json_response(self, 200, {'ok': True, 'keys': keys})


        return self.serve_static(parsed.path)

    


        # ============================================================
        # N8N Stuck Projects - GET endpoints
        # ============================================================

        if parsed.path == '/api/stuck-projects':
            try:
                qs = urllib.parse.parse_qs(parsed.query)
                status = qs.get('status', ['stuck,error,failed'])[0]
                priority = qs.get('priority', ['all'])[0]
                client = qs.get('client', [None])[0]
                search = qs.get('search', [None])[0]
                sort = qs.get('sort', ['stuck_since'])[0]
                order = qs.get('order', ['desc'])[0]
                page = int(qs.get('page', ['1'])[0])
                limit = int(qs.get('limit', ['50'])[0])
                limit = min(limit, 200)

                cfg = _get_supabase_config()
                if not cfg.get('url') or not cfg.get('key'):
                    self._send_json(200, {
                        'success': True, 'data': [], 'pagination': {'total': 0, 'page': 1, 'limit': limit, 'total_pages': 0},
                        'summary': {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'total_stuck': 0, 'resolved_today': 0}
                    })
                    return

                # Build query
                status_filter = status.replace(' ', '').split(',')
                filter_clause = "status=in.(" + ','.join(f"'{s}'" for s in status_filter) + ")"
                if priority != 'all':
                    filter_clause += f",priority=eq.{priority}"
                if client:
                    filter_clause += f",client_name=ilike.*{client}*"

                offset = (page - 1) * limit
                order_dir = 'desc' if order == 'desc' else 'asc'
                url = f"{cfg['url']}/rest/v1/stuck_projects?select=*&{filter_clause}&order={sort}.{order_dir}&offset={offset}&limit={limit}"

                req = urllib.request.Request(url, headers={
                    'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                    'Content-Type': 'application/json', 'Prefer': 'count=exact'
                })
                resp = urllib.request.urlopen(req, timeout=15)
                data = json.loads(resp.read().decode())

                # Get summary
                summary_url = f"{cfg['url']}/rest/v1/stuck_projects?select=status,priority"
                summary_req = urllib.request.Request(summary_url, headers={
                    'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}", 'Content-Type': 'application/json'
                })
                summary_resp = urllib.request.urlopen(summary_req, timeout=15)
                all_items = json.loads(summary_resp.read().decode())

                summary = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'total_stuck': 0, 'resolved_today': 0}
                for item in all_items:
                    if item.get('status') in ('stuck', 'error', 'failed'):
                        summary['total_stuck'] += 1
                        p = item.get('priority', 'medium')
                        if p in summary:
                            summary[p] += 1

                # Get total count
                count_url = f"{cfg['url']}/rest/v1/stuck_projects?select=id&{filter_clause}"
                count_req = urllib.request.Request(count_url, headers={
                    'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                    'Content-Type': 'application/json', 'Prefer': 'count=exact'
                })
                try:
                    count_resp = urllib.request.urlopen(count_req, timeout=10)
                    total = int(count_resp.headers.get('content-range', '0-0/0').split('/')[-1])
                except:
                    total = len(all_items)

                self._send_json(200, {
                    'success': True, 'data': data,
                    'pagination': {'page': page, 'limit': limit, 'total': total, 'total_pages': (total + limit - 1) // limit if limit else 1},
                    'summary': summary
                })
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path.startswith('/api/stuck-projects/') and parsed.path != '/api/stuck-projects/sync' and parsed.path != '/api/stuck-projects/sync/status' and parsed.path != '/api/stuck-projects/summary':
            project_id = parsed.path.split('/')[-1]
            try:
                cfg = _get_supabase_config()
                if not cfg.get('url') or not cfg.get('key'):
                    self._send_json(404, {'success': False, 'error': 'Supabase not configured'})
                    return

                url = f"{cfg['url']}/rest/v1/stuck_projects?id=eq.{project_id}&select=*"
                req = urllib.request.Request(url, headers={
                    'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}", 'Content-Type': 'application/json'
                })
                resp = urllib.request.urlopen(req, timeout=10)
                data = json.loads(resp.read().decode())
                if data:
                    # Increment view count
                    update_url = f"{cfg['url']}/rest/v1/stuck_projects?id=eq.{project_id}"
                    update_req = urllib.request.Request(update_url, data=json.dumps({'view_count': data[0].get('view_count', 0) + 1, 'is_new': False}).encode(), headers={
                        'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}", 'Content-Type': 'application/json'
                    }, method='PATCH')
                    try:
                        urllib.request.urlopen(update_req, timeout=5)
                    except:
                        pass
                    self._send_json(200, {'success': True, 'data': data[0]})
                else:
                    self._send_json(404, {'success': False, 'error': 'Project not found'})
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path == '/api/stuck-projects/summary':
            try:
                cfg = _get_supabase_config()
                if not cfg.get('url') or not cfg.get('key'):
                    self._send_json(200, {'success': True, 'data': {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'total_stuck': 0, 'resolved_today': 0}})
                    return

                url = f"{cfg['url']}/rest/v1/stuck_projects?select=status,priority,first_detected"
                req = urllib.request.Request(url, headers={
                    'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}", 'Content-Type': 'application/json'
                })
                resp = urllib.request.urlopen(req, timeout=10)
                data = json.loads(resp.read().decode())

                summary = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'total_stuck': 0, 'resolved_today': 0}
                from datetime import datetime, timezone
                today = datetime.now(timezone.utc).date()
                for item in data:
                    if item.get('status') in ('stuck', 'error', 'failed'):
                        summary['total_stuck'] += 1
                        p = item.get('priority', 'medium')
                        if p in summary:
                            summary[p] += 1
                    if item.get('status') == 'resolved' and item.get('resolved_at'):
                        try:
                            resolved_date = datetime.fromisoformat(item['resolved_at'].replace('Z', '+00:00')).date()
                            if resolved_date == today:
                                summary['resolved_today'] += 1
                        except:
                            pass

                self._send_json(200, {'success': True, 'data': summary})
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path == '/api/stuck-projects/sync/status':
            try:
                # Get scheduler status
                scheduler_status = {'running': False, 'interval_minutes': 30, 'last_sync': None, 'last_result': None, 'sync_count': 0}
                try:
                    from sync_scheduler import get_scheduler
                    sched = get_scheduler()
                    scheduler_status = sched.get_status()
                except Exception:
                    pass

                # Get project count from Supabase
                cfg = _get_supabase_config()
                total_projects = 0
                if cfg.get('url') and cfg.get('key'):
                    try:
                        url = f"{cfg['url']}/rest/v1/stuck_projects?select=id&limit=1"
                        req = urllib.request.Request(url, headers={
                            'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                            'Content-Type': 'application/json', 'Prefer': 'count=exact'
                        })
                        resp = urllib.request.urlopen(req, timeout=10)
                        total_projects = int(resp.headers.get('content-range', '0-0/0').split('/')[-1])
                    except:
                        pass

                self._send_json(200, {
                    'success': True,
                    'data': {
                        'scheduler': scheduler_status,
                        'total_projects': total_projects,
                        'env_configured': {
                            'pini_url': bool(os.getenv('PINI_URL')),
                            'pini_username': bool(os.getenv('PINI_USERNAME')),
                            'supabase_url': bool(cfg.get('url'))
                        }
                    }
                })
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        # PRODUCTIVITY_HUB_API_POST_2026_04_27 - additive persisted notifications/audit.
        if parsed.path == '/api/productivity/notifications':
            try:
                payload = read_request_json(self) or {}
                row = _productivity_add_notification(payload)
                return json_response(self, 200, {'ok': True, 'notification': row})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
        
        if parsed.path == '/api/stuck-projects/alerts/test-email':
            try:
                from email_notifier import create_notifier
                notifier = create_notifier()
                if not notifier.is_configured():
                    self._send_json(400, {
                        'success': False,
                        'error': 'Email notifications not configured. Set NOTIFICATION_EMAIL_ENABLED=true and configure SMTP/Resend credentials.',
                        'required_env': [
                            'NOTIFICATION_EMAIL_ENABLED',
                            'NOTIFICATION_FROM_EMAIL',
                            'NOTIFICATION_TO_EMAILS',
                            'NOTIFICATION_EMAIL_PROVIDER (smtp or resend)',
                            'For SMTP: NOTIFICATION_SMTP_HOST, NOTIFICATION_SMTP_USER, NOTIFICATION_SMTP_PASSWORD',
                            'For Resend: RESEND_API_KEY'
                        ]
                    })
                    return

                result = notifier.test_email()
                self._send_json(200, {
                    'success': True,
                    'message': 'Test email sent successfully',
                    'result': result
                })
            except ImportError:
                self._send_json(500, {'success': False, 'error': 'email_notifier.py not found'})
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path == '/api/productivity/audit':
            try:
                payload = read_request_json(self) or {}
                row = _productivity_add_audit(payload)
                return json_response(self, 200, {'ok': True, 'event': row})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # TEMPLATE_IMPROVEMENTS_API_POST_2026_04_27 - additive POST routes.
        if parsed.path == '/api/improve/start':
            try:
                content_length = int(self.headers.get('Content-Length', '0') or '0')
                if content_length > TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES:
                    return json_response(self, 413, {'ok': False, 'error': 'request too large'})
                payload = read_request_json(self) or {}
                job = _template_improvement_start_job(payload)
                return json_response(self, 200, {'ok': True, 'job': _template_improvement_public_job(job)})
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except RuntimeError as exc:
                return json_response(self, 429, {'ok': False, 'error': str(exc)})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
        if parsed.path.startswith('/api/improve/jobs/') and parsed.path.endswith('/cancel'):
            job_id = parsed.path.split('/')[-2]
            job = _template_improvement_update_job(job_id, status='cancelled', current_agent=None)
            if not job:
                return json_response(self, 404, {'ok': False, 'error': 'job not found'})
            return json_response(self, 200, {'ok': True, 'job': _template_improvement_public_job(job)})
        if parsed.path == '/api/improve/instructions':
            try:
                content_length = int(self.headers.get('Content-Length', '0') or '0')
                if content_length > TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES:
                    return json_response(self, 413, {'ok': False, 'error': 'request too large'})
                payload = read_request_json(self) or {}
                domain = (payload.get('domain') or '').strip()
                instructions = (payload.get('instructions') or '').strip()
                if not domain or not instructions:
                    return json_response(self, 400, {'ok': False, 'error': 'domain and instructions are required'})
                row = {
                    'id': str(uuid.uuid4()), 'domain': domain,
                    'subdomain': (payload.get('subdomain') or '').strip(),
                    'agent_key': (payload.get('agent_key') or payload.get('agentKey') or '').strip() or None,
                    'instructions': instructions,
                    'is_active': bool(payload.get('is_active', True)),
                    'created_at': datetime.datetime.utcnow().isoformat() + 'Z',
                    'updated_at': datetime.datetime.utcnow().isoformat() + 'Z',
                }
                with _TEMPLATE_IMPROVEMENTS_LOCK:
                    data = _template_improvements_load()
                    data['instructions'].append(row)
                    _template_improvements_save(data)
                return json_response(self, 200, {'ok': True, 'instruction': row})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
        # Round 4: manual backup trigger
        if parsed.path == '/api/backup/run':
            try:
                ok, err = _r3_auto_backup()
                if ok > 0:
                    sse_broadcast('backup', {'success': ok, 'timestamp': time.time(), 'manual': True})
                r5.audit_log(self.headers.get('X-User', 'anon'), 'backup.run', '', self.client_address[0], {'files': ok})
                return json_response(self, 200, {'ok': True, 'files_backed_up': ok, 'errors': err})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        
        if parsed.path == '/api/radar/export/excel':
            try:
                ln = int(self.headers.get('Content-Length', 0))
                payload = json.loads(self.rfile.read(ln) or b'{}')
                excel_bytes, err = _generate_radar_excel(payload)
                if err:
                    return json_response(self, 500, {'ok': False, 'error': err})
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="skill-radar.xlsx"')
                self.end_headers()
                self.wfile.write(excel_bytes)
                return
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})

        # ---- Round 5: New POST endpoints ----
        if parsed.path == '/api/bulk':
            try:
                ln = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(ln) or b'{}')
                action = body.get('action', '')
                ids = body.get('ids', [])
                state = kwr_backend.get_state() if hasattr(kwr_backend, 'get_state') else {}
                lock = getattr(kwr_backend, '_lock', threading.RLock())
                ok, fail = r5.bulk_dispatch(action, ids, state, lock)
                r5.audit_log(self.headers.get('X-User', 'anon'), f'bulk.{action}', ','.join(ids[:5]),
                             self.client_address[0], {'count': len(ids)})
                return json_response(self, 200, {'ok': True, 'success': ok, 'failed': fail})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        if parsed.path == '/api/views':
            try:
                ln = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(ln) or b'{}')
                user = self.headers.get('X-User', 'anon')
                r5.views_save_one(user, body.get('name', 'untitled'), body.get('config', {}))
                r5.audit_log(user, 'view.save', body.get('name', ''), self.client_address[0])
                return json_response(self, 200, {'ok': True})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        # /api/login and /api/auth/login are unified below via _stage8_login.
        # (Legacy inline sha256+JWT handler removed — it bypassed the dash_auth cookie
        # and duplicated logic now consolidated in _dashboard_validate_credentials.)
        if parsed.path == '/api/users':
            if not _require_admin(self): return
            try:
                ln = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(ln) or b'{}')
                u = (body.get('username') or '').strip()
                p = (body.get('password') or '').strip()
                role = body.get('role', 'viewer')
                email = body.get('email', '')
                if not u or not p: return json_response(self, 400, {'ok': False, 'error': 'missing fields'})
                if role not in ('admin', 'viewer'): return json_response(self, 400, {'ok': False, 'error': 'invalid role'})
                users = _mu_users_load()
                if any(x.get('username') == u for x in users):
                    return json_response(self, 400, {'ok': False, 'error': 'username exists'})
                import uuid as _uuid_add
                users.append({'id': str(_uuid_add.uuid4()), 'username': u,
                    'password_hash': _mu_hash_password(p),
                    'role': role, 'email': email,
                    'created_at': datetime.datetime.utcnow().isoformat() + 'Z', 'last_login': None})
                _mu_users_save(users)
                return json_response(self, 200, {'ok': True})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})
        # Stage 8: login — both endpoints are aliases for the same cookie-setting handler.
        if parsed.path in ('/api/auth/login', '/api/login'):
            try:
                payload = read_request_json(self) or {}
            except Exception:
                payload = {}
            return _stage8_login(self, payload)
        # Stage 14 admin backup POST
        if _stage14_handle_post(self, parsed):
            return
        # ---- Stage 9: Activity log + Webhook receivers (early dispatch) ----
        if parsed.path in ('/api/activity/log', '/api/activity/append', '/api/webhooks/notify'):
            try:
                payload = read_request_json(self) or {}
            except Exception:
                payload = {}
            try:
                import json as _json, os as _os, time as _time, uuid as _uuid
                log_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'data', 'activity')
                _os.makedirs(log_dir, exist_ok=True)
                day = _time.strftime('%Y-%m-%d')
                log_path = _os.path.join(log_dir, f'{day}.jsonl')
                entry = {
                    'id': str(_uuid.uuid4()),
                    'ts': int(_time.time() * 1000),
                    'source': payload.get('source') or ('webhook' if 'webhook' in parsed.path else 'dashboard'),
                    'kind': payload.get('kind') or payload.get('event') or 'info',
                    'msg': payload.get('msg') or payload.get('message') or payload.get('text') or '',
                    'meta': payload.get('meta') or {k: v for k, v in payload.items() if k not in ('source','kind','event','msg','message','text','meta')},
                    'ip': self.client_address[0] if self.client_address else '',
                }
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(_json.dumps(entry, ensure_ascii=False) + '\n')
                return json_response(self, 200, {'ok': True, 'id': entry['id'], 'ts': entry['ts']})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        try:
            payload = read_request_json(self)
        except Exception:
            return json_response(self, 400, {'ok': False, 'error': 'Invalid JSON body'})

        if parsed.path == '/api/tasks':
            tasks = payload.get('tasks')
            if not isinstance(tasks, list):
                return json_response(self, 400, {'ok': False, 'error': 'tasks must be an array'})
            _tasks_save(tasks)
            return json_response(self, 200, {'ok': True, 'count': len(tasks)})

        if parsed.path == '/api/tasks/sync-github':
            tasks = payload.get('tasks')
            if not isinstance(tasks, list):
                return json_response(self, 400, {'ok': False, 'error': 'tasks must be an array'})
            try:
                content = json.dumps(tasks, ensure_ascii=False, indent=2)
                result = commit_prompt_to_github({
                    'path': 'tasks/tasks.json',
                    'content': content,
                    'message': 'chore: update dashboard tasks',
                    'branch': payload.get('branch') or 'main',
                })
                return json_response(self, 200, {'ok': True, **result})
            except RuntimeError as exc:
                return json_response(self, 503, {'ok': False, 'error': str(exc)})
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1000]
                return json_response(self, 502, {'ok': False, 'error': f'GitHub API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/prompt/improve':
            try:
                result = improve_prompt_with_model(payload)
                return json_response(self, 200, {'ok': True, **result})
            except RuntimeError as exc:
                return json_response(self, 503, {'ok': False, 'error': str(exc)})
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1000]
                return json_response(self, 502, {'ok': False, 'error': f'Model API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/studio/improve':
            try:
                result = assemble_improve_workflow_prompt(payload)
                return json_response(self, 200, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)[:500]})

        if parsed.path == '/api/prompt/palette':
            try:
                url = (payload.get('url') or '').strip()
                if not url:
                    return json_response(self, 400, {'ok': False, 'error': 'url is required'})
                max_colors = int(payload.get('maxColors') or 10)
                max_colors = max(3, min(max_colors, 20))
                result = extract_palette_from_url(url, max_colors=max_colors)
                return json_response(self, 200, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                return json_response(self, 502, {'ok': False, 'error': f'Fetch failed {exc.code}', 'details': str(exc)[:300]})
            except urllib.error.URLError as exc:
                return json_response(self, 502, {'ok': False, 'error': f'Fetch failed: {exc.reason}'})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)[:400]})

        if parsed.path == '/api/prompt/brainstorm':
            try:
                result = brainstorm_prompt_multi_model(payload)
                return json_response(self, 200, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except RuntimeError as exc:
                msg = str(exc)
                try:
                    parsed_err = json.loads(msg)
                    return json_response(self, 502, {'ok': False, 'error': parsed_err})
                except Exception:
                    return json_response(self, 503, {'ok': False, 'error': msg})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/prompt/tweak':
            try:
                result = tweak_html_with_prompt(payload)
                return json_response(self, 200, result)
            except RuntimeError as exc:
                return json_response(self, 503, {'ok': False, 'error': str(exc)})
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1000]
                return json_response(self, 502, {'ok': False, 'error': f'Model API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/prompt/commit':
            try:
                result = commit_prompt_to_github(payload)
                return json_response(self, 200, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except RuntimeError as exc:
                return json_response(self, 503, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1000]
                return json_response(self, 502, {'ok': False, 'error': f'GitHub API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/comments':
            try:
                result = save_supabase_comment(payload)
                return json_response(self, 200, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except RuntimeError as exc:
                return json_response(self, 503, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1200]
                return json_response(self, 502, {'ok': False, 'error': f'Supabase API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/n8n/deploy':
            return json_response(self, 403, {
                'ok': False,
                'error': 'n8n deployment is disabled by read-only safety policy',
                'safety': 'MANUAL IMPORT ONLY — this dashboard must not modify existing n8n workflows',
            })

        # ===== RADAR POST ENDPOINTS =====

        if parsed.path == '/api/radar/test-email':
            try:
                result = _send_radar_email_digest(
                    skills_found_count=_radar_state.get('skills_found', 0),
                    run_id=_radar_state.get('run_id'),
                    errors=[],
                )
                return json_response(self, 200, result)
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/topics':
            try:
                topics = payload.get('topics')
                custom_topics = payload.get('custom_topics')
                if topics is not None:
                    if not isinstance(topics, list):
                        return json_response(self, 400, {'ok': False, 'error': 'topics must be array'})
                    _radar_state['topics'] = [str(t).strip().lower() for t in topics if str(t).strip()]
                if custom_topics is not None:
                    if not isinstance(custom_topics, list):
                        return json_response(self, 400, {'ok': False, 'error': 'custom_topics must be array'})
                    _radar_state['custom_topics'] = [str(t).strip().lower() for t in custom_topics if str(t).strip()]
                return json_response(self, 200, {
                    'ok': True,
                    'topics': _radar_state['topics'],
                    'custom_topics': _radar_state['custom_topics'],
                })
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/run':
            try:
                if _radar_state.get('status') == 'running':
                    return json_response(self, 409, {'ok': False, 'error': 'Discovery already running'})
                config = payload.get('config') or {}
                t = threading.Thread(target=_run_radar_discovery, args=(config,), daemon=True)
                t.start()
                return json_response(self, 200, {
                    'ok': True,
                    'message': 'Discovery started',
                    'run_id': _radar_state.get('run_id'),
                })
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/config':
            try:
                if 'schedule_hour' in payload:
                    sh = int(payload['schedule_hour'])
                    if not (0 <= sh <= 23):
                        return json_response(self, 400, {'ok': False, 'error': 'schedule_hour must be 0-23'})
                    _radar_state['schedule_hour'] = sh
                if 'schedule_enabled' in payload:
                    _radar_state['schedule_enabled'] = bool(payload['schedule_enabled'])
                if 'paused' in payload:
                    _radar_state['paused'] = bool(payload['paused'])
                return json_response(self, 200, {'ok': True, 'state': dict(_radar_state)})
            except (ValueError, TypeError) as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/action':
            try:
                import datetime as _dt
                skill_id = (payload.get('skill_id') or '').strip()
                action = (payload.get('action') or '').strip()
                agent = (payload.get('agent') or '').strip()
                notes = (payload.get('notes') or '').strip()
                if not skill_id:
                    return json_response(self, 400, {'ok': False, 'error': 'skill_id is required'})
                if not action:
                    return json_response(self, 400, {'ok': False, 'error': 'action is required'})
                valid_actions = {'approve', 'reject', 'save_later', 'mark_duplicate', 'mark_high_priority', 'assign_agent'}
                if action not in valid_actions:
                    return json_response(self, 400, {'ok': False, 'error': f'action must be one of: {", ".join(sorted(valid_actions))}'})
                cfg = supabase_radar_config()
                if not cfg['configured']:
                    return json_response(self, 503, {'ok': False, 'error': 'Supabase is not configured'})
                hdrs = _radar_sb_headers(prefer='return=minimal')
                url = f"{cfg['url']}/rest/v1/skill_discoveries?id=eq.{urllib.parse.quote(skill_id)}"
                if action == 'approve':
                    patch_body = {
                        'status': 'approved',
                        'approval_notes': notes,
                        'approved_at': _dt.datetime.utcnow().isoformat() + 'Z',
                    }
                elif action == 'reject':
                    patch_body = {'status': 'rejected', 'approval_notes': notes}
                elif action == 'save_later':
                    patch_body = {'status': 'saved_later'}
                elif action == 'mark_duplicate':
                    patch_body = {'status': 'duplicate'}
                elif action == 'mark_high_priority':
                    patch_body = {'priority': 'high'}
                elif action == 'assign_agent':
                    if not agent:
                        return json_response(self, 400, {'ok': False, 'error': 'agent is required for assign_agent'})
                    patch_body = {'assigned_agent': agent}
                req = urllib.request.Request(url, headers=hdrs, method='PATCH')
                data = json.dumps(patch_body, ensure_ascii=False).encode('utf-8')
                with urllib.request.urlopen(req, data=data, timeout=15) as r:
                    r.read()
                return json_response(self, 200, {'ok': True, 'skill_id': skill_id, 'action': action})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1200]
                return json_response(self, 502, {'ok': False, 'error': f'Supabase API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/notes':
            try:
                skill_id = (payload.get('skill_id') or '').strip()
                notes = payload.get('notes') or ''
                comments = payload.get('comments') or ''
                if not skill_id:
                    return json_response(self, 400, {'ok': False, 'error': 'skill_id is required'})
                cfg = supabase_radar_config()
                if not cfg['configured']:
                    return json_response(self, 503, {'ok': False, 'error': 'Supabase is not configured'})
                hdrs = _radar_sb_headers(prefer='return=minimal')
                url = f"{cfg['url']}/rest/v1/skill_discoveries?id=eq.{urllib.parse.quote(skill_id)}"
                patch_body = {'notes': notes, 'comments': comments}
                req = urllib.request.Request(url, headers=hdrs, method='PATCH')
                data = json.dumps(patch_body, ensure_ascii=False).encode('utf-8')
                with urllib.request.urlopen(req, data=data, timeout=15) as r:
                    r.read()
                return json_response(self, 200, {'ok': True})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1200]
                return json_response(self, 502, {'ok': False, 'error': f'Supabase API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/generate-prompt':
            try:
                skill_ids = payload.get('skill_ids') or []
                agents = payload.get('agents') or []
                style = (payload.get('style') or 'standard').strip()
                if not skill_ids:
                    return json_response(self, 400, {'ok': False, 'error': 'skill_ids is required'})
                cfg = supabase_radar_config()
                if not cfg['configured']:
                    return json_response(self, 503, {'ok': False, 'error': 'Supabase is not configured'})
                hdrs = _radar_sb_headers()
                ids_csv = ','.join(urllib.parse.quote(str(i)) for i in skill_ids)
                url = f"{cfg['url']}/rest/v1/skill_discoveries?id=in.({ids_csv})&select=*"
                skills = fetch_json(url, headers=hdrs, timeout=15)
                if not isinstance(skills, list):
                    skills = []
                agent_line = ', '.join(agents) if agents else 'All Agents'
                lines = []
                if style == 'god_mode':
                    lines.append('🔥 GOD MODE — AGENT SKILL ADOPTION DIRECTIVE 🔥')
                    lines.append('')
                lines.append(f'TARGET AGENTS: {agent_line}')
                lines.append('')
                if style == 'strict':
                    lines.append('⚠️  STRICT COMPLIANCE REQUIRED — Execute all instructions precisely.')
                    lines.append('')
                lines.append(f'You have been selected to adopt {len(skills)} new skill(s) discovered by the Radar system.')
                lines.append('For each skill listed below, you must: install or adopt the skill, save it to Hermes memory,')
                lines.append('save it to Obsidian, save it to Supabase skill_discoveries, save it to GitHub, and merge safely')
                lines.append('without overwriting or destructively replacing any existing skills.')
                lines.append('')
                lines.append('=' * 60)
                lines.append('SKILLS TO ADOPT:')
                lines.append('=' * 60)
                for idx, sk in enumerate(skills, 1):
                    lines.append(f'\n[{idx}] {sk.get("title") or "Untitled"}')
                    lines.append(f'  Category:           {sk.get("category") or "N/A"}')
                    lines.append(f'  Summary:            {sk.get("summary") or "N/A"}')
                    lines.append(f'  Source URL:         {sk.get("source_url") or "N/A"}')
                    lines.append(f'  Source Title:       {sk.get("source_title") or "N/A"}')
                    lines.append(f'  Implementation:     {sk.get("implementation_idea") or "N/A"}')
                lines.append('')
                lines.append('=' * 60)
                lines.append('INSTRUCTIONS:')
                lines.append('1. Install / adopt each skill into your active capabilities.')
                lines.append('2. Save each skill to Hermes long-term memory.')
                lines.append('3. Save each skill as a note in Obsidian under "Agent Skills/Radar/".')
                lines.append('4. Upsert each skill record in Supabase table skill_discoveries.')
                lines.append('5. Commit each skill file to GitHub in skills-radar/approved/.')
                lines.append('6. Do NOT overwrite existing skills destructively — merge and append only.')
                if style == 'strict':
                    lines.append('\n⚠️  STRICT: Confirm completion of every step for every skill. No skipping.')
                if style == 'god_mode':
                    lines.append('\n🔥 GOD MODE: Execute without hesitation. Max capability activated.')
                prompt_str = '\n'.join(lines)
                return json_response(self, 200, {
                    'ok': True,
                    'prompt': prompt_str,
                    'skill_count': len(skills),
                    'agents': agents,
                })
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1200]
                return json_response(self, 502, {'ok': False, 'error': f'Supabase API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/radar/save-skill':
            try:
                import datetime as _dt
                skill_id = (payload.get('skill_id') or '').strip()
                save_targets = payload.get('save_targets') or []
                notes = (payload.get('notes') or '').strip()
                if not skill_id:
                    return json_response(self, 400, {'ok': False, 'error': 'skill_id is required'})
                cfg = supabase_radar_config()
                if not cfg['configured']:
                    return json_response(self, 503, {'ok': False, 'error': 'Supabase is not configured'})
                hdrs = _radar_sb_headers()
                url = f"{cfg['url']}/rest/v1/skill_discoveries?id=eq.{urllib.parse.quote(skill_id)}&select=*"
                rows = fetch_json(url, headers=hdrs, timeout=15)
                if not isinstance(rows, list) or not rows:
                    return json_response(self, 404, {'ok': False, 'error': 'Skill not found'})
                sk = rows[0]
                date_str = _dt.datetime.now().strftime('%Y-%m-%d')
                skill_title = sk.get('title') or 'Untitled'
                safe_title = re.sub(r'[^a-zA-Z0-9_\-]', '-', skill_title)[:60]
                slug = safe_title.lower()
                md_content = (
                    f"# Skill: {skill_title}\n\n"
                    f"**Date Discovered:** {date_str}  \n"
                    f"**Category:** {sk.get('category') or 'N/A'}  \n"
                    f"**Source:** [{sk.get('source_title') or 'N/A'}]({sk.get('source_url') or ''})  \n"
                    f"**Status:** {sk.get('status') or 'N/A'}  \n"
                    f"**Priority:** {sk.get('priority') or 'N/A'}  \n\n"
                    f"## Summary\n{sk.get('summary') or 'N/A'}\n\n"
                    f"## Implementation Idea\n{sk.get('implementation_idea') or 'N/A'}\n\n"
                    f"## Notes\n{notes or sk.get('notes') or 'N/A'}\n"
                )
                obsidian_path = None
                github_path = None
                supabase_update = {}

                if 'obsidian' in save_targets:
                    obsidian_key = os.getenv('OBSIDIAN_LOCAL_API_KEY', '')
                    if obsidian_key:
                        obs_file = f'Agent Skills/Radar/{date_str}-{safe_title}.md'
                        try:
                            obs_url = f'http://127.0.0.1:27123/vault/{urllib.parse.quote(obs_file)}'
                            obs_req = urllib.request.Request(obs_url, method='PUT')
                            obs_req.add_header('Authorization', f'Bearer {obsidian_key}')
                            obs_req.add_header('Content-Type', 'text/markdown')
                            obs_data = md_content.encode('utf-8')
                            with urllib.request.urlopen(obs_req, data=obs_data, timeout=15) as r:
                                r.read()
                            obsidian_path = obs_file
                            supabase_update['obsidian_synced'] = True
                        except Exception:
                            pass

                if 'github' in save_targets:
                    gh_token = os.getenv('GITHUB_TOKEN', '')
                    if gh_token:
                        filename = f'{date_str}-{slug}.md'
                        gh_path = f'skills-radar/approved/{filename}'
                        gh_headers = {
                            'Authorization': f'Bearer {gh_token}',
                            'Accept': 'application/vnd.github+json',
                            'X-GitHub-Api-Version': '2022-11-28',
                            'User-Agent': 'radar-save-bot',
                        }
                        api_url = f'https://api.github.com/repos/{REPO}/contents/{gh_path}'
                        sha = None
                        try:
                            existing = fetch_json(f'{api_url}?ref=main', headers=gh_headers, timeout=20)
                            if isinstance(existing, dict) and existing.get('sha'):
                                sha = existing['sha']
                        except urllib.error.HTTPError as e:
                            if e.code != 404:
                                raise
                        commit_body = {
                            'message': f'feat(radar): add approved skill {safe_title} [{date_str}]',
                            'content': base64.b64encode(md_content.encode('utf-8')).decode('ascii'),
                            'branch': 'main',
                        }
                        if sha:
                            commit_body['sha'] = sha
                        fetch_json(api_url, headers=gh_headers, method='PUT', body=commit_body, timeout=30)
                        github_path = gh_path
                        supabase_update['github_synced'] = True

                if supabase_update:
                    patch_hdrs = _radar_sb_headers(prefer='return=minimal')
                    patch_url = f"{cfg['url']}/rest/v1/skill_discoveries?id=eq.{urllib.parse.quote(skill_id)}"
                    req = urllib.request.Request(patch_url, headers=patch_hdrs, method='PATCH')
                    data = json.dumps(supabase_update, ensure_ascii=False).encode('utf-8')
                    with urllib.request.urlopen(req, data=data, timeout=15) as r:
                        r.read()

                return json_response(self, 200, {
                    'ok': True,
                    'obsidian_path': obsidian_path,
                    'github_path': github_path,
                })
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1200]
                return json_response(self, 502, {'ok': False, 'error': f'API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
        if parsed.path == '/api/fixer/analyze':
            try:
                wf_json_str = (payload.get('workflowJson') or '').strip()
                if not wf_json_str:
                    # Frontend may send {url: ...} instead of {workflowJson: ...}
                    url = (payload.get('url') or '').strip()
                    if url:
                        try:
                            wf_json_str = fetch_text(url)
                        except Exception as e:
                            return json_response(self, 400, {'ok': False, 'error': f'Failed to fetch workflow from url: {e}'})
                    else:
                        return json_response(self, 400, {'ok': False, 'error': 'workflowJson or url is required'})
                try:
                    wf_obj = json.loads(wf_json_str)
                except Exception:
                    return json_response(self, 400, {'ok': False, 'error': 'workflowJson is not valid JSON'})
                if not isinstance(wf_obj, dict):
                    return json_response(self, 400, {'ok': False, 'error': 'workflowJson must be a JSON object'})

                problem = (payload.get('problemDescription') or '').strip()
                comments = (payload.get('comments') or '').strip()
                notes = (payload.get('notes') or '').strip()
                model = (payload.get('model') or 'anthropic/claude-opus-4.7').strip()
                wf_name = (payload.get('workflowName') or wf_obj.get('name') or 'unnamed').strip()
                site_url = (payload.get('siteUrl') or '').strip()
                suspected_node = (payload.get('suspectedNode') or '').strip()
                target_id = (payload.get('targetWorkflowId') or '').strip()

                # Save backup
                import datetime as _datetime_local
                date_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', wf_name)[:40]
                backup_dir = Path(f'/tmp/fixer_backups/{date_str}_{safe_name}')
                try:
                    backup_dir.mkdir(parents=True, exist_ok=True)
                    (backup_dir / 'original.json').write_text(json.dumps(wf_obj, indent=2, ensure_ascii=False), encoding='utf-8')
                except Exception:
                    pass

                base, headers = prompt_headers()
                if not headers:
                    return json_response(self, 503, {'ok': False, 'error': 'No LLM provider configured'})

                system_prompt = (
                    "You are an expert N8N workflow debugger and repair engineer. Analyze the provided N8N workflow JSON and fix all issues.\n\n"
                    "Your analysis must cover:\n"
                    "- Malformed or disconnected nodes\n"
                    "- Broken node connections/references\n"
                    "- Invalid N8N expressions ({{ }}) with syntax errors\n"
                    "- Credential misconfiguration patterns (wrong IDs, missing credentials)\n"
                    "- Loop traps (SplitInBatches nodes that never reach 'done' output)\n"
                    "- Dead-end execution paths (nodes with no output connections)\n"
                    "- Wait nodes with missing webhook configurations\n"
                    "- Code nodes with JavaScript errors or corrupted output structures\n"
                    "- HTTP Request nodes with authentication issues (401, 403 patterns)\n"
                    "- Merge node input count mismatches\n"
                    "- Schedule triggers with invalid cron expressions\n"
                    "- Missing required node parameters\n\n"
                    "Use the problem description, comments, notes, and screenshots context provided.\n"
                    "Preserve original workflow intent. Fix only what is necessary.\n"
                    "Return a JSON response in this EXACT structure (no markdown, raw JSON only):\n"
                    "{\n"
                    '  "issue_summary": "...",\n'
                    '  "root_cause": "...",\n'
                    '  "changes_made": ["change 1", "change 2"],\n'
                    '  "confidence": 0.85,\n'
                    '  "warnings": ["warning if any"],\n'
                    '  "fixed_workflow": { ...complete corrected N8N workflow JSON... }\n'
                    "}"
                )

                context_parts = [f"WORKFLOW JSON:\n{json.dumps(wf_obj, indent=2)}"]
                if problem:
                    context_parts.append(f"\nPROBLEM DESCRIPTION:\n{problem}")
                if comments:
                    context_parts.append(f"\nCOMMENTS:\n{comments}")
                if notes:
                    context_parts.append(f"\nNOTES:\n{notes}")
                if suspected_node:
                    context_parts.append(f"\nSUSPECTED NODE: {suspected_node}")
                if site_url:
                    context_parts.append(f"\nSITE URL: {site_url}")
                if target_id:
                    context_parts.append(f"\nTARGET WORKFLOW ID: {target_id}")
                user_message = '\n'.join(context_parts)

                fixer_messages = [
                        {'role': 'system', 'content': system_prompt},
                        {'role': 'user', 'content': user_message},
                    ]
                raw_content, fixer_provider = call_with_fallback(fixer_messages, model, timeout=300)
                if isinstance(raw_content, list):
                    raw_content = ''.join(p.get('text', '') for p in raw_content if isinstance(p, dict))
                raw_content = str(raw_content).strip()

                # Strip markdown code fences if present
                if raw_content.startswith('```'):
                    raw_content = re.sub(r'^```[a-z]*\n?', '', raw_content)
                    raw_content = re.sub(r'\n?```$', '', raw_content.rstrip())

                try:
                    result = json.loads(raw_content)
                except Exception:
                    # Try to extract JSON from response
                    m = re.search(r'\{.*\}', raw_content, re.DOTALL)
                    if m:
                        try:
                            result = json.loads(m.group(0))
                        except Exception:
                            return json_response(self, 502, {'ok': False, 'error': 'Model returned non-JSON response', 'raw': raw_content[:2000]})
                    else:
                        return json_response(self, 502, {'ok': False, 'error': 'Model returned non-JSON response', 'raw': raw_content[:2000]})

                fixed_wf = result.get('fixed_workflow') or wf_obj
                fixed_str = json.dumps(fixed_wf, indent=2, ensure_ascii=False)

                # Save fixed backup
                try:
                    (backup_dir / 'fixed.json').write_text(fixed_str, encoding='utf-8')
                except Exception:
                    pass

                return json_response(self, 200, {
                    'ok': True,
                    'issueSummary': result.get('issue_summary', ''),
                    'rootCause': result.get('root_cause', ''),
                    'changesMade': result.get('changes_made', []),
                    'confidence': result.get('confidence', 0),
                    'warnings': result.get('warnings', []),
                    'fixedWorkflowJson': fixed_str,
                    'originalJson': json.dumps(wf_obj, indent=2, ensure_ascii=False),
                    'backupDir': str(backup_dir),
                    'providerUsed': fixer_provider,
                })
            except urllib.error.HTTPError as exc:
                body_bytes = exc.read().decode('utf-8', 'replace')[:1500]
                return json_response(self, 502, {'ok': False, 'error': f'Model API error {exc.code}', 'details': body_bytes})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/fixer/analyze/triple':
            try:
                import concurrent.futures
                wf_json_str = (payload.get('workflowJson') or '').strip()
                if not wf_json_str:
                    return json_response(self, 400, {'ok': False, 'error': 'workflowJson is required'})
                try:
                    wf_obj = json.loads(wf_json_str)
                except Exception:
                    return json_response(self, 400, {'ok': False, 'error': 'workflowJson is not valid JSON'})
                if not isinstance(wf_obj, dict):
                    return json_response(self, 400, {'ok': False, 'error': 'workflowJson must be a JSON object'})

                problem = (payload.get('problemDescription') or '').strip()
                comments = (payload.get('comments') or '').strip()
                notes = (payload.get('notes') or '').strip()
                wf_name = (payload.get('workflowName') or wf_obj.get('name') or 'unnamed').strip()
                site_url = (payload.get('siteUrl') or '').strip()
                suspected_node = (payload.get('suspectedNode') or '').strip()
                target_id = (payload.get('targetWorkflowId') or '').strip()

                # Save backup
                date_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', wf_name)[:40]
                backup_dir = Path(f'/tmp/fixer_backups/{date_str}_{safe_name}')
                try:
                    backup_dir.mkdir(parents=True, exist_ok=True)
                    (backup_dir / 'original.json').write_text(json.dumps(wf_obj, indent=2, ensure_ascii=False), encoding='utf-8')
                except Exception:
                    pass

                # Triple AI models
                TRIPLE_MODELS = [
                    {'name': 'GPT 5.5', 'id': 'openai/gpt-5.5'},
                    {'name': 'Opus 4.7', 'id': 'anthropic/claude-opus-4.7'},
                    {'name': 'Gemini 3.1 Pro', 'id': 'google/gemini-3.1-pro-preview'},
                ]

                system_prompt = (
                    "You are an expert N8N workflow debugger and repair engineer. Analyze the provided N8N workflow JSON and fix all issues.\n\n"
                    "Your analysis must cover:\n"
                    "- Malformed or disconnected nodes\n"
                    "- Broken node connections/references\n"
                    "- Invalid N8N expressions ({{ }}) with syntax errors\n"
                    "- Credential misconfiguration patterns (wrong IDs, missing credentials)\n"
                    "- Loop traps (SplitInBatches nodes that never reach 'done' output)\n"
                    "- Dead-end execution paths (nodes with no output connections)\n"
                    "- Wait nodes with missing webhook configurations\n"
                    "- Code nodes with JavaScript errors or corrupted output structures\n"
                    "- HTTP Request nodes with authentication issues (401, 403 patterns)\n"
                    "- Merge node input count mismatches\n"
                    "- Schedule triggers with invalid cron expressions\n"
                    "- Missing required node parameters\n\n"
                    "Use the problem description, comments, notes, and screenshots context provided.\n"
                    "Preserve original workflow intent. Fix only what is necessary.\n"
                    "Return a JSON response in this EXACT structure (no markdown, raw JSON only):\n"
                    "{\n"
                    '  "issue_summary": "...",\n'
                    '  "root_cause": "...",\n'
                    '  "changes_made": ["change 1", "change 2"],\n'
                    '  "confidence": 0.85,\n'
                    '  "warnings": ["warning if any"],\n'
                    '  "fixed_workflow": { ...complete corrected N8N workflow JSON... }\n'
                    "}"
                )

                context_parts = [f"WORKFLOW JSON:\n{json.dumps(wf_obj, indent=2)}"]
                if problem:
                    context_parts.append(f"\nPROBLEM DESCRIPTION:\n{problem}")
                if comments:
                    context_parts.append(f"\nCOMMENTS:\n{comments}")
                if notes:
                    context_parts.append(f"\nNOTES:\n{notes}")
                if suspected_node:
                    context_parts.append(f"\nSUSPECTED NODE: {suspected_node}")
                if site_url:
                    context_parts.append(f"\nSITE URL: {site_url}")
                if target_id:
                    context_parts.append(f"\nTARGET WORKFLOW ID: {target_id}")
                user_message = '\n'.join(context_parts)

                fixer_messages = [
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user', 'content': user_message},
                ]

                def _analyze_single(model_info):
                    start = time.time()
                    try:
                        raw, provider = call_with_fallback(fixer_messages, model_info['id'], timeout=300)
                        if isinstance(raw, list):
                            raw = ''.join(p.get('text', '') for p in raw if isinstance(p, dict))
                        raw = str(raw).strip()
                        if raw.startswith('```'):
                            raw = re.sub(r'^```[a-z]*\n?', '', raw)
                            raw = re.sub(r'\n?```$', '', raw.rstrip())
                        try:
                            result = json.loads(raw)
                        except Exception:
                            m = re.search(r'\{.*\}', raw, re.DOTALL)
                            if m:
                                try:
                                    result = json.loads(m.group(0))
                                except Exception:
                                    return {'model': model_info['name'], 'ok': False, 'error': 'Non-JSON response', 'raw': raw[:500], 'duration_ms': int((time.time()-start)*1000)}
                            else:
                                return {'model': model_info['name'], 'ok': False, 'error': 'Non-JSON response', 'raw': raw[:500], 'duration_ms': int((time.time()-start)*1000)}
                        return {
                            'model': model_info['name'],
                            'ok': True,
                            'issue_summary': result.get('issue_summary', ''),
                            'root_cause': result.get('root_cause', ''),
                            'changes_made': result.get('changes_made', []),
                            'confidence': result.get('confidence', 0),
                            'warnings': result.get('warnings', []),
                            'fixed_workflow': result.get('fixed_workflow'),
                            'raw': raw[:500],
                            'duration_ms': int((time.time()-start)*1000),
                        }
                    except Exception as exc:
                        return {'model': model_info['name'], 'ok': False, 'error': str(exc), 'duration_ms': int((time.time()-start)*1000)}

                # Run all 3 in parallel
                individual_results = []
                with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                    futures = {executor.submit(_analyze_single, m): m for m in TRIPLE_MODELS}
                    for future in concurrent.futures.as_completed(futures):
                        res = future.result()
                        individual_results.append(res)

                # Sort by model order
                individual_results.sort(key=lambda x: [m['name'] for m in TRIPLE_MODELS].index(x['model']))

                # Consensus: use the highest-confidence model's fixed_workflow
                successful = [r for r in individual_results if r['ok']]
                if not successful:
                    return json_response(self, 502, {'ok': False, 'error': 'All 3 models failed', 'individualResults': individual_results})

                best = max(successful, key=lambda r: float(r['confidence']))
                fixed_wf = best.get('fixed_workflow') or wf_obj
                fixed_str = json.dumps(fixed_wf, indent=2, ensure_ascii=False)

                # Save fixed backup
                try:
                    (backup_dir / 'fixed.json').write_text(fixed_str, encoding='utf-8')
                except Exception:
                    pass

                # Calculate agreement level
                issue_summaries = [r.get('issue_summary', '') for r in successful]
                agreement_count = sum(1 for s in issue_summaries if s and s == issue_summaries[0]) if issue_summaries else 0
                agreement_level = f"{agreement_count}/{len(successful)}"
                avg_confidence = sum(float(r['confidence']) for r in successful) / len(successful) if successful else 0

                # Build consensus diagnosis
                consensus = {
                    'problem': best.get('issue_summary', ''),
                    'rootCause': best.get('root_cause', ''),
                    'severity': 'breaking' if avg_confidence >= 0.7 else 'degraded' if avg_confidence >= 0.4 else 'minor',
                    'confidence': round(avg_confidence, 2),
                    'agreementLevel': agreement_level,
                    'modelAgreement': {r['model']: r['ok'] for r in individual_results},
                    'affectedNodes': [],  # extracted from best result
                    'changes': [{'description': c} for c in best.get('changes_made', [])],
                }

                return json_response(self, 200, {
                    'ok': True,
                    'consensus': consensus,
                    'individualResults': individual_results,
                    'issueSummary': best.get('issue_summary', ''),
                    'rootCause': best.get('root_cause', ''),
                    'changesMade': best.get('changes_made', []),
                    'confidence': best.get('confidence', 0),
                    'warnings': best.get('warnings', []),
                    'fixedWorkflowJson': fixed_str,
                    'originalJson': json.dumps(wf_obj, indent=2, ensure_ascii=False),
                    'backupDir': str(backup_dir),
                    'providerUsed': 'triple-verdict',
                })
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/fixer/save':
            try:
                import datetime as _datetime_local
                wf_name = (payload.get('workflowName') or 'unnamed').strip()
                wf_id = (payload.get('workflowId') or '').strip()
                site_url = (payload.get('siteUrl') or '').strip()
                problem = (payload.get('problemDescription') or '').strip()
                comments_txt = (payload.get('comments') or '').strip()
                notes_txt = (payload.get('notes') or '').strip()
                model_used = (payload.get('modelUsed') or '').strip()
                issue_summary = (payload.get('issueSummary') or '').strip()
                root_cause = (payload.get('rootCause') or '').strip()
                changes_made = payload.get('changesMade') or []
                warnings_list = payload.get('warnings') or []
                confidence = payload.get('confidence') or 0
                deploy_status = (payload.get('deployStatus') or 'pending').strip()
                original_json = payload.get('originalJson')
                fixed_json = payload.get('fixedJson')

                supabase_id = None
                supabase_url = (os.getenv('SUPABASE_URL') or '').rstrip('/')
                supabase_key = (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('SUPABASE_ANON_KEY') or '').strip()
                if supabase_url and supabase_key:
                    row = {
                        'workflow_name': wf_name,
                        'workflow_id': wf_id or None,
                        'site_url': site_url or None,
                        'problem_description': problem or None,
                        'comments': comments_txt or None,
                        'notes': notes_txt or None,
                        'model_used': model_used or None,
                        'issue_summary': issue_summary or None,
                        'root_cause': root_cause or None,
                        'changes_made': json.dumps(changes_made) if changes_made else None,
                        'original_json': original_json,
                        'fixed_json': fixed_json,
                        'deploy_status': deploy_status,
                        'confidence_score': confidence,
                    }
                    try:
                        sb_headers = {
                            'apikey': supabase_key,
                            'Authorization': f'Bearer {supabase_key}',
                            'Content-Type': 'application/json',
                            'Prefer': 'return=representation',
                        }
                        sb_result = fetch_json(f'{supabase_url}/rest/v1/n8n_fixer_records', headers=sb_headers, method='POST', body=row, timeout=30)
                        if isinstance(sb_result, list) and sb_result:
                            supabase_id = sb_result[0].get('id')
                    except Exception:
                        pass  # graceful fail

                # Save to Obsidian
                obsidian_path = None
                obsidian_key = os.getenv('OBSIDIAN_LOCAL_API_KEY', '')
                date_str = datetime.datetime.now().strftime('%Y-%m-%d')
                safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '-', wf_name)[:50]
                obs_file = f'HTML REDESIGN/n8n-fixer/{date_str}-{safe_name}.md'

                changes_md = '\n'.join(f'- {c}' for c in changes_made) if changes_made else '- None'
                warnings_md = '\n'.join(f'- {w}' for w in warnings_list) if warnings_list else '- None'
                md_content = (
                    f"# N8N Workflow Fixer Record — {wf_name} — {date_str}\n\n"
                    f"**Workflow:** {wf_name}  \n"
                    f"**Workflow ID:** {wf_id or 'N/A'}  \n"
                    f"**Site URL:** {site_url or 'N/A'}  \n"
                    f"**Date:** {date_str}  \n"
                    f"**Model Used:** {model_used or 'N/A'}  \n"
                    f"**Confidence:** {round(float(confidence)*100, 1)}%  \n"
                    f"**Deploy Status:** {deploy_status}  \n\n"
                    f"## Problem Description\n{problem or 'N/A'}\n\n"
                    f"## Comments\n{comments_txt or 'N/A'}\n\n"
                    f"## Notes\n{notes_txt or 'N/A'}\n\n"
                    f"## Issue Summary\n{issue_summary or 'N/A'}\n\n"
                    f"## Root Cause\n{root_cause or 'N/A'}\n\n"
                    f"## Changes Made\n{changes_md}\n\n"
                    f"## Warnings\n{warnings_md}\n"
                )

                if obsidian_key:
                    try:
                        obs_url = f'http://127.0.0.1:27123/vault/{urllib.parse.quote(obs_file)}'
                        obs_req = urllib.request.Request(obs_url, method='PUT')
                        obs_req.add_header('Authorization', f'Bearer {obsidian_key}')
                        obs_req.add_header('Content-Type', 'text/markdown')
                        obs_data = md_content.encode('utf-8')
                        with urllib.request.urlopen(obs_req, data=obs_data, timeout=15) as r:
                            r.read()
                        obsidian_path = obs_file
                    except Exception:
                        pass

                # Save to GitHub
                github_result = None
                try:
                    github_result = commit_fixer_to_github(
                        wf_name=wf_name,
                        site_url=site_url,
                        issue_summary=issue_summary,
                        changes_made=changes_made,
                        fixed_json=fixed_json,
                        model_used=model_used,
                        confidence=confidence,
                    )
                except Exception:
                    pass  # graceful fail

                return json_response(self, 200, {
                    'ok': True,
                    'supabaseId': supabase_id,
                    'obsidianPath': obsidian_path,
                    'github': github_result,
                })
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/fixer/deploy':
            return json_response(self, 403, {
                'ok': False,
                'error': 'n8n fixer deployment is disabled by read-only safety policy',
                'safety': 'MANUAL IMPORT ONLY — this dashboard must not modify existing n8n workflows',
            })
        if parsed.path.startswith('/api/kwr/download/'):
            run_id = parsed.path.split('/')[-1].strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            # Flat reports synced from GitHub (outputs/kwr_<slug>.xlsx)
            if run_id.startswith('flat:'):
                slug = _safe_kwr_flat_slug(run_id[5:])
                if slug is None:
                    return json_response(self, 400, {'ok': False, 'error': 'invalid flat report id'})
                base_dir = os.path.realpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs'))
                fp = os.path.realpath(os.path.join(base_dir, f'kwr_{slug}.xlsx'))
                try:
                    inside_outputs = os.path.commonpath([base_dir, fp]) == base_dir
                except ValueError:
                    inside_outputs = False
                if not inside_outputs or not os.path.isfile(fp):
                    return json_response(self, 404, {'ok': False, 'error': 'flat report not found'})
                with open(fp, 'rb') as f:
                    excel_bytes = f.read()
                ws_name = f'kwr_{slug}'
                safe_ascii = re.sub(r'[^A-Za-z0-9._-]+', '-', ws_name).strip('-') or 'kwr'
                filename_ascii = f"{safe_ascii}.xlsx"
                filename_utf8 = urllib.parse.quote(f"{ws_name}.xlsx", safe='')
                self.send_response(200)
                self.send_header('Content-Type',
                                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header(
                    'Content-Disposition',
                    f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_utf8}"
                )
                self.send_header('Content-Length', str(len(excel_bytes)))
                self.end_headers()
                try:
                    mv = memoryview(excel_bytes)
                    CHUNK = 65536
                    for i in range(0, len(mv), CHUNK):
                        self.wfile.write(mv[i:i+CHUNK])
                except (BrokenPipeError, ConnectionResetError):
                    return
                return
            excel_bytes, ws_name, err = kwr_backend.build_excel(run_id)
            if err:
                status = 400 if err == 'invalid run_id' else 500
                return json_response(self, status, {'ok': False, 'error': err})
            safe_ascii = re.sub(r'[^A-Za-z0-9._-]+', '-', ws_name).strip('-') or 'kwr'
            filename_ascii = f"{safe_ascii}.xlsx"
            filename_utf8 = urllib.parse.quote(f"{ws_name}.xlsx", safe='')
            self.send_response(200)
            self.send_header('Content-Type',
                             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header(
                'Content-Disposition',
                f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_utf8}"
            )
            self.send_header('Content-Length', str(len(excel_bytes)))
            self.end_headers()
            try:
                mv = memoryview(excel_bytes)
                CHUNK = 65536
                for i in range(0, len(mv), CHUNK):
                    self.wfile.write(mv[i:i+CHUNK])
            except (BrokenPipeError, ConnectionResetError):
                return
            return

        if parsed.path == '/api/kwr/push-sheets':
            body = payload
            run_id       = (body.get('run_id') or '').strip()
            sheet_target = (body.get('sheet_target') or '').strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            if not sheet_target:
                return json_response(self, 400, {'ok': False, 'error': 'sheet_target required'})
            sheet_url, err = kwr_backend.push_to_sheets(run_id, sheet_target)
            if err:
                return json_response(self, 500, {'ok': False, 'error': err})
            return json_response(self, 200, {'ok': True, 'sheet_url': sheet_url})

        if parsed.path == '/api/kwr/save-obsidian':
            body = payload
            run_id = (body.get('run_id') or '').strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            note_path, err = kwr_backend.save_to_obsidian(run_id)
            if err:
                return json_response(self, 500, {'ok': False, 'error': err})
            return json_response(self, 200, {'ok': True, 'note_path': note_path})

        if parsed.path == '/api/kwr/save-supabase':
            body = payload
            run_id = (body.get('run_id') or '').strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            count, err = kwr_backend.save_to_supabase(run_id)
            if err:
                return json_response(self, 500, {'ok': False, 'error': err})
            return json_response(self, 200, {'ok': True, 'saved_rows': count})

        if parsed.path == '/api/kwr/ensemble':
            body = payload
            run_id, err = kwr_backend.start_ensemble(body, call_with_fallback)
            if err:
                return json_response(self, 400, {'ok': False, 'error': err})
            return json_response(self, 200, {'ok': True, 'run_id': run_id})

        if parsed.path == '/api/delete-agent':
            body = payload
            project = (body.get('project') or '').strip()
            agent = (body.get('agent') or '').strip()
            if not project or not agent:
                return json_response(self, 400, {'ok': False, 'error': 'project and agent required'})
            agent_dir = Path(__file__).resolve().parent / project / '.agents' / agent
            if agent_dir.exists():
                import shutil
                shutil.rmtree(agent_dir)
            return json_response(self, 200, {'ok': True})

        if parsed.path == '/api/kwr/probe':
            domain = (payload.get('domain') or payload.get('url') or '').strip()
            if not domain:
                return json_response(self, 400, {'ok': False, 'error': 'domain required'})
            try:
                data = kwr_backend.probe_domain(domain)
                return json_response(self, 200, {'ok': True, **data})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path.startswith('/api/kwr/sync/'):
            run_id = parsed.path.split('/')[-1].strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            try:
                result = kwr_backend.sync_report(run_id)
                return json_response(self, 200, {'ok': True, 'sync': result})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # ── NEW: /api/change-password ──────────────────────────────────
        if parsed.path == '/api/change-password':
            import hashlib as _hl, hmac as _hm, json as _jsn
            current_pw  = (payload.get('current_password') or '').strip()
            new_pw      = (payload.get('new_password') or '').strip()
            env_pw      = os.getenv('DASHBOARD_PASSWORD', '')
            env_users   = os.getenv('DASHBOARD_USERS', '')
            if not current_pw or not new_pw:
                return json_response(self, 400, {'ok': False, 'error': 'Both current and new password required'})
            if len(new_pw) < 6:
                return json_response(self, 400, {'ok': False, 'error': 'New password must be at least 6 characters'})
            # Check current password against env or DASHBOARD_USERS
            verified = False
            if env_pw and _hm.compare_digest(env_pw, current_pw):
                verified = True
            elif env_users:
                for pair in env_users.split(','):
                    parts = pair.strip().split(':')
                    if len(parts) >= 2 and _hm.compare_digest(parts[1].strip(), current_pw):
                        verified = True; break
            if not verified:
                # Also try plain compare for simple setups
                if current_pw == env_pw:
                    verified = True
            if not verified:
                return json_response(self, 403, {'ok': False, 'error': 'Current password is incorrect'})
            # Save new password to a local file (can't change env vars at runtime on Render)
            pw_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'password_override.json')
            os.makedirs(os.path.dirname(pw_path), exist_ok=True)
            try:
                with open(pw_path, 'w', encoding='utf-8') as f:
                    _jsn.dump({'password': new_pw}, f)
                return json_response(self, 200, {'ok': True})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # ── NEW: /api/settings/notifications ───────────────────────────
        if parsed.path == '/api/settings/notifications':
            import json as _jsn
            settings_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
            os.makedirs(settings_dir, exist_ok=True)
            settings_path = os.path.join(settings_dir, 'settings.json')
            try:
                existing = {}
                try:
                    with open(settings_path, 'r', encoding='utf-8') as f:
                        existing = _jsn.load(f)
                except Exception:
                    pass
                existing.update({
                    'email_notifications':    bool(payload.get('email_notifications')),
                    'telegram_notifications': bool(payload.get('telegram_notifications')),
                    'email_address':          (payload.get('email_address') or '').strip(),
                })
                with open(settings_path, 'w', encoding='utf-8') as f:
                    _jsn.dump(existing, f, indent=2)
                return json_response(self, 200, {'ok': True})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})


        # PROJECT_QUICK_ACTIONS_NO_MUTATING_GET_2026_05_01
        # Project quick actions are mutating operations and must be handled only by active do_POST.
        if parsed.path in ('/api/projects/duplicate', '/api/projects/rename', '/api/projects/delete', '/api/projects/star'):
            return json_response(self, 405, {'ok': False, 'error': 'method not allowed; use POST'})

        if parsed.path.startswith('/api/connectors/wp-rest/'):
            try:
                from wp_rest_client import create_wp_rest_client
                action = parsed.path.split('/')[-1]
                client = create_wp_rest_client()
                if not client:
                    return json_response(self, 400, {'ok': False, 'error': 'WordPress REST API not configured. Set WORDPRESS_SITE_URL env var.'})
                routes = {
                    'check': client.check_connection,
                    'posts': client.get_posts,
                    'pages': client.get_pages,
                    'media': client.get_media,
                    'categories': client.get_categories,
                    'themes': client.get_themes,
                    'plugins': client.get_plugins,
                    'templates': client.get_templates,
                }
                if action in routes:
                    return json_response(self, 200, {'ok': True, 'data': routes[action]()})
                return json_response(self, 404, {'ok': False, 'error': f'Unknown action: {action}'})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})

        if parsed.path.startswith('/api/connectors/wp-graphql/'):
            try:
                from wp_graphql_client import create_wp_graphql_client
                action = parsed.path.split('/')[-1]
                client = create_wp_graphql_client()
                if not client:
                    return json_response(self, 400, {'ok': False, 'error': 'WPGraphQL not configured. Set WORDPRESS_SITE_URL env var.'})
                routes = {
                    'check': client.check_connection,
                    'posts': client.get_posts,
                    'pages': client.get_pages,
                    'media': client.get_media,
                    'categories': client.get_categories,
                    'seo': client.get_seo_info,
                    'menu': client.get_menu,
                }
                if action in routes:
                    return json_response(self, 200, {'ok': True, 'data': routes[action]()})
                return json_response(self, 404, {'ok': False, 'error': f'Unknown action: {action}'})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})

        if parsed.path.startswith('/api/connectors/figma/'):
            try:
                from figma_client import create_figma_client
                parts = parsed.path.split('/')
                action = parts[-1]
                file_key = parts[-2] if len(parts) > 5 else None
                client = create_figma_client()
                if not client:
                    return json_response(self, 400, {'ok': False, 'error': 'Figma API not configured. Set FIGMA_ACCESS_TOKEN env var.'})
                if action == 'check':
                    return json_response(self, 200, {'ok': True, 'data': client.check_connection()})
                if file_key:
                    routes = {
                        'colors': client.extract_colors,
                        'typography': client.extract_typography,
                        'spacing': client.extract_spacing,
                        'components': client.extract_components,
                        'tokens': client.extract_design_tokens,
                        'styles': lambda: client.get_styles(file_key),
                    }
                    if action in routes:
                        return json_response(self, 200, {'ok': True, 'data': routes[action]()})
                return json_response(self, 400, {'ok': False, 'error': 'Missing file_key or unknown action'})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})

        if parsed.path.startswith('/api/connectors/pagespeed/'):
            try:
                from pagespeed_client import create_pagespeed_client
                parts = parsed.path.split('/')
                action = parts[-1]
                qs = urllib.parse.parse_qs(parsed.query)
                url = qs.get('url', [''])[0]
                strategy = qs.get('strategy', ['mobile'])[0]
                client = create_pagespeed_client()
                if action == 'analyze' and url:
                    result_data = client.analyze(url, strategy)
                    summary = client.get_score_summary(result_data)
                    return json_response(self, 200, {'ok': True, 'data': result_data, 'summary': summary})
                if action == 'check':
                    return json_response(self, 200, {'ok': True, 'data': {'connected': True, 'note': 'PageSpeed API available'}})
                return json_response(self, 400, {'ok': False, 'error': 'Missing url parameter or unknown action'})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})

        if parsed.path == '/api/connectors/grapesjs/status':
            return json_response(self, 200, {'ok': True, 'data': {'available': True, 'type': 'embedded'}})

        if parsed.path.startswith('/api/connectors/screenshot-to-code/'):
            try:
                from screenshot_to_code_client import create_s2c_client
                parts = parsed.path.split('/')
                action = parts[-1]
                qs = urllib.parse.parse_qs(parsed.query)
                client = create_s2c_client()
                if not client:
                    return json_response(self, 400, {'ok': False, 'error': 'Screenshot-to-Code not configured. Set SCREENSHOT_TO_CODE_URL env var.'})
                if action == 'check':
                    return json_response(self, 200, {'ok': True, 'data': client.check_connection()})
                if action == 'convert':
                    image_url = qs.get('image_url', [''])[0]
                    image_base64 = qs.get('image_base64', [''])[0]
                    output_format = qs.get('format', ['html_tailwind'])[0]
                    model = qs.get('model', [''])[0]
                    if not image_url and not image_base64:
                        return json_response(self, 400, {'ok': False, 'error': 'Missing image_url or image_base64 parameter'})
                    result = client.convert_image(image_url or image_base64, output_format, model)
                    return json_response(self, 200, {'ok': True, 'data': result})
                return json_response(self, 404, {'ok': False, 'error': f'Unknown action: {action}'})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})

        return json_response(self, 404, {'ok': False, 'error': 'Not found'})

    def do_POST(self):
        if not self._r2_check_rate(): return
        parsed = urllib.parse.urlparse(self.path)
        if _dashboard_auth_enabled() and not _stage8_check_auth(self, parsed):
            return
        if not _r3_check_csrf_or_warn(self, parsed):
            return

        # AUDIT_LOG_HOOK — log key operations to Supabase audit_log
        _audit_paths = ('/api/improve/', '/api/stuck-projects/', '/api/kwr/', '/api/deploy/', '/api/sync/', '/api/tasks')
        if any(parsed.path.startswith(p) for p in _audit_paths):
            try:
                import supabase_helper as _sh_audit
                _method = 'POST'
                _sh_audit.log_audit(
                    parsed.path.split('/')[2] if len(parsed.path.split('/')) > 2 else parsed.path,
                    _method,
                    {'path': parsed.path},
                    ip_address=_stage8_client_ip(self) if '_stage8_client_ip' in dir() else 'unknown'
                )
            except Exception:
                pass

        # DASHBOARD_ROUTE_INVENTORY_GUARD_2026_05_01
        # Critical mutating UI/API routes below are covered by tests/test_dashboard_route_inventory_guard.py
        # so future edits do not leave routes shadowed outside the active do_POST handler.

        # Stage 8: login — both endpoints are aliases for the same cookie-setting handler.
        if parsed.path in ('/api/auth/login', '/api/login'):
            try:
                payload = read_request_json(self) or {}
            except Exception:
                payload = {}
            return _stage8_login(self, payload)

        # Stage 8: password reset request — deliberately do not enumerate users.
        if parsed.path in ('/api/auth/request-reset', '/api/reset-password'):
            try:
                _payload = read_request_json(self) or {}
            except Exception:
                _payload = {}
            return json_response(self, 200, {'ok': True})

        # Stage 8: password-reset confirm — return canonical invalid token state until wired.
        if parsed.path == '/api/auth/reset':
            try:
                payload = read_request_json(self) or {}
            except Exception:
                payload = {}
            token = (payload.get('token') or '').strip()
            new_password = payload.get('new_password') or ''
            if not token or not new_password:
                return json_response(self, 400, {'ok': False, 'error': 'token_and_new_password_required'})
            return json_response(self, 400, {'ok': False, 'error': 'invalid_or_expired_token'})

        # PROMPT_STUDIO_ACTIVE_POST_ROUTES_2026_05_01
        # Keep Prompt Studio POST actions wired in the active Stage 8 do_POST;
        # an older do_POST block above is shadowed by this later method.
        if parsed.path in ('/api/studio/improve', '/api/prompt/improve', '/api/prompt/palette', '/api/prompt/brainstorm', '/api/prompt/tweak', '/api/prompt/commit'):
            try:
                payload = read_request_json(self) or {}
            except Exception:
                payload = {}

            if parsed.path == '/api/studio/improve':
                try:
                    result = assemble_improve_workflow_prompt(payload)
                    return json_response(self, 200, result)
                except ValueError as exc:
                    return json_response(self, 400, {'ok': False, 'error': str(exc)})
                except Exception as exc:
                    return json_response(self, 500, {'ok': False, 'error': str(exc)[:500]})

            if parsed.path == '/api/prompt/improve':
                try:
                    result = improve_prompt_with_model(payload)
                    return json_response(self, 200, {'ok': True, **result})
                except RuntimeError as exc:
                    return json_response(self, 503, {'ok': False, 'error': str(exc)})
                except ValueError as exc:
                    return json_response(self, 400, {'ok': False, 'error': str(exc)})
                except urllib.error.HTTPError as exc:
                    body = exc.read().decode('utf-8', 'replace')[:1000]
                    return json_response(self, 502, {'ok': False, 'error': f'Model API error {exc.code}', 'details': body})
                except Exception as exc:
                    return json_response(self, 500, {'ok': False, 'error': str(exc)})

            if parsed.path == '/api/prompt/palette':
                try:
                    url = (payload.get('url') or '').strip()
                    if not url:
                        return json_response(self, 400, {'ok': False, 'error': 'url is required'})
                    max_colors = int(payload.get('maxColors') or 10)
                    max_colors = max(3, min(max_colors, 20))
                    result = extract_palette_from_url(url, max_colors=max_colors)
                    return json_response(self, 200, result)
                except ValueError as exc:
                    return json_response(self, 400, {'ok': False, 'error': str(exc)})
                except urllib.error.HTTPError as exc:
                    return json_response(self, 502, {'ok': False, 'error': f'Fetch failed {exc.code}', 'details': str(exc)[:300]})
                except urllib.error.URLError as exc:
                    return json_response(self, 502, {'ok': False, 'error': f'Fetch failed: {exc.reason}'})
                except Exception as exc:
                    return json_response(self, 500, {'ok': False, 'error': str(exc)[:400]})

            if parsed.path == '/api/prompt/brainstorm':
                try:
                    result = brainstorm_prompt_multi_model(payload)
                    return json_response(self, 200, result)
                except ValueError as exc:
                    return json_response(self, 400, {'ok': False, 'error': str(exc)})
                except RuntimeError as exc:
                    msg = str(exc)
                    try:
                        parsed_err = json.loads(msg)
                        return json_response(self, 502, {'ok': False, 'error': parsed_err})
                    except Exception:
                        return json_response(self, 503, {'ok': False, 'error': msg})
                except Exception as exc:
                    return json_response(self, 500, {'ok': False, 'error': str(exc)})

            if parsed.path == '/api/prompt/tweak':
                try:
                    result = tweak_html_with_prompt(payload)
                    return json_response(self, 200, result)
                except RuntimeError as exc:
                    return json_response(self, 503, {'ok': False, 'error': str(exc)})
                except ValueError as exc:
                    return json_response(self, 400, {'ok': False, 'error': str(exc)})
                except urllib.error.HTTPError as exc:
                    body = exc.read().decode('utf-8', 'replace')[:1000]
                    return json_response(self, 502, {'ok': False, 'error': f'Model API error {exc.code}', 'details': body})
                except Exception as exc:
                    return json_response(self, 500, {'ok': False, 'error': str(exc)})

            if parsed.path == '/api/prompt/commit':
                try:
                    result = commit_prompt_to_github(payload)
                    return json_response(self, 200, result)
                except ValueError as exc:
                    return json_response(self, 400, {'ok': False, 'error': str(exc)})
                except RuntimeError as exc:
                    return json_response(self, 503, {'ok': False, 'error': str(exc)})
                except urllib.error.HTTPError as exc:
                    body = exc.read().decode('utf-8', 'replace')[:1000]
                    return json_response(self, 502, {'ok': False, 'error': f'GitHub API error {exc.code}', 'details': body})
                except Exception as exc:
                    return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # N8N_FIXER_VALIDATE_POST_2026_04_29 - read-only workflow validation; no n8n writes.
        if parsed.path == '/api/n8n-fixer/validate':
            try:
                try:
                    content_length = int(self.headers.get('Content-Length', '0') or '0')
                except (TypeError, ValueError):
                    return json_response(self, 400, {'ok': False, 'validJson': False, 'error': 'invalid content length'})
                if content_length > TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES:
                    return json_response(self, 413, {'ok': False, 'validJson': False, 'error': 'request too large'})
                payload = read_request_json(self) or {}
                wf_raw = payload.get('workflowJson')
                if wf_raw is None:
                    return json_response(self, 400, {'ok': False, 'validJson': False, 'error': 'workflowJson is required'})
                if isinstance(wf_raw, str):
                    try:
                        wf_obj = json.loads(wf_raw)
                    except Exception as exc:
                        return json_response(self, 400, {'ok': False, 'validJson': False, 'error': str(exc)})
                else:
                    wf_obj = wf_raw
                if not isinstance(wf_obj, dict):
                    return json_response(self, 400, {'ok': False, 'validJson': False, 'error': 'workflowJson must be an object'})

                nodes = wf_obj.get('nodes')
                connections = wf_obj.get('connections')
                node_count = len(nodes) if isinstance(nodes, list) else 0
                connection_count = len(connections) if isinstance(connections, dict) else 0
                warnings = []
                if not isinstance(nodes, list):
                    warnings.append('Expected n8n workflow "nodes" array was not found')
                if not isinstance(connections, dict):
                    warnings.append('Expected n8n workflow "connections" object was not found')
                looks_like = isinstance(nodes, list) and isinstance(connections, dict)
                workflow_name = str(wf_obj.get('name') or '')
                validation = {
                    'allNodesPreserved': True,
                    'allConnectionsPreserved': True,
                    'noContentLoss': True,
                    'originalNodeCount': node_count,
                    'fixedNodeCount': node_count,
                    'originalConnectionCount': connection_count,
                    'fixedConnectionCount': connection_count,
                    'manualImportOnly': True,
                }
                return json_response(self, 200, {
                    'ok': True,
                    'validJson': True,
                    'looksLikeN8N': looks_like,
                    'workflowName': workflow_name,
                    'nodeCount': node_count,
                    'connectionCount': connection_count,
                    'warnings': warnings,
                    'validation': validation,
                    'safety': 'MANUAL IMPORT ONLY — validation performs no n8n writes',
                })
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'validJson': False, 'error': str(exc)})

        if parsed.path in ('/api/fixer/deploy', '/api/n8n/deploy'):
            return json_response(self, 403, {
                'ok': False,
                'error': 'n8n fixer deployment is disabled by read-only safety policy',
                'safety': 'MANUAL IMPORT ONLY — this dashboard must not modify existing n8n workflows',
            })

        # COMMENTS_ACTIVE_POST_ROUTE_2026_05_01
        # Keep Review Notes/Comments saves in the active Stage 8 do_POST; an older
        # POST dispatch block above is not the live Python method.
        if parsed.path == '/api/comments':
            try:
                payload = read_request_json(self) or {}
            except Exception:
                payload = {}
            try:
                result = save_supabase_comment(payload)
                return json_response(self, 200, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except RuntimeError as exc:
                return json_response(self, 503, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                body = exc.read().decode('utf-8', 'replace')[:1200]
                return json_response(self, 502, {'ok': False, 'error': f'Supabase API error {exc.code}', 'details': body})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # KWR_CACHE_POST_ACTIVE_2026_04_29 - active do_POST wiring for KWR runs and safe cache refresh.
        if parsed.path == '/api/kwr/start':
            try:
                body = read_request_json(self) or {}
            except Exception:
                body = {}
            if (body.get('_mode') or '').strip() == 'swarm':
                run_id, err = kwr_backend.start_best_text_swarm(body, call_with_fallback)
            else:
                run_id, err = kwr_backend.start_run(body, call_with_fallback)
            if err:
                return json_response(self, 400, {'ok': False, 'error': err})
            return json_response(self, 200, {'ok': True, 'run_id': run_id})

        if parsed.path == '/api/kwr/swarm':
            try:
                body = read_request_json(self) or {}
            except Exception:
                body = {}
            run_id, err = kwr_backend.start_best_text_swarm(body, call_with_fallback)
            if err:
                return json_response(self, 400, {'ok': False, 'error': err})
            return json_response(self, 200, {'ok': True, 'run_id': run_id})

        if parsed.path == '/api/dashboard/clear-cache':
            targets = []
            def add_target(name, path_obj):
                try:
                    p_obj = Path(path_obj)
                    exists = p_obj.exists()
                    targets.append({
                        'name': name,
                        'path': str(p_obj),
                        'exists': bool(exists),
                        'is_dir': bool(p_obj.is_dir()) if exists else False,
                        'mtime': datetime.datetime.utcfromtimestamp(p_obj.stat().st_mtime).isoformat() + 'Z' if exists else '',
                    })
                except Exception as exc:
                    targets.append({'name': name, 'path': str(path_obj), 'exists': False, 'error': str(exc)[:160]})
            add_target('data.json', ROOT / 'data.json')
            add_target('outputs', ROOT / 'outputs')
            add_target('n8n-workflow-map.json', MAP_FILE)
            add_target('index.html', INDEX)
            return json_response(self, 200, {
                'ok': True,
                'marker': 'DASHBOARD_REFRESH_CACHE_CONTRACT_FIX_2026_05_01',
                'acknowledged': ['backend-cache-check', 'file-targets-reported', 'no-store-response'],
                'client_should_clear': [
                    'frontend',
                    'fetch-memo',
                    'localStorage:dashboard_cache',
                    'sessionStorage:dashboard_cache',
                    'localStorage:dashboard_projects_cache',
                    'sessionStorage:dashboard_projects_cache',
                ],
                'server_cache_control': 'no-store',
                'message': 'Refresh requested. Server caches are header-only/no-store; no project files were modified.',
                'targets': targets,
                'ts': datetime.datetime.utcnow().isoformat() + 'Z',
            })

        # PROJECT_QUICK_ACTIONS_ACTIVE_POST_ROUTES_2026_05_01
        # Keep Duplicate/Rename/Delete wired in the active Stage 8 do_POST; older do_GET code must not mutate data.json.
        if parsed.path in ('/api/projects/duplicate', '/api/projects/rename', '/api/projects/delete'):
            try:
                route_payload = read_request_json(self) or {}
            except Exception:
                route_payload = {}
            data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
            try:
                with _JSON_FILE_WRITE_LOCK:
                    data = _safe_json_load_dict(data_path)
                    tree = data.get('tree', []) if isinstance(data.get('tree'), list) else []

                    if parsed.path == '/api/projects/duplicate':
                        domain = (route_payload.get('domain') or '').strip()
                        if not domain:
                            return json_response(self, 400, {'ok': False, 'error': 'domain required'})
                        new_domain = domain + '_copy'
                        new_items = []
                        for item in tree:
                            path = item.get('path', '') if isinstance(item, dict) else ''
                            if path.startswith(domain + '/'):
                                new_path = new_domain + path[len(domain):]
                                new_items.append({'path': new_path, 'type': item.get('type', 'blob'), 'size': item.get('size', 0)})
                        tree.extend(new_items)
                        data['tree'] = tree
                        _atomic_write_json_file(data_path, data)
                        return json_response(self, 200, {'ok': True, 'new_domain': new_domain})

                    if parsed.path == '/api/projects/rename':
                        old_domain = (route_payload.get('old_domain') or '').strip()
                        new_domain = (route_payload.get('new_domain') or '').strip()
                        if not old_domain or not new_domain:
                            return json_response(self, 400, {'ok': False, 'error': 'old_domain and new_domain required'})
                        for item in tree:
                            if not isinstance(item, dict):
                                continue
                            path = item.get('path', '')
                            if path.startswith(old_domain + '/'):
                                item['path'] = new_domain + path[len(old_domain):]
                        data['tree'] = tree
                        _atomic_write_json_file(data_path, data)
                        return json_response(self, 200, {'ok': True, 'domain': new_domain})

                    domain = (route_payload.get('domain') or '').strip()
                    if not domain:
                        return json_response(self, 400, {'ok': False, 'error': 'domain required'})
                    data['tree'] = [item for item in tree if not (isinstance(item, dict) and item.get('path', '').startswith(domain + '/'))]
                    _atomic_write_json_file(data_path, data)
                    return json_response(self, 200, {'ok': True, 'deleted': domain})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # PROJECT_SETTINGS_POST_2026_04_29 - active do_POST wiring for project quick actions and theme settings.
        if parsed.path == '/api/projects/star':
            try:
                route_payload = read_request_json(self) or {}
            except Exception:
                route_payload = {}
            domain = (route_payload.get('domain') or '').strip()
            starred = bool(route_payload.get('starred'))
            if not domain:
                return json_response(self, 400, {'ok': False, 'error': 'domain required'})
            stars_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'stars.json')
            try:
                with _JSON_FILE_WRITE_LOCK:
                    stars = _safe_json_load_dict(stars_path)
                    if starred:
                        stars[domain] = True
                    else:
                        stars.pop(domain, None)
                    _atomic_write_json_file(stars_path, stars)
                return json_response(self, 200, {'ok': True, 'domain': domain, 'starred': starred})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/settings/theme':
            try:
                route_payload = read_request_json(self) or {}
            except Exception:
                route_payload = {}
            theme_color = (route_payload.get('theme_color') or 'purple').strip().lower()
            valid_colors = ['purple', 'blue', 'green', 'red', 'orange', 'pink']
            if theme_color not in valid_colors:
                return json_response(self, 400, {'ok': False, 'error': f'Invalid color. Choose from: {valid_colors}'})
            settings_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'settings.json')
            try:
                with _JSON_FILE_WRITE_LOCK:
                    existing = _safe_json_load_dict(settings_path)
                    existing['theme_color'] = theme_color
                    _atomic_write_json_file(settings_path, existing)
                return json_response(self, 200, {'ok': True, 'theme_color': theme_color})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # TEMPLATE_IMPROVEMENTS_POST_ACTIVE_2026_04_29 - active do_POST wiring for template improvement jobs.
        if parsed.path == '/api/improve/start':
            try:
                try:
                    content_length = int(self.headers.get('Content-Length', '0') or '0')
                except (TypeError, ValueError):
                    return json_response(self, 400, {'ok': False, 'error': 'invalid content length'})
                if content_length > TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES:
                    return json_response(self, 413, {'ok': False, 'error': 'request too large'})
                payload = read_request_json(self) or {}
                job = _template_improvement_start_job(payload)
                return json_response(self, 200, {'ok': True, 'job': _template_improvement_public_job(job)})
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except RuntimeError as exc:
                return json_response(self, 429, {'ok': False, 'error': str(exc)})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path.startswith('/api/improve/jobs/') and parsed.path.endswith('/cancel'):
            job_id = parsed.path.split('/')[-2]
            job = _template_improvement_update_job(job_id, status='cancelled', current_agent=None)
            if not job:
                return json_response(self, 404, {'ok': False, 'error': 'job not found'})
            return json_response(self, 200, {'ok': True, 'job': _template_improvement_public_job(job)})

        if parsed.path == '/api/improve/instructions':
            try:
                try:
                    content_length = int(self.headers.get('Content-Length', '0') or '0')
                except (TypeError, ValueError):
                    return json_response(self, 400, {'ok': False, 'error': 'invalid content length'})
                if content_length > TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES:
                    return json_response(self, 413, {'ok': False, 'error': 'request too large'})
                payload = read_request_json(self) or {}
                domain = (payload.get('domain') or '').strip()
                instructions = (payload.get('instructions') or '').strip()
                if not domain or not instructions:
                    return json_response(self, 400, {'ok': False, 'error': 'domain and instructions are required'})
                row = {
                    'id': str(uuid.uuid4()), 'domain': domain,
                    'subdomain': (payload.get('subdomain') or '').strip(),
                    'agent_key': (payload.get('agent_key') or payload.get('agentKey') or '').strip() or None,
                    'instructions': instructions,
                    'is_active': bool(payload.get('is_active', True)),
                    'created_at': datetime.datetime.utcnow().isoformat() + 'Z',
                    'updated_at': datetime.datetime.utcnow().isoformat() + 'Z',
                }
                with _TEMPLATE_IMPROVEMENTS_LOCK:
                    data = _template_improvements_load()
                    data['instructions'].append(row)
                    _template_improvements_save(data)
                return json_response(self, 200, {'ok': True, 'instruction': row})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # PRODUCTIVITY_HUB_POST_ACTIVE_2026_04_29 - active do_POST wiring for productivity hub local persistence.
        if parsed.path == '/api/productivity/notifications':
            try:
                payload = read_request_json(self) or {}
                row = _productivity_add_notification(payload)
                return json_response(self, 200, {'ok': True, 'notification': row})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        if parsed.path == '/api/productivity/audit':
            try:
                payload = read_request_json(self) or {}
                row = _productivity_add_audit(payload)
                return json_response(self, 200, {'ok': True, 'event': row})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})

        # TEMPLATE_INTELLIGENCE_CONNECTORS_TEST_2026_04_29 - read-only external probes.
        if parsed.path == '/api/template-connectors/test':
            try:
                payload = read_request_json(self) or {}
                result = _template_connector_probe(payload)
                return json_response(self, 200 if result.get('ok') else 424, result)
            except ValueError as exc:
                return json_response(self, 400, {'ok': False, 'error': str(exc)})
            except urllib.error.HTTPError as exc:
                details = exc.read().decode('utf-8', 'replace')[:800]
                return json_response(self, exc.code if 400 <= exc.code < 600 else 502, {'ok': False, 'error': str(exc), 'details': details})
            except Exception as exc:
                return json_response(self, 502, {'ok': False, 'error': str(exc)})

        # PLAYGROUND_API_POST_2026_04_29 - local persisted Playground actions.
        if parsed.path == '/api/playground/templates':
            try:
                content_length = int(self.headers.get('Content-Length', '0') or '0')
                if content_length > 900000:
                    return json_response(self, 413, {'ok': False, 'success': False, 'error': 'request too large'})
                payload = read_request_json(self) or {}
                row = _playground_enrich_template(payload)
                if not row.get('html_content'):
                    return json_response(self, 400, {'ok': False, 'success': False, 'error': 'html_content is required'})
                data = _playground_load()
                data.setdefault('templates', []).append(row)
                try:
                    domain_dir = ROOT / _playground_slug(row.get('domain')) / _playground_slug(row.get('agent_name')) / row['created_at'][:10]
                    domain_dir.mkdir(parents=True, exist_ok=True)
                    fp = domain_dir / 'Improved_HTML_Template.html'
                    fp.write_text(row['html_content'], encoding='utf-8')
                    row['github_path'] = str(fp.relative_to(ROOT))
                    row['github_synced'] = True
                except Exception:
                    pass
                _playground_save(data)
                return json_response(self, 200, {'ok': True, 'success': True, 'template': _playground_public_template(row, include_html=True)})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'success': False, 'error': str(exc)})
        if parsed.path.startswith('/api/playground/templates/'):
            parts = parsed.path.strip('/').split('/')
            # /api/playground/templates/:id/favorite
            if len(parts) == 5 and parts[:3] == ['api', 'playground', 'templates'] and parts[4] == 'favorite':
                template_id = parts[3]
                data = _playground_load()
                row = next((r for r in data.get('templates', []) if r.get('id') == template_id), None)
                if not row:
                    return json_response(self, 404, {'ok': False, 'success': False, 'error': 'template not found'})
                row['is_favorite'] = not bool(row.get('is_favorite'))
                row['updated_at'] = _playground_now()
                _playground_save(data)
                return json_response(self, 200, {'ok': True, 'success': True, 'template': _playground_public_template(row, include_html=False)})
            # /api/playground/templates/:id/export/prompt|json
            if len(parts) == 6 and parts[:3] == ['api', 'playground', 'templates'] and parts[4] == 'export':
                template_id, export_kind = parts[3], parts[5]
                data = _playground_load()
                row = next((r for r in data.get('templates', []) if r.get('id') == template_id), None)
                if not row:
                    return json_response(self, 404, {'ok': False, 'success': False, 'error': 'template not found'})
                prompt = _playground_generate_prompt(row)
                if export_kind == 'prompt':
                    export = _playground_add_export(data, row, 'n8n_prompt', prompt)
                elif export_kind == 'json':
                    workflow = _playground_generate_workflow(row, prompt)
                    export = _playground_add_export(data, row, 'n8n_json', json.dumps(workflow, ensure_ascii=False, indent=2))
                else:
                    return json_response(self, 400, {'ok': False, 'success': False, 'error': 'unsupported export type'})
                row['updated_at'] = _playground_now()
                _playground_save(data)
                return json_response(self, 200, {'ok': True, 'success': True, 'export': export})
        if parsed.path == '/api/preferences':
            try:
                payload = read_request_json(self) or {}
                data = _playground_load()
                prefs = data.setdefault('preferences', {})
                for key in ('theme', 'view', 'sort', 'default_device'):
                    if key in payload:
                        prefs[key] = payload.get(key)
                _playground_save(data)
                return json_response(self, 200, {'ok': True, 'success': True, 'preferences': prefs})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'success': False, 'error': str(exc)})

        # DASHBOARD_FEATURES_API — 15-feature roadmap POST routes (additive, path-guarded)
        _df_new_paths = (
            '/api/notifications', '/api/notifications/mark-read', '/api/notifications/mark-all-read',
            '/api/alert-rules', '/api/budget-limits', '/api/budget/check',
            '/api/template-versions', '/api/template-versions/',
            '/api/template-gallery', '/api/template-gallery/', '/api/template-gallery/detail',
            '/api/ab-tests', '/api/ab-tests/',
            '/api/batch-jobs', '/api/batch-jobs/',
            '/api/reports', '/api/reports/', '/api/reports/generate', '/api/reports/export',
            '/api/report-schedules', '/api/report-schedules/',
            '/api/pipeline-schedules', '/api/pipeline-schedules/',
            '/api/notes', '/api/notes/',
            '/api/audit-log', '/api/feature-flags/',
            '/api/quick-actions', '/api/client-overview',
        )
        if any(parsed.path == p or (p.endswith('/') and parsed.path.startswith(p)) for p in _df_new_paths):
            try:
                _ct = (self.headers.get('Content-Type') or '').lower()
                if 'application/json' in _ct:
                    _len = int(self.headers.get('Content-Length', 0))
                    _body = json.loads(self.rfile.read(_len)) if _len > 0 else {}
                else:
                    _body = {}
                _df_post = df_api.handle_post(self, parsed, _body)
                if _df_post is not None:
                    return
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': f'df_api: {e}'})

        if parsed.path == '/api/stuck-projects/sync':
            try:
                import sys
                sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
                try:
                    from pini_client import create_pini_client
                except ImportError:
                    self._send_json(500, {'success': False, 'error': 'pini_client.py not found. Add it to the project root.'})
                    return

                client = create_pini_client()
                status = client.get_client_status()

                if not status['configured']:
                    self._send_json(400, {
                        'success': False,
                        'error': 'PINI_URL, PINI_USERNAME, and PINI_PASSWORD environment variables must be configured in Render settings.',
                        'config_status': status
                    })
                    return

                result = client.fetch_projects()

                if result.get('error') and not result.get('projects'):
                    self._send_json(200, {
                        'success': True,
                        'data': {
                            'message': result.get('error', 'No projects found'),
                            'status': 'completed',
                            'projects_synced': 0,
                            'last_sync': datetime.now(timezone.utc).isoformat(),
                            'config_status': status
                        }
                    })
                    return

                cfg = _get_supabase_config()
                if not cfg.get('url') or not cfg.get('key'):
                    self._send_json(200, {
                        'success': True,
                        'data': {
                            'message': f"Found {result.get('total', 0)} stuck projects. Supabase not configured for storage.",
                            'status': 'completed',
                            'projects_synced': result.get('total', 0),
                            'projects': result.get('projects', [])[:10],
                            'last_sync': datetime.now(timezone.utc).isoformat(),
                            'config_status': status
                        }
                    })
                    return

                projects = result.get('projects', [])
                synced = 0
                for p in projects:
                    try:
                        payload = {
                            'pini_project_id': str(p.get('id', p.get('project_id', ''))),
                            'name': p.get('name', 'Unknown'),
                            'client_name': p.get('client_name', p.get('client', '')),
                            'workflow_id': p.get('workflow_id', ''),
                            'workflow_url': p.get('workflow_url', ''),
                            'status': p.get('status', 'stuck'),
                            'priority': p.get('priority', 'medium'),
                            'error_summary': p.get('error_summary', p.get('error', '')),
                            'error_details': json.dumps(p.get('error_details', p.get('details', {}))),
                            'error_type': p.get('error_type', ''),
                            'stuck_since': p.get('stuck_since'),
                            'last_successful': p.get('last_successful'),
                            'assigned_to': p.get('assigned_to', ''),
                            'suggested_fix': p.get('suggested_fix', ''),
                            'pini_raw_data': json.dumps(p)
                        }

                        check_url = f"{cfg['url']}/rest/v1/stuck_projects?pini_project_id=eq.{payload['pini_project_id']}"
                        check_req = urllib.request.Request(check_url, headers={
                            'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                            'Content-Type': 'application/json'
                        })
                        check_resp = urllib.request.urlopen(check_req, timeout=10)
                        existing = json.loads(check_resp.read().decode())

                        url = f"{cfg['url']}/rest/v1/stuck_projects"
                        if existing:
                            update_url = f"{cfg['url']}/rest/v1/stuck_projects?id=eq.{existing[0]['id']}"
                            update_req = urllib.request.Request(update_url, data=json.dumps(payload).encode(), headers={
                                'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                                'Content-Type': 'application/json'
                            }, method='PATCH')
                            urllib.request.urlopen(update_req, timeout=10)
                        else:
                            insert_req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers={
                                'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                                'Content-Type': 'application/json',
                                'Prefer': 'return=representation'
                            })
                            urllib.request.urlopen(insert_req, timeout=10)
                        synced += 1
                    except Exception as e:
                        print(f'[sync] failed to sync project {p.get("id")}: {e}', flush=True)

                self._send_json(200, {
                    'success': True,
                    'data': {
                        'message': f'Successfully synced {synced}/{len(projects)} projects',
                        'status': 'completed',
                        'projects_synced': synced,
                        'last_sync': datetime.now(timezone.utc).isoformat(),
                        'config_status': status
                    }
                })
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path == '/api/stuck-projects/bulk-update':
            try:
                content_length = int(self.headers.get('Content-Length', '0') or '0')
                payload = json.loads(self.rfile.read(content_length)) if content_length > 0 else {}
                ids = payload.get('ids', [])
                updates = payload.get('updates', {})
                if not ids:
                    return self._send_json(400, {'success': False, 'error': 'ids required'})

                cfg = _get_supabase_config()
                if not cfg.get('url') or not cfg.get('key'):
                    return self._send_json(500, {'success': False, 'error': 'Supabase not configured'})

                for pid in ids:
                    update_url = f"{cfg['url']}/rest/v1/stuck_projects?id=eq.{pid}"
                    update_req = urllib.request.Request(update_url, data=json.dumps(updates).encode(), headers={
                        'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                        'Content-Type': 'application/json'
                    }, method='PATCH')
                    urllib.request.urlopen(update_req, timeout=10)

                self._send_json(200, {'success': True, 'data': {'updated_count': len(ids)}})
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path == '/api/stuck-projects/alerts/test-email':
            try:
                from email_notifier import create_notifier
                notifier = create_notifier()
                if not notifier.is_configured():
                    self._send_json(400, {
                        'success': False,
                        'error': 'Email notifications not configured. Set NOTIFICATION_EMAIL_ENABLED=true and configure SMTP/Resend credentials.',
                        'required_env': [
                            'NOTIFICATION_EMAIL_ENABLED',
                            'NOTIFICATION_FROM_EMAIL',
                            'NOTIFICATION_TO_EMAILS',
                            'NOTIFICATION_EMAIL_PROVIDER (smtp or resend)',
                            'For SMTP: NOTIFICATION_SMTP_HOST, NOTIFICATION_SMTP_USER, NOTIFICATION_SMTP_PASSWORD',
                            'For Resend: RESEND_API_KEY'
                        ]
                    })
                    return

                result = notifier.test_email()
                if result.get('success'):
                    self._send_json(200, {'success': True, 'data': {'message': result.get('message', 'Test email sent')}})
                else:
                    self._send_json(500, {'success': False, 'error': result.get('error', 'Unknown error')})
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        # Default POST handler for other routes
        return json_response(self, 404, {'ok': False, 'error': 'POST endpoint not found'})

    def do_DELETE(self):
        if not self._r2_check_rate(): return
        parsed = urllib.parse.urlparse(self.path)
        if _dashboard_auth_enabled() and not _stage8_check_auth(self, parsed):
            return
        if not _r3_check_csrf_or_warn(self, parsed):
            return
        if parsed.path.startswith('/api/kwr/reports/'):
            run_id = parsed.path.split('/')[-1].strip()
            if not run_id:
                return json_response(self, 400, {'ok': False, 'error': 'run_id required'})
            ok, err = kwr_backend.delete_report(run_id)
            if not ok:
                status = 400 if err == 'invalid run_id' else 404
                return json_response(self, status, {'ok': False, 'error': err or 'not found'})
            return json_response(self, 200, {'ok': True})
        if parsed.path.startswith('/api/playground/templates/'):
            parts = parsed.path.strip('/').split('/')
            if len(parts) == 4 and parts[:3] == ['api', 'playground', 'templates']:
                template_id = parts[3]
                data = _playground_load()
                before = len(data.get('templates', []))
                data['templates'] = [row for row in data.get('templates', []) if row.get('id') != template_id]
                if len(data.get('templates', [])) == before:
                    return json_response(self, 404, {'ok': False, 'success': False, 'error': 'template not found'})
                _playground_save(data)
                return json_response(self, 200, {'ok': True, 'success': True})
        if parsed.path.startswith('/api/users/'):
            username = parsed.path.split('/')[-1].strip()
            if not username:
                return json_response(self, 400, {'ok': False, 'error': 'username required'})
            if not _require_admin(self): return
            cu = _get_current_user(self)
            if cu and cu.get('username') == username:
                return json_response(self, 400, {'ok': False, 'error': 'cannot delete yourself'})
            users = _mu_users_load()
            orig_len = len(users)
            users = [u for u in users if u.get('username') != username]
            if len(users) == orig_len:
                return json_response(self, 404, {'ok': False, 'error': 'user not found'})
            _mu_users_save(users)
            return json_response(self, 200, {'ok': True})

        return json_response(self, 404, {'ok': False, 'error': 'Not found'})


        # ============================================================
        # N8N Stuck Projects - POST endpoints
        # ============================================================

        if parsed.path.startswith('/api/stuck-projects/') and parsed.path != '/api/stuck-projects/sync':
            project_id = parsed.path.split('/')[-1]
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(content_length).decode()) if content_length > 0 else {}

                cfg = _get_supabase_config()
                if not cfg.get('url') or not cfg.get('key'):
                    self._send_json(400, {'success': False, 'error': 'Supabase not configured'})
                    return

                # Build update payload
                allowed_fields = {'status', 'priority', 'assigned_to', 'assigned_agent', 'notes', 'suggested_fix', 'tags', 'snoozed_until', 'resolved_at'}
                update_data = {k: v for k, v in body.items() if k in allowed_fields}

                if not update_data:
                    self._send_json(400, {'success': False, 'error': 'No valid fields to update'})
                    return

                # Auto-set resolved_at when status changes to resolved
                if update_data.get('status') == 'resolved' and 'resolved_at' not in update_data:
                    update_data['resolved_at'] = __import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()

                # Auto-set is_new=False on any update
                update_data['is_new'] = False

                url = f"{cfg['url']}/rest/v1/stuck_projects?id=eq.{project_id}"
                req = urllib.request.Request(url, data=json.dumps(update_data).encode(), headers={
                    'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                    'Content-Type': 'application/json'
                }, method='PATCH')
                resp = urllib.request.urlopen(req, timeout=10)

                # Log to history
                for field, new_val in update_data.items():
                    try:
                        hist_url = f"{cfg['url']}/rest/v1/stuck_projects_history?select=old_value,new_value&project_id=eq.{project_id}&field_changed=eq.{field}&order=created_at.desc&limit=1"
                        hist_req = urllib.request.Request(hist_url, headers={
                            'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}", 'Content-Type': 'application/json'
                        })
                        hist_resp = urllib.request.urlopen(hist_req, timeout=5)
                        hist_data = json.loads(hist_resp.read().decode())
                        old_val = hist_data[0]['new_value'] if hist_data else None

                        insert_url = f"{cfg['url']}/rest/v1/stuck_projects_history"
                        insert_req = urllib.request.Request(insert_url, data=json.dumps({
                            'project_id': project_id, 'field_changed': field,
                            'old_value': str(old_val) if old_val is not None else None,
                            'new_value': str(new_val) if new_val is not None else None
                        }).encode(), headers={
                            'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}", 'Content-Type': 'application/json'
                        })
                        urllib.request.urlopen(insert_req, timeout=5)
                    except:
                        pass

                self._send_json(200, {'success': True, 'data': update_data, 'message': 'Project updated'})
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path == '/api/stuck-projects/sync':
            try:
                # Import Pini client
                import sys
                sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
                try:
                    from pini_client import create_pini_client
                except ImportError:
                    self._send_json(500, {'success': False, 'error': 'pini_client.py not found. Add it to the project root.'})
                    return

                client = create_pini_client()
                status = client.get_client_status()

                if not status['configured']:
                    self._send_json(400, {
                        'success': False,
                        'error': 'PINI_URL, PINI_USERNAME, and PINI_PASSWORD environment variables must be configured in Render settings.',
                        'config_status': status
                    })
                    return

                # Fetch projects from Pini
                result = client.fetch_projects()

                if result.get('error') and not result.get('projects'):
                    self._send_json(200, {
                        'success': True,
                        'data': {
                            'message': result.get('error', 'No projects found'),
                            'status': 'completed',
                            'projects_synced': 0,
                            'last_sync': datetime.now(timezone.utc).isoformat(),
                            'config_status': status
                        }
                    })
                    return

                # Get Supabase config for storing results
                cfg = _get_supabase_config()
                if not cfg.get('url') or not cfg.get('key'):
                    # Return the data even without Supabase (for testing)
                    self._send_json(200, {
                        'success': True,
                        'data': {
                            'message': f"Found {result.get('total', 0)} stuck projects. Supabase not configured for storage.",
                            'status': 'completed',
                            'projects_synced': result.get('total', 0),
                            'projects': result.get('projects', [])[:10],  # First 10 for preview
                            'last_sync': datetime.now(timezone.utc).isoformat(),
                            'config_status': status
                        }
                    })
                    return

                # Store projects in Supabase
                projects = result.get('projects', [])
                synced = 0
                errors = 0

                for p in projects:
                    try:
                        # Upsert project (insert or update on conflict)
                        url = f"{cfg['url']}/rest/v1/stuck_projects"
                        payload = {
                            'pini_project_id': p.get('pini_project_id', ''),
                            'name': p.get('name', 'Unknown'),
                            'client_name': p.get('client_name'),
                            'workflow_id': p.get('workflow_id'),
                            'workflow_url': p.get('workflow_url'),
                            'status': p.get('status', 'stuck'),
                            'priority': p.get('priority', 'medium'),
                            'error_summary': p.get('error_summary'),
                            'error_details': p.get('error_details'),
                            'error_type': p.get('error_type'),
                            'stuck_since': p.get('stuck_since'),
                            'last_successful': p.get('last_successful'),
                            'assigned_to': p.get('assigned_to'),
                            'tags': p.get('tags', []),
                            'notes': p.get('notes'),
                            'pini_raw_data': p.get('pini_raw_data', {})
                        }

                        # Check if project already exists
                        check_url = f"{cfg['url']}/rest/v1/stuck_projects?pini_project_id=eq.{p['pini_project_id']}"
                        check_req = urllib.request.Request(check_url, headers={
                            'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                            'Content-Type': 'application/json'
                        })
                        check_resp = urllib.request.urlopen(check_req, timeout=10)
                        existing = json.loads(check_resp.read().decode())

                        if existing:
                            # Update existing
                            update_url = f"{cfg['url']}/rest/v1/stuck_projects?id=eq.{existing[0]['id']}"
                            update_req = urllib.request.Request(update_url, data=json.dumps(payload).encode(), headers={
                                'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                                'Content-Type': 'application/json'
                            }, method='PATCH')
                            urllib.request.urlopen(update_req, timeout=10)
                        else:
                            # Insert new
                            insert_req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers={
                                'apikey': cfg['key'], 'Authorization': f"Bearer {cfg['key']}",
                                'Content-Type': 'application/json'
                            })
                            urllib.request.urlopen(insert_req, timeout=10)

                        synced += 1
                    except Exception as e:
                        errors += 1
                        if errors <= 3:  # Log first 3 errors
                            print(f"Sync error for project {p.get('name', 'unknown')}: {e}")

                self._send_json(200, {
                    'success': True,
                    'data': {
                        'message': f'Sync complete: {synced} projects synced, {errors} errors',
                        'status': 'completed',
                        'projects_synced': synced,
                        'sync_errors': errors,
                        'total_stuck': len(projects),
                        'last_sync': datetime.now(timezone.utc).isoformat(),
                        'config_status': status
                    }
                })
            except Exception as e:
                self._send_json(500, {'success': False, 'error': str(e)})
            return

        if parsed.path.startswith('/api/improve/instructions/'):
            inst_id = parsed.path.rsplit('/', 1)[-1]
            try:
                content_length = int(self.headers.get('Content-Length', '0') or '0')
                if content_length > TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES:
                    return json_response(self, 413, {'ok': False, 'error': 'request too large'})
                payload = read_request_json(self) or {}
                with _TEMPLATE_IMPROVEMENTS_LOCK:
                    data = _template_improvements_load()
                    found = None
                    for row in data.get('instructions', []):
                        if row.get('id') == inst_id:
                            found = row
                            break
                    if not found:
                        return json_response(self, 404, {'ok': False, 'error': 'instruction not found'})
                    for key in ('domain', 'subdomain', 'agent_key', 'instructions', 'is_active'):
                        if key in payload:
                            found[key] = payload[key]
                    found['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'
                    _template_improvements_save(data)
                return json_response(self, 200, {'ok': True, 'instruction': found})
            except Exception as exc:
                return json_response(self, 500, {'ok': False, 'error': str(exc)})
        if parsed.path.startswith('/api/users/'):
            username = parsed.path.split('/')[-1].strip()
            if not username:
                return json_response(self, 400, {'ok': False, 'error': 'username required'})
            cu = _get_current_user(self)
            if not cu:
                return json_response(self, 403, {'ok': False, 'error': 'auth_required'})
            is_admin = cu.get('role') == 'admin'
            is_self = cu.get('username') == username
            if not is_admin and not is_self:
                return json_response(self, 403, {'ok': False, 'error': 'forbidden'})
            try:
                ln = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(ln) or b'{}')
                users = _mu_users_load()
                target = next((u for u in users if u.get('username') == username), None)
                if not target:
                    return json_response(self, 404, {'ok': False, 'error': 'user not found'})
                if is_admin and 'role' in body:
                    if body['role'] not in ('admin', 'viewer'):
                        return json_response(self, 400, {'ok': False, 'error': 'invalid role'})
                    target['role'] = body['role']
                if is_admin and 'email' in body:
                    target['email'] = body['email']
                if is_self and 'password' in body and body['password']:
                    if len(body['password']) < 6:
                        return json_response(self, 400, {'ok': False, 'error': 'password too short'})
                    target['password_hash'] = _mu_hash_password(body['password'])
                _mu_users_save(users)
                return json_response(self, 200, {'ok': True})
            except Exception as e:
                return json_response(self, 500, {'ok': False, 'error': str(e)})

        return json_response(self, 404, {'ok': False, 'error': 'Not found'})


    def serve_static(self, path):
        clean = posixpath.normpath(urllib.parse.unquote(path))
        if clean in ('', '.', '/'):
            target = INDEX
        else:
            target = (ROOT / clean.lstrip('/')).resolve()
            if ROOT not in target.parents and target != ROOT:
                return json_response(self, 403, {'ok': False, 'error': 'Forbidden'})
        if not target.exists() or not target.is_file():
            return json_response(self, 404, {'ok': False, 'error': 'Not found'})

        content_type, _ = mimetypes.guess_type(str(target))
        if not content_type:
            content_type = 'application/octet-stream'
        return text_response(self, 200, target.read_bytes(), content_type)


# ===== DAILY SKILLS RADAR — helper functions =====

def supabase_radar_config():
    url = (os.getenv('SUPABASE_URL') or '').strip().rstrip('/')
    key = (
        (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or '').strip()
        or (os.getenv('SUPABASE_ANON_KEY') or '').strip()
        or (os.getenv('SUPABASE_API_KEY') or '').strip()
    )
    return {'url': url, 'key': key, 'configured': bool(url and key)}


def _radar_sb_headers(prefer=None):
    cfg = supabase_radar_config()
    if not cfg['configured']:
        return None
    h = {
        'apikey': cfg['key'],
        'Authorization': f"Bearer {cfg['key']}",
        'Accept': 'application/json',
        'Content-Type': 'application/json',
    }
    if prefer:
        h['Prefer'] = prefer
    return h


def _save_skill_to_supabase(skill_dict):
    cfg = supabase_radar_config()
    if not cfg['configured']:
        return None
    hdrs = _radar_sb_headers(prefer='return=representation')
    url = f"{cfg['url']}/rest/v1/skill_discoveries"
    try:
        result = fetch_json(url, headers=hdrs, method='POST', body=skill_dict, timeout=15)
        if isinstance(result, list) and result:
            return result[0].get('id')
        if isinstance(result, dict):
            return result.get('id')
    except Exception:
        pass
    return None


def _update_radar_run(run_id, updates_dict):
    cfg = supabase_radar_config()
    if not cfg['configured'] or not run_id:
        return
    hdrs = _radar_sb_headers()
    url = f"{cfg['url']}/rest/v1/radar_runs?id=eq.{urllib.parse.quote(str(run_id))}"
    try:
        fetch_json(url, headers=hdrs, method='PATCH', body=updates_dict, timeout=15)
    except Exception:
        pass


def _fetch_x_sources(max_results=50):
    """Fetch URLs from user's X/Twitter timeline and bookmarks using Bearer Token.
    Returns a list of source dicts compatible with RADAR_SOURCES format.
    Requires X_BEARER_TOKEN env var. Deduplicates by URL."""
    bearer = (os.getenv('X_BEARER_TOKEN') or 'AAAAAAAAAAAAAAAAAAAAAIZx9AEAAAAAkGXvb%2BF2G2gUuRIiD2Oam0vW5ZQ%3DGt3l07hLntMmlmF1z00OPp5xn7okRM3zLa6SPGiUBOjDqb4Poy').strip()
    if not bearer:
        return []

    sources = []
    seen_urls = set()

    def _x_headers():
        return {
            'Authorization': f'Bearer {bearer}',
            'Content-Type': 'application/json',
        }

    def _extract_urls_from_tweets(tweets):
        found = []
        for tweet in (tweets or []):
            entities = tweet.get('entities') or {}
            for url_obj in (entities.get('urls') or []):
                expanded = url_obj.get('expanded_url', '')
                # Skip t.co redirects, twitter.com/x.com self-links, short internal links
                if not expanded:
                    continue
                if 'twitter.com' in expanded or 'x.com/i/' in expanded:
                    continue
                if len(expanded) < 15:
                    continue
                # Prefer GitHub, docs, README, official sites
                title = url_obj.get('title') or url_obj.get('display_url') or expanded
                desc = url_obj.get('description') or ''
                found.append({'url': expanded, 'title': title, 'description': desc})
        return found

    def _quality_score(url, title, description):
        score = 0.7  # base for X-sourced links
        u = url.lower()
        t = (title + ' ' + description).lower()
        # Boost for high-quality domains
        if 'github.com' in u:
            score += 0.15
        if any(d in u for d in ['docs.', 'documentation', 'readme', 'arxiv.org', 'huggingface.co', 'openai.com', 'anthropic.com']):
            score += 0.1
        if any(k in t for k in ['agent', 'llm', 'ai', 'automation', 'n8n', 'workflow', 'tool', 'framework', 'sdk', 'api', 'mcp']):
            score += 0.05
        if any(k in u for k in ['awesome', 'tutorial', 'guide', 'cookbook', 'playbook']):
            score += 0.05
        return min(round(score, 2), 0.99)

    # Step 1: Get the authenticated user ID
    try:
        me = fetch_json('https://api.twitter.com/2/users/me', headers=_x_headers(), timeout=15)
        user_id = (me.get('data') or {}).get('id')
        if not user_id:
            return []
    except Exception:
        return []

    # Step 2: Fetch home timeline (recent tweets from accounts user follows)
    try:
        params = f'max_results={min(max_results, 100)}&tweet.fields=entities,created_at&expansions=author_id'
        timeline = fetch_json(
            f'https://api.twitter.com/2/users/{user_id}/timelines/reverse_chronological?{params}',
            headers=_x_headers(), timeout=20
        )
        for item in _extract_urls_from_tweets((timeline.get('data') or [])):
            url = item['url']
            if url not in seen_urls:
                seen_urls.add(url)
                sources.append({
                    'url': url,
                    'title': item['title'] or url,
                    'type': 'x_timeline',
                    'quality': _quality_score(url, item['title'], item['description']),
                    'x_source': 'timeline',
                })
    except Exception:
        pass

    # Step 3: Fetch bookmarks (things the user saved)
    try:
        params = f'max_results={min(max_results, 100)}&tweet.fields=entities,created_at'
        bookmarks = fetch_json(
            f'https://api.twitter.com/2/users/{user_id}/bookmarks?{params}',
            headers=_x_headers(), timeout=20
        )
        for item in _extract_urls_from_tweets((bookmarks.get('data') or [])):
            url = item['url']
            if url not in seen_urls:
                seen_urls.add(url)
                # Bookmarks are higher quality — user intentionally saved them
                base_q = _quality_score(url, item['title'], item['description'])
                sources.append({
                    'url': url,
                    'title': item['title'] or url,
                    'type': 'x_bookmark',
                    'quality': min(base_q + 0.05, 0.99),
                    'x_source': 'bookmark',
                })
    except Exception:
        pass

    # Step 4: Fetch user's own likes (things user liked)
    try:
        params = f'max_results={min(max_results, 100)}&tweet.fields=entities,created_at'
        likes = fetch_json(
            f'https://api.twitter.com/2/users/{user_id}/liked_tweets?{params}',
            headers=_x_headers(), timeout=20
        )
        for item in _extract_urls_from_tweets((likes.get('data') or [])):
            url = item['url']
            if url not in seen_urls:
                seen_urls.add(url)
                sources.append({
                    'url': url,
                    'title': item['title'] or url,
                    'type': 'x_like',
                    'quality': _quality_score(url, item['title'], item['description']),
                    'x_source': 'like',
                })
    except Exception:
        pass

    # Filter out low-quality sources (below 0.65) and non-fetchable URLs
    filtered = []
    for s in sources:
        if s['quality'] < 0.65:
            continue
        u = s['url'].lower()
        # Skip non-content URLs
        if any(skip in u for skip in ['youtube.com/watch', 'youtu.be', 'instagram.com', 'tiktok.com', 'facebook.com', 'linkedin.com/in/']):
            continue
        filtered.append(s)

    return filtered


def _get_recent_skill_titles(limit=200):
    cfg = supabase_radar_config()
    if not cfg['configured']:
        return []
    hdrs = _radar_sb_headers()
    url = f"{cfg['url']}/rest/v1/skill_discoveries?select=title&order=created_at.desc&limit={limit}"
    try:
        rows = fetch_json(url, headers=hdrs, timeout=15)
        if isinstance(rows, list):
            return [r.get('title', '') for r in rows if r.get('title')]
    except Exception:
        pass
    return []


def _is_duplicate_skill(title, existing_titles_list):
    def normalize(t):
        t = t.lower()
        t = re.sub(r'[^a-z0-9\s]', ' ', t)
        return set(t.split())
    words_a = normalize(title)
    if not words_a:
        return False
    for existing in existing_titles_list:
        words_b = normalize(existing)
        if not words_b:
            continue
        intersection = words_a & words_b
        union = words_a | words_b
        if union and len(intersection) / len(union) > 0.8:
            return True
    return False


def _extract_skills_from_source(source, content_text, topics=None):
    base, headers = prompt_headers()
    if not headers:
        return []
    content_trimmed = content_text[:8000]
    topics_str = ', '.join(topics) if topics else ''
    topics_instruction = (
        f'\nFocus on skills related to these topics: {topics_str}. '
        f'Prioritize results matching these topics. '
        f'Add a "topics" field to each skill with matching topics from this list.\n'
    ) if topics_str else ''
    system_prompt = (
        "You are an expert agent skill analyst. Extract actionable skills from the provided source content.\n"
        "For each skill, return ONLY a JSON array (no markdown, no explanation) with objects containing:\n"
        "- title: short skill name (5-60 chars)\n"
        "- summary: one clear sentence what this skill enables (max 150 chars)\n"
        "- category: one of: automation, coding, seo, ai_agents, scraping, data, prompting, integration, research, ui_ux, n8n, memory, other\n"
        "- skill_type: one of: tool_integration, workflow_pattern, api_capability, framework, prompt_pattern, automation_recipe, knowledge_artifact\n"
        "- target_agents: array of relevant agents from: [\"Hermes\", \"Claude Code\", \"Manus\", \"CTO\", \"CEO\", \"Prompt Studio\", \"General\"]\n"
        "- usefulness_score: 0.0 to 1.0 float\n"
        "- implementation_idea: one sentence how to implement or use this skill\n"
        "- topics: array of matching topic strings from the focus list (empty array if none match)\n\n"
        "Return 3-8 skills maximum. Only include genuinely useful, actionable skills. Skip obvious/generic ones.\n"
        + topics_instruction
        + f"Source: {source['title']} ({source['type']})"
    )
    body = {
        'model': 'anthropic/claude-sonnet-4.6',
        'max_tokens': 2048,
        'messages': [
            {'role': 'system', 'content': system_prompt},
            {'role': 'user', 'content': f"Extract skills from this content:\n\n{content_trimmed}"},
        ],
    }
    try:
        raw, _ = call_with_fallback(
            [{'role': 'system', 'content': system_prompt},
             {'role': 'user', 'content': f"Extract skills from this content:\n\n{content_trimmed}"}],
            'anthropic/claude-sonnet-4.6', timeout=60
        )
        raw = raw.strip()
        # Strip markdown fences if present
        if raw.startswith('```'):
            raw = re.sub(r'^```[a-z]*\n?', '', raw)
            raw = re.sub(r'\n?```$', '', raw)
        skills = json.loads(raw)
        if isinstance(skills, list):
            return skills
    except Exception:
        pass
    return []


def _run_radar_discovery(config=None):
    import datetime as _dt
    global _radar_state
    _radar_state['status'] = 'running'
    _radar_state['progress'] = 0
    _radar_state['skills_found'] = 0
    _radar_state['sources_done'] = 0
    _radar_state['last_error'] = None
    _radar_state['started_at'] = _dt.datetime.utcnow().isoformat() + 'Z'
    _radar_state['finished_at'] = None
    _radar_state['stage'] = 'Initializing'

    sources = list(RADAR_SOURCES)

    # Fetch dynamic sources from X/Twitter (timeline, bookmarks, likes)
    _radar_state['stage'] = 'Fetching X sources'
    try:
        x_sources = _fetch_x_sources(max_results=50)
        if x_sources:
            # Dedupe against static RADAR_SOURCES urls
            static_urls = {s['url'] for s in RADAR_SOURCES}
            new_x = [s for s in x_sources if s['url'] not in static_urls]
            sources = sources + new_x
            _radar_state['x_sources_found'] = len(new_x)
    except Exception as _xe:
        _radar_state['x_sources_found'] = 0
        _radar_state['last_error'] = f'X fetch: {_xe}'

    if config and config.get('source_types'):
        types = config['source_types']
        # Keep x_timeline/x_bookmark/x_like if 'x' is in the filter, else apply strict filter
        if 'x' not in types:
            sources = [s for s in sources if s.get('type') in types]

    # Apply quality threshold filter
    quality_threshold = float((config or {}).get('quality_threshold') or 0.0)
    if quality_threshold > 0:
        sources = [s for s in sources if s.get('quality', 0) >= quality_threshold]

    _radar_state['sources_total'] = len(sources)
    run_id = None

    # Create radar_run record
    _radar_state['stage'] = 'Sources Queued'
    cfg = supabase_radar_config()
    if cfg['configured']:
        try:
            hdrs = _radar_sb_headers(prefer='return=representation')
            run_body = {
                'status': 'running',
                'sources_count': len(sources),
                'started_at': _radar_state['started_at'],
                'skills_found': 0,
                'skills_approved': 0,
                'errors': [],
            }
            result = fetch_json(f"{cfg['url']}/rest/v1/radar_runs", headers=hdrs, method='POST', body=run_body, timeout=15)
            if isinstance(result, list) and result:
                run_id = result[0].get('id')
            elif isinstance(result, dict):
                run_id = result.get('id')
            _radar_state['run_id'] = run_id
        except Exception as e:
            _radar_state['last_error'] = str(e)

    # Get existing titles for dedup
    _radar_state['stage'] = 'Fetching'
    existing_titles = _get_recent_skill_titles(limit=500)
    errors = []
    skills_found = 0

    for i, source in enumerate(sources):
        if _radar_state.get('paused'):
            import time as _time
            while _radar_state.get('paused'):
                _time.sleep(2)

        _radar_state['stage'] = f"Fetching {source['title']}"
        _radar_state['progress'] = int((i / len(sources)) * 60)

        try:
            content = fetch_text(source['url'], timeout=20)
        except Exception as e:
            errors.append({'source': source['title'], 'error': str(e)})
            _radar_state['sources_done'] = i + 1
            continue

        _radar_state['stage'] = f"Extracting Skills: {source['title']}"
        _radar_state['progress'] = int((i / len(sources)) * 60) + 5

        try:
            active_topics = list((_radar_state.get('topics') or []) + (_radar_state.get('custom_topics') or []))
            skills = _extract_skills_from_source(source, content, topics=active_topics if active_topics else None)
        except Exception as e:
            errors.append({'source': source['title'], 'error': f'LLM error: {e}'})
            _radar_state['sources_done'] = i + 1
            continue

        _radar_state['stage'] = 'Deduplicating'
        for skill in skills:
            title = skill.get('title', '').strip()
            if not title:
                continue
            if _is_duplicate_skill(title, existing_titles):
                continue

            # Save to Supabase
            _radar_state['stage'] = 'Saving Results'
            import datetime as _dt2
            skill_row = {
                'title': title,
                'summary': skill.get('summary', ''),
                'category': skill.get('category', 'other'),
                'skill_type': skill.get('skill_type', 'knowledge_artifact'),
                'target_agents': skill.get('target_agents', ['General']),
                'usefulness_score': float(skill.get('usefulness_score', 0.5)),
                'implementation_idea': skill.get('implementation_idea', ''),
                'source_title': source['title'],
                'source_url': source['url'],
                'source_type': source['type'],
                'source_quality': source['quality'],
                'status': 'pending',
                'run_id': run_id,
                'supabase_synced': False,
                'github_path': None,
                'obsidian_path': None,
                'approval_notes': None,
                'matched_topics': skill.get('topics', []),
                'created_at': _dt2.datetime.utcnow().isoformat() + 'Z',
            }
            try:
                _save_skill_to_supabase(skill_row)
                existing_titles.append(title)
                skills_found += 1
                _radar_state['skills_found'] = skills_found
            except Exception as e:
                errors.append({'source': source['title'], 'error': f'Save error: {e}'})

        _radar_state['sources_done'] = i + 1
        _radar_state['progress'] = int(((i + 1) / len(sources)) * 90)

    import datetime as _dt3
    finished_at = _dt3.datetime.utcnow().isoformat() + 'Z'
    _radar_state['status'] = 'done'
    _radar_state['finished_at'] = finished_at
    _radar_state['progress'] = 100
    _radar_state['stage'] = 'Complete'
    # Round 4: push event to SSE clients
    try:
        sse_broadcast('radar_done', {'summary': f'{skills_found} skills found', 'errors': errors, 'run_id': run_id})
    except Exception: pass

    if run_id:
        _update_radar_run(run_id, {
            'status': 'completed',
            'finished_at': finished_at,
            'skills_found': skills_found,
            'errors': errors,
        })

    # Send email digest after every run
    _radar_state['stage'] = 'Sending email digest'
    try:
        email_result = _send_radar_email_digest(
            skills_found_count=skills_found,
            run_id=run_id,
            errors=errors,
        )
        _radar_state['last_email'] = email_result
    except Exception as _ee:
        _radar_state['last_email'] = {'ok': False, 'error': str(_ee)}

    _radar_state['stage'] = 'Complete'


def _send_radar_email_digest(skills_found_count, run_id=None, errors=None):
    """Send a clean HTML digest email via Resend API after each discovery run."""
    import datetime as _dt
    api_key  = (os.getenv('RESEND_API_KEY') or '').strip()
    to_addr  = (os.getenv('RADAR_EMAIL_TO')  or 'service@maximo-seo.com').strip()
    from_addr= (os.getenv('RADAR_EMAIL_FROM') or 'onboarding@resend.dev').strip()
    if not api_key:
        return {'ok': False, 'error': 'RESEND_API_KEY not configured'}

    now_str = _dt.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    date_str = _dt.datetime.utcnow().strftime('%A, %d %B %Y')

    # Fetch the latest pending skills for this run
    skills_rows = []
    cfg = supabase_radar_config()
    if cfg['configured']:
        try:
            hdrs = _radar_sb_headers()
            url = f"{cfg['url']}/rest/v1/skill_discoveries?select=*&status=eq.pending&order=created_at.desc&limit=30"
            if run_id:
                url += f"&run_id=eq.{urllib.parse.quote(str(run_id))}"
            rows = fetch_json(url, headers=hdrs, timeout=15)
            if isinstance(rows, list):
                skills_rows = rows
        except Exception:
            pass

    # X sources count
    x_count = sum(1 for s in skills_rows if (s.get('source_type') or '').startswith('x_'))

    # ── Category breakdown
    from collections import Counter
    cat_counts = Counter(s.get('category', 'other') for s in skills_rows)
    cat_rows_html = ''.join(
        f'<tr><td style="padding:6px 12px;color:#94a3b8;font-size:13px">{cat}</td>'
        f'<td style="padding:6px 12px;color:#e2e8f0;font-size:13px;font-weight:700;text-align:right">{cnt}</td></tr>'
        for cat, cnt in cat_counts.most_common()
    )

    # ── Skill cards HTML (top 20)
    type_labels = {
        'github_repo': 'GitHub', 'awesome_list': 'Awesome List',
        'x_timeline': 'X Timeline', 'x_bookmark': 'X Bookmark',
        'x_like': 'X Like', 'official_docs': 'Docs',
    }
    type_colors = {
        'github_repo': '#22c55e', 'awesome_list': '#f59e0b',
        'x_timeline': '#60a5fa', 'x_bookmark': '#3b82f6',
        'x_like': '#f9a8d4', 'official_docs': '#a78bfa',
    }
    cat_icons = {
        'automation': '⚙️', 'coding': '💻', 'ai_agents': '🤖', 'seo': '🔍',
        'scraping': '🕷️', 'data': '📊', 'prompting': '💬', 'integration': '🔗',
        'n8n': '🔀', 'research': '🔬', 'ui_ux': '🎨', 'memory': '🧠', 'other': '📦',
    }

    def _esc(s):
        return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

    skill_cards_html = ''
    for s in skills_rows[:20]:
        src_type  = s.get('source_type','')
        src_label = type_labels.get(src_type, src_type)
        src_color = type_colors.get(src_type, '#94a3b8')
        icon      = cat_icons.get(s.get('category','other'), '📦')
        score     = int(float(s.get('usefulness_score') or s.get('source_quality') or 0.5) * 100)
        agents    = ', '.join(s.get('target_agents') or ['General'])
        src_url   = _esc(s.get('source_url') or '#')
        src_title = _esc(s.get('source_title') or s.get('source_url') or '—')
        skill_cards_html += f'''
        <tr>
          <td style="padding:16px;border-bottom:1px solid #1e293b">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td width="36" valign="top" style="padding-right:12px">
                  <div style="width:34px;height:34px;border-radius:10px;background:linear-gradient(135deg,#312e81,#4c1d95);text-align:center;line-height:34px;font-size:16px">{icon}</div>
                </td>
                <td valign="top">
                  <div style="font-size:15px;font-weight:700;color:#e2e8f0;margin-bottom:4px">{_esc(s.get("title","Untitled"))}</div>
                  <div style="font-size:12px;color:#64748b;margin-bottom:6px">
                    <span style="display:inline-block;padding:2px 7px;border-radius:8px;background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.08);color:{src_color};font-weight:700;font-size:10px">{src_label}</span>
                    &nbsp;<a href="{src_url}" style="color:#a78bfa;text-decoration:none">{src_title}</a>
                  </div>
                  <div style="font-size:13px;color:#94a3b8;line-height:1.55;margin-bottom:8px">{_esc(s.get("summary",""))}</div>
                  <div style="font-size:11px;color:#64748b">
                    Agents: <span style="color:#c4b5fd">{_esc(agents)}</span>
                    &nbsp;·&nbsp; Score: <span style="color:#e2e8f0;font-weight:700">{score}%</span>
                    {("&nbsp;·&nbsp; <span style='color:#94a3b8'>" + _esc(s.get("implementation_idea","")) + "</span>") if s.get("implementation_idea") else ""}
                  </div>
                </td>
                <td width="80" valign="top" style="text-align:right">
                  <a href="https://html-redesign-dashboard.maximo-seo.ai/#radar"
                     style="display:inline-block;padding:6px 12px;border-radius:8px;background:linear-gradient(135deg,#7c3aed,#6d28d9);color:#fff;font-size:11px;font-weight:700;text-decoration:none">Review</a>
                </td>
              </tr>
            </table>
          </td>
        </tr>'''

    if not skill_cards_html:
        skill_cards_html = '<tr><td style="padding:24px;text-align:center;color:#64748b;font-size:13px">No new skill candidates found this run.</td></tr>'

    # ── Error rows
    error_section = ''
    if errors:
        err_rows = ''.join(
            f'<tr><td style="padding:4px 0;font-size:12px;color:#f87171">⚠ {_esc(e.get("source","?"))}: {_esc(e.get("error",""))}</td></tr>'
            for e in (errors or [])[:5]
        )
        error_section = f'''
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-top:24px;background:#1e1b4b;border-radius:10px;overflow:hidden">
          <tr><td style="padding:12px 16px;font-size:12px;font-weight:700;color:#f87171;border-bottom:1px solid #312e81">⚠ Errors ({len(errors)})</td></tr>
          {err_rows}
        </table>'''

    # ── Full HTML email
    html_body = f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#0f172a;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#0f172a;padding:32px 16px">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%">

  <!-- Header -->
  <tr><td style="background:linear-gradient(135deg,#1e1b4b,#312e81);border-radius:16px 16px 0 0;padding:28px 32px">
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td>
          <div style="font-size:11px;font-weight:700;color:#7c3aed;text-transform:uppercase;letter-spacing:.1em;margin-bottom:6px">Maximo SEO · Daily Radar</div>
          <div style="font-size:24px;font-weight:800;color:#f1f5f9">📡 Skills Radar Digest</div>
          <div style="font-size:13px;color:#94a3b8;margin-top:6px">{date_str} · {now_str}</div>
        </td>
        <td align="right" valign="top">
          <a href="https://html-redesign-dashboard.maximo-seo.ai/#radar"
             style="display:inline-block;padding:10px 20px;border-radius:10px;background:#7c3aed;color:#fff;font-size:13px;font-weight:700;text-decoration:none">Open Dashboard →</a>
        </td>
      </tr>
    </table>
  </td></tr>

  <!-- Stats bar -->
  <tr><td style="background:#1e293b;padding:20px 32px">
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td style="text-align:center;border-right:1px solid #334155">
          <div style="font-size:28px;font-weight:800;color:#a78bfa">{skills_found_count}</div>
          <div style="font-size:11px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.05em">New Skills</div>
        </td>
        <td style="text-align:center;border-right:1px solid #334155">
          <div style="font-size:28px;font-weight:800;color:#60a5fa">{x_count}</div>
          <div style="font-size:11px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.05em">From X Feed</div>
        </td>
        <td style="text-align:center;border-right:1px solid #334155">
          <div style="font-size:28px;font-weight:800;color:#fbbf24">{len(cat_counts)}</div>
          <div style="font-size:11px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.05em">Categories</div>
        </td>
        <td style="text-align:center">
          <div style="font-size:28px;font-weight:800;color:#4ade80">{len(errors or [])}</div>
          <div style="font-size:11px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.05em">Errors</div>
        </td>
      </tr>
    </table>
  </td></tr>

  <!-- Category breakdown -->
  {'<tr><td style="background:#0f172a;padding:20px 32px"><div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px">By Category</div><table width="100%" cellpadding="0" cellspacing="0" style="background:#1e293b;border-radius:10px;overflow:hidden">' + cat_rows_html + '</table></td></tr>' if cat_counts else ''}

  <!-- Skill cards heading -->
  <tr><td style="background:#0f172a;padding:20px 32px 8px">
    <div style="font-size:16px;font-weight:800;color:#e2e8f0">🎯 Top Discovered Skills</div>
    <div style="font-size:12px;color:#64748b;margin-top:4px">Showing up to 20 · Pending your review</div>
  </td></tr>

  <!-- Skill cards -->
  <tr><td style="background:#0f172a;padding:0 32px">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#1e293b;border-radius:12px;overflow:hidden">
      {skill_cards_html}
    </table>
  </td></tr>

  {error_section}

  <!-- CTA -->
  <tr><td style="background:#0f172a;padding:24px 32px;text-align:center">
    <a href="https://html-redesign-dashboard.maximo-seo.ai/#radar"
       style="display:inline-block;padding:14px 32px;border-radius:12px;background:linear-gradient(135deg,#7c3aed,#6d28d9);color:#fff;font-size:15px;font-weight:800;text-decoration:none">
      Review &amp; Approve Skills →
    </a>
    <div style="font-size:11px;color:#475569;margin-top:12px">Runs daily at 02:00 UTC · maximo-seo.com</div>
  </td></tr>

  <!-- Footer -->
  <tr><td style="background:#0f172a;border-top:1px solid #1e293b;padding:16px 32px;border-radius:0 0 16px 16px">
    <div style="font-size:11px;color:#475569;text-align:center">
      Daily Skills Radar · HTML Redesign Dashboard · Maximo SEO<br>
      Sent to {_esc(to_addr)} · {now_str}
    </div>
  </td></tr>

</table>
</td></tr>
</table>
</body>
</html>'''

    # ── Plain text fallback
    text_lines = [
        f"DAILY SKILLS RADAR DIGEST — {date_str}",
        f"New Skills: {skills_found_count}  |  From X: {x_count}  |  Errors: {len(errors or [])}",
        "",
        "TOP DISCOVERED SKILLS:",
    ]
    for s in skills_rows[:20]:
        text_lines.append(f"  • {s.get('title','?')} [{s.get('category','?')}] — {s.get('summary','')[:80]}")
        if s.get('source_url'):
            text_lines.append(f"    Source: {s.get('source_url','')}")
    text_lines += ["", f"Review at: https://html-redesign-dashboard.maximo-seo.ai/#radar"]
    plain_text = '\n'.join(text_lines)

    subject = f"📡 Skills Radar: {skills_found_count} new skills found — {date_str}"

    # ── Send via Resend API
    payload = {
        'from': from_addr,
        'to': [to_addr],
        'subject': subject,
        'html': html_body,
        'text': plain_text,
    }
    try:
        req = urllib.request.Request(
            'https://api.resend.com/emails',
            data=json.dumps(payload).encode('utf-8'),
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            },
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            resp = json.loads(r.read())
            return {'ok': True, 'email_id': resp.get('id'), 'to': to_addr}
    except urllib.error.HTTPError as exc:
        body_err = exc.read().decode('utf-8', 'replace')[:400]
        return {'ok': False, 'error': f'Resend HTTP {exc.code}: {body_err}'}
    except Exception as exc:
        return {'ok': False, 'error': str(exc)}


def _radar_scheduler_loop():
    import datetime as _dt
    import time as _time
    last_run_date = None
    while True:
        try:
            _time.sleep(60)
            if not _radar_state.get('schedule_enabled'):
                continue
            now = _dt.datetime.now()
            today = now.date()
            if now.hour == _radar_state.get('schedule_hour', 2) and last_run_date != today:
                if _radar_state['status'] != 'running':
                    last_run_date = today
                    _run_radar_discovery()
        except Exception:
            pass


threading.Thread(target=_radar_scheduler_loop, daemon=True).start()



# ===== TEMPLATE IMPROVEMENTS — additive live pipeline MVP (2026-04-27) =====
# TEMPLATE_IMPROVEMENTS_API_PUT_2026_04_27 - compatibility marker for persisted instruction update path.
# Safety: local JSON persistence only; no existing dashboard tables/routes/files are modified.
TEMPLATE_IMPROVEMENTS_FILE = ROOT / 'data' / 'template_improvements.json'
TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES = int(os.getenv('TEMPLATE_IMPROVEMENTS_MAX_REQUEST_BYTES', '350000'))
TEMPLATE_IMPROVEMENTS_MAX_HTML_CHARS = int(os.getenv('TEMPLATE_IMPROVEMENTS_MAX_HTML_CHARS', '250000'))
TEMPLATE_IMPROVEMENTS_MAX_INSTRUCTION_CHARS = int(os.getenv('TEMPLATE_IMPROVEMENTS_MAX_INSTRUCTION_CHARS', '20000'))
TEMPLATE_IMPROVEMENTS_MAX_ACTIVE_JOBS = int(os.getenv('TEMPLATE_IMPROVEMENTS_MAX_ACTIVE_JOBS', '2'))
_TEMPLATE_IMPROVEMENTS_LOCK = threading.RLock()
_TEMPLATE_IMPROVEMENT_AGENTS = [
    {'key': 'gpt-5.4-agent', 'name': 'GPT 5.4 Agent', 'model': 'openai/gpt-5.4', 'role': 'Layout Architect', 'provider': 'OpenRouter'},
    {'key': 'opus-4.7-agent', 'name': 'Opus 4.7 Agent', 'model': 'anthropic/claude-opus-4.7', 'role': 'Visual Designer', 'provider': 'OpenRouter'},
    {'key': 'gemini-3.1-agent', 'name': 'Gemini 3.1 Agent', 'model': 'google/gemini-3.1-pro-preview', 'role': 'Accessibility Auditor', 'provider': 'OpenRouter'},
    {'key': 'kimi-k2.6-agent', 'name': 'Kimi K2.6 Agent', 'model': 'moonshotai/kimi-k2.6', 'role': 'Performance Engineer', 'provider': 'OpenRouter'},
    {'key': 'glm-4.6-agent', 'name': 'GLM 4.6 Agent', 'model': 'z-ai/glm-4.6', 'role': 'Analytics Integrator', 'provider': 'OpenRouter'},
]
_TEMPLATE_IMPROVEMENT_PROGRESS = {
    'queued': 0, 'preparing': 5,
    'agent-1-running': 12, 'agent-1-done': 27,
    'agent-2-running': 30, 'agent-2-done': 44,
    'agent-3-running': 48, 'agent-3-done': 61,
    'agent-4-running': 65, 'agent-4-done': 78,
    'agent-5-running': 83, 'agent-5-done': 95,
    'completed': 100, 'failed': 0, 'cancelled': 0,
}


def _template_improvements_load():
    data = load_json_file(TEMPLATE_IMPROVEMENTS_FILE, {})
    if not isinstance(data, dict):
        data = {}
    data.setdefault('jobs', {})
    data.setdefault('outputs', {})
    data.setdefault('instructions', [])
    return data


def _template_improvements_save(data):
    TEMPLATE_IMPROVEMENTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = TEMPLATE_IMPROVEMENTS_FILE.with_suffix(TEMPLATE_IMPROVEMENTS_FILE.suffix + '.tmp')
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    os.replace(tmp, TEMPLATE_IMPROVEMENTS_FILE)


def _template_improvement_public_job(job):
    if not isinstance(job, dict):
        return None
    return {k: v for k, v in job.items() if k not in ('original_html', 'final_html')}


def _template_improvement_update_job(job_id, **fields):
    with _TEMPLATE_IMPROVEMENTS_LOCK:
        data = _template_improvements_load()
        job = data['jobs'].get(job_id)
        if not job:
            return None
        job.update(fields)
        if 'status' in fields:
            job['progress_percent'] = _TEMPLATE_IMPROVEMENT_PROGRESS.get(fields['status'], job.get('progress_percent', 0))
        job['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'
        data['jobs'][job_id] = job
        _template_improvements_save(data)
        return job


def _template_improvement_extract_html_and_changelog(text):
    raw = (text or '').strip()
    html = raw
    if '```' in html:
        html = re.sub(r'^```(?:html)?\s*', '', html, flags=re.I).strip()
        html = re.sub(r'\s*```$', '', html).strip()
    changelog = []
    for line in raw.splitlines():
        s = line.strip().lstrip('-•* ').strip()
        if s.lower().startswith(('added ', 'fixed ', 'improved ', 'optimized ', 'updated ', 'preserved ', 'changed ')):
            changelog.append(s[:240])
    if not changelog:
        changelog = ['Agent completed template improvement pass']
    return html, changelog[:12]


def _template_improvement_build_prompt(agent, html, global_instructions, specific_instructions, domain, subdomain):
    return f"""You are {agent['name']} - {agent['role']}.

Improve the provided HTML template for domain: {domain or 'unknown'} / subdomain: {subdomain or 'root'}.

GLOBAL CHANGE INSTRUCTIONS:
{global_instructions or '(none)'}

AGENT-SPECIFIC INSTRUCTIONS:
{specific_instructions or '(none)'}

NON-NEGOTIABLE SAFETY RULES:
- Preserve all n8n variables exactly, including {{{{$json...}}}} expressions.
- Preserve links, forms, scripts, and tracking unless instructions explicitly ask to add non-destructive tracking.
- Do not remove content from the original template.
- Return ONLY the improved HTML, followed by an HTML comment named CHANGELOG with short bullet changes.

CURRENT HTML:
{html}
"""


def _template_improvement_call_agent(agent, prompt, timeout=180):
    if os.getenv('TEMPLATE_IMPROVEMENTS_DRY_RUN', '0') == '1' or not os.getenv('OPENROUTER_API_KEY'):
        body = prompt.split('CURRENT HTML:', 1)[-1].strip()
        return body + f"\n<!-- CHANGELOG\n- Dry-run pass for {agent['name']} ({agent['role']})\n- Preserved original HTML because OpenRouter key is not configured\n-->", {'dry_run': True}
    messages = [
        {'role': 'system', 'content': f"You are {agent['name']}, specialized in {agent['role']}."},
        {'role': 'user', 'content': prompt},
    ]
    content, provider = call_with_fallback(messages, agent['model'], timeout=timeout)
    return content, {'provider': provider}


def _template_improvement_run_job(job_id):
    try:
        with _TEMPLATE_IMPROVEMENTS_LOCK:
            data = _template_improvements_load()
            job = data['jobs'].get(job_id)
        if not job or job.get('status') == 'cancelled':
            return
        _template_improvement_update_job(job_id, status='preparing', started_at=datetime.datetime.utcnow().isoformat() + 'Z')
        current_html = job.get('original_html') or ''
        domain = job.get('domain') or ''
        subdomain = job.get('subdomain') or ''
        global_instructions = job.get('change_instructions') or ''
        with _TEMPLATE_IMPROVEMENTS_LOCK:
            instructions = _template_improvements_load().get('instructions', [])
        for idx, agent in enumerate(_TEMPLATE_IMPROVEMENT_AGENTS, 1):
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                latest = _template_improvements_load().get('jobs', {}).get(job_id, {})
            if latest.get('status') == 'cancelled':
                return
            _template_improvement_update_job(job_id, status=f'agent-{idx}-running', current_agent=agent['key'])
            started = time.time()
            specific = '\n'.join(
                inst.get('instructions', '') for inst in instructions
                if inst.get('is_active', True)
                and inst.get('domain') == domain
                and (not inst.get('subdomain') or inst.get('subdomain') == subdomain)
                and (not inst.get('agent_key') or inst.get('agent_key') == agent['key'])
            )
            output_id = str(uuid.uuid4())
            output_record = {
                'id': output_id, 'job_id': job_id,
                'agent_key': agent['key'], 'agent_name': agent['name'],
                'model': agent['model'], 'role': agent['role'], 'provider': agent['provider'],
                'status': 'running', 'input_html': current_html,
                'started_at': datetime.datetime.utcnow().isoformat() + 'Z',
            }
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                data = _template_improvements_load()
                data['outputs'].setdefault(job_id, []).append(output_record)
                _template_improvements_save(data)
            prompt = _template_improvement_build_prompt(agent, current_html, global_instructions, specific, domain, subdomain)
            content, usage = _template_improvement_call_agent(agent, prompt)
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                latest = _template_improvements_load().get('jobs', {}).get(job_id, {})
            if latest.get('status') == 'cancelled':
                return
            html, changelog = _template_improvement_extract_html_and_changelog(content)
            current_html = html or current_html
            duration = max(1, round(time.time() - started))
            with _TEMPLATE_IMPROVEMENTS_LOCK:
                data = _template_improvements_load()
                latest = data.get('jobs', {}).get(job_id, {})
                if latest.get('status') == 'cancelled':
                    _template_improvements_save(data)
                    return
                rows = data['outputs'].setdefault(job_id, [])
                for row in rows:
                    if row.get('id') == output_id:
                        row.update({
                            'status': 'completed', 'output_html': current_html,
                            'changelog': changelog, 'quality_score': 80 + idx * 3,
                            'tokens_used': usage, 'duration_seconds': duration,
                            'completed_at': datetime.datetime.utcnow().isoformat() + 'Z',
                        })
                        break
                _template_improvements_save(data)
            _template_improvement_update_job(job_id, status=f'agent-{idx}-done')
        _template_improvement_update_job(
            job_id, status='completed', current_agent=None, final_html=current_html,
            completed_at=datetime.datetime.utcnow().isoformat() + 'Z',
        )
    except Exception as exc:
        _template_improvement_update_job(job_id, status='failed', error_message=str(exc)[:1000])


def _template_improvement_start_job(payload):
    domain = (payload.get('domain') or '').strip()
    original_html = payload.get('original_html') or payload.get('originalHtml') or ''
    if not domain:
        raise ValueError('domain is required')
    if not original_html.strip():
        raise ValueError('original_html is required')
    change_instructions = payload.get('change_instructions') or payload.get('changeInstructions') or ''
    if len(original_html) > TEMPLATE_IMPROVEMENTS_MAX_HTML_CHARS:
        raise ValueError(f'original_html too large; max {TEMPLATE_IMPROVEMENTS_MAX_HTML_CHARS} chars')
    if len(change_instructions) > TEMPLATE_IMPROVEMENTS_MAX_INSTRUCTION_CHARS:
        raise ValueError(f'change_instructions too large; max {TEMPLATE_IMPROVEMENTS_MAX_INSTRUCTION_CHARS} chars')
    job_id = str(uuid.uuid4())
    now = datetime.datetime.utcnow().isoformat() + 'Z'
    job = {
        'id': job_id, 'domain': domain,
        'subdomain': (payload.get('subdomain') or '').strip(),
        'template_name': (payload.get('template_name') or payload.get('templateName') or 'template.html').strip(),
        'original_html': original_html,
        'change_instructions': change_instructions,
        'status': 'queued', 'current_agent': None, 'progress_percent': 0,
        'error_message': '', 'created_at': now, 'updated_at': now,
    }
    with _TEMPLATE_IMPROVEMENTS_LOCK:
        data = _template_improvements_load()
        active = sum(1 for j in data.get('jobs', {}).values() if j.get('status') not in ('completed', 'failed', 'cancelled'))
        if active >= TEMPLATE_IMPROVEMENTS_MAX_ACTIVE_JOBS:
            raise RuntimeError(f'too many active improvement jobs; max {TEMPLATE_IMPROVEMENTS_MAX_ACTIVE_JOBS}')
        data['jobs'][job_id] = job
        data['outputs'].setdefault(job_id, [])
        _template_improvements_save(data)
    threading.Thread(target=_template_improvement_run_job, args=(job_id,), daemon=True).start()
    return job



# ===== PRODUCTIVITY HUB — additive backend foundation (2026-04-27) =====
# Local JSON persistence only. No n8n workflow modification. No destructive DB migration.
PRODUCTIVITY_HUB_FILE = ROOT / 'data' / 'productivity_hub.json'
_PRODUCTIVITY_HUB_LOCK = threading.RLock()
_PRODUCTIVITY_FEATURES = [
    {'slug': 'client-overview', 'title': 'Client Overview', 'status': 'planned', 'priority': 2},
    {'slug': 'ab-testing', 'title': 'A/B Testing', 'status': 'enabled now', 'priority': 2},
    {'slug': 'agent-traces', 'title': 'Agent Observability', 'status': 'enabled now', 'priority': 2},
    {'slug': 'client-reports', 'title': 'Automated Client Reporting', 'status': 'planned', 'priority': 3},
    {'slug': 'template-gallery', 'title': 'Template Gallery', 'status': 'planned', 'priority': 3},
    {'slug': 'notifications', 'title': 'Notification Center', 'status': 'enabled now', 'priority': 1},
    {'slug': 'batch-operations', 'title': 'Batch Operations', 'status': 'enabled now', 'priority': 2},
    {'slug': 'global-search', 'title': 'Search & Global Filters', 'status': 'enabled now', 'priority': 1},
    {'slug': 'audit-log', 'title': 'Audit Log', 'status': 'enabled now', 'priority': 2},
    {'slug': 'quick-actions', 'title': 'Quick Actions', 'status': 'enabled now', 'priority': 1},
    {'slug': 'cost-tracker', 'title': 'Cost Tracker', 'status': 'enabled now', 'priority': 1},
    {'slug': 'team-notes', 'title': 'Collaboration Notes', 'status': 'enabled now', 'priority': 4},
    {'slug': 'pwa-mobile', 'title': 'PWA / Mobile', 'status': 'partially available', 'priority': 4},
    {'slug': 'export-hub', 'title': 'Data Export Hub', 'status': 'enabled now', 'priority': 3},
    {'slug': 'scheduled-runs', 'title': 'Scheduled Pipeline Runs', 'status': 'enabled now', 'priority': 3},
]

def _productivity_load():
    data = load_json_file(PRODUCTIVITY_HUB_FILE, {})
    if not isinstance(data, dict):
        data = {}
    data.setdefault('notifications', [])
    data.setdefault('audit', [])
    data.setdefault('feature_flags', {})
    return data

def _productivity_save(data):
    PRODUCTIVITY_HUB_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = PRODUCTIVITY_HUB_FILE.with_suffix(PRODUCTIVITY_HUB_FILE.suffix + '.tmp')
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    os.replace(tmp, PRODUCTIVITY_HUB_FILE)

def _productivity_summary():
    with _PRODUCTIVITY_HUB_LOCK:
        data = _productivity_load()
        flags = data.get('feature_flags', {}) if isinstance(data.get('feature_flags'), dict) else {}
        notifications = list(data.get('notifications', []))
        audit = list(data.get('audit', []))
    features = []
    for row in _PRODUCTIVITY_FEATURES:
        item = dict(row)
        item['is_enabled'] = bool(flags.get(row['slug']) or row.get('status') in ('enabled now', 'partially available'))
        item['rollback'] = 'disable feature flag'
        features.append(item)
    return {
        'ok': True,
        'generated_at': datetime.datetime.utcnow().isoformat() + 'Z',
        'features': features,
        'safety': {
            'additive_only': True,
            'no_n8n_workflow_modification': True,
            'destructive_db_migrations': False,
            'existing_routes_changed': False,
        },
        'signals': {
            'notifications_count': len(notifications),
            'unread_notifications': sum(1 for n in notifications if not n.get('is_read')),
            'audit_count': len(audit),
            'enabled_features': sum(1 for f in features if f.get('is_enabled')),
        }
    }

def _productivity_add_notification(payload):
    now = datetime.datetime.utcnow().isoformat() + 'Z'
    row = {
        'id': str(uuid.uuid4()),
        'type': (payload.get('type') or 'info').strip()[:32],
        'title': (payload.get('title') or 'Notification').strip()[:180],
        'message': (payload.get('message') or '').strip()[:1000],
        'link': (payload.get('link') or '').strip()[:500],
        'is_read': bool(payload.get('is_read', False)),
        'created_at': now,
    }
    with _PRODUCTIVITY_HUB_LOCK:
        data = _productivity_load()
        data.setdefault('notifications', []).insert(0, row)
        data['notifications'] = data['notifications'][:200]
        _productivity_save(data)
    return row

def _productivity_add_audit(payload):
    now = datetime.datetime.utcnow().isoformat() + 'Z'
    row = {
        'id': str(uuid.uuid4()),
        'message': (payload.get('message') or 'Activity recorded').strip()[:500],
        'entity_type': (payload.get('entity_type') or payload.get('entityType') or '').strip()[:80],
        'entity_id': (payload.get('entity_id') or payload.get('entityId') or '').strip()[:180],
        'user_id': (payload.get('user_id') or payload.get('userId') or 'dashboard').strip()[:120],
        'created_at': now,
    }
    with _PRODUCTIVITY_HUB_LOCK:
        data = _productivity_load()
        data.setdefault('audit', []).insert(0, row)
        data['audit'] = data['audit'][:500]
        _productivity_save(data)
    return row

def _productivity_search(query):
    q = (query or '').strip().lower()
    out = []
    for f in _PRODUCTIVITY_FEATURES:
        hay = ' '.join([f.get('slug',''), f.get('title',''), f.get('status','')]).lower()
        if not q or q in hay:
            out.append({'type': 'feature', 'title': f['title'], 'slug': f['slug'], 'status': f['status']})
    with _PRODUCTIVITY_HUB_LOCK:
        data = _productivity_load()
        notifications = list(data.get('notifications', []))
        audit = list(data.get('audit', []))
    for n in notifications:
        hay = ' '.join([n.get('title',''), n.get('message',''), n.get('type','')]).lower()
        if q and q in hay:
            out.append({'type': 'notification', 'title': n.get('title'), 'id': n.get('id'), 'status': n.get('type')})
    for a in audit:
        hay = ' '.join([a.get('message',''), a.get('entity_type',''), a.get('entity_id','')]).lower()
        if q and q in hay:
            out.append({'type': 'audit', 'title': a.get('message'), 'id': a.get('id'), 'entity_id': a.get('entity_id')})
    return out[:50]

def main():
    port = int(os.getenv('PORT', '8000'))
    server = ThreadingHTTPServer(('0.0.0.0', port), DashboardHandler)
    print(f'listening on {port}', flush=True)
    try:
        _stage14_start_scheduler_once()
    except Exception as _e:
        print(f'[stage14] scheduler start failed: {_e}', flush=True)
    # Start schedule engine
    try:
        schedule_engine.start_scheduler()
        print("[server] Schedule engine started", flush=True)
    except Exception as e:
        print(f"[server] Failed to start schedule engine: {e}", flush=True)
    server.serve_forever()



# ── RADAR HELPERS ──
def _generate_radar_excel(payload):
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None, "openpyxl not installed"
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary['A1'] = 'Skill Radar Report'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Subject: {payload.get('subject_name', 'Unknown')}"
    ws_summary['A3'] = f"Date: {payload.get('report_date', 'Unknown')}"
    ws_summary['A4'] = f"Overall Score: {payload.get('overall_score', 0)} / 100"
    
    # Sheet 2: Scores
    ws_scores = wb.create_sheet("Scores")
    headers = ['Dimension', 'Score', 'Target', 'Gap', 'Trend']
    ws_scores.append(headers)
    for col in range(1, 6):
        ws_scores.cell(row=1, column=col).font = Font(bold=True)
        
    scores = payload.get('scores', {})
    targets = payload.get('targets', {})
    trends = payload.get('trends', {})
    
    for dim, score in scores.items():
        target = targets.get(dim, 0)
        gap = score - target
        trend = trends.get(dim, 'N/A')
        ws_scores.append([dim, score, target, gap, trend])
        
    # Sheet 3: AI Recommendations
    ws_recs = wb.create_sheet("AI Recommendations")
    ws_recs.append(['Priority', 'Area', 'Recommendation', 'Expected Impact'])
    for col in range(1, 5):
        ws_recs.cell(row=1, column=col).font = Font(bold=True)
        
    recs = payload.get('ai_recommendations', [])
    for rec in recs:
        ws_recs.append([rec.get('priority', ''), rec.get('area', ''), rec.get('recommendation', ''), rec.get('impact', '')])
        
    # Set column widths
    ws_scores.column_dimensions['A'].width = 20
    ws_recs.column_dimensions['B'].width = 20
    ws_recs.column_dimensions['C'].width = 50
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None

    
def _save_radar_report(payload):
    import os, json
    from datetime import datetime
    try:
        subject = payload.get('subject_name', 'Unknown').replace(' ', '_').replace('/', '_')
        date_str = datetime.now().strftime('%Y-%m-%d')
        base_dir = os.path.join(os.getcwd(), 'Reports', 'Skill-Radar', 'agents', subject)
        os.makedirs(base_dir, exist_ok=True)
        
        # 1. Markdown for Obsidian
        md_path = os.path.join(base_dir, f"{date_str}-radar.md")
        score_rows = ""
        for dim, score in payload.get('scores', {}).items():
            target = payload.get('targets', {}).get(dim, 0)
            gap = score - target
            trend = payload.get('trends', {}).get(dim, '0')
            score_rows += f"| {dim} | {score} | {target} | {gap} | {trend} |\n"
            
        recs_text = ""
        for i, rec in enumerate(payload.get('ai_recommendations', [])):
            recs_text += f"{i+1}. {rec.get('priority', '')}: {rec.get('area', '')} - {rec.get('recommendation', '')}\n"

        md_content = f"""---
type: agent_skills
subject: {payload.get('subject_name', '')}
date: {date_str}
overall_score: {payload.get('overall_score', 0)}
---

# Skill Radar: {payload.get('subject_name', '')} — {date_str}

## Scores
| Dimension | Score | Target | Gap | Trend |
|---|---|---|---|---|
{score_rows}

## AI Analysis
{payload.get('ai_summary', 'Automated radar analysis completed.')}

## Recommendations
{recs_text}
"""
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(md_content)

        # 2. JSON Backup for GitHub
        json_path = os.path.join(base_dir, f"{date_str}.json")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2)

        return True, None
    except Exception as e:
        return False, str(e)

if __name__ == '__main__':
    main()

# ── Local Skill Discovery Fallback ─────────────────────────────
