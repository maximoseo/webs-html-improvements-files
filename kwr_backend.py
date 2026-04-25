"""
Keyword Research Automation backend.
General-purpose: works for any site - no hardcoded domains.

State machine stages:
  queued -> validating -> analyzing_site -> parsing_sitemap ->
  researching_competitors -> generating_rows -> deduplicating ->
  ready | error | cancelled

Thread safety: all _state mutations go through _lock (RLock).
No imports from server.py - call_llm injected as argument.
"""

import datetime
import io
import json
import logging
import os
import re
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid


# #9 Structured logger for production debugging on Render
logger = logging.getLogger('kwr')
if not logger.handlers:
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter(
        '%(asctime)s [%(levelname)s] kwr: %(message)s',
        datefmt='%Y-%m-%dT%H:%M:%S'
    ))
    logger.addHandler(_h)
    logger.setLevel(os.environ.get('KWR_LOG_LEVEL', 'INFO').upper())
    logger.propagate = False

_state = {}          # run_id -> job dict
_lock = threading.RLock()

# Disk persistence for KWR runs (survives Render restarts).
OUTPUTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')
try:
    os.makedirs(OUTPUTS_DIR, exist_ok=True)
except Exception:
    pass


def _run_dir(run_id: str) -> str:
    return os.path.join(OUTPUTS_DIR, run_id)


def _atomic_write_json(path: str, data) -> None:
    """Write JSON atomically: write to .tmp then os.replace (POSIX atomic).
    Prevents partial/corrupt files if the process is killed mid-write (e.g. Render OOM).
    """
    tmp = path + '.tmp'
    try:
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            try:
                f.flush()
                os.fsync(f.fileno())
            except Exception:
                pass
        os.replace(tmp, path)
    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        raise


def _persist_job(run_id: str) -> None:
    """Write job JSON + meta.json to outputs/{run_id}/ for cross-restart access."""
    try:
        with _lock:
            job = _state.get(run_id)
            if job is None:
                return
            job_copy = dict(job)
        rd = _run_dir(run_id)
        os.makedirs(rd, exist_ok=True)
        _atomic_write_json(os.path.join(rd, 'job.json'), job_copy)
        # Write a small meta.json for the reports listing
        try:
            inputs = job_copy.get('inputs') or {}
            website = (inputs.get('website_url') or '').strip()
            domain = website.replace('https://', '').replace('http://', '').split('/')[0]
            meta = {
                'run_id': run_id,
                'domain': domain,
                'website_url': website,
                'brand': inputs.get('brand_name', ''),
                'worksheet_name': job_copy.get('worksheet_name') or _worksheet_name(job_copy, ''),
                'created_at': job_copy.get('created_at', ''),
                'updated_at': job_copy.get('updated_at', ''),
                'status': job_copy.get('status', ''),
                'row_count': int(job_copy.get('row_count') or len(job_copy.get('rows') or [])),
            }
            _atomic_write_json(os.path.join(rd, 'meta.json'), meta)
        except Exception:
            pass
    except Exception:
        pass


def _load_job_from_disk(run_id: str) -> dict | None:
    try:
        path = os.path.join(_run_dir(run_id), 'job.json')
        if not os.path.exists(path):
            return None
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def get_state():
    """Return a copy of all runs for analytics. Loads from disk if _state is empty."""
    with _lock:
        if _state:
            return dict(_state)
    # Try to load from disk
    try:
        import glob
        runs = {}
        for meta_path in glob.glob(os.path.join(OUTPUTS_DIR, '*/meta.json')):
            try:
                with open(meta_path, 'r', encoding='utf-8') as f:
                    meta = json.load(f)
                rid = meta.get('run_id')
                if rid:
                    runs[rid] = {
                        'status': meta.get('status', 'unknown'),
                        'created_at_ts': None,
                        'updated_at_ts': None,
                        'duration_seconds': None,
                        'error': None,
                        'deploy_error': None,
                    }
            except Exception:
                pass
        return runs
    except Exception:
        return {}


def clean_keywords(cell: str, brand: str = '') -> str:
    """
    Split a comma-separated keyword cell, strip whitespace, dedup case-insensitively
    while preserving order, and (if brand given) ensure brand is the first entry.
    Returns a clean comma-joined string. Also normalizes whitespace (collapses
    multiple spaces) and strips surrounding quotes/brackets that LLMs sometimes add.
    """
    if not cell:
        return (brand or '').strip()
    raw = str(cell).strip().strip('"\'[](){}').replace('\n', ',').replace(';', ',')
    parts = [re.sub(r'\s+', ' ', p).strip(' "\'') for p in raw.split(',')]
    parts = [p for p in parts if p]
    seen = set()
    out = []
    for p in parts:
        k = p.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    brand = (brand or '').strip()
    if brand:
        bk = brand.casefold()
        # Remove any existing brand occurrence (any case) then put clean brand first.
        out = [p for p in out if p.casefold() != bk]
        out.insert(0, brand)
    return ', '.join(out)


def _kw_set(cell: str) -> set:
    """Helper: return casefolded set of individual keywords from a cleaned cell."""
    return {p.strip().casefold() for p in (cell or '').split(',') if p.strip()}


_SEMANTIC_STOPWORDS = {
    'a', 'an', 'the', 'and', 'or', 'for', 'to', 'of', 'in', 'on', 'with', 'near',
    'best', 'top', 'service', 'services', 'company', 'companies', 'expert', 'experts',
}


def _keyword_tokens(text: str, brand: str = '') -> list:
    text = re.sub(r'[^\w\s-]', ' ', (text or '').casefold())
    text = re.sub(r'[_\-/]+', ' ', text)
    parts = [p for p in text.split() if p]
    brand_tokens = {p for p in re.sub(r'[^\w\s-]', ' ', (brand or '').casefold()).split() if p}
    out = []
    for token in parts:
        if token in brand_tokens or token in _SEMANTIC_STOPWORDS:
            continue
        if len(token) > 4 and token.endswith('ies'):
            token = token[:-3] + 'y'
        elif len(token) > 4 and token.endswith('es'):
            token = token[:-2]
        elif len(token) > 3 and token.endswith('s'):
            token = token[:-1]
        out.append(token)
    return out


def _semantic_fingerprint(text: str, brand: str = '') -> tuple:
    tokens = sorted(set(_keyword_tokens(text, brand)))
    if tokens:
        return tuple(tokens)
    fallback = re.sub(r'\s+', ' ', (text or '').casefold()).strip()
    return (fallback,) if fallback else tuple()


def _slug_fingerprint(path: str) -> tuple:
    slug = (path or '').strip('/').split('/')[-1]
    slug = slug.replace('-', ' ')
    return _semantic_fingerprint(slug)


def _quality_score_row(row: dict, brand: str = '', existing_page_fps=None, pillar_primary_map=None) -> tuple:
    existing_page_fps = existing_page_fps or set()
    pillar_primary_map = pillar_primary_map or {}
    keyword = (row.get('primary_keyword') or '').strip()
    keywords = clean_keywords(row.get('keywords') or '', brand)
    intent = (row.get('intent') or '').strip().lower()
    pillar = (row.get('pillar') or '').strip()
    fp = _semantic_fingerprint(keyword, brand)
    keyword_terms = [p for p in keyword.split() if p]
    support_terms = [p.strip() for p in keywords.split(',') if p.strip()]

    score = 70
    reasons = []

    if 2 <= len(keyword_terms) <= 5:
        score += 8
        reasons.append('keyword length is long-tail friendly')
    elif len(keyword_terms) <= 1:
        score -= 18
        reasons.append('keyword is too broad')
    elif len(keyword_terms) >= 7:
        score -= 8
        reasons.append('keyword may be too long or unnatural')

    if support_terms and support_terms[0].casefold() == (brand or '').strip().casefold():
        score += 4
        reasons.append('brand-first keyword list is preserved')
    else:
        score -= 6
        reasons.append('brand-first keyword list is missing')

    if len(set(support_terms)) >= 4:
        score += 6
        reasons.append('supporting keywords add topic depth')
    elif len(set(support_terms)) <= 2:
        score -= 8
        reasons.append('supporting keywords are too thin')

    if intent == 'pillar':
        score += 4
        reasons.append('pillar topic can support multiple clusters')
    elif fp and fp in existing_page_fps:
        score -= 28
        reasons.append('keyword looks close to existing site coverage')

    pillar_fp = _semantic_fingerprint((pillar_primary_map.get(pillar) or ''), brand) if pillar else tuple()
    if intent != 'pillar' and fp and pillar_fp and fp == pillar_fp:
        score -= 32
        reasons.append('cluster looks too similar to its pillar keyword')

    score = max(0, min(100, int(round(score))))
    tier = 'high' if score >= 80 else 'medium' if score >= 60 else 'low'
    if not reasons:
        reasons.append('balanced keyword opportunity')
    return score, tier, reasons[:4]


def prepare_kwr_rows(raw_rows: list, existing_pages=None, brand: str = '', source_model: str = '') -> tuple:
    existing_pages = existing_pages or []
    normalized_rows = []
    pillar_primary_map = {}
    brand_name = (brand or '').strip()

    for row in raw_rows or []:
        cleaned_kw = clean_keywords((row.get('keywords') or '').strip(), brand_name)
        clean_row = {
            'existing_parent_page': (row.get('existing_parent_page') or '-').strip(),
            'pillar': (row.get('pillar') or '').strip(),
            'cluster': (row.get('cluster') or '').strip(),
            'intent': (row.get('intent') or 'informational').strip(),
            'primary_keyword': (row.get('primary_keyword') or '').strip(),
            'keywords': cleaned_kw,
            'source_model': (row.get('source_model') or source_model or '').strip(),
        }
        normalized_rows.append(clean_row)
        if clean_row['intent'].strip().lower() == 'pillar' and clean_row['pillar']:
            pillar_primary_map[clean_row['pillar']] = clean_row['primary_keyword']

    seen_kw = set()
    seen_fp = set()
    cluster_fp_by_pillar = {}
    existing_page_fps = {_slug_fingerprint(p) for p in existing_pages if p}
    existing_page_fps.discard(tuple())
    clean_rows = []
    quality_counts = {'high': 0, 'medium': 0, 'low': 0}
    score_total = 0
    stats = {
        'raw_count': len(raw_rows or []),
        'exact_duplicates_removed': 0,
        'semantic_duplicates_removed': 0,
        'existing_coverage_removed': 0,
        'quality_summary': {'total': 0, 'avg_score': 0, 'high': 0, 'medium': 0, 'low': 0},
    }

    for row in normalized_rows:
        pk = (row.get('primary_keyword') or '').strip()
        if not pk:
            continue
        pk_cf = pk.casefold()
        fp = _semantic_fingerprint(pk, brand_name)
        intent = (row.get('intent') or '').strip().lower()
        pillar = (row.get('pillar') or '').strip()

        if pk_cf in seen_kw:
            stats['exact_duplicates_removed'] += 1
            continue
        if fp and fp in seen_fp:
            stats['semantic_duplicates_removed'] += 1
            continue
        if fp and intent != 'pillar' and fp in existing_page_fps:
            stats['existing_coverage_removed'] += 1
            continue
        pillar_fp = _semantic_fingerprint(pillar_primary_map.get(pillar, ''), brand_name) if pillar else tuple()
        if fp and intent != 'pillar' and pillar_fp and fp == pillar_fp:
            stats['semantic_duplicates_removed'] += 1
            continue
        if fp and intent != 'pillar':
            bucket = cluster_fp_by_pillar.setdefault(pillar, set())
            if fp in bucket:
                stats['semantic_duplicates_removed'] += 1
                continue
            bucket.add(fp)

        seen_kw.add(pk_cf)
        if fp:
            seen_fp.add(fp)

        score, tier, reasons = _quality_score_row(
            row,
            brand=brand_name,
            existing_page_fps=existing_page_fps,
            pillar_primary_map=pillar_primary_map,
        )
        row['quality_score'] = score
        row['quality_tier'] = tier
        row['quality_reasons'] = reasons
        clean_rows.append(row)
        quality_counts[tier] += 1
        score_total += score

    total = len(clean_rows)
    stats['quality_summary'] = {
        'total': total,
        'avg_score': round(score_total / total, 1) if total else 0,
        'high': quality_counts['high'],
        'medium': quality_counts['medium'],
        'low': quality_counts['low'],
    }
    return clean_rows, stats


def build_model_participation(child_jobs: list, merged_rows: list) -> list:
    kept_counts = {}
    for row in merged_rows or []:
        model = (row.get('source_model') or '').strip() or 'unknown'
        kept_counts[model] = kept_counts.get(model, 0) + 1

    summary = []
    for job in child_jobs or []:
        inputs = job.get('inputs') or {}
        model = (inputs.get('model') or inputs.get('_model') or '').strip() or 'unknown'
        total_rows = len(job.get('rows') or [])
        kept_rows = kept_counts.get(model, 0)
        summary.append({
            'model': model,
            'total_rows': total_rows,
            'kept_rows': kept_rows,
            'kept_pct': round((kept_rows / total_rows) * 100, 1) if total_rows else 0,
        })

    summary.sort(key=lambda item: (-item.get('kept_rows', 0), item.get('model', '')))
    return summary


def validate_kwr_rows(rows: list, brand: str = '', log=None) -> dict:
    """
    Conservative validator — runs AFTER clean_keywords. Does NOT mutate rows.
    Returns a dict with counts + list of warnings. Logs each warning if log() given.

    Checks:
      1. Pillar URL convention: pillar rows should have existing_parent_page='-'
      2. Cluster URL convention: cluster rows should have '/pillar-slug/' format
      3. Anti-cannibalization: cluster keywords overlapping their pillar's keywords
      4. Brand-first: brand appears as first KW in every row
    """
    warnings = []
    log = log or (lambda *_: None)

    # Build pillar -> keyword-set map
    pillar_kws = {}
    for r in rows:
        if (r.get('intent') or '').strip().lower() == 'pillar':
            pname = (r.get('pillar') or '').strip()
            if pname:
                pillar_kws[pname] = _kw_set(r.get('keywords', ''))

    brand_cf = (brand or '').strip().casefold()

    for i, r in enumerate(rows):
        intent = (r.get('intent') or '').strip().lower()
        epp = (r.get('existing_parent_page') or '').strip()
        pillar = (r.get('pillar') or '').strip()
        kw_cell = (r.get('keywords') or '').strip()

        # Brand-first check
        if brand_cf:
            first_kw = kw_cell.split(',')[0].strip().casefold() if kw_cell else ''
            if first_kw != brand_cf:
                warnings.append(f"row {i} ({pillar}): brand not first in keywords")

        # URL convention checks
        if intent == 'pillar':
            if epp != '-':
                warnings.append(f"row {i} pillar '{pillar}': existing_parent_page should be '-' (got '{epp}')")
        else:
            if epp == '-' or not epp.startswith('/') or not epp.endswith('/'):
                warnings.append(f"row {i} cluster under '{pillar}': bad URL format '{epp}' (expected '/pillar-slug/')")

        # Anti-cannibalization (warning only, no mutation)
        if intent != 'pillar' and pillar in pillar_kws:
            cluster_kws = _kw_set(kw_cell) - {brand_cf}
            overlap = cluster_kws & (pillar_kws[pillar] - {brand_cf})
            if overlap:
                warnings.append(
                    f"row {i} cluster under '{pillar}': cannibalization risk — "
                    f"shares KWs with pillar: {sorted(overlap)[:3]}"
                )

    for w in warnings:
        log(f"[KWR validate] WARN: {w}")

    return {
        'warnings_count': len(warnings),
        'warnings': warnings[:50],  # cap to avoid bloat
        'pillars_validated': len(pillar_kws),
        'rows_validated': len(rows),
    }

# ---------------------------------------------------------------------------
# Public API (called by server.py route handlers)
# ---------------------------------------------------------------------------

def start_run(payload: dict, call_llm) -> tuple:
    """
    Validate payload, create a new run, launch threaded runner.
    Returns (run_id, None) on success or (None, error_str) on validation failure.
    """
    required = ['website_url', 'sitemap_url', 'about_url', 'brand_name',
                'target_language', 'target_market']
    missing = [f for f in required if not (payload.get(f) or '').strip()]
    if missing:
        return None, f"Missing required fields: {', '.join(missing)}"

    # Validate URLs
    for field in ['website_url', 'sitemap_url', 'about_url']:
        val = payload.get(field, '').strip()
        if val and not val.startswith(('http://', 'https://')):
            return None, f"{field} must start with http:// or https://"

    run_id = str(uuid.uuid4())
    now = datetime.datetime.utcnow().isoformat() + 'Z'

    with _lock:
        _state[run_id] = {
            'run_id': run_id,
            'status': 'queued',
            'current_stage': 'queued',
            'finished_stages': [],
            'error_stages': [],
            'progress': 0,
            'logs': [],
            'rows': [],
            'existing_pages': [],
            'row_count': 0,
            'honest_count_note': '',
            'quality_summary': {'total': 0, 'avg_score': 0, 'high': 0, 'medium': 0, 'low': 0},
            'dedup_summary': {'raw_count': 0, 'exact_duplicates_removed': 0, 'semantic_duplicates_removed': 0, 'existing_coverage_removed': 0},
            'model_participation': [],
            'preview_edited': False,
            'created_at': now,
            'updated_at': now,
            'inputs': payload,
            'cancel_requested': False,
            'sheet_url': None,
            'deploy_error': None,
            'deployed_at': None,
        }

    t = threading.Thread(target=_runner, args=(run_id, call_llm), daemon=True)
    t.start()
    return run_id, None


def get_status(run_id: str) -> dict | None:
    with _lock:
        job = _state.get(run_id)
        if job is not None:
            return dict(job)   # shallow copy is fine for JSON serialization
    # Fallback: read from disk (survives server restarts)
    disk = _load_job_from_disk(run_id)
    if disk is not None:
        return disk
    return None


def cancel_run(run_id: str) -> bool:
    with _lock:
        job = _state.get(run_id)
        if job is None:
            return False
        if job['status'] in ('ready', 'complete', 'error', 'cancelled'):
            return False
        job['cancel_requested'] = True
        return True


COLUMNS = ['עמוד אב קיים', 'Pillar', 'Cluster', 'Intent', 'Primary Keyword', 'Keywords']


def _worksheet_name(job: dict, sheet_prefix: str) -> str:
    brand = (job.get('inputs') or {}).get('brand_name', 'kwr')
    safe_brand = _slugify(brand)[:30]
    date_str = datetime.datetime.utcnow().strftime('%Y-%m-%d')
    prefix = _slugify(sheet_prefix)[:20] if sheet_prefix else ''
    return (f"{prefix}-{safe_brand}-{date_str}".strip('-')
            if prefix else f"kwr-{safe_brand}-{date_str}")


def build_excel(run_id: str) -> tuple:
    """
    Build an .xlsx file in memory from the current rows.
    Returns (bytes, worksheet_name, None) or (None, None, error_str).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, None, "openpyxl not installed - cannot build Excel file"

    with _lock:
        job = _state.get(run_id)
        if job is not None:
            rows = list(job.get('rows') or [])
            ws_name = _worksheet_name(job, '')
        else:
            job = None
            rows = None
            ws_name = None

    if job is None:
        # Try disk fallback (in-memory state gone after restart).
        # First try cached xlsx.
        rd = _run_dir(run_id)
        xlsx_path = os.path.join(rd, 'file.xlsx')
        if os.path.exists(xlsx_path):
            try:
                with open(xlsx_path, 'rb') as f:
                    data = f.read()
                # Recover worksheet name from job.json if present
                disk = _load_job_from_disk(run_id) or {}
                ws_name = disk.get('worksheet_name') or _worksheet_name(disk or {'inputs': {}}, '')
                return data, ws_name, None
            except Exception as exc:
                return None, None, f"Disk cache read failed: {exc}"
        disk = _load_job_from_disk(run_id)
        if disk is None:
            return None, None, f"Run {run_id} not found"
        rows = list(disk.get('rows') or [])
        ws_name = disk.get('worksheet_name') or _worksheet_name(disk, '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ws_name[:31]  # Excel tab name limit

    # Header style
    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Pillar / cluster row styles — one pastel color per pillar block (matches reference template)
    PILLAR_PALETTE = [
        'E6F3FF',  # light blue
        'FFE6E6',  # light red/pink
        'E8F5E9',  # light green
        'FFF3E0',  # light amber
        'F3E5F5',  # light purple
        'E0F7FA',  # light teal
        'FFF9C4',  # light yellow
        'FCE4EC',  # light pink
        'E8EAF6',  # light indigo
        'EFEBE9',  # light brown
        'F1F8E9',  # light lime
        'E0F2F1',  # light mint
        'FFEBEE',  # blush
        'EDE7F6',  # lavender
        'FFF8E1',  # cream
        'E1F5FE',  # sky
        'F9FBE7',  # pale lime
        'FBE9E7',  # peach
        'F0F4C3',  # pale chartreuse
        'D7CCC8',  # taupe
        'C8E6C9',  # sage
        'B3E5FC',  # baby blue
        'F8BBD0',  # rose
        'D1C4E9',  # lilac
    ]
    LINK_COLOR = '0563C1'
    pillar_idx = -1
    current_pillar_key = None
    current_pillar_name = ''

    for r_idx, row in enumerate(rows, 2):
        # Support both schemas: snake_case (current pipeline) and col_a..col_f (legacy)
        col_a = row.get('existing_parent_page', row.get('col_a', '')) or ''
        col_b = row.get('pillar',               row.get('col_b', '')) or ''
        col_c = row.get('cluster',              row.get('col_c', '')) or ''
        col_d = row.get('intent',               row.get('col_d', '')) or ''
        col_e = row.get('primary_keyword',      row.get('col_e', '')) or ''
        col_f = row.get('keywords',             row.get('col_f', '')) or ''

        # Pillar detection: intent=='pillar' OR col_a == '-' OR cluster empty/equals pillar
        is_pillar = (
            str(col_d).strip().lower() == 'pillar'
            or str(col_a).strip() in ('-', '')
            or not str(col_c).strip()
            or str(col_c).strip() == str(col_b).strip()
        )

        # Track the pillar block this row belongs to
        # Only advance palette when we hit a new pillar header.
        # Cluster rows often have empty pillar name → inherit from current block.
        pillar_key = str(col_b).strip().lower()
        if is_pillar:
            if pillar_key != current_pillar_key:
                pillar_idx = (pillar_idx + 1) % len(PILLAR_PALETTE)
                current_pillar_key = pillar_key
                current_pillar_name = str(col_b).strip()
        else:
            # Cluster row — inherit pillar name if empty
            if not str(col_b).strip() and current_pillar_name:
                col_b = current_pillar_name
            if pillar_idx < 0:
                pillar_idx = 0

        block_color = PILLAR_PALETTE[pillar_idx if pillar_idx >= 0 else 0]
        fill = PatternFill('solid', fgColor=block_color)

        # Pillar row: bold, cluster column repeats pillar name (block header)
        # Cluster row: regular weight
        if is_pillar:
            font_default = Font(bold=True, size=11, color='000000')
            if not str(col_c).strip():
                col_c = col_b  # cluster column = pillar name on header rows
            # Pillar URL: '-' if empty/dash, else hyperlink
            url_value = str(col_a).strip() or '-'
        else:
            font_default = Font(size=10, color='000000')
            url_value = str(col_a).strip() or '-'

        values = [url_value, col_b, col_c, col_d, col_e, col_f]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='right' if c_idx in (2, 3, 5, 6) else 'left')
            cell.border = border
            # Hyperlink styling for URL column when it's an actual URL
            if c_idx == 1 and url_value and url_value != '-':
                cell.hyperlink = url_value
                cell.font = Font(size=11 if is_pillar else 10, bold=is_pillar, color=LINK_COLOR, underline='single')
            else:
                cell.font = font_default

    # Column widths
    col_widths = [30, 25, 30, 15, 35, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    # Free workbook ASAP — openpyxl holds large in-memory structures
    try:
        wb.close()
    except Exception:
        pass
    del wb, buf
    try:
        import gc as _gc
        _gc.collect()
    except Exception:
        pass
    # Cache xlsx to disk so /api/kwr/download survives Render restarts.
    try:
        rd = _run_dir(run_id)
        os.makedirs(rd, exist_ok=True)
        with open(os.path.join(rd, 'file.xlsx'), 'wb') as f:
            f.write(data)
    except Exception:
        pass
    return data, ws_name, None


def approve_and_save(run_id: str, edited_rows: list, sheet_prefix: str) -> tuple:
    """
    Mark run as approved, save edited rows, build Excel bytes.
    Returns (excel_bytes, worksheet_name, None) or (None, None, error_str).
    Does NOT write anywhere externally — caller decides where to save.
    """
    with _lock:
        job = _state.get(run_id)
        if job is None:
            return None, None, f"Run {run_id} not found"
        if job['status'] not in ('ready', 'complete'):
            return None, None, f"Run is in state '{job['status']}' — can only approve 'ready' runs"
        if not edited_rows:
            return None, None, "No rows to save"
        job['rows'] = edited_rows
        job['row_count'] = len(edited_rows)

    excel_bytes, ws_name, err = build_excel(run_id)
    if err:
        return None, None, err

    deployed_at = datetime.datetime.utcnow().isoformat() + 'Z'
    with _lock:
        _state[run_id]['status'] = 'complete'
        _state[run_id]['deployed_at'] = deployed_at
        _state[run_id]['worksheet_name'] = ws_name
        _state[run_id]['deploy_error'] = None
        # Round 4: SSE broadcast on KWR completion
        try:
            import sys
            _srv = sys.modules.get('server') or sys.modules.get('__main__')
            if _srv and hasattr(_srv, 'sse_broadcast'):
                _srv.sse_broadcast('kwr_done', {'run_id': run_id, 'summary': f'Deployed to {ws_name}', 'worksheet': ws_name})
        except Exception: pass

    # Persist approved state so downloads survive restarts.
    _persist_job(run_id)
    # Best-effort multi-destination sync (GitHub + Obsidian + Supabase)
    try:
        threading.Thread(target=sync_report, args=(run_id,), daemon=True).start()
    except Exception:
        pass

    return excel_bytes, ws_name, None


def push_to_sheets(run_id: str, sheet_target: str) -> tuple:
    """
    Push approved rows to Google Sheets using gspread + service account.
    Requires GOOGLE_SERVICE_ACCOUNT_JSON env var (JSON string of the SA key).
    Returns (sheet_url, None) or (None, error_str).
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return None, "gspread / google-auth not installed"

    sa_json = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON', '').strip()
    if not sa_json:
        return None, "GOOGLE_SERVICE_ACCOUNT_JSON env var not set"

    with _lock:
        job = _state.get(run_id)
        if job is None:
            return None, f"Run {run_id} not found"
        rows = list(job.get('rows') or [])
        ws_name = job.get('worksheet_name') or _worksheet_name(job, '')

    if not rows:
        return None, "No rows to push — approve first"
    if not sheet_target or not sheet_target.strip():
        return None, "sheet_target (Google Sheet URL or ID) is required"

    try:
        sa_info = json.loads(sa_json)
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive',
        ]
        creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
        gc = gspread.authorize(creds)

        # Open by URL or ID
        if 'docs.google.com' in sheet_target:
            sh = gc.open_by_url(sheet_target)
        else:
            sh = gc.open_by_key(sheet_target)

        # Always create a NEW worksheet — never overwrite
        try:
            ws = sh.add_worksheet(title=ws_name, rows=len(rows) + 5, cols=6)
        except Exception:
            # If name collision, append timestamp
            ws_name = ws_name + '-' + datetime.datetime.utcnow().strftime('%H%M%S')
            ws = sh.add_worksheet(title=ws_name, rows=len(rows) + 5, cols=6)

        # Write header + rows
        data = [COLUMNS]
        for row in rows:
            data.append([
                row.get('col_a', ''),
                row.get('col_b', ''),
                row.get('col_c', ''),
                row.get('col_d', ''),
                row.get('col_e', ''),
                row.get('col_f', ''),
            ])
        ws.update(data, value_input_option='RAW')

        # Bold header
        ws.format('A1:F1', {'textFormat': {'bold': True},
                             'backgroundColor': {'red': 0.12, 'green': 0.22, 'blue': 0.39}})

        sheet_url = sh.url
        with _lock:
            _state[run_id]['sheet_url'] = sheet_url
            _state[run_id]['worksheet_name'] = ws_name

        return sheet_url, None

    except json.JSONDecodeError:
        return None, "GOOGLE_SERVICE_ACCOUNT_JSON is not valid JSON"
    except Exception as exc:
        return None, str(exc)[:500]


def get_note_content(run_id: str) -> tuple:
    """
    Build the Obsidian markdown note content and note path for a run.
    Returns (note_content_str, note_path_str, None) or (None, None, error_str).
    The caller (browser JS) is responsible for PUT-ing to the Obsidian local REST API
    at http://127.0.0.1:27123 — the server never calls Obsidian directly because
    Obsidian runs on the user's local machine, not on Render.
    """
    with _lock:
        job = _state.get(run_id)
        if job is None:
            return None, None, f"Run {run_id} not found"
        rows = list(job.get('rows') or [])
        inputs = job.get('inputs') or {}
        ws_name = job.get('worksheet_name', run_id)
        created_at = job.get('created_at', '')

    brand = inputs.get('brand_name', 'unknown')
    website = inputs.get('website_url', '')
    market = inputs.get('target_market', '')
    language = inputs.get('target_language', '')
    note_path = f"html-redesign-dashboard/kwr-results/{ws_name}.md"

    lines = [
        f"# KWR: {brand}",
        f"",
        f"- **Site:** {website}",
        f"- **Market:** {market}",
        f"- **Language:** {language}",
        f"- **Run date:** {created_at[:10] if created_at else 'unknown'}",
        f"- **Rows:** {len(rows)}",
        f"- **Worksheet:** {ws_name}",
        f"",
        f"| Existing Parent Page | Pillar | Cluster | Intent | Primary Keyword | Keywords |",
        f"|---|---|---|---|---|---|",
    ]
    for row in rows:
        def esc(v):
            return str(v or '').replace('|', '\\|')
        lines.append(
            f"| {esc(row.get('col_a',''))} "
            f"| {esc(row.get('col_b',''))} "
            f"| {esc(row.get('col_c',''))} "
            f"| {esc(row.get('col_d',''))} "
            f"| {esc(row.get('col_e',''))} "
            f"| {esc(row.get('col_f',''))} |"
        )

    note_content = '\n'.join(lines)
    return note_content, note_path, None


def save_to_obsidian(run_id: str) -> tuple:
    """
    DEPRECATED: called save_to_obsidian from server-side, which doesn't work on Render
    because Obsidian runs on the user's local machine.
    Kept for backwards compat — delegates to get_note_content and returns the path.
    The browser JS should use /api/kwr/note-content/<run_id> instead.
    """
    note_content, note_path, err = get_note_content(run_id)
    if err:
        return None, err
    # Server-side Obsidian write only works in local dev (not on Render).
    # Return path so the caller knows the intended vault path.
    return note_path, None


# --- Internal helper preserved for save_to_obsidian legacy signature ---
def _save_to_obsidian_local(run_id: str) -> tuple:
    """
    Actually attempts the server-side Obsidian PUT (only works in local dev).
    """
    obsidian_url = os.getenv('OBSIDIAN_LOCAL_HTTP_URL', 'http://127.0.0.1:27123').rstrip('/')
    obsidian_key = os.getenv('OBSIDIAN_LOCAL_API_KEY', '').strip()
    if not obsidian_key:
        return None, "OBSIDIAN_LOCAL_API_KEY env var not set"

    note_content, note_path, err = get_note_content(run_id)
    if err:
        return None, err

def update_rows(run_id: str, rows: list) -> bool:
    """Allow user to save edits to the preview without deploying."""
    with _lock:
        job = _state.get(run_id)
        if job is None:
            return False
        job['rows'] = rows
        job['preview_edited'] = True
        job['row_count'] = len(rows)
        return True


def list_recent(limit: int = 10) -> list:
    with _lock:
        jobs = sorted(_state.values(), key=lambda j: j.get('created_at', ''), reverse=True)
        return [_job_summary(j) for j in jobs[:limit]]


# ---------------------------------------------------------------------------
# Internal runner
# ---------------------------------------------------------------------------

class _Cancelled(Exception):
    pass


def _runner(run_id: str, call_llm):
    def log(msg: str):
        ts = datetime.datetime.utcnow().strftime('%H:%M:%S')
        with _lock:
            _state[run_id]['logs'].append(f"[{ts}] {msg}")
            _state[run_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'

    def set_stage(stage: str, progress: int):
        with _lock:
            prev = _state[run_id]['current_stage']
            if prev and prev != stage and prev not in _state[run_id]['finished_stages']:
                _state[run_id]['finished_stages'].append(prev)
            _state[run_id]['current_stage'] = stage
            _state[run_id]['progress'] = progress
            _state[run_id]['status'] = 'running'
            _state[run_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'

    def check_cancel():
        with _lock:
            if _state[run_id].get('cancel_requested'):
                raise _Cancelled()

    def fail(stage: str, error: str):
        with _lock:
            _state[run_id]['status'] = 'error'
            _state[run_id]['error'] = str(error)
            try:
                import sys as _s4
                _srv4 = _s4.modules.get('server')
                if _srv4 and hasattr(_srv4, 'sse_broadcast'):
                    _srv4.sse_broadcast('kwr_error', {'run_id': run_id, 'error': str(error)[:200]})
            except Exception: pass
            _state[run_id]['current_stage'] = stage
            _state[run_id]['error_stages'].append(stage)
            _state[run_id]['logs'].append(f"[ERROR] {error}")
            _state[run_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'

    try:
        with _lock:
            inputs = dict(_state[run_id]['inputs'])

        website_url   = inputs['website_url'].rstrip('/')
        sitemap_url   = inputs['sitemap_url']
        about_url     = inputs['about_url']
        brand_name    = inputs['brand_name']
        target_lang   = inputs['target_language']
        target_market = inputs['target_market']
        sheet_target  = inputs.get('spreadsheet_target', '')
        competitor_urls = [u.strip() for u in (inputs.get('competitor_urls') or '').split('\n') if u.strip()]
        exclusions    = inputs.get('notes_exclusions', '')
        # Model override from UI (optional — defaults to OpenRouter-first GPT-5.4)
        llm_model     = (inputs.get('model') or inputs.get('_model') or 'anthropic/claude-opus-4.7').strip()

        # ── Stage 1: validating ─────────────────────────────────────────────
        set_stage('validating', 5)
        check_cancel()
        log(f"Validating inputs for {website_url}")

        # Quick HEAD check on website_url
        try:
            req = urllib.request.Request(website_url, method='HEAD', headers={'User-Agent': 'KWR-Bot/1.0'})
            urllib.request.urlopen(req, timeout=10)
            log(f"Site reachable: {website_url}")
        except Exception as e:
            log(f"Warning: site HEAD check failed ({e}) - continuing anyway")

        # ── Stage 2: analyzing_site ─────────────────────────────────────────
        set_stage('analyzing_site', 15)
        check_cancel()
        log("Fetching home page and about page...")

        home_text = _fetch_page_text(website_url)
        about_text = _fetch_page_text(about_url)
        log(f"Home page: {len(home_text)} chars, About page: {len(about_text)} chars")

        # Summarize site with LLM
        check_cancel()
        site_summary_prompt = f"""Analyze this website for keyword research purposes.
Brand: {brand_name}
Target Language: {target_lang}
Target Market: {target_market}
Home page content (truncated):
{home_text[:3000]}

About page content (truncated):
{about_text[:3000]}

Return a JSON object with:
- "services": array of strings (actual services/products offered)
- "topics": array of strings (main topic areas)
- "tone": string (brand tone)
- "description": string (1-2 sentence site description)
Only include real services/products visible in the content. Do not invent offerings."""

        site_msgs = [
            {'role': 'system', 'content': 'You are an SEO analyst. Return only valid JSON.'},
            {'role': 'user', 'content': site_summary_prompt}
        ]
        site_summary_raw, _ = call_llm(site_msgs, llm_model)
        try:
            site_summary = json.loads(_extract_json(site_summary_raw))
        except Exception:
            site_summary = {'services': [], 'topics': [], 'tone': 'professional', 'description': site_summary_raw[:200]}
        log(f"Site analysis done. Services: {site_summary.get('services', [])[:3]}")

        # ── Stage 3: parsing_sitemap ─────────────────────────────────────────
        set_stage('parsing_sitemap', 30)
        check_cancel()
        log(f"Parsing sitemap: {sitemap_url}")

        existing_pages = _parse_sitemap(sitemap_url)
        log(f"Found {len(existing_pages)} existing pages in sitemap")
        with _lock:
            _state[run_id]['existing_pages'] = existing_pages[:200]  # cap for UI

        # ─�� Stage 4: researching_competitors ────────────────────────────────
        set_stage('researching_competitors', 45)
        check_cancel()
        log("Researching competitor topics...")

        comp_topics = []
        # Auto-discover competitors via DuckDuckGo if none provided
        if not competitor_urls:
            try:
                topics_for_search = (site_summary.get('topics') or site_summary.get('services') or [])[:3]
                seed_domain = urllib.parse.urlparse(website_url).netloc.lower()
                discovered = []
                for topic in topics_for_search:
                    if not topic: continue
                    q = f"{topic} {target_market}".strip()
                    try:
                        search_url = f"https://duckduckgo.com/html/?q={urllib.parse.quote(q)}"
                        req = urllib.request.Request(search_url, headers={'User-Agent': 'Mozilla/5.0'})
                        html = urllib.request.urlopen(req, timeout=15).read().decode('utf-8', errors='ignore')
                        for m in re.finditer(r'href="(https?://[^"?]+)"', html):
                            url = m.group(1)
                            dom = urllib.parse.urlparse(url).netloc.lower()
                            if dom and seed_domain not in dom and 'duckduckgo' not in dom and 'google' not in dom and dom not in [urllib.parse.urlparse(x).netloc.lower() for x in discovered]:
                                discovered.append(f"https://{dom}")
                                if len(discovered) >= 3: break
                        if len(discovered) >= 3: break
                    except Exception as e:
                        log(f"Competitor discovery failed for '{topic}': {e}")
                competitor_urls = discovered
                log(f"Auto-discovered competitors: {competitor_urls}")
            except Exception as e:
                log(f"Competitor auto-discovery error: {e}")

        if competitor_urls:
            for cu in competitor_urls[:3]:
                check_cancel()
                comp_text = _fetch_page_text(cu)
                log(f"Fetched competitor: {cu} ({len(comp_text)} chars)")
                comp_topics.append({'url': cu, 'text': comp_text[:2000]})

        # ── Stage 5: generating_rows ─────────────────────────────────────────
        set_stage('generating_rows', 60)
        check_cancel()
        log("Generating keyword research rows with LLM...")

        existing_paths = '\n'.join(existing_pages[:100]) if existing_pages else '(none found)'
        comp_context = '\n'.join([f"Competitor {i+1} ({c['url']}): {c['text'][:500]}" for i,c in enumerate(comp_topics)]) if comp_topics else '(none provided)'
        exclusion_note = f"Exclusions: {exclusions}" if exclusions else ''

        kwr_prompt = f"""You are an expert SEO content strategist. Generate a comprehensive keyword research plan for:
Brand: {brand_name}
Website: {website_url}
Target Language: {target_lang}
Target Market: {target_market}
Services/Products: {json.dumps(site_summary.get('services', []))}
Main Topics: {json.dumps(site_summary.get('topics', []))}
{exclusion_note}

Existing pages on the site (DO NOT create pillars/clusters that duplicate these — these are already covered):
{existing_paths}

Competitor research context (use these to find topical gaps the site is missing):
{comp_context}

═══════════════════════════════════════════════════════════════════════════
STRUCTURE RULES — read carefully, this is the most important section:
═══════════════════════════════════════════════════════════════════════════

1. CREATE 10-15 NEW PILLARS — topics NOT already covered by the existing pages above. Pillars are the strategic content gaps that will drive new traffic.

2. EACH PILLAR = ONE "PILLAR ROW" + 10-15 "CLUSTER ROWS":
   • PILLAR ROW: existing_parent_page="-", pillar=<topic>, cluster=<same as pillar>, intent="pillar", primary_keyword=<main KW for pillar>, keywords="<BRAND_NAME>, <3-4 supporting KWs>"
   • CLUSTER ROWS: existing_parent_page="/<pillar-slug>/" (slug of the NEW pillar, e.g. "/garage-door-repair/"), pillar=<same pillar name>, cluster=<unique sub-topic>, intent=<one of: informational|commercial|transactional|navigational>, primary_keyword=<unique KW>, keywords="<BRAND_NAME>, <3-4 unique long-tail KWs>"

3. URL RULES (STRICT):
   • Pillar rows ALWAYS use "-" as existing_parent_page (these are NEW pages to be created).
   • Cluster rows use a slug path "/<pillar-slug>/" pointing to their NEW pillar (not a real domain — the slug only). Example: "/garage-door-spring-repair/".
   • DO NOT invent full URLs with the brand domain. DO NOT use real existing-page URLs (those pages are already covered).

4. KEYWORDS FORMAT (CRITICAL):
   • Brand/business name MUST always appear FIRST in the keywords field, then 3-4 supporting keywords separated by commas.
   • Example: "{brand_name}, garage door repair, broken spring fix, emergency garage service"

5. ANTI-CANNIBALIZATION:
   • Cluster keywords must be DIFFERENT from their pillar's keywords (not even LSI/synonyms).
   • Each primary_keyword must be UNIQUE across the entire output.
   • Each cluster targets a distinct sub-intent / long-tail variation.

6. PILLAR/CLUSTER NAMING:
   • DO NOT include city names in pillar names or cluster names.
   • City names are ALLOWED only inside primary_keyword and keywords fields (e.g. "garage door repair San Antonio").

7. RELEVANCE:
   • Only include topics genuinely related to what the business actually does (per the about page and existing pages).
   • Use competitor context to identify legitimate topic gaps — but only adopt topics the business CAN credibly cover.
   • Prefer terms real people search (Google Auto-Complete, People Also Ask, Related Searches mindset).

8. VOLUME:
   • Target 200-250 total rows (10-15 pillars × 11-16 rows each including the pillar row itself).
   • Mix long-tail and short-tail keywords.

9. LANGUAGE: Use {target_lang} for all pillar/cluster/keyword text (except brand name which stays as-is).

═══════════════════════════════════════════════════════════════════════════
OUTPUT FORMAT:
Return ONLY a JSON array. Each object has exactly these keys:
  existing_parent_page, pillar, cluster, intent, primary_keyword, keywords
No explanation, no markdown, no preamble — just the raw JSON array."""

        kwr_msgs = [
            {'role': 'system', 'content': 'You are an expert SEO strategist. Return only valid JSON arrays.'},
            {'role': 'user', 'content': kwr_prompt}
        ]
        kwr_raw, provider = call_llm(kwr_msgs, llm_model)
        log(f"LLM response received via {provider} (model={llm_model}, {len(kwr_raw)} chars)")

        # ── Stage 6: deduplicating ────────────────────────────────────────────
        set_stage('deduplicating', 80)
        check_cancel()
        log("Parsing and deduplicating rows...")

        try:
            raw_rows = json.loads(_extract_json(kwr_raw))
            if not isinstance(raw_rows, list):
                raise ValueError("LLM did not return an array")
        except Exception as e:
            fail('deduplicating', f"Failed to parse LLM output as JSON: {e}\nRaw (first 500): {kwr_raw[:500]}")
            return

        # Normalize + deduplicate with semantic filtering + quality scoring
        brand_name = (_state.get(run_id, {}).get('inputs', {}) or {}).get('brand_name', '') if run_id in _state else ''
        clean_rows, prep_stats = prepare_kwr_rows(
            raw_rows,
            existing_pages=existing_pages,
            brand=brand_name,
            source_model=llm_model,
        )
        log(
            f"After deduplication: {len(clean_rows)} rows (from {len(raw_rows)} raw, "
            f"exact dupes removed={prep_stats.get('exact_duplicates_removed', 0)}, "
            f"semantic dupes removed={prep_stats.get('semantic_duplicates_removed', 0)}, "
            f"existing coverage removed={prep_stats.get('existing_coverage_removed', 0)})"
        )

        # Conservative validation — warnings only, no mutation
        try:
            v = validate_kwr_rows(clean_rows, brand_name, log=log)
            if v['warnings_count']:
                log(f"[KWR validate] {v['warnings_count']} warning(s) across {v['rows_validated']} rows ({v['pillars_validated']} pillars)")
            else:
                log(f"[KWR validate] OK — {v['rows_validated']} rows, {v['pillars_validated']} pillars, no issues")
        except Exception as _ve:
            log(f"[KWR validate] skipped: {_ve}")

        quality_summary = prep_stats.get('quality_summary') or {'total': len(clean_rows), 'avg_score': 0, 'high': 0, 'medium': 0, 'low': 0}
        log(
            f"Quality mix — avg {quality_summary.get('avg_score', 0)}, "
            f"high {quality_summary.get('high', 0)}, medium {quality_summary.get('medium', 0)}, low {quality_summary.get('low', 0)}"
        )

        honest_note = ''
        if len(clean_rows) < 200:
            honest_note = (
                f"Produced {len(clean_rows)} valid net-new rows. "
                f"The site's existing coverage and deduplication constraints "
                f"did not support expansion to 200 rows without adding irrelevant filler."
            )
            log(f"Note: {honest_note}")

        with _lock:
            _state[run_id]['rows'] = clean_rows
            _state[run_id]['row_count'] = len(clean_rows)
            _state[run_id]['honest_count_note'] = honest_note
            _state[run_id]['quality_summary'] = quality_summary
            _state[run_id]['dedup_summary'] = {
                'raw_count': prep_stats.get('raw_count', len(raw_rows)),
                'exact_duplicates_removed': prep_stats.get('exact_duplicates_removed', 0),
                'semantic_duplicates_removed': prep_stats.get('semantic_duplicates_removed', 0),
                'existing_coverage_removed': prep_stats.get('existing_coverage_removed', 0),
            }
            _state[run_id]['status'] = 'ready'
            _state[run_id]['current_stage'] = 'ready'
            _state[run_id]['progress'] = 100
            _state[run_id]['finished_stages'].append('deduplicating')
            _state[run_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'

        log(f"Pipeline complete. {len(clean_rows)} rows ready for review.")

        # Persist to disk so /api/kwr/download survives Render restarts.
        _persist_job(run_id)
        try:
            build_excel(run_id)  # caches outputs/{run_id}/file.xlsx
        except Exception:
            pass

    except _Cancelled:
        with _lock:
            _state[run_id]['status'] = 'cancelled'
            _state[run_id]['current_stage'] = 'cancelled'
            _state[run_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'
        _log_safe(run_id, "Run cancelled by user.")

    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        with _lock:
            stage = _state[run_id].get('current_stage', 'unknown')
            _state[run_id]['status'] = 'error'
            _state[run_id]['error'] = str(exc)
            try:
                import sys as _s4
                _srv4 = _s4.modules.get('server')
                if _srv4 and hasattr(_srv4, 'sse_broadcast'):
                    _srv4.sse_broadcast('kwr_error', {'run_id': run_id, 'error': str(exc)[:200]})
            except Exception: pass
            _state[run_id]['error_stages'].append(stage)
            _state[run_id]['logs'].append(f"[ERROR] Unexpected: {exc}")
            _state[run_id]['logs'].append(tb[-800:])
            _state[run_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _log_safe(run_id, msg):
    ts = datetime.datetime.utcnow().strftime('%H:%M:%S')
    with _lock:
        if run_id in _state:
            _state[run_id]['logs'].append(f"[{ts}] {msg}")


# ---------------------------------------------------------------------------
# Supabase KWR save
# ---------------------------------------------------------------------------

def _supabase_kwr_config():
    url = (os.getenv('SUPABASE_URL') or '').strip().rstrip('/')
    key = (
        (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or '').strip()
        or (os.getenv('SUPABASE_ANON_KEY') or '').strip()
        or (os.getenv('SUPABASE_API_KEY') or '').strip()
    )
    table = (os.getenv('SUPABASE_KWR_TABLE') or 'kwr_results').strip()
    schema = (os.getenv('SUPABASE_SCHEMA') or 'public').strip()
    return {'url': url, 'key': key, 'table': table, 'schema': schema, 'configured': bool(url and key)}


def save_to_supabase(run_id: str) -> tuple:
    """
    Save approved KWR rows to Supabase table `kwr_results`.
    Each row becomes one record with the full job metadata.
    Returns (record_count, None) or (None, error_str).
    Table schema (auto-created via insert):\n      id, run_id, domain, brand_name, run_date,
      existing_parent_page, pillar, cluster, intent,
      primary_keyword, keywords, model_used
    """
    cfg = _supabase_kwr_config()
    if not cfg['configured']:
        return None, 'SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY/SUPABASE_ANON_KEY are not configured'

    with _lock:
        job = _state.get(run_id)
        if job is None:
            return None, f'Run {run_id} not found'
        rows = list(job.get('rows') or [])
        inputs = dict(job.get('inputs') or {})
        created_at = job.get('created_at', '')
        status = job.get('status', '')

    if not rows:
        return None, 'No rows to save — approve first'
    if status not in ('ready', 'complete'):
        return None, f"Run is in state '{status}' — approve first"

    brand = inputs.get('brand_name', '')
    website = inputs.get('website_url', '')
    # Extract domain from URL
    domain = website.replace('https://', '').replace('http://', '').split('/')[0]
    model_used = inputs.get('model', '') or inputs.get('_model', '') or 'default'
    run_date = created_at[:10] if created_at else datetime.datetime.utcnow().strftime('%Y-%m-%d')

    url = cfg['url']
    key = cfg['key']
    table = cfg['table']
    schema = cfg['schema']

    headers = {
        'apikey': key,
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'Accept-Profile': schema,
        'Content-Profile': schema,
        'Prefer': 'return=minimal',
    }

    records = []
    for row in rows:
        records.append({
            'run_id': run_id,
            'domain': domain,
            'brand_name': brand,
            'run_date': run_date,
            'existing_parent_page': row.get('col_a', '') or row.get('existing_parent_page', ''),
            'pillar': row.get('col_b', '') or row.get('pillar', ''),
            'cluster': row.get('col_c', '') or row.get('cluster', ''),
            'intent': row.get('col_d', '') or row.get('intent', ''),
            'primary_keyword': row.get('col_e', '') or row.get('primary_keyword', ''),
            'keywords': row.get('col_f', '') or row.get('keywords', ''),
            'model_used': model_used,
        })

    # Insert in batches of 100
    batch_size = 100
    total_saved = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        body_bytes = json.dumps(batch).encode('utf-8')
        req = urllib.request.Request(
            f'{url}/rest/v1/{table}',
            data=body_bytes,
            headers=headers,
            method='POST'
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                _ = resp.read()
                total_saved += len(batch)
        except urllib.error.HTTPError as exc:
            err_body = exc.read().decode('utf-8', 'replace')[:400]
            return None, f'Supabase HTTP {exc.code}: {err_body}'
        except Exception as exc:
            return None, f'Supabase error: {exc}'

    return total_saved, None




# ---------------------------------------------------------------------------
# Ensemble / swarm mode — run multiple models, merge + deduplicate results
# ---------------------------------------------------------------------------

ENSEMBLE_MODELS = [
    'anthropic/claude-opus-4.7',
    'openai/gpt-5.4',
    'google/gemini-2.5-pro',
]

BEST_TEXT_SWARM_MODELS = [
    'anthropic/claude-opus-4.7',
    'openai/gpt-5.4',
    'google/gemini-2.5-pro',
    'x-ai/grok-4.20-multi-agent',
    'z-ai/glm-5.1',
    'moonshotai/kimi-k2-thinking',
]


def _normalize_model_list(models) -> list:
    if not isinstance(models, list):
        return []
    normalized = []
    seen = set()
    for raw in models:
        model = str(raw or '').strip()
        if not model or model in seen:
            continue
        seen.add(model)
        normalized.append(model)
    return normalized


def start_ensemble(payload: dict, call_llm, default_models=None, mode_label: str = 'ensemble') -> tuple:
    """
    Launch ensemble run: fire N single-model runs, then merge + deduplicate.
    Returns (ensemble_run_id, None) or (None, error_str).
    """
    models = _normalize_model_list(payload.get('ensemble_models') or default_models or ENSEMBLE_MODELS)
    if not models:
        models = list(default_models or ENSEMBLE_MODELS)

    run_prefix = 'swarm' if mode_label == 'swarm' else 'ensemble'
    # Create a coordinator run_id
    ensemble_id = run_prefix + '-' + str(uuid.uuid4())
    now = datetime.datetime.utcnow().isoformat() + 'Z'
    child_run_ids = []
    stored_inputs = dict(payload)
    stored_inputs['ensemble_models'] = list(models)
    stored_inputs['_mode'] = mode_label

    with _lock:
        _state[ensemble_id] = {
            'run_id': ensemble_id,
            'status': 'running',
            'current_stage': 'launching',
            'finished_stages': [],
            'error_stages': [],
            'progress': 0,
            'logs': [f'[{datetime.datetime.utcnow().strftime("%H:%M:%S")}] {mode_label.title()}: launching {len(models)} models'],
            'rows': [],
            'existing_pages': [],
            'row_count': 0,
            'honest_count_note': '',
            'quality_summary': {'total': 0, 'avg_score': 0, 'high': 0, 'medium': 0, 'low': 0},
            'dedup_summary': {'raw_count': 0, 'exact_duplicates_removed': 0, 'semantic_duplicates_removed': 0, 'existing_coverage_removed': 0},
            'model_participation': [],
            'preview_edited': False,
            'created_at': now,
            'updated_at': now,
            'inputs': stored_inputs,
            'cancel_requested': False,
            'sheet_url': None,
            'deploy_error': None,
            'deployed_at': None,
            'ensemble_models': models,
            'child_run_ids': child_run_ids,
        }

    # Launch one child run per model in parallel threads
    for model in models:
        child_payload = dict(payload)
        child_payload['model'] = model
        child_payload.pop('ensemble_models', None)
        child_id, err = start_run(child_payload, call_llm)
        if err:
            _log_safe(ensemble_id, f'WARN: failed to start model {model}: {err}')
            continue
        child_run_ids.append(child_id)
        with _lock:
            _state[ensemble_id]['child_run_ids'] = list(child_run_ids)

    if not child_run_ids:
        with _lock:
            _state[ensemble_id]['status'] = 'error'
            _state[ensemble_id]['logs'].append('[ERROR] All child runs failed to start')
        return ensemble_id, None  # return ID so UI can poll

    # Launch merger thread
    t = threading.Thread(target=_ensemble_merger, args=(ensemble_id, child_run_ids), daemon=True)
    t.start()
    return ensemble_id, None


def start_best_text_swarm(payload: dict, call_llm) -> tuple:
    swarm_payload = dict(payload or {})
    swarm_payload.pop('ensemble_models', None)
    swarm_payload.pop('model', None)
    swarm_payload['model'] = 'best-text-swarm'
    swarm_payload['_mode'] = 'swarm'
    return start_ensemble(
        swarm_payload,
        call_llm,
        default_models=BEST_TEXT_SWARM_MODELS,
        mode_label='swarm',
    )


def _ensemble_merger(ensemble_id: str, child_run_ids: list):
    """Wait for all child runs to finish, then merge + deduplicate rows."""
    def log(msg):
        ts = datetime.datetime.utcnow().strftime('%H:%M:%S')
        with _lock:
            if ensemble_id in _state:
                _state[ensemble_id]['logs'].append(f'[{ts}] {msg}')
                _state[ensemble_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'

    max_wait = 900  # 15 minutes
    poll_interval = 5
    elapsed = 0

    while elapsed < max_wait:
        time.sleep(poll_interval)
        elapsed += poll_interval

        with _lock:
            if _state[ensemble_id].get('cancel_requested'):
                _state[ensemble_id]['status'] = 'cancelled'
                _state[ensemble_id]['current_stage'] = 'cancelled'
                return

        # Check if all children are done
        all_done = True
        ready_ids = []
        with _lock:
            for cid in child_run_ids:
                job = _state.get(cid)
                if job is None:
                    continue
                s = job.get('status', '')
                if s in ('ready', 'complete'):
                    ready_ids.append(cid)
                elif s in ('error', 'cancelled'):
                    log(f'Child {cid} ended with status={s}')
                else:
                    all_done = False

        total = len(child_run_ids)
        done_count = len([c for c in child_run_ids if _state.get(c, {}).get('status') in ('ready','complete','error','cancelled')])
        pct = int(10 + (done_count / max(total, 1)) * 70)
        with _lock:
            _state[ensemble_id]['progress'] = pct
            _state[ensemble_id]['current_stage'] = f'waiting ({done_count}/{total} models done)'

        if all_done or done_count == total:
            break

    mode_label = (_state.get(ensemble_id, {}).get('inputs') or {}).get('_mode') or 'ensemble'
    log(f'All child runs complete — merging {mode_label} results…')
    with _lock:
        _state[ensemble_id]['current_stage'] = 'merging'
        _state[ensemble_id]['progress'] = 85

    all_rows = []
    ready_jobs = []
    with _lock:
        existing_pages_union = []
        ep_seen = set()
        for cid in ready_ids:
            job = _state.get(cid)
            if not job:
                continue
            ready_jobs.append(dict(job))
            for row in (job.get('rows') or []):
                all_rows.append(dict(row))
            for ep in (job.get('existing_pages') or []):
                if ep not in ep_seen:
                    ep_seen.add(ep)
                    existing_pages_union.append(ep)

    brand_name = ((_state.get(ensemble_id, {}).get('inputs') or {}).get('brand_name') or '').strip()
    merged_rows, prep_stats = prepare_kwr_rows(
        all_rows,
        existing_pages=existing_pages_union,
        brand=brand_name,
    )
    participation = build_model_participation(ready_jobs, merged_rows)
    log(
        f"Merged {len(merged_rows)} unique rows from {len(ready_ids)} models "
        f"(semantic dupes removed={prep_stats.get('semantic_duplicates_removed', 0)})"
    )

    summary_label = 'Best text swarm' if mode_label == 'swarm' else 'Ensemble'
    quality_summary = prep_stats.get('quality_summary') or {'total': len(merged_rows), 'avg_score': 0, 'high': 0, 'medium': 0, 'low': 0}
    with _lock:
        _state[ensemble_id]['rows'] = merged_rows
        _state[ensemble_id]['row_count'] = len(merged_rows)
        _state[ensemble_id]['existing_pages'] = existing_pages_union[:200]
        _state[ensemble_id]['quality_summary'] = quality_summary
        _state[ensemble_id]['dedup_summary'] = {
            'raw_count': prep_stats.get('raw_count', len(all_rows)),
            'exact_duplicates_removed': prep_stats.get('exact_duplicates_removed', 0),
            'semantic_duplicates_removed': prep_stats.get('semantic_duplicates_removed', 0),
            'existing_coverage_removed': prep_stats.get('existing_coverage_removed', 0),
        }
        _state[ensemble_id]['model_participation'] = participation
        _state[ensemble_id]['status'] = 'ready'
        _state[ensemble_id]['current_stage'] = 'ready'
        _state[ensemble_id]['progress'] = 100
        _state[ensemble_id]['finished_stages'].append('merging')
        _state[ensemble_id]['honest_count_note'] = (
            f'{summary_label} of {len(ready_ids)} models. '
            f'{len(merged_rows)} unique rows after deduplication. '
            f'Quality avg {quality_summary.get("avg_score", 0)}.'
        )
        _state[ensemble_id]['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'

    if participation:
        top = ', '.join(
            f"{p.get('model')} kept {p.get('kept_rows')}/{p.get('total_rows')}"
            for p in participation[:4]
        )
        log(f'{summary_label} participation — {top}')
    log(
        f'{summary_label} complete. Quality avg {quality_summary.get("avg_score", 0)} '
        f'(high {quality_summary.get("high", 0)}, medium {quality_summary.get("medium", 0)}, low {quality_summary.get("low", 0)}).'
    )


def _fetch_page_text(url: str, timeout: int = 15) -> str:
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'KWR-Bot/1.0'})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode('utf-8', 'replace')
        # Strip HTML tags
        text = re.sub(r'<[^>]+>', ' ', raw)
        text = re.sub(r'\s+', ' ', text).strip()
        return text[:8000]
    except Exception as e:
        return f"(fetch failed: {e})"


def _parse_sitemap(sitemap_url: str) -> list:
    try:
        req = urllib.request.Request(sitemap_url, headers={'User-Agent': 'KWR-Bot/1.0'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            raw = resp.read().decode('utf-8', 'replace')
        # Handle sitemap index (links to sub-sitemaps)
        sub_maps = re.findall(r'<loc>\s*(https?://[^\s<]+sitemap[^\s<]*)\s*</loc>', raw, re.IGNORECASE)
        urls = re.findall(r'<loc>\s*(https?://[^\s<]+)\s*</loc>', raw, re.IGNORECASE)
        # If only sub-maps found, fetch first few
        if sub_maps and len(urls) <= len(sub_maps):
            all_urls = []
            for sm in sub_maps[:3]:
                try:
                    req2 = urllib.request.Request(sm, headers={'User-Agent': 'KWR-Bot/1.0'})
                    with urllib.request.urlopen(req2, timeout=10) as r2:
                        sub_raw = r2.read().decode('utf-8', 'replace')
                    all_urls += re.findall(r'<loc>\s*(https?://[^\s<]+)\s*</loc>', sub_raw, re.IGNORECASE)
                except Exception:
                    pass
            return all_urls[:300]
        return [u for u in urls if 'sitemap' not in u.lower()][:300]
    except Exception:
        return []


def _extract_json(text: str) -> str:
    """Extract JSON from LLM output that may contain prose, fenced blocks, or JSONL."""
    # Try fenced block first
    m = re.search(r'```(?:json|jsonl)?\s*(.+?)\s*```', text, re.DOTALL)
    if m:
        text = m.group(1)
    # If it parses cleanly as-is, return it
    stripped = text.strip()
    try:
        json.loads(stripped)
        return stripped
    except Exception:
        pass
    # Try to extract a balanced top-level JSON array
    arr_match = re.search(r'\[\s*\{.*\}\s*\]', stripped, re.DOTALL)
    if arr_match:
        candidate = arr_match.group(0)
        try:
            json.loads(candidate)
            return candidate
        except Exception:
            pass
    # JSONL fallback: collect each {...} object on its own line(s) and wrap as array
    objects = []
    decoder = json.JSONDecoder()
    idx = 0
    s = stripped
    while idx < len(s):
        # skip whitespace and commas/newlines between objects
        while idx < len(s) and s[idx] in ' \t\r\n,':
            idx += 1
        if idx >= len(s):
            break
        if s[idx] != '{':
            # skip until next '{'
            nxt = s.find('{', idx)
            if nxt == -1:
                break
            idx = nxt
        try:
            obj, end = decoder.raw_decode(s, idx)
            objects.append(obj)
            idx = end
        except Exception:
            break
    if objects:
        return json.dumps(objects, ensure_ascii=False)
    return stripped


def _slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text.strip('-')


def _job_summary(job: dict) -> dict:
    return {
        'run_id':       job.get('run_id'),
        'status':       job.get('status'),
        'current_stage': job.get('current_stage'),
        'progress':     job.get('progress', 0),
        'row_count':    job.get('row_count', 0),
        'created_at':   job.get('created_at'),
        'updated_at':   job.get('updated_at'),
        'website_url':  (job.get('inputs') or {}).get('website_url', ''),
        'brand_name':   (job.get('inputs') or {}).get('brand_name', ''),
        'sheet_url':    job.get('sheet_url'),
        'deploy_error': job.get('deploy_error'),
        'honest_count_note': job.get('honest_count_note', ''),
    }


# ===========================================================================
# TASK A — Auto-fill probe (real verified URLs)
# ===========================================================================

def _http_request(url, method='GET', timeout=6, max_bytes=None):
    """Tiny urllib helper. Returns (status, body_bytes, final_url) or raises."""
    req = urllib.request.Request(
        url, method=method,
        headers={
            'User-Agent': 'Mozilla/5.0 (KWR-Probe/1.0)',
            'Accept': '*/*',
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = b''
        if method != 'HEAD':
            if max_bytes:
                body = resp.read(max_bytes)
            else:
                body = resp.read()
        return resp.status, body, resp.geturl()


def _head_ok(url, timeout=5):
    """True if HEAD or GET returns 2xx on the URL."""
    try:
        st, _, _ = _http_request(url, method='HEAD', timeout=timeout)
        if 200 <= st < 300:
            return True
    except Exception:
        pass
    # Fallback to small GET (some servers reject HEAD)
    try:
        st, _, _ = _http_request(url, method='GET', timeout=timeout, max_bytes=512)
        return 200 <= st < 300
    except Exception:
        return False


def probe_domain(domain: str) -> dict:
    """
    Real-verified auto-fill data for a domain.
    Returns {website_url, sitemap_url, about_url, language, market, brand_name}.
    Never throws; falls back to empty strings on failure.
    """
    raw = (domain or '').strip()
    if not raw:
        return {'website_url': '', 'sitemap_url': '', 'about_url': '',
                'language': '', 'market': '', 'brand_name': ''}
    if not raw.startswith(('http://', 'https://')):
        raw = 'https://' + raw

    out = {'website_url': '', 'sitemap_url': '', 'about_url': '',
           'language': '', 'market': '', 'brand_name': ''}

    try:
        parsed = urllib.parse.urlparse(raw)
        host = parsed.netloc or parsed.path.split('/')[0]
        origin = f"{parsed.scheme}://{host}"
    except Exception:
        return out

    # Verify website reachable; prefer https, fall back to http.
    website_url = ''
    for candidate in (origin, origin.replace('https://', 'http://')):
        try:
            st, _, final = _http_request(candidate, method='GET', timeout=6, max_bytes=4096)
            if 200 <= st < 400:
                # Use final URL (handles redirects to www.)
                fu = urllib.parse.urlparse(final)
                website_url = f"{fu.scheme}://{fu.netloc}"
                origin = website_url
                host = fu.netloc or host
                break
        except Exception:
            continue
    if not website_url:
        website_url = origin
    out['website_url'] = website_url

    # Brand from hostname's first label
    brand_label = host.replace('www.', '').split('.')[0] or host
    out['brand_name'] = brand_label[:1].upper() + brand_label[1:]

    # 1. Sitemap from robots.txt (canonical)
    sitemap_candidates = []
    try:
        st, body, _ = _http_request(origin + '/robots.txt', method='GET', timeout=5, max_bytes=200_000)
        if 200 <= st < 300 and body:
            text = body.decode('utf-8', 'replace')
            for line in text.splitlines():
                line = line.strip()
                if line.lower().startswith('sitemap:'):
                    sm = line.split(':', 1)[1].strip()
                    if sm:
                        sitemap_candidates.append(sm)
    except Exception:
        pass
    sitemap_candidates += [
        origin + '/sitemap.xml',
        origin + '/sitemap_index.xml',
        origin + '/wp-sitemap.xml',
        origin + '/sitemap-index.xml',
    ]
    for sm in sitemap_candidates:
        if _head_ok(sm, timeout=5):
            out['sitemap_url'] = sm
            break

    # 2. About page — HEAD-check candidates
    about_candidates = ['/about', '/about-us', '/about/', '/about-us/',
                        '/%D7%90%D7%95%D7%93%D7%95%D7%AA',  # אודות
                        '/he/about', '/en/about', '/about.html']
    for path in about_candidates:
        url = origin + path
        if _head_ok(url, timeout=4):
            out['about_url'] = url
            break

    # 3. Language from <html lang="..."> + brand from og:site_name
    homepage_html = ''
    try:
        st, body, _ = _http_request(origin, method='GET', timeout=8, max_bytes=200_000)
        if 200 <= st < 400 and body:
            try:
                homepage_html = body.decode('utf-8', 'replace')
            except Exception:
                homepage_html = body.decode('latin-1', 'replace')
    except Exception:
        pass

    language = ''
    if homepage_html:
        m = re.search(r'<html[^>]*\blang\s*=\s*["\']?([A-Za-z-]{2,10})', homepage_html, re.IGNORECASE)
        if m:
            code = m.group(1).lower()
            if code.startswith('he') or code == 'iw':
                language = 'Hebrew'
            elif code.startswith('en'):
                language = 'English'
            elif code.startswith('ar'):
                language = 'Arabic'
            elif code.startswith('ru'):
                language = 'Russian'
            elif code.startswith('es'):
                language = 'Spanish'
            elif code.startswith('fr'):
                language = 'French'
            else:
                language = code.upper()
        if not language:
            head = homepage_html[:4000]
            if re.search(r'[\u0590-\u05FF]', head):
                language = 'Hebrew'
            elif re.search(r'[\u0600-\u06FF]', head):
                language = 'Arabic'
            else:
                language = 'English'

        # og:site_name → brand override
        m = re.search(
            r'<meta[^>]+property\s*=\s*["\']og:site_name["\'][^>]*content\s*=\s*["\']([^"\']+)["\']',
            homepage_html, re.IGNORECASE)
        if not m:
            m = re.search(
                r'<meta[^>]+content\s*=\s*["\']([^"\']+)["\'][^>]*property\s*=\s*["\']og:site_name["\']',
                homepage_html, re.IGNORECASE)
        if m:
            site_name = m.group(1).strip()
            if site_name and len(site_name) <= 80:
                out['brand_name'] = site_name

    if not language:
        # TLD heuristic fallback
        tld = host.lower()
        if tld.endswith('.il') or '.co.il' in tld:
            language = 'Hebrew'
        else:
            language = 'English'
    out['language'] = language

    # Market guess
    if language == 'Hebrew':
        out['market'] = 'Israel'
    elif host.endswith('.uk') or host.endswith('.co.uk'):
        out['market'] = 'United Kingdom'
    elif host.endswith('.au') or host.endswith('.com.au'):
        out['market'] = 'Australia'
    elif host.endswith('.ca'):
        out['market'] = 'Canada'
    elif host.endswith('.de'):
        out['market'] = 'Germany'
    elif language == 'English':
        out['market'] = 'United States'
    else:
        out['market'] = ''

    # 4. Discover competitors via DuckDuckGo + LLM (best-effort)
    try:
        out['competitors'] = _discover_competitors_for_probe(
            host=host,
            brand=out.get('brand_name') or '',
            language=out.get('language') or '',
            market=out.get('market') or '',
            homepage_html=homepage_html,
        )
    except Exception:
        out['competitors'] = []

    return out


def _discover_competitors_for_probe(host: str, brand: str, language: str,
                                    market: str, homepage_html: str = '',
                                    max_results: int = 5) -> list:
    """
    Find up to N competitor URLs using:
      1. Page topics (h1/title) + market → DuckDuckGo HTML search
      2. LLM call (best frontier model via call_with_fallback) to suggest
         likely competitor domains given brand + language + market.
    Returns list of {url, source}.
    """
    seed = (host or '').lower().replace('www.', '')
    found = []           # list[(url, source)]
    seen = set()
    seen.add(seed)

    # Build search topics from homepage
    topics = []
    if homepage_html:
        m = re.search(r'<title[^>]*>([^<]{4,120})</title>', homepage_html, re.IGNORECASE)
        if m:
            topics.append(re.sub(r'\s+', ' ', m.group(1)).strip())
        for h in re.findall(r'<h1[^>]*>([^<]{4,120})</h1>', homepage_html, re.IGNORECASE)[:2]:
            topics.append(re.sub(r'\s+', ' ', h).strip())
    if brand:
        topics.append(brand)
    topics = [t for t in topics if t][:3]

    # 1. DuckDuckGo HTML search
    market_q = market or ''
    queries = []
    for t in topics:
        q = f"{t} {market_q}".strip()
        queries.append(q)
    if brand and market_q:
        queries.append(f"competitors of {brand} {market_q}")
    for q in queries:
        if len(found) >= max_results:
            break
        try:
            search_url = f"https://duckduckgo.com/html/?q={urllib.parse.quote(q)}"
            req = urllib.request.Request(search_url, headers={'User-Agent': 'Mozilla/5.0'})
            html = urllib.request.urlopen(req, timeout=10).read().decode('utf-8', errors='ignore')
            for m in re.finditer(r'href="(https?://[^"?]+)"', html):
                url = m.group(1)
                dom = urllib.parse.urlparse(url).netloc.lower().replace('www.', '')
                if not dom or dom in seen:
                    continue
                if any(x in dom for x in ('duckduckgo', 'google.', 'bing.', 'wikipedia',
                                          'youtube.', 'facebook.', 'linkedin.',
                                          'twitter.', 'instagram.', 'pinterest.',
                                          'amazon.', 'tripadvisor.', 'yelp.', 'reddit.')):
                    continue
                seen.add(dom)
                found.append({'url': f'https://{dom}', 'source': 'duckduckgo'})
                if len(found) >= max_results:
                    break
        except Exception:
            continue

    # 2. LLM-based suggestion (frontier model via call_with_fallback)
    if len(found) < max_results:
        try:
            from llm_router import call_with_fallback  # type: ignore
        except Exception:
            call_with_fallback = None  # type: ignore
        if call_with_fallback:
            try:
                prompt = (
                    f"List up to 8 real competitor websites (just bare domains, "
                    f"one per line, no commentary, no numbering) for a business "
                    f"with these details:\n"
                    f"- Brand: {brand or '(unknown)'}\n"
                    f"- Domain: {host}\n"
                    f"- Language: {language or '(unknown)'}\n"
                    f"- Market: {market or '(unknown)'}\n"
                    f"- Homepage signals: {(topics[0] if topics else '')[:200]}\n\n"
                    f"Return only domain names like example.com, no URLs, no http://."
                )
                resp = call_with_fallback(
                    messages=[{'role': 'user', 'content': prompt}],
                    max_tokens=300,
                    temperature=0.4,
                )
                txt = ''
                if isinstance(resp, dict):
                    txt = (resp.get('content') or resp.get('text') or '') or ''
                elif isinstance(resp, str):
                    txt = resp
                for line in (txt or '').splitlines():
                    line = re.sub(r'[`*\-\u2022\d\.\s]+', ' ', line).strip().lower()
                    line = line.split(' ')[0]
                    if not line or '.' not in line:
                        continue
                    line = line.replace('http://', '').replace('https://', '').replace('www.', '').strip('/.,')
                    if line in seen:
                        continue
                    seen.add(line)
                    found.append({'url': f'https://{line}', 'source': 'llm'})
                    if len(found) >= max_results:
                        break
            except Exception:
                pass

    return found[:max_results]


# ===========================================================================
# TASK B — Multi-destination sync (GitHub + Obsidian + Supabase storage)
# ===========================================================================

def _domain_slug_for_run(run_id: str) -> str:
    """Build domain slug from job state (or disk)."""
    job = None
    with _lock:
        job = _state.get(run_id)
    if job is None:
        job = _load_job_from_disk(run_id) or {}
    inputs = job.get('inputs') or {}
    website = (inputs.get('website_url') or '').strip()
    domain = website.replace('https://', '').replace('http://', '').split('/')[0]
    return _slugify(domain) or 'site'


def _read_run_xlsx(run_id: str) -> bytes | None:
    rd = _run_dir(run_id)
    p = os.path.join(rd, 'file.xlsx')
    if os.path.exists(p):
        try:
            with open(p, 'rb') as f:
                return f.read()
        except Exception:
            return None
    # Build it on demand
    data, _, err = build_excel(run_id)
    if err:
        return None
    return data


def _sync_github(run_id: str, xlsx_bytes: bytes, domain_slug: str) -> dict:
    """Commit outputs/kwr_{slug}.xlsx to the repo.

    Primary: GitHub Contents REST API (works from Render). Requires GITHUB_PAT.
    Fallback: local git CLI (works from the dev box).
    """
    import base64, json as _json
    import urllib.request, urllib.error
    pat = (os.getenv('GITHUB_PAT') or os.getenv('GH_TOKEN') or os.getenv('GITHUB_TOKEN') or '').strip()
    owner = os.getenv('GITHUB_OWNER', 'maximoseo')
    repo  = os.getenv('GITHUB_REPO',  'webs-html-improvements-files')
    branch = os.getenv('GITHUB_BRANCH', 'main')
    target_rel = f'outputs/kwr_{domain_slug}.xlsx'

    if pat:
        try:
            api = f'https://api.github.com/repos/{owner}/{repo}/contents/{target_rel}'
            headers = {
                'Authorization': f'token {pat}',
                'Accept': 'application/vnd.github+json',
                'User-Agent': 'kwr-bot',
            }
            # Check existing file to get sha (required for update)
            sha = None
            try:
                req = urllib.request.Request(f'{api}?ref={branch}', headers=headers)
                with urllib.request.urlopen(req, timeout=15) as resp:
                    sha = _json.loads(resp.read().decode('utf-8')).get('sha')
            except urllib.error.HTTPError as e:
                if e.code != 404:
                    raise
            body = {
                'message': f'kwr: report for {domain_slug} ({run_id[:8]})',
                'content': base64.b64encode(xlsx_bytes).decode('ascii'),
                'branch': branch,
                'committer': {'name': 'KWR Bot', 'email': 'kwr-bot@maximo-seo.ai'},
            }
            if sha:
                body['sha'] = sha
            data = _json.dumps(body).encode('utf-8')
            req = urllib.request.Request(api, data=data, method='PUT', headers={**headers, 'Content-Type': 'application/json'})
            with urllib.request.urlopen(req, timeout=30) as resp:
                out = _json.loads(resp.read().decode('utf-8'))
                commit_sha = ((out.get('commit') or {}).get('sha') or '')[:7]
                return {'ok': True, 'path': target_rel, 'pushed': True, 'commit': commit_sha, 'via': 'api'}
        except Exception as exc:
            # fall through to git CLI
            api_err = str(exc)[:200]
    else:
        api_err = 'no GITHUB_PAT'

    # Fallback: local git CLI
    import subprocess
    repo_dir = os.path.dirname(os.path.abspath(__file__))
    if not os.path.isdir(os.path.join(repo_dir, '.git')):
        return {'ok': False, 'skipped': True, 'reason': f'API failed ({api_err}); not a git repo'}
    target_abs = os.path.join(repo_dir, target_rel)
    try:
        os.makedirs(os.path.dirname(target_abs), exist_ok=True)
        with open(target_abs, 'wb') as f:
            f.write(xlsx_bytes)
        env = os.environ.copy()
        env.setdefault('GIT_AUTHOR_NAME', 'KWR Bot')
        env.setdefault('GIT_AUTHOR_EMAIL', 'kwr-bot@maximo-seo.ai')
        env.setdefault('GIT_COMMITTER_NAME', 'KWR Bot')
        env.setdefault('GIT_COMMITTER_EMAIL', 'kwr-bot@maximo-seo.ai')
        subprocess.run(['git', 'add', target_rel], cwd=repo_dir, env=env, check=False, capture_output=True, timeout=30)
        r = subprocess.run(['git', 'commit', '-m', f'kwr: report for {domain_slug}'],
                           cwd=repo_dir, env=env, check=False, capture_output=True, timeout=30)
        if r.returncode != 0 and b'nothing to commit' not in (r.stdout + r.stderr):
            return {'ok': False, 'error': (r.stderr or r.stdout).decode('utf-8', 'replace')[:200]}
        if pat:
            push_url = f'https://maximoseo:{pat}@github.com/{owner}/{repo}.git'
            r2 = subprocess.run(['git', 'push', push_url, f'HEAD:{branch}'],
                                cwd=repo_dir, env=env, check=False,
                                capture_output=True, timeout=60)
            if r2.returncode != 0:
                return {'ok': False, 'error': (r2.stderr or r2.stdout).decode('utf-8', 'replace')[:200]}
            return {'ok': True, 'path': target_rel, 'pushed': True, 'via': 'cli'}
        return {'ok': True, 'path': target_rel, 'pushed': False, 'note': 'no PAT', 'via': 'cli'}
    except Exception as exc:
        return {'ok': False, 'error': str(exc)[:200]}


def _sync_obsidian(run_id: str, xlsx_bytes: bytes, domain_slug: str) -> dict:
    """Copy xlsx to <vault>/SEO/KWR/{slug}__{date}.xlsx + maintain index.md."""
    vault = (os.getenv('OBSIDIAN_VAULT_PATH') or '').strip()
    if not vault:
        # Default to known Windows vault if mounted (works only on the dev box)
        guess = '/mnt/c/Obsidian/HTML REDESIGN/HTML REDESIGN'
        if os.path.isdir(guess):
            vault = guess
    if not vault or not os.path.isdir(vault):
        return {'ok': False, 'skipped': True, 'reason': 'OBSIDIAN_VAULT_PATH not set / not mounted'}
    try:
        seo_dir = os.path.join(vault, 'SEO', 'KWR')
        os.makedirs(seo_dir, exist_ok=True)
        date_str = datetime.datetime.utcnow().strftime('%Y-%m-%d')
        fname = f'{domain_slug}__{date_str}.xlsx'
        fpath = os.path.join(seo_dir, fname)
        with open(fpath, 'wb') as f:
            f.write(xlsx_bytes)
        # Maintain index.md
        index_path = os.path.join(seo_dir, 'index.md')
        try:
            files = sorted(
                [f for f in os.listdir(seo_dir) if f.endswith('.xlsx')],
                reverse=True,
            )
            lines = ['# KWR Reports', '', f'_Last updated: {datetime.datetime.utcnow().isoformat()}Z_', '']
            for ff in files:
                lines.append(f'- [[{ff}]]')
            with open(index_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
        except Exception:
            pass
        return {'ok': True, 'path': fpath}
    except Exception as exc:
        return {'ok': False, 'error': str(exc)[:200]}


def _sync_supabase_storage(run_id: str, xlsx_bytes: bytes, domain_slug: str) -> dict:
    """Upload to Supabase storage bucket 'kwr-reports' via REST."""
    cfg_url = (os.getenv('SUPABASE_URL') or '').strip().rstrip('/')
    key = (os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('SUPABASE_ANON_KEY') or '').strip()
    if not (cfg_url and key):
        return {'ok': False, 'skipped': True, 'reason': 'SUPABASE creds not set'}
    bucket = (os.getenv('SUPABASE_KWR_BUCKET') or 'kwr-reports').strip()
    object_path = f'{domain_slug}/{run_id}.xlsx'
    upload_url = f'{cfg_url}/storage/v1/object/{bucket}/{object_path}'
    headers = {
        'Authorization': f'Bearer {key}',
        'apikey': key,
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'x-upsert': 'true',
        'Cache-Control': 'max-age=3600',
    }
    try:
        req = urllib.request.Request(upload_url, data=xlsx_bytes, headers=headers, method='POST')
        with urllib.request.urlopen(req, timeout=30) as resp:
            _ = resp.read()
        public_url = f'{cfg_url}/storage/v1/object/public/{bucket}/{object_path}'
        return {'ok': True, 'path': object_path, 'public_url': public_url}
    except urllib.error.HTTPError as exc:
        body = exc.read().decode('utf-8', 'replace')[:400]
        # If bucket missing, try creating it then retry once.
        if exc.code in (400, 404) and 'Bucket not found' in body:
            try:
                create_req = urllib.request.Request(
                    f'{cfg_url}/storage/v1/bucket',
                    data=json.dumps({'id': bucket, 'name': bucket, 'public': True}).encode('utf-8'),
                    headers={'Authorization': f'Bearer {key}', 'apikey': key,
                             'Content-Type': 'application/json'},
                    method='POST',
                )
                urllib.request.urlopen(create_req, timeout=15).read()
                req2 = urllib.request.Request(upload_url, data=xlsx_bytes, headers=headers, method='POST')
                urllib.request.urlopen(req2, timeout=30).read()
                return {'ok': True, 'path': object_path, 'created_bucket': True}
            except Exception as exc2:
                return {'ok': False, 'error': f'bucket-create then upload failed: {exc2}'}
        return {'ok': False, 'error': f'HTTP {exc.code}: {body}'}
    except Exception as exc:
        return {'ok': False, 'error': str(exc)[:200]}


def sync_report(run_id: str) -> dict:
    """Sync the run's xlsx to GitHub + Obsidian + Supabase storage. Best-effort."""
    xlsx = _read_run_xlsx(run_id)
    if not xlsx:
        return {'github': {'ok': False, 'error': 'no xlsx'},
                'obsidian': {'ok': False, 'error': 'no xlsx'},
                'supabase': {'ok': False, 'error': 'no xlsx'}}
    slug = _domain_slug_for_run(run_id)
    result = {
        'github':   _sync_github(run_id, xlsx, slug),
        'obsidian': _sync_obsidian(run_id, xlsx, slug),
        'supabase': _sync_supabase_storage(run_id, xlsx, slug),
    }
    # Persist sync result alongside meta.json
    try:
        rd = _run_dir(run_id)
        os.makedirs(rd, exist_ok=True)
        _atomic_write_json(
            os.path.join(rd, 'sync.json'),
            {'updated_at': datetime.datetime.utcnow().isoformat() + 'Z', **result},
        )
    except Exception:
        pass
    return result


# ===========================================================================
# TASK C — Reports listing/delete
# ===========================================================================

def list_reports() -> list:
    """Scan outputs/ for persisted reports. Returns newest-first list of dicts."""
    out = []
    try:
        if not os.path.isdir(OUTPUTS_DIR):
            return out
        for entry in os.listdir(OUTPUTS_DIR):
            rd = os.path.join(OUTPUTS_DIR, entry)
            if not os.path.isdir(rd):
                continue
            xlsx = os.path.join(rd, 'file.xlsx')
            if not os.path.exists(xlsx):
                continue
            meta = {}
            mp = os.path.join(rd, 'meta.json')
            if os.path.exists(mp):
                try:
                    with open(mp, 'r', encoding='utf-8') as f:
                        meta = json.load(f)
                except Exception:
                    meta = {}
            sync = {}
            sp = os.path.join(rd, 'sync.json')
            if os.path.exists(sp):
                try:
                    with open(sp, 'r', encoding='utf-8') as f:
                        sync = json.load(f)
                except Exception:
                    sync = {}
            try:
                size_kb = round(os.path.getsize(xlsx) / 1024.0, 1)
            except Exception:
                size_kb = 0
            try:
                mtime = datetime.datetime.utcfromtimestamp(os.path.getmtime(xlsx)).isoformat() + 'Z'
            except Exception:
                mtime = ''
            out.append({
                'run_id': meta.get('run_id') or entry,
                'domain': meta.get('domain') or '',
                'website_url': meta.get('website_url') or '',
                'brand': meta.get('brand') or '',
                'worksheet_name': meta.get('worksheet_name') or 'kwr',
                'created_at': meta.get('created_at') or mtime,
                'updated_at': meta.get('updated_at') or mtime,
                'row_count': meta.get('row_count') or 0,
                'file_size_kb': size_kb,
                'sync_status': {
                    'github':   bool((sync.get('github')   or {}).get('ok')),
                    'obsidian': bool((sync.get('obsidian') or {}).get('ok')),
                    'supabase': bool((sync.get('supabase') or {}).get('ok')),
                },
            })
        # Also include flat xlsx files synced from GitHub (e.g. outputs/kwr_<slug>.xlsx)
        seen_slugs = set()
        for r in out:
            dom = (r.get('domain') or '').replace('.', '_').replace('-', '_').lower()
            if dom:
                seen_slugs.add(dom)
        for fn in os.listdir(OUTPUTS_DIR):
            if not fn.startswith('kwr_') or not fn.lower().endswith('.xlsx'):
                continue
            fp = os.path.join(OUTPUTS_DIR, fn)
            if not os.path.isfile(fp):
                continue
            slug = fn[4:-5]  # strip 'kwr_' and '.xlsx'
            if slug in seen_slugs:
                continue
            try:
                size_kb = round(os.path.getsize(fp) / 1024.0, 1)
            except Exception:
                size_kb = 0
            try:
                mtime = datetime.datetime.utcfromtimestamp(os.path.getmtime(fp)).isoformat() + 'Z'
            except Exception:
                mtime = ''
            domain_display = slug.replace('_', '.')
            try:
                from openpyxl import load_workbook
                wb = load_workbook(fp, read_only=True)
                ws = wb.active
                row_count = max(0, ws.max_row - 1)
                wb.close()
            except Exception:
                row_count = 0
            out.append({
                'run_id': f'flat:{slug}',
                'domain': domain_display,
                'website_url': f'https://{domain_display}',
                'brand': '',
                'worksheet_name': f'kwr_{slug}',
                'created_at': mtime,
                'updated_at': mtime,
                'row_count': row_count,
                'file_size_kb': size_kb,
                'flat_file': fn,
                'sync_status': {
                    'github':   True,   # served from repo
                    'obsidian': False,
                    'supabase': False,
                },
            })
        out.sort(key=lambda x: x.get('created_at') or '', reverse=True)
    except Exception:
        pass
    return out


def delete_report(run_id: str) -> tuple:
    """Remove outputs/{run_id}/ directory. Returns (ok, error)."""
    import shutil
    rd = _run_dir(run_id)
    if not os.path.isdir(rd):
        return False, 'not found'
    try:
        shutil.rmtree(rd)
        with _lock:
            _state.pop(run_id, None)
        return True, None
    except Exception as exc:
        return False, str(exc)
